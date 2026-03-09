
import streamlit as st
from supabase import create_client, Client
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import datetime
import time
from PIL import Image
import io
import json
import hashlib
import bcrypt
import shutil
import os
import random
import streamlit.components.v1 as components

# Move to the absolute top to ensure it's the first streamlit command
st.set_page_config(page_title="מערכת בקרה רבנות פיקוד מרכז", page_icon="✡️")

# טעינת הכפתור המעוצב מתוך התיקייה שיצרנו
custom_location_button = components.declare_component("whatsapp_loc_button", path="wa_loc_component")


import math
from typing import Tuple, Optional, List, Dict, Any
import folium
from streamlit_folium import st_folium
try:
    from streamlit_drawable_canvas import st_canvas
except ImportError:
    st_canvas = None



# WhatsApp phones for Hatmar Rabbis


# ===== פונקציות עזר למיקום וחישוב מרחקים =====

# קואורדינטות בסיסים ידועים
BASE_COORDINATES = {
    "מחנה עופר": (32.1089, 35.1911),
    "בית אל": (31.9333, 35.2167),
    "פסגות": (31.9667, 35.2000),
    "מחנה שומרון": (32.2167, 35.2833),
    "אריאל": (32.1039, 35.1794),
    "קדומים": (32.1667, 35.2000),
    "גוש עציון": (31.6500, 35.1333),
    "אפרת": (31.6500, 35.1333),
    "בית לחם": (31.7050, 35.2061),
    "מחנה עציון": (31.6500, 35.1333),
    "אלון שבות": (31.6500, 35.1500),
    "מוצב אפרים": (32.0500, 35.3000),
    "מוצב מנשה": (32.3000, 35.1800),
    "מוצב הבקעה": (31.8500, 35.4500),
}

# קודי גישה לרבני חטמ"ר - מועבר ל-DB (פונקציות עזר להלן)

BASE_BARCODES = {
    "מחנה עופר": "RB_OFER_99", "בית אל": "RB_BETEL_88", "פסגות": "RB_PSAGOT_77",
    "מחנה שומרון": "RB_SHOMRON_66", "אריאל": "RB_ARIEL_55", "קדומים": "RB_KEDUMIM_44",
    "גוש עציון": "RB_ETZION_33", "אפרת": "RB_EFRAT_22", "בית לחם": "RB_BLEHEM_11",
    "מחנה עציון": "RB_ETZION_BASE", "אלון שבות": "RB_ALON_SHEVUT", "מוצב אפרים": "RB_EFRAIM_POS",
    "מוצב מנשה": "RB_MENASHE_POS", "מוצב הבקעה": "RB_BIKA_POS",
}

def haversine_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """חישוב מרחק בין שתי נקודות על פני כדור הארץ (ק\"מ)"""
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
    c = 2 * math.asin(math.sqrt(a))
    return c * 6371

def find_nearest_base(lat: float, lon: float) -> Tuple[str, float]:
    """מציאת הבסיס הקרוב ביותר"""
    min_distance = float('inf')
    nearest_base = "לא ידוע"
    for base_name, (base_lat, base_lon) in BASE_COORDINATES.items():
        distance = haversine_distance(lat, lon, base_lat, base_lon)
        if distance < min_distance:
            min_distance = distance
            nearest_base = base_name
    return nearest_base, min_distance

# ════════════════════════════════════════════════════════════
# פיצ'ר 4 — WhatsApp / SMS התראות אוטומטיות
# ════════════════════════════════════════════════════════════
def send_email_alerts(report_data: dict, unit: str):
    """
    שולח דוא"ל לרב החטמ"ר, רב האוגדה ורב הפיקוד כשיש בעיה קריטית.
    """
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    # בדיקת בעיות קריטיות
    alerts = []
    if report_data.get("e_status") == "פסול":
        alerts.append(f"🚧 עירוב פסול במוצב {report_data.get('base','?')}")
    if report_data.get("k_cert") == "לא":
        alerts.append(f"🍽️ כשרות חסרה במוצב {report_data.get('base','?')}")
    mezuzot = int(report_data.get("r_mezuzot_missing", 0) or 0)
    if mezuzot >= 5:
        alerts.append(f"📜 {mezuzot} מזוזות חסרות במוצב {report_data.get('base','?')}")
    if report_data.get("p_mix") == "כן":
        alerts.append(f"🔴 ערבוב כלים במוצב {report_data.get('base','?')}")

    if not alerts:
        return

    # איסוף כתובות אימייל לפי ההיררכיה
    recipients = []
    
    def add_unit_email(u):
        try:
            res = supabase.table("unit_emails").select("email").eq("unit", u).execute()
            if res.data:
                recipients.append(res.data[0]['email'])
        except:
            pass

    # 1. אימייל של היחידה עצמה
    add_unit_email(unit)
    
    # 2. אימייל של היחידות מעל (אוגדה, פיקוד)
    try:
        current_u = unit
        for _ in range(2): # עד 2 רמות מעל
            res = supabase.table("hierarchy").select("parent_unit").eq("child_unit", current_u).execute()
            if res.data:
                parent = res.data[0]['parent_unit']
                add_unit_email(parent)
                current_u = parent
            else:
                break
    except:
        pass

    recipients = list(set([r for r in recipients if r])) # הסרת כפילויות וריקים
    if not recipients:
        return

    # הגדרות SMTP מה-secrets
    try:
        smtp_server = st.secrets["email"]["smtp_server"]
        smtp_port = st.secrets["email"]["smtp_port"]
        smtp_user = st.secrets["email"]["smtp_user"]
        smtp_pass = st.secrets["email"]["smtp_pass"]
        from_email = st.secrets["email"].get("from_email", smtp_user)
    except Exception:
        print("⚠️ חסרים secrets של Email (SMTP)")
        return

    subject = f"🛡️ התראה קריטית - מערכת רבנות פקמ\"ז - {unit}"
    body = f"""
    שלום רב,
    
    זוהי התראה אוטומטית ממערכת רבנות פקמ"ז לגבי דיווח חדש:
    
    📋 יחידה: {unit}
    👤 מבקר: {report_data.get('inspector','?')}
    📍 מוצב: {report_data.get('base','?')}
    📅 תאריך: {report_data.get('date','?')[:10]}
    
    ⚠️ בעיות שדווחו:
    """ + "\n".join(f"  • {a}" for a in alerts) + f"""
    
    🔗 לפרטים מלאים וצפייה בתמונות, היכנס למערכת.
    
    בברכה,
    חמ"ל רבנות פקמ"ז
    """

    try:
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = ", ".join(recipients)
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.send_message(msg)
        server.quit()
        
        st.toast(f"📧 התראה נשלחה בדוא\"ל ל-{len(recipients)} נמענים", icon="✅")
        # שליחת WhatsApp במקביל
        try:
            send_whatsapp_alert("🚨 התראה - " + unit + ": " + " | ".join(alerts), unit)
        except Exception:
            pass
        log_audit_event("EMAIL_ALERT", unit,
                        details={"alerts": alerts, "recipients": recipients, "base": report_data.get("base")},
                        severity="warning")
    except Exception as e:
        print(f"⚠️ שגיאה בשליחת Email: {e}")


def send_whatsapp_alert(message: str, unit: str):
    """שולח WhatsApp לרבני חטמ״ר דרך CallMeBot - ללא Twilio, ללא עלות"""
    try:
        result = supabase.table("whatsapp_numbers")\
            .select("phone,callmebot_key")\
            .eq("unit", unit)\
            .eq("active", True)\
            .execute()
        if not result.data:
            return
        import urllib.parse, urllib.request as _ureq
        for row in result.data:
            phone = row['phone']
            key = row['callmebot_key']
            encoded = urllib.parse.quote(message)
            url = f"https://api.callmebot.com/whatsapp.php?phone={phone}&text={encoded}&apikey={key}"
            _ureq.urlopen(url, timeout=5)
    except Exception as e:
        print(f"WhatsApp error: {e}")


# ════════════════════════════════════════════════════════════
# פיצ'ר 5 — History Card לכל מוצב
# ════════════════════════════════════════════════════════════
def render_base_history_card(base: str, unit: str):
    """
    🕵️ מבלש כרטיס היסטוריה למוצב שנבחר ושאלות מעקב חכמות (בלש חוקר).
    """
    if not base or len(base) < 2:
        return

    try:
        result = (
            supabase.table("reports")
            .select("*")
            .eq("unit", unit)
            .ilike("base", f"%{base}%")
            .order("date", desc=True)
            .limit(2)
            .execute()
        )
        reports = result.data
    except Exception:
        return

    if not reports:
        st.info(f"📍 **{base}** — אין ביקורים קודמים רשומים")
        return

    last = reports[0]
    try:
        last_date = pd.to_datetime(last["date"]).tz_localize(None) if pd.to_datetime(last["date"]).tzinfo else pd.to_datetime(last["date"])
        days_ago  = (pd.Timestamp.now() - last_date).days
    except Exception:
        days_ago = "?"

    findings = []
    if last.get("e_status") == "פסול":
        findings.append(("🚧", "עירוב פסול", "#ef4444", "eruv"))
    if last.get("k_cert") == "לא":
        findings.append(("🍽️", "כשרות חסרה", "#f59e0b", "kashrut"))
    mezuzot = int(last.get("r_mezuzot_missing") or 0)
    if mezuzot > 0:
        findings.append(("📜", f"{mezuzot} מזוזות חסרות", "#3b82f6", "mezuzot"))
    if last.get("p_mix") == "כן":
        findings.append(("🔴", "ערבוב כלים", "#dc2626", "mix"))
    if last.get("k_issues") == "כן":
        desc = last.get("k_issues_description", "")
        findings.append(("⚠️", f"תקלת כשרות: {desc[:40]}..." if len(desc) > 40 else f"תקלת כשרות: {desc}", "#f97316", "kashrut_issue"))

    urgency_color = "#ef4444" if days_ago != "?" and days_ago > 14 else \
                    "#f59e0b" if days_ago != "?" and days_ago > 7  else "#10b981"

    findings_html = ""
    if findings:
        findings_html = "<div style='margin-top:10px;'><strong>ממצאים מהביקור הקודם:</strong><br/>"
        for icon, text, color, _ in findings:
            findings_html += f"<span style='background:{color}22;color:{color};padding:3px 8px;border-radius:4px;margin:3px 2px;display:inline-block;font-size:13px;'>{icon} {text}</span>"
        findings_html += "</div>"
    else:
        findings_html = "<div style='margin-top:8px;color:#10b981;font-size:13px;'>✅ לא נמצאו בעיות בביקור הקודם</div>"

    inspector_prev = last.get("inspector", "לא ידוע")

    st.markdown(f"""
    <div style='background: linear-gradient(135deg, #f8fafc 0%, #e2e8f0 100%); border-right: 5px solid {urgency_color}; border-radius: 12px; padding: 16px 20px; margin: 10px 0 18px 0; box-shadow: 0 2px 8px rgba(0,0,0,0.07);'>
        <div style='display:flex; justify-content:space-between; align-items:center;'>
            <div>
                <span style='font-size:18px; font-weight:700; color:#1e3a8a;'>📍 {base}</span>
                <span style='margin-right:10px; font-size:13px; color:#64748b;'>ביקור אחרון ע"י {inspector_prev}</span>
            </div>
            <div style='background:{urgency_color}; color:white; padding:6px 14px; border-radius:20px; font-weight:700; font-size:14px;'>⏱️ לפני {days_ago} ימים</div>
        </div>
        {findings_html}
    </div>
    """, unsafe_allow_html=True)

    # ===== 🕵️ Smart Continuity — שאלות מעקב בלש חוקר =====
    if findings:
        st.markdown("##### 🕵️ סגירת מעגלים מהביקור הקודם — סגירת מעגלים")
        for icon, text, color, key_suffix in findings:
            is_fixed = st.checkbox(
                f"{icon} {text} — האם תוקן?",
                key=f"fixed_{key_suffix}_{base}",
                value=False
            )
            if is_fixed:
                st.toast(f"✅ סגירת מעגל — {text} סומן כמותקן! העם דע בחוסרים אם צריך", icon="✅")

    # ===== 💡 Crowdsourced Tip — טיפ מהמבקר הקודם =====
    last_tip = last.get("inspector_tip", "") if last else ""
    if last_tip and str(last_tip).strip():
        tip_days = days_ago if days_ago != "?" else "?"
        st.markdown(f"""
        <div style='background:linear-gradient(135deg,#fefce8,#fef9c3); border-right:4px solid #eab308;
                    border-radius:10px; padding:12px 16px; margin:8px 0;'>
            <span style='font-size:13px; color:#92400e; font-weight:600;'>💡 טיפ מאת {inspector_prev} (לפני {tip_days} ימים):</span><br/>
            <span style='font-size:15px; color:#78350f;'>"{last_tip.strip()}"</span>
        </div>
        """, unsafe_allow_html=True)

    if len(reports) >= 2:
        prev = reports[1]
        changes = []
        if last.get("e_status") != prev.get("e_status"):
            arrow = "✅ תוקן" if last.get("e_status") == "תקין" else "⬇️ הורע"
            changes.append(f"עירוב: {arrow}")
        if last.get("k_cert") != prev.get("k_cert"):
            arrow = "✅ תוקן" if last.get("k_cert") == "כן" else "⬇️ הורע"
            changes.append(f"כשרות: {arrow}")
        prev_mez = int(prev.get("r_mezuzot_missing") or 0)
        if mezuzot != prev_mez:
            diff = prev_mez - mezuzot
            arrow = f"✅ -{diff} הושלמו" if diff > 0 else f"⬇️ +{abs(diff)} חדשות"
            changes.append(f"מזוזות: {arrow}")
        if changes:
            st.caption("📊 שינויים מהביקור הקודם: " + " | ".join(changes))


# ===== פונקציות לניהול קודי גישה דינמיים =====

def get_commander_code(unit: str) -> str:
    """מחזיר את קוד הגישה של היחידה מה-DB, ברירת מחדל 1111"""
    try:
        res = supabase.table("commander_settings").select("access_code").eq("unit", unit).execute()
        if res.data and res.data[0].get('access_code'):
            return str(res.data[0]['access_code'])
    except Exception:
        pass
    return "1111"

def set_commander_code(unit: str, new_code: str):
    """מעדכן את קוד הגישה של היחידה ב-DB"""
    try:
        # השתמש ב-upsert כדי ליצור אם לא קיים
        supabase.table("commander_settings").upsert({
            "unit": unit,
            "access_code": str(new_code),
            "updated_at": datetime.datetime.now().isoformat()
        }, on_conflict="unit").execute()
        return True
    except Exception as e:
        print(f"Error setting commander code: {e}")
        return False


# ════════════════════════════════════════════════════════════
# Wave 2.5 — Advanced Anti-Fraud Intelligence
# ════════════════════════════════════════════════════════════

def get_previous_open_findings(base: str, unit: str) -> list[dict]:
    """שולף מהדוח הקודם כל ממצא שלא סומן כ'תוקן'."""
    try:
        result = (
            supabase.table("reports")
            .select("*")
            .eq("unit", unit)
            .ilike("base", f"%{base}%")
            .order("date", desc=True)
            .limit(1)
            .execute()
        )
        if not result.data:
            return []
        
        last = result.data[0]
        findings = []
        
        # בדיקת רכיבי כשרות ועירוב
        if last.get("e_status") == "פסול":
            findings.append({
                "key": "e_status",
                "label": "עירוב פסול",
                "icon": "🔴",
                "question": "בביקור הקודם דווח שהעירוב פסול — מה המצב כיום?",
                "options": ["תוקן ✅", "עדיין פסול ❌", "לא בדקתי"]
            })
        
        if last.get("k_cert") == "לא":
            findings.append({
                "key": "k_cert",
                "label": "תעודת כשרות חסרה",
                "icon": "🟡",
                "question": "בביקור הקודם דווח שאין תעודת כשרות — האם הוסדר?",
                "options": ["כן, יש תעודה ✅", "עדיין חסרה ❌", "לא בדקתי"]
            })
        
        try:
            mezuzot = int(last.get("r_mezuzot_missing") or 0)
        except:
            mezuzot = 0
            
        if mezuzot > 0:
            findings.append({
                "key": "mezuzot",
                "label": f"{mezuzot} מזוזות חסרות",
                "icon": "🔵",
                "question": f"בביקור הקודם דווח על {mezuzot} מזוזות חסרות — מה המצב?",
                "options": ["הוחלפו/הורכבו כולן ✅", "חלק הוסדר, חלק עדיין חסר", "עדיין חסרות ❌", "לא בדקתי"]
            })
        
        if last.get("p_mix") == "כן":
            findings.append({
                "key": "p_mix",
                "label": "ערבוב כלים",
                "icon": "🟠",
                "question": "בביקור הקודם דווח על ערבוב כלים — האם הופרדו?",
                "options": ["כן, הופרד ✅", "עדיין מעורבב ❌", "לא בדקתי"]
            })
        
        return findings
    except Exception:
        return []

def render_continuity_questions(base: str, unit: str) -> dict:
    """מציג שאלות מעקב על ממצאים פתוחים מהביקור הקודם."""
    findings = get_previous_open_findings(base, unit)
    if not findings:
        return {}
    
    st.markdown("---")
    st.markdown("""
    <div style='background: linear-gradient(135deg, #fff7ed, #fed7aa);
                border-right: 4px solid #f97316;
                border-radius: 10px; padding: 14px 18px; margin: 10px 0;'>
        <span style='font-size:15px; font-weight:700; color:#9a3412;'>
            🔄 מעקב ממצאים מביקור קודם
        </span><br/>
        <span style='font-size:13px; color:#7c2d12;'>
            נמצאו ממצאים פתוחים מהביקור האחרון — אנא דווח על מצבם הנוכחי
        </span>
    </div>
    """, unsafe_allow_html=True)
    
    continuity_answers = {}
    for finding in findings:
        answer = st.radio(
            f"{finding['icon']} {finding['question']}",
            options=finding["options"],
            key=f"cont_{finding['key']}_{base}",
            index=None
        )
        continuity_answers[f"continuity_{finding['key']}"] = answer
        if answer and "לא בדקתי" in answer:
            st.warning(f"⚠️ שים לב: ממצא פתוח לא נבדק. הדוח יסומן לתשומת לב.")
            if "continuity_flags" not in st.session_state: st.session_state.continuity_flags = []
            st.session_state.continuity_flags.append(finding["label"])
    
    st.markdown("---")
    return continuity_answers

# --- Honeypot Logic (Wave 2.15) ---
HONEYPOT_POOL = [
    {
        "question": "האם מכונת הקפה בטרקלין קיבלה הכשר מיוחד לפסח?",
        "trap": ["כן", "תקין"],
        "context": "רוב המוצבים אין להם מכונת קפה בטרקלין"
    },
    {
        "question": "האם פח הגניזה ב'חדר השמרים' ריק?",
        "trap": ["כן", "תקין"],
        "context": "רוב המוצבים אין להם חדר שמרים"
    },
    {
        "question": "האם קפיטריית הקבע מפרידה כלים בשריים וחלביים?",
        "trap": ["כן", "תקין"],
        "context": "רוב המוצבים אין קפיטריית קבע"
    },
    {
        "question": "האם מדיח הכלים הגדול בחדר האוכל מכושר?",
        "trap": ["כן", "תקין"],
        "context": "לא בכל מוצב יש מדיח"
    },
    {
        "question": "האם הוכשר הדוד הקהילתי לחמים לשבת?",
        "trap": ["כן", "תקין"],
        "context": "לא בכל מוצב יש דוד קהילתי"
    },
    {
        "question": "האם ספריית הקודש בחדר הפלוגה מסודרת?",
        "trap": ["כן"],
        "context": "זו שאלה לגיטימית — בודקת אם עונים בלי לבדוק"
    },
    {
        "question": "האם חדר הבידוד מצויד בסידורים?",
        "trap": ["כן", "תקין"],
        "context": "לא בכל מוצב יש חדר בידוד"
    },
    {
        "question": "האם הוכשר מכשיר ה-Nespresso בחדר הפיקוד?",
        "trap": ["כן", "תקין"],
        "context": "לא בכל מוצב יש Nespresso בפיקוד"
    },
]

def get_honeypot_question(base: str) -> dict:
    import hashlib, datetime
    week = datetime.date.today().isocalendar()[1]
    year = datetime.date.today().year
    seed = f"{base}_{week}_{year}"
    idx = int(hashlib.md5(seed.encode()).hexdigest(), 16) % len(HONEYPOT_POOL)
    return HONEYPOT_POOL[idx]

def render_honeypot_question(base: str) -> None:
    if not base or len(base) < 2:
        return
    honeypot = get_honeypot_question(base)
    key = f"honeypot_{base}"
    answer = st.radio(
        honeypot["question"],  # רק השאלה — בלי שום כותרת חשודה
        options=["כן", "לא", "לא קיים במוצב זה"],
        key=key,
        index=None,
        horizontal=True
    )
    if answer in honeypot["trap"]:
        st.session_state["honeypot_triggered"] = True
        st.session_state["honeypot_detail"] = {
            "question": honeypot["question"],
            "answer": answer,
            "context": honeypot["context"]
        }
        st.session_state["reliability_score"] = 0

def render_gps_checkpoint(checkpoint_num: int, base: str):
    done_key = f"gps_done_{checkpoint_num}"
    data_key = f"gps_data_{checkpoint_num}"
    
    labels = {
        1: ("🟢 נקודת פתיחה", "כשנכנסת למוצב"),
        2: ("🟡 נקודת אמצע", "בבדיקת המטבח או ביהכ\"נ"),
        3: ("🔵 נקודת סיום", "כשאתה עומד לעזוב")
    }
    label, instruction = labels.get(checkpoint_num, ("📍", ""))
    
    # אם המיקום כבר נשמר, נציג וי ירוק ונסיים
    if st.session_state.get(done_key):
        st.success(f"{label} — מיקום נקלט ונשמר בהצלחה ✅", icon="📍")
        return True
    
    st.markdown(f"**{label}** — {instruction}")
    
    # שימוש ברכיב ה-GPS המובנה (אמין ב-100% בעיר)
    loc_data = custom_location_button(key=f"gps_cp_{checkpoint_num}_{base}")
    
    # אם התקבלו נתונים (המשתמש לחץ ואישר)
    if loc_data and "lat" in loc_data:
        st.session_state[data_key] = {
            "latitude": loc_data["lat"],
            "longitude": loc_data["lon"],
            "timestamp": time.time(),
            "accuracy": loc_data.get("acc", 0)
        }
        st.session_state[done_key] = True
        st.rerun()

    # במקרה נדיר של חוסר קליטה מוחלט, נשאיר גם אופציה לבחור מוצב מהרשימה
    with st.expander("⚙️ ה-GPS חסום? בחר ידנית מהרשימה"):
        if base in BASE_COORDINATES:
            if st.button(f"📍 השתמש במיקום של '{base}'", key=f"use_dict_{checkpoint_num}", use_container_width=True):
                base_lat, base_lon = BASE_COORDINATES[base]
                st.session_state[data_key] = {"latitude": base_lat, "longitude": base_lon, "timestamp": time.time(), "accuracy": 999}
                st.session_state[done_key] = True
                st.rerun()
        else:
            known_bases = ["-- בחר מוצב מוכר --"] + list(BASE_COORDINATES.keys())
            selected_base = st.selectbox("בחר מיקום:", known_bases, key=f"sel_dict_{checkpoint_num}")
            if st.button("💾 שמור", key=f"save_dict_{checkpoint_num}", use_container_width=True) and selected_base != "-- בחר מוצב מוכר --":
                base_lat, base_lon = BASE_COORDINATES[selected_base]
                st.session_state[data_key] = {"latitude": base_lat, "longitude": base_lon, "timestamp": time.time(), "accuracy": 999}
                st.session_state[done_key] = True
                st.rerun()

    return False

def analyze_gps_fingerprint() -> dict:
    """מנתח את 3 נקודות ה-GPS ומוודא תנועה."""
    points = [st.session_state.get(f"gps_data_{i}") for i in range(1, 4) if st.session_state.get(f"gps_data_{i}")]
    if len(points) < 2: return {"suspicious": False, "points": len(points)}
    
    def dist_m(p1, p2):
        lat1, lon1 = math.radians(p1["latitude"]), math.radians(p1["longitude"])
        lat2, lon2 = math.radians(p2["latitude"]), math.radians(p2["longitude"])
        d = 2 * 6371000 * math.asin(math.sqrt(math.sin((lat2-lat1)/2)**2 + math.cos(lat1)*math.cos(lat2)*math.sin((lon2-lon1)/2)**2))
        return d
    
    max_d = max([dist_m(points[i], points[i+1]) for i in range(len(points)-1)]) if len(points) > 1 else 0
    suspicious = max_d < 15
    return {"suspicious": suspicious, "max_m": max_d, "points": len(points)}

CONTRADICTION_RULES = [
    {"id": "no_kitchen_but_kashrut", "severity": "high", "msg": "סימן 'אין מטבח' אך מילא פרטי כשרות", "check": lambda d: d.get("has_kitchen") == "לא" and (d.get("k_cert") or d.get("p_mix"))},
    {"id": "eruv_ok_but_no_perimeter", "severity": "high", "msg": "עירוב תקין אך לא ביקר בהיקף המוצב", "check": lambda d: d.get("e_status") == "תקין" and d.get("visited_perimeter") == "לא"},
    {"id": "kashrut_ok_no_cert", "severity": "medium", "msg": "דיווח 'כשרות תקינה' ללא תעודה", "check": lambda d: d.get("k_issues") == "לא" and d.get("k_cert") == "לא"},
    {"id": "mix_but_ok_kashrut", "severity": "medium", "msg": "ערבוב כלים אך סימן 'אין תקלות'", "check": lambda d: d.get("p_mix") == "כן" and d.get("k_issues") == "לא"}
]

def check_internal_consistency(data: dict) -> list[dict]:
    res = []
    for r in CONTRADICTION_RULES:
        try:
            if r["check"](data): res.append({"id": r["id"], "severity": r["severity"], "message": r["msg"]})
        except: pass
    return res

def calculate_reliability_score(data: dict) -> int:
    score = 100
    if st.session_state.get("honeypot_triggered"): return 0
    gps = analyze_gps_fingerprint()
    if gps["suspicious"]: score -= 35
    if gps["points"] < 3: score -= (3 - gps["points"]) * 15
    for c in check_internal_consistency(data):
        score -= {"high": 25, "medium": 15, "low": 5}.get(c["severity"], 10)
    return max(0, score)

def save_report_with_analysis(data: dict, unit: str):
    score = calculate_reliability_score(data)
    contradictions = check_internal_consistency(data)
    
    data.update({
        "reliability_score": score,
        "contradictions": contradictions,
        "gps_points": [st.session_state.get(f"gps_data_{i}") for i in range(1, 4)],
        "honeypot_triggered": st.session_state.get("honeypot_triggered", False),
        "review_status": "blocked" if score == 0 else "suspicious" if score < 50 else "review" if score < 75 else "ok"
    })
    
    try:
        supabase.table("reports").insert(data).execute()
        st.success("✅ הדוח נשמר בהצלחה!")
        log_audit_event("REPORT_SAVED", unit, details={"score": score, "status": data["review_status"]})
        return True
    except Exception as e:
        st.error(f"שגיאה בשמירה: {e}")
        return False


# ════════════════════════════════════════════════════════════
# פיצ'ר 8 — חתימה דיגיטלית
# ════════════════════════════════════════════════════════════
def render_signature_pad() -> str | None:
    """
    מציג לוח חתימה דיגיטלי.
    """
    if st_canvas is None:
        st.warning("⚠️ streamlit-drawable-canvas לא מותקן.")
        return None

    st.markdown("### ✍️ חתימת המבקר")
    st.caption("חתום בתיבה למטה עם האצבע / עכבר לאישור הדוח")
    col_canvas, col_actions = st.columns([3, 1])
    with col_canvas:
        canvas_result = st_canvas(
            fill_color="rgba(0, 0, 0, 0)", stroke_width=3, stroke_color="#1e3a8a",
            background_color="#f8fafc", height=120, width=400, drawing_mode="freedraw",
            key="signature_canvas", display_toolbar=False,
        )
    with col_actions:
        st.write("")
        st.write("")
        if st.button("🗑️ נקה", key="clear_signature", use_container_width=True):
            if "signature_url" in st.session_state:
                del st.session_state["signature_url"]
            st.rerun()

    if canvas_result.image_data is not None:
        import numpy as np
        img_array = canvas_result.image_data
        alpha_channel = img_array[:, :, 3]
        if bool(np.any(alpha_channel > 10)):
            st.success("✅ חתימה נקלטה")
            canvas_hash = str(hash(img_array.tobytes()))
            if st.session_state.get("_last_sig_hash") != canvas_hash:
                st.session_state["_last_sig_hash"] = canvas_hash
                from PIL import Image as PILImage
                import io as _io
                sig_img = PILImage.fromarray(img_array.astype("uint8"), "RGBA")
                white_bg = PILImage.new("RGB", sig_img.size, (255, 255, 255))
                white_bg.paste(sig_img, mask=sig_img.split()[3])
                buffer = _io.BytesIO()
                white_bg.save(buffer, format="PNG")
                buffer.seek(0)
                try:
                    import uuid as _uuid, time as _time
                    file_path = f"signatures/sig_{int(_time.time())}_{str(_uuid.uuid4())[:6]}.png"
                    supabase.storage.from_("report-photos").upload(file_path, buffer.getvalue(), {"content-type": "image/png"})
                    project_url = st.secrets["supabase"]["url"].rstrip("/")
                    sig_url = f"{project_url}/storage/v1/object/public/report-photos/{file_path}"
                    st.session_state["signature_url"] = sig_url
                    return sig_url
                except Exception:
                    import base64 as _b64
                    buffer.seek(0)
                    return f"data:image/png;base64,{_b64.b64encode(buffer.read()).decode()}"
            return st.session_state.get("signature_url")
        else:
            st.caption("☝️ טרם נחתם — נא לחתום לפני השליחה")
    return None


# ════════════════════════════════════════════════════════════
# פיצ'ר 9 — "מה השתנה" — Diff אוטומטי
# ════════════════════════════════════════════════════════════
def render_report_diff(new_data: dict, unit: str, base: str):
    """
    משווה את הדוח החדש לדוח הקודם.
    """
    try:
        result = (supabase.table("reports").select("*").eq("unit", unit).ilike("base", f"%{base}%").order("date", desc=True).limit(10).execute())
        # סינון מחיקה רכה ב-Python
        prev_reports = [d for d in result.data if not str(d.get('inspector', '')).startswith('🗑️')]
    except Exception:
        return

    if len(prev_reports) < 2:
        st.info("📋 זהו הביקור הראשון במוצב זה — אין השוואה קודמת")
        return

    prev = prev_reports[1]
    COMPARE_FIELDS = {
        "e_status": ("🚧 עירוב", "תקין", "פסול"),
        "k_cert": ("🍽️ כשרות", "כן", "לא"),
        "r_mezuzot_missing": ("📜 מזוזות חסרות", 0, None),
        "p_mix": ("🔴 ערבוב כלים", "לא", "כן"),
        "k_issues": ("⚠️ תקלות כשרות", "לא", "כן"),
        "s_clean": ("🧹 ניקיון", "מצוין", None),
        "k_shabbat_supervisor": ("👤 נאמן שבת", "כן", "לא"),
    }
    improved, worsened, unchanged = [], [], []
    for field, (label, good_val, bad_val) in COMPARE_FIELDS.items():
        new_val, prev_val = new_data.get(field), prev.get(field)
        if bad_val is None:
            try:
                new_v, prev_v = int(new_val or 0), int(prev_val or 0)
                if new_v < prev_v: improved.append(f"{label}: ירד מ-{prev_v} ל-{new_v} (−{prev_v-new_v})")
                elif new_v > prev_v: worsened.append(f"{label}: עלה מ-{prev_v} ל-{new_v} (+{new_v-prev_v})")
                elif new_v == prev_v and new_v != 0: unchanged.append(f"{label}: {new_v} (ללא שינוי)")
            except Exception: pass
        else:
            if new_val == prev_val:
                if new_val == good_val: unchanged.append(f"{label}: {new_val} ✅")
                elif new_val == bad_val: unchanged.append(f"{label}: {new_val} ⚠️")
            elif new_val == good_val and prev_val == bad_val: improved.append(f"{label}: תוקן! ({prev_val} → {new_val})")
            elif new_val == bad_val and prev_val == good_val: worsened.append(f"{label}: הורע ({prev_val} → {new_val})")

    try:
        all_data = load_reports_cached(None) or []
        df_all = pd.DataFrame(all_data)
        new_score = calculate_unit_score(df_all[df_all["unit"] == unit]) if not df_all.empty else 0
    except Exception: new_score = 0

    st.markdown("---")
    st.markdown("## 📊 מה השתנה מהביקור הקודם?")
    score_color = "#10b981" if new_score >= 80 else "#f59e0b" if new_score >= 60 else "#ef4444"
    c1, c2, c3 = st.columns(3)
    c1.markdown(f"<div style='background:#d1fae5;border-radius:10px;padding:14px;text-align:center;'><div style='font-size:28px;font-weight:800;color:#065f46;'>{len(improved)}</div><div style='color:#065f46;font-weight:600;'>✅ תוקן / השתפר</div></div>", unsafe_allow_html=True)
    c2.markdown(f"<div style='background:#fee2e2;border-radius:10px;padding:14px;text-align:center;'><div style='font-size:28px;font-weight:800;color:#991b1b;'>{len(worsened)}</div><div style='color:#991b1b;font-weight:600;'>🔴 הורע / חדש</div></div>", unsafe_allow_html=True)
    c3.markdown(f"<div style='background:{score_color}22;border-radius:10px;padding:14px;text-align:center;'><div style='font-size:28px;font-weight:800;color:{score_color};'>{new_score:.0f}</div><div style='color:{score_color};font-weight:600;'>ציון עדכני</div></div>", unsafe_allow_html=True)
    if improved:
        with st.expander(f"✅ {len(improved)} נושאים שהשתפרו", expanded=True):
            for item in improved: st.markdown(f"<div style='background:#d1fae5;border-right:4px solid #10b981;padding:10px 14px;border-radius:6px;margin-bottom:6px;color:#064e3b;'>✅ {item}</div>", unsafe_allow_html=True)
    if worsened:
        with st.expander(f"🔴 {len(worsened)} נושאים שהורעו", expanded=True):
            for item in worsened: st.markdown(f"<div style='background:#fee2e2;border-right:4px solid #ef4444;padding:10px 14px;border-radius:6px;margin-bottom:6px;color:#7f1d1d;'>🔴 {item}</div>", unsafe_allow_html=True)
    if unchanged:
        with st.expander(f"➡️ {len(unchanged)} ללא שינוי", expanded=False):
            for item in unchanged: st.markdown(f"<div style='background:#f1f5f9;border-right:4px solid #94a3b8;padding:8px 14px;border-radius:6px;margin-bottom:4px;color:#475569;font-size:14px;'>➡️ {item}</div>", unsafe_allow_html=True)
    if not improved and not worsened: st.info("➡️ אין שינויים משמעותיים לעומת הביקור הקודם")
    st.markdown("---")

def calculate_clusters(df: pd.DataFrame, radius_km: float = 2.0) -> pd.DataFrame:
    """קיבוץ דיווחים קרובים"""
    if df.empty or 'latitude' not in df.columns or 'longitude' not in df.columns:
        return df
    df = df.copy()
    df['cluster_id'] = -1
    cluster_id = 0
    for idx, row in df.iterrows():
        if df.loc[idx, 'cluster_id'] != -1:
            continue
        df.loc[idx, 'cluster_id'] = cluster_id
        for idx2, row2 in df.iterrows():
            if idx == idx2 or df.loc[idx2, 'cluster_id'] != -1:
                continue
            distance = haversine_distance(
                row['latitude'], row['longitude'],
                row2['latitude'], row2['longitude']
            )
            if distance <= radius_km:
                df.loc[idx2, 'cluster_id'] = cluster_id
        cluster_id += 1
    return df

def get_cluster_stats(df: pd.DataFrame) -> List[Dict]:
    """חישוב סטטיסטיקות לכל cluster"""
    if 'cluster_id' not in df.columns:
        return []
    stats = []
    for cluster_id in df['cluster_id'].unique():
        if cluster_id == -1:
            continue
        cluster_df = df[df['cluster_id'] == cluster_id]
        center_lat = cluster_df['latitude'].mean()
        center_lon = cluster_df['longitude'].mean()
        most_common_base = cluster_df['base'].mode()[0] if 'base' in cluster_df.columns and not cluster_df['base'].mode().empty else "לא ידוע"
        most_common_unit = cluster_df['unit'].mode()[0] if 'unit' in cluster_df.columns and not cluster_df['unit'].mode().empty else "לא ידוע"
        stats.append({
            'cluster_id': int(cluster_id),
            'count': len(cluster_df),
            'center_lat': center_lat,
            'center_lon': center_lon,
            'base': most_common_base,
            'unit': most_common_unit
        })
    return stats

# ===== פונקציות Folium למפות ברמת רחוב =====

def secure_location_offset(lat: float, lon: float, unique_id: str, offset_meters: int = 300) -> Tuple[float, float]:
    """
    מזיז מיקום בצורה קבועה לפי מזהה ייחודי (ביטחון מידע)
    - אותו unique_id = תמיד אותה הזזה
    - לא ניתן לנחש את המיקום המקורי
    - ההזזה היא 300 מטר בכיוון אקראי (אבל קבוע)
    """
    # ✅ תיקון: השתמש רק ב-unit+base ללא תאריך (כדי שהמיקום יישאר קבוע)
    try:
        stable_id = f"{unique_id.split('_')[0]}_{unique_id.split('_')[1]}" if '_' in unique_id else unique_id
    except:
        stable_id = unique_id
    
    # יצירת seed קבוע מהמזהה
    seed = int(hashlib.sha256(stable_id.encode()).hexdigest(), 16) % (10**8)
    
    # ✅ שמירת המצב הנוכחי של random
    current_random_state = random.getstate()
    
    # יצירת random generator נפרד
    rng = random.Random(seed)
    
    # המרה למעלות (111km = 1 מעלה)
    offset_deg = offset_meters / 111000
    
    # זווית ומרחק אקראיים (אבל קבועים לאותו ID)
    angle = rng.uniform(0, 2 * math.pi)
    dist = rng.uniform(offset_deg * 0.7, offset_deg)
    
    # ✅ שחזור המצב של random
    random.setstate(current_random_state)
    
    # חישוב offset
    lat_offset = dist * math.cos(angle)
    lon_offset = dist * math.sin(angle) / math.cos(math.radians(lat))
    
    return lat + lat_offset, lon + lon_offset

def create_street_level_map(center=(31.9, 35.2), zoom_start=12):
    """יוצר מפה ברמת רחוב עם שכבות מרובות"""
    m = folium.Map(
        location=center,
        zoom_start=zoom_start,
        max_zoom=20,
        control_scale=True,
        tiles=None,
        prefer_canvas=True
    )
    
    # שכבת רחובות עברית (CartoDB Positron - מציג עברית מצוין)
    folium.TileLayer(
        tiles="CartoDB positron",
        name="מפת רחובות",
        max_zoom=20,
        attr="© CartoDB © OpenStreetMap"
    ).add_to(m)
    
    # שכבת לווין Google
    folium.TileLayer(
        tiles="https://mt1.google.com/vt/lyrs=s&x={x}&y={y}&z={z}",
        name="תצלום לווין",
        attr="© Google",
        max_zoom=20
    ).add_to(m)
    
    # בקרת שכבות
    folium.LayerControl(position='topleft').add_to(m)
    
    return m

def add_unit_marker_to_folium(m, row, unit_colors):
    """מוסיף סימון ליחידה עם offset ביטחוני"""
    # הזזה ביטחונית קבועה (300 מטר)
    lat, lon = secure_location_offset(
        row.get("latitude", 31.9),
        row.get("longitude", 35.2),
        unique_id=f"{row.get('unit', 'unknown')}_{row.get('base', 'unknown')}_{row.get('date', '')}"
    )
    
    # צבע לפי יחידה
    color = unit_colors.get(row.get('unit', ''), '#808080')
    
    # גודל לפי בעיות
    has_issues = (row.get('e_status') == 'פסול' or row.get('k_cert') == 'לא')
    radius = 10 if has_issues else 7
    
    # popup בעברית RTL
    popup_html = f"""
    <div dir="rtl" style="text-align:right; font-family:Arial; font-size:14px; min-width:200px;">
        <b style="color:#1e3a8a; font-size:16px;">📍 {row.get('base', 'לא ידוע')}</b><br><br>
        <b>יחידה:</b> {row.get('unit', 'לא ידוע')}<br>
        <b>מבקר:</b> {row.get('inspector', 'לא ידוע')}<br>
        <b>עירוב:</b> <span style="color:{'#ef4444' if row.get('e_status')=='פסול' else '#10b981'};">{row.get('e_status', 'לא ידוע')}</span><br>
        <b>כשרות:</b> <span style="color:{'#ef4444' if row.get('k_cert')=='לא' else '#10b981'};">{row.get('k_cert', 'לא ידוע')}</span><br>
        <b>תאריך:</b> {row.get('date', 'לא ידוע')}
    </div>
    """
    
    folium.CircleMarker(
        location=[lat, lon],
        radius=radius,
        color=color,
        fill=True,
        fillColor=color,
        fillOpacity=0.7,
        weight=2,
        popup=folium.Popup(popup_html, max_width=300),
        tooltip=f"📍 {row.get('base', 'מוצב')}"
    ).add_to(m)

# --- 1. הגדרת עמוד ---
st.set_page_config(
    page_title="מערכת בקרה ושליטה רבנות פקמ״ז", 
    layout="wide", 
    initial_sidebar_state="collapsed", 
    page_icon="🛡️"
)

# CSS למובייל - אופטימיזציה מלאה
st.markdown("""
<style>
    /* RTL Support - יישור לימין לעברית */
    .main, .block-container, [data-testid="stAppViewContainer"], [data-testid="stApp"] {
        direction: rtl !important;
        text-align: right !important;
    }

    /* RTL for st.info / st.warning / st.success / st.error */
    [data-testid="stAlert"], [data-testid="stAlertContainer"],
    div[class*="stAlert"] {
        direction: rtl !important;
        text-align: right !important;
    }

    /* RTL for st.metric */
    [data-testid="stMetric"], [data-testid="stMetricLabel"],
    [data-testid="stMetricValue"], [data-testid="stMetricDelta"] {
        direction: rtl !important;
        text-align: right !important;
    }

    /* כפיית כיווניות עברית על כל רכיבי הטקסט */
    .stMarkdown, .stText, [data-testid="stMarkdownContainer"] {
        direction: rtl !important;
        text-align: right !important;
    }
    
    /* סידור רשימות (Bullets) בצד ימין */
    [data-testid="stMarkdownContainer"] ul {
        direction: rtl !important;
        text-align: right !important;
        padding-right: 1.8rem !important;
        padding-left: 0 !important;
        list-style-position: inside !important;
    }
    
    [data-testid="stMarkdownContainer"] li {
        direction: rtl !important;
        text-align: right !important;
        margin-bottom: 5px !important;
    }

    /* מניעת היפוך מספרים וסימנים מיוחדים */
    span, p, li {
        unicode-bidi: plaintext !important;
    }

    /* RTL for st.expander */
    [data-testid="stExpander"], .streamlit-expanderHeader,
    .streamlit-expanderContent {
        direction: rtl !important;
        text-align: right !important;
    }

    /* RTL for st.tabs */
    [data-testid="stTabs"], [data-testid="stTabContent"],
    [data-baseweb="tab-list"], [data-baseweb="tab-panel"] {
        direction: rtl !important;
        text-align: right !important;
    }

    /* RTL for st.selectbox / st.text_input labels */
    [data-testid="stSelectbox"] label, [data-testid="stTextInput"] label,
    [data-testid="stTextArea"] label, [data-testid="stCheckbox"] label,
    [data-testid="stRadio"] label {
        direction: rtl !important;
        text-align: right !important;
    }

    /* RTL for all markdown paragraphs and lists */
    [data-testid="stMarkdown"] p, [data-testid="stMarkdown"] li,
    [data-testid="stMarkdown"] ul, [data-testid="stMarkdown"] ol {
        direction: rtl !important;
        text-align: right !important;
    }

    /* נקודות רשימה - הזזה לימין */
    .stMarkdown ul, .stMarkdown ol {
        direction: rtl !important;
        text-align: right !important;
        padding-right: 2rem !important;
        padding-left: 0 !important;
    }

    .stMarkdown li {
        direction: rtl !important;
        text-align: right !important;
    }

    /* כל תוכן Markdown Container מיושר לימין */
    div[data-testid="stMarkdownContainer"] {
        direction: rtl !important;
        text-align: right !important;
    }
    
    /* כותרות - יישור לימין */
    h1, h2, h3, h4, h5, h6 {
        direction: rtl !important;
        text-align: right !important;
    }
    
    /* הסתרת sidebar מלאה */
    [data-testid="stSidebar"], [data-testid="collapsedControl"] {
        display: none !important;
    }
    
    /* עיצוב Header עליון במקום סיידבר */
    .top-header {
        background: white;
        padding: 10px 20px;
        border-bottom: 1px solid #e2e8f0;
        display: flex;
        justify-content: space-between;
        align-items: center;
        margin-bottom: 20px;
        border-radius: 8px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.1);
    }
    
    /* מניעת היפוך מספרים וסימנים מיוחדים */
    
    /* במובייל */
    @media (max-width: 768px) {
        
        /* כותרות - צבע כהה וקריא + כיוון מימין לשמאל */
        h1, h2, h3, h4, h5, h6 {
            color: #1e293b !important;
            font-weight: 700 !important;
            direction: rtl !important;
            text-align: right !important;
        }
        
        /* טקסט רגיל - צבע כהה */
        p, span, div, label {
            color: #334155 !important;
        }
        
        /* כפתורים - גדולים יותר למובייל */
        button {
            min-height: 48px !important;
            font-size: 16px !important;
            font-weight: 600 !important;
        }
        
        /* שדות קלט - גדולים וקריאים */
        input, textarea, select {
            min-height: 48px !important;
            font-size: 16px !important;
            color: #1e293b !important;
            background-color: white !important;
            border: 2px solid #cbd5e1 !important;
            border-radius: 8px !important;
            padding: 12px !important;
        }
        
        /* תיבות בחירה - גדולות יותר */
        [data-testid="stRadio"] label {
            font-size: 16px !important;
            color: #1e293b !important;
            padding: 12px !important;
        }
        
        /* מדדים (metrics) - קריאים יותר */
        [data-testid="stMetric"] {
            background-color: white !important;
            border: 2px solid #e2e8f0 !important;
            border-radius: 12px !important;
            padding: 16px !important;
        }
        
        [data-testid="stMetricLabel"] {
            font-size: 14px !important;
            color: #64748b !important;
            font-weight: 600 !important;
        }
        
        [data-testid="stMetricValue"] {
            font-size: 24px !important;
            color: #1e293b !important;
            font-weight: 700 !important;
        }
        
        /* טבלאות - קריאות יותר */
        table {
            font-size: 14px !important;
        }
        
        table th {
            background-color: #1e293b !important;
            color: white !important;
            font-weight: 700 !important;
            padding: 12px !important;
        }
        
        table td {
            color: #334155 !important;
            padding: 12px !important;
            border-bottom: 1px solid #e2e8f0 !important;
        }
        
        /* כרטיסים - ניגודיות טובה */
        [data-testid="stExpander"] {
            background-color: white !important;
            border: 2px solid #e2e8f0 !important;
            border-radius: 12px !important;
            margin-bottom: 16px !important;
        }
        
        /* התראות - צבעים ברורים */
        .stAlert {
            font-size: 16px !important;
            padding: 16px !important;
            border-radius: 8px !important;
        }
        
        /* הודעות מידע */
        [data-baseweb="notification"] {
            background-color: #dbeafe !important;
            color: #1e40af !important;
            border: 2px solid #3b82f6 !important;
        }
        
        /* הודעות הצלחה */
        .element-container:has(.stSuccess) {
            background-color: #d1fae5 !important;
            color: #065f46 !important;
            border: 2px solid #10b981 !important;
        }
        
        /* הודעות שגיאה */
        .element-container:has(.stError) {
            background-color: #fee2e2 !important;
            color: #991b1b !important;
            border: 2px solid #ef4444 !important;
        }
        
        /* טאבים - גדולים וקריאים */
        [data-baseweb="tab-list"] {
            gap: 8px !important;
        }
        
        [data-baseweb="tab"] {
            min-height: 48px !important;
            font-size: 15px !important;
            font-weight: 600 !important;
            color: #475569 !important;
            background-color: #f1f5f9 !important;
            border-radius: 8px !important;
            padding: 12px 16px !important;
        }
        
        [data-baseweb="tab"][aria-selected="true"] {
            background-color: #3b82f6 !important;
            color: white !important;
        }
        
        /* גרפים - גודל מותאם */
        [data-testid="stPlotlyChart"] {
            height: auto !important;
            min-height: 300px !important;
        }
        
        /* מרווחים */
        .main .block-container {
            padding: 16px !important;
            max-width: 100% !important;
        }
        
        /* כותרת ראשית */
        .main h1:first-of-type {
            font-size: 24px !important;
            margin-bottom: 16px !important;
        }
        
        /* תמונות - מותאמות */
        img {
            max-width: 100% !important;
            height: auto !important;
            border-radius: 8px !important;
        }
        
        /* dataframe - גלילה אופקית */
        [data-testid="stDataFrame"] {
            overflow-x: auto !important;
        }
        
        /* הסתרת footer של streamlit */
        footer {
            display: none !important;
        }
        
        /* הסתרת תפריט */
        #MainMenu {
            display: none !important;
        }
        
        /* כפתור העלאת קובץ */
        [data-testid="stFileUploader"] {
            background-color: white !important;
            border: 2px dashed #cbd5e1 !important;
            border-radius: 12px !important;
            padding: 24px !important;
        }
        
        [data-testid="stFileUploader"] label {
            font-size: 16px !important;
            color: #1e293b !important;
            font-weight: 600 !important;
        }
    }
    
    /* שיפורים כלליים לכל המכשירים */
    /* שיפורים כלליים לכל המכשירים - תיקון אייקונים */
    html, body, [class*="css"] {
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    }
    
    /* RTL for all Alerts */
    [data-testid="stAlert"], .stAlert {
        direction: rtl !important;
        text-align: right !important;
    }
</style>
""", unsafe_allow_html=True)

# ===== מצב לילה/יום – לאחר 17:00 רקע חום חמים וכתב שחור =====
st.markdown("""
<style>
body.night-mode,
body.night-mode .main,
body.night-mode [data-testid="stAppViewContainer"],
body.night-mode [data-testid="stApp"] {
    background-color: #3D2B1F !important;
    color: #0a0a0a !important;
}
body.night-mode h1, body.night-mode h2, body.night-mode h3,
body.night-mode h4, body.night-mode h5, body.night-mode h6,
body.night-mode p, body.night-mode span, body.night-mode div,
body.night-mode label, body.night-mode .stMarkdown {
    color: #0a0a0a !important;
}
body.night-mode .block-container {
    background-color: #4A3328 !important;
}
body.night-mode input, body.night-mode textarea, body.night-mode select {
    background-color: #5C3D2E !important;
    color: #0a0a0a !important;
    border-color: #7A5040 !important;
}
body.night-mode [data-baseweb="tab"] {
    background-color: #5C3D2E !important;
    color: #0a0a0a !important;
}
body.night-mode [data-baseweb="tab"][aria-selected="true"] {
    background-color: #8B5E3C !important;
    color: #000000 !important;
}
body.night-mode [data-testid="stAlert"] {
    background-color: #5C3D2E !important;
    color: #0a0a0a !important;
}
body.night-mode button {
    background-color: #7A5040 !important;
    color: #000000 !important;
}
body.night-mode [data-testid="stMetric"] {
    background-color: #5C3D2E !important;
}
</style>
<script>
(function() {
    function applyTheme() {
        var hour = new Date().getHours();
        if (hour >= 17 || hour < 6) {
            document.body.classList.add('night-mode');
        } else {
            document.body.classList.remove('night-mode');
        }
    }
    applyTheme();
    setInterval(applyTheme, 60000);
})();
</script>
""", unsafe_allow_html=True)


# --- 2. חיבור ל-Supabase ---
try:
    url = st.secrets["supabase"]["url"]
    key = st.secrets["supabase"]["key"]
    supabase: Client = create_client(url, key)
except:
    st.error("שגיאה קריטית: אין חיבור למסד הנתונים. וודא קובץ Secrets.")
    st.stop()

# --- 3. קונפיגורציה ---
REGIONAL_UNITS = [
    "חטמ״ר בנימין", "חטמ״ר שומרון", "חטמ״ר יהודה",
    "חטמ״ר עציון", "חטמ״ר אפרים", "חטמ״ר מנשה", "חטמ״ר הבקעה"
]
REGULAR_UNITS = ["חטיבה 35", "חטיבה 89", "חטיבה 900"]
HATMAR_UNITS = REGIONAL_UNITS + REGULAR_UNITS
# חטיבות ללא טרקלין ויקוק
NO_LOUNGE_WECOOK_UNITS = {"חטיבה 35", "חטיבה 89", "חטיבה 900"}
COMMAND_UNITS = ["אוגדת 877", "אוגדת 96", "אוגדת 98", "פיקוד מרכז"]
ALL_UNITS = HATMAR_UNITS + COMMAND_UNITS

UNIT_ID_MAP = {
    "חטמ״ר בנימין": "binyamin", "חטמ״ר שומרון": "shomron", "חטמ״ר יהודה": "yehuda",
    "חטמ״ר עציון": "etzion", "חטמ״ר אפרים": "efraim", "חטמ״ר מנשה": "menashe",
    "חטמ״ר הבקעה": "habikaa",
    "חטיבה 35": "hativa_35", "חטיבה 89": "hativa_89", "חטיבה 900": "hativa_900",
    "אוגדת 877": "ugdat_877", "אוגדת 96": "ugda_96", "אוגדת 98": "ugda_98",
    "פיקוד מרכז": "pikud"
}

BASES_LIST = [
    "מחנה עופר", "בית אל", "חטיבת יהודה", "קדומים", "שבי שומרון", 
    "מבוא דותן", "בקעות", "אריאל", "מצודת כפיר", "תפוח", "נווה צוף"
]

COLORS = {
    "primary": "#1e3a8a", "secondary": "#3b82f6", "success": "#10b981",
    "warning": "#f59e0b", "danger": "#ef4444", "bg": "#f8fafc", "dark": "#0f172a"
}

# --- 4. פונקציות מערכת ---
def init_db():
    try: supabase.table("reports").select("id").limit(1).execute()
    except: pass

def init_hierarchy_table():
    """יצירת טבלת היררכיה אם לא קיימת"""
    try:
        # ניסיון לקרוא מהטבלה
        supabase.table("hierarchy").select("*").limit(1).execute()
    except:
        # אם הטבלה לא קיימת, ננסה ליצור אותה
        try:
            # יצירת רשומה ראשונית ומחיקתה מיד (כדי ליצור את הטבלה)
            supabase.table("hierarchy").insert({
                "parent_unit": "אוגדת 877",
                "child_unit": "חטמ״ר בנימין"
            }).execute()
        except:
            pass
            
def init_strategic_tables():
    """יצירת טבלאות אסטרטגיות אם לא קיימות (Maintenance & Alerts)"""
    # בדיקת טבלת כרטיסי עבודה
    try:
        supabase.table("maintenance_tickets").select("id").limit(1).execute()
    except:
        pass # הטבלה תיווצר ידנית או ע"י סקריפט SQL חיצוני

    # בדיקת טבלת התראות מפקד
    try:
        supabase.table("command_alerts").select("id").limit(1).execute()
    except:
        pass

if "db_checked" not in st.session_state:
    init_db()
    init_hierarchy_table()
    init_strategic_tables()
    st.session_state.db_checked = True

def hash_password(password):
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(stored_password, input_password):
    """✅ אימות סיסמה מאובטח — bcrypt בלבד. ללא דלת אחורית."""
    try:
        if stored_password and stored_password.startswith("$2b$"):
            return bcrypt.checkpw(
                input_password.encode(),
                stored_password.encode()
            )
    except Exception:
        pass
    return False

def get_logo_url(unit_name):
    project_url = st.secrets['supabase']['url'].rstrip("/")
    english_name = UNIT_ID_MAP.get(unit_name, "default")
    return f"{project_url}/storage/v1/object/public/logos/{english_name}.png?t={int(time.time())}"

def get_user_role(unit_name):
    if unit_name == "פיקוד מרכז": return "pikud"
    # בדיקה לאוגדה - גם "אוגדה" וגם "אוגדת"
    if "אוגדה" in unit_name or "אוגדת" in unit_name: return "ugda"
    # חטיבות נכנסות כחטמ"ר (hatmar)
    if unit_name in NO_LOUNGE_WECOOK_UNITS: return "hatmar"
    try:
        res = supabase.table("unit_passwords").select("role").eq("unit_name", unit_name).execute()
        if res.data and res.data[0].get("role"): return res.data[0]["role"]
    except: pass
    return "hatmar"

def get_accessible_units(unit_name, role):
    if role == "pikud": return ALL_UNITS
    if role == "ugda":
        try:
            res = supabase.table("hierarchy").select("child_unit").eq("parent_unit", unit_name).execute()
            children = [row["child_unit"] for row in res.data]
            return [unit_name] + children
        except: return [unit_name]
    return [unit_name]

@st.cache_data(ttl=60)
def load_reports_cached(accessible_units=None):
    try:
        data = supabase.table("reports").select("*").execute().data
        if not data: return []
        
        # סינון "מחיקה רכה" - דוחות שהמפקח שלהם מתחיל באייקון פח
        filtered_data = [d for d in data if not (str(d.get('inspector', '')).startswith('🗑️'))]
        
        if accessible_units:
            return [d for d in filtered_data if d['unit'] in accessible_units]
        return filtered_data
    except: return []

def clear_cache(): load_reports_cached.clear()

def log_api_usage(unit_name: str, api_type: str = "text"):
    """רישום קריאת API ב-Supabase למניעת חריגת תקציב"""
    try:
        supabase.table("api_usage").insert({
            "unit": unit_name,
            "api_type": api_type
        }).execute()
    except Exception as e:
        print(f"⚠️ שגיאה ברישום שימוש API: {e}")

def check_api_quota_safety(daily_limit: int = 500) -> bool:
    """
    בדיקה האם עברנו את המכסה היומית הכללית.
    מחזיר True אם בטוח להמשיך, False אם הגענו לקצה.
    """
    try:
        # קבלת תחילת היום הנוכחי בפורמט ISO
        today = datetime.date.today().isoformat()
        
        # ספירת השורות מהיום
        result = supabase.table("api_usage") \
            .select("id", count="exact") \
            .gte("created_at", today) \
            .execute()
        
        current_usage = result.count if result.count is not None else 0
        
        if current_usage >= daily_limit:
            return False
        return True
    except Exception as e:
        # במקרה של שגיאת תקשורת, נחמיר ונחזיר True כדי לא לתקוע את המערכת
        print(f"⚠️ שגיאה בבדיקת מכסה: {e}")
        return True

def upload_report_photo(photo_bytes, unit_name, base_name):
    """העלאת תמונה ל-Supabase Storage עם שם קובץ בטוח (ASCII בלבד)"""
    try:
        # המרת התמונה ל-JPEG
        img = Image.open(io.BytesIO(photo_bytes)).convert('RGB')
        output = io.BytesIO()
        img.save(output, format='JPEG', quality=80)
        
        # יצירת שם קובץ בטוח לחלוטין - רק תווים באנגלית ומספרים
        # שימוש ב-UUID וזמן יוניקס למניעת כל סיכוי לבעיות קידוד
        import uuid
        file_ext = "jpg"
        safe_filename = f"report_{int(time.time())}_{str(uuid.uuid4())[:8]}.{file_ext}"
        
        # נתיב הקובץ
        file_path = f"reports/{safe_filename}"
        
        # העלאה ל-Supabase Storage
        supabase.storage.from_("report-photos").upload(
            file_path, 
            output.getvalue(), 
            {"content-type": "image/jpeg"}
        )
        
        # יצירת URL ציבורי
        project_url = st.secrets['supabase']['url'].rstrip("/")
        public_url = f"{project_url}/storage/v1/object/public/report-photos/{file_path}"
        
        return public_url
        
    except Exception as e:
        # הדפסת שגיאה מפורטת ללוג
        print(f"Upload error: {str(e)}")
        st.error(f"❌ שגיאה בהעלאת תמונה: {str(e)}")
        if "InvalidKey" in str(e):
             st.warning("💡 השגיאה נובעת משם קובץ לא תקין. הקוד החדש אמור לפתור זאת.")
        return None

def apply_custom_css():
    """החלת עיצוב CSS מותאם אישית"""
    st.markdown("""
        <style>
        /* יישור לימין לכל האפליקציה */
        .stApp {
            direction: rtl;
            text-align: right;
        }
        
        /* כפיית צבע טקסט כהה עבור נראות במחשב - כולל שאלונים והודעות */
        .stMarkdown, .stText, h1, h2, h3, h4, h5, h6, .stMetricLabel, .stMetricValue, 
        .stRadio label, .stCheckbox label, .stTextInput label, .stSelectbox label, 
        .stTextArea label, .stFileUploader label, .stAlert {
            color: #1e293b !important;
        }
        
        /* צבע טקסט בתוך התיבות עצמן */
        .stTextInput input, .stTextArea textarea, .stSelectbox select {
            color: #1e293b !important;
        }
        
        /* רקע בהיר לאפליקציה */
        .stApp {
            background-color: #f8fafc;
        }
        
        /* הודעות (Alerts) */
        .stAlert {
            background-color: white; /* רקע לבן להודעות כדי שהטקסט יבלוט */
            border: 1px solid #e2e8f0;
        }
        
        /* כרטיסים מעוצבים */
        .css-1r6slb0, .stCard {
            background-color: white;
            padding: 1.5rem;
            border-radius: 0.5rem;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
            border: 1px solid #e2e8f0;
        }
        
        /* כפתורים */
        .stButton button {
            width: 100%;
            border-radius: 0.5rem;
            font-weight: bold;
        }
        
        /* מדדים */
        div[data-testid="stMetricValue"] {
            font-size: 2rem;
            font-weight: bold;
            color: #1e3a8a !important; /* כחול כהה */
        }
        
        div[data-testid="stMetricLabel"] {
            font-size: 1rem;
            color: #64748b !important; /* אפור כהה */
        }
        
        /* טבלאות */
        table {
            color: #1e293b !important;
        }
        </style>
    """, unsafe_allow_html=True)

def apply_dark_theme_css():
    """החלת עיצוב כהה מקצועי (Dark Mode)"""
    st.markdown("""
        <style>
        /* רקע כללי כהה */
        .stApp, [data-testid="stAppViewContainer"] {
            background-color: #0f172a !important; /* slate-950 */
            color: #f8fafc !important; /* slate-50 */
        }
        
        /* טקסט וכותרות */
        h1, h2, h3, h4, h5, h6, p, span, label, .stMarkdown, .stText, .stAlert, 
        .stMetricLabel, .stMetricValue, [data-testid="stMarkdownContainer"] p,
        [data-testid="stMarkdownContainer"] li {
            color: #f1f5f9 !important; /* slate-100 */
        }
        
        /* כרטיסים ו-Expanders */
        [data-testid="stExpander"], .stCard, .css-1r6slb0, div[style*="background: white"], 
        div[style*="background-color: white"], div[style*="background:white"] {
            background-color: #1e293b !important; /* slate-800 */
            border-color: #334155 !important;
            color: #f1f5f9 !important;
        }
        
        /* שדות קלט בסביבה כהה */
        input, textarea, select, [data-baseweb="select"] div, .stTextInput input, .stTextArea textarea {
            background-color: #1e293b !important;
            color: #f8fafc !important;
            border-color: #475569 !important;
        }
        
        /* תיקון צבע סלקטבוקס במצב כהה */
        [data-baseweb="select"] {
            background-color: #1e293b !important;
        }
        
        /* התאמות סיידבר */
        [data-testid="stSidebar"] {
            background-color: #020617 !important; /* slate-950 */
            border-left: 1px solid #1e293b;
        }
        
        /* תיקון צבעים למדדים (Metrics) */
        [data-testid="stMetric"] {
            background-color: #1e293b !important;
            border: 1px solid #334155 !important;
        }
        
        div[data-testid="stMetricValue"] {
            color: #60a5fa !important; /* כחול בהיר */
        }
        
        div[data-testid="stMetricLabel"] {
            color: #94a3b8 !important; /* slate-400 */
        }
        
        /* טאבים */
        button[data-baseweb="tab"] {
            color: #94a3b8 !important;
        }
        
        button[data-baseweb="tab"][aria-selected="true"] {
            color: #3b82f6 !important;
            border-bottom-color: #3b82f6 !important;
        }

        /* RTL & Professional Polish */
        .stApp {
            direction: rtl;
        }
        
        /* כפתורים - התאמה לצבע כהה */
        .stButton button {
            background-color: #3b82f6 !important;
            border: none !important;
            color: white !important;
        }

        /* מסך ניצחון (Impact Screen) */
        div[style*="background:linear-gradient(135deg,#f0fdf4,#dcfce7)"] {
            background: linear-gradient(135deg, #064e3b, #022c22) !important;
            border-color: #10b981 !important;
        }
        
        div[style*="color:#374151"] {
            color: #d1fae5 !important;
        }

        /* בועת טיפ (Inspector Tip) */
        div[style*="background:linear-gradient(135deg,#fefce8,#fef9c3)"] {
            background: linear-gradient(135deg, #713f12, #422006) !important;
            border-color: #eab308 !important;
        }
        
        span[style*="color:#78350f"] {
            color: #fef9c3 !important;
        }
        </style>
    """, unsafe_allow_html=True)

def upload_logo_to_supabase(unit_name, image_bytes):
    """העלאת לוגו חדש לסופהבייס"""
    try:
        img = Image.open(io.BytesIO(image_bytes)).convert('RGBA')
        output = io.BytesIO()
        img.save(output, format='PNG')
        english_name = UNIT_ID_MAP.get(unit_name, "default")
        file_path = f"{english_name}.png"
        
        # מחיקת לוגו קיים
        try:
            supabase.storage.from_("logos").remove([file_path])
        except: pass
        
        # העלאה חדשה
        supabase.storage.from_("logos").upload(file_path, output.getvalue(), {"content-type": "image/png", "upsert": "true"})
        clear_cache()
        return True
    except Exception as e:
        st.error(f"שגיאה בהעלאת לוגו: {e}")
        return False

def update_unit_password(unit_name, new_password):
    """עדכון סיסמה ליחידה"""
    try:
        hashed = hash_password(new_password)
        role = get_user_role(unit_name)
        result = supabase.table("unit_passwords").upsert({
            "unit_name": unit_name, 
            "password": hashed, 
            "role": role
        }, on_conflict="unit_name").execute()
        return True, "הסיסמה עודכנה בהצלחה"
    except Exception as e:
        error_msg = str(e)
        return False, f"שגיאה: {error_msg}"


def add_gps_privacy_offset(lat: float, lon: float, offset_meters: int = 300) -> Tuple[float, float]:
    """
    מוסיף רעש אקראי למיקום GPS לצורכי אבטחה
    מזיז את המיקום ב-~300 מטר כדי שלא לחשוף את המיקום המדויק של המוצב
    
    Args:
        lat: קו רוחב
        lon: קו אורך  
        offset_meters: מרחק מקסימלי במטרים (ברירת מחדל: 300)
    
    Returns:
        tuple: (lat_with_offset, lon_with_offset)
    """
    # המרה ממטרים לדרגות (קירוב: 1 מעלה = ~111km)
    offset_degrees = offset_meters / 111000.0
    
    # רעש אקראי בכיוון אקראי
    random_angle = random.uniform(0, 2 * math.pi)
    random_distance = random.uniform(0, offset_degrees)
    
    # חישוב ההסטה
    lat_offset = random_distance * math.cos(random_angle)
    lon_offset = random_distance * math.sin(random_angle) / math.cos(math.radians(lat))
    
    return (lat + lat_offset, lon + lon_offset)


# ===== מעקב חוסרים =====

def detect_and_track_deficits(report_data: dict, report_id: str, unit: str):
    """
    🔧 תיקון: זיהוי אוטומטי חכם של חוסרים עם סנכרון מלא
    - מזהה חוסרים חדשים לפי מוצב (ולא רק יחידה)
    - מעדכן חוסרים קיימים אם הכמות השתנתה
    - סוגר אוטומטית חוסרים שהושלמו (10→0)
    """
    try:
        base = report_data.get('base', 'לא ידוע')  # ✅ עכשיו לפי מוצב!
        current_date = datetime.datetime.now().isoformat()
        
        # רשימת כל סוגי החוסרים לבדיקה
        deficit_checks = [
            ('mezuzot', int(report_data.get('r_mezuzot_missing', 0))),
            ('eruv_kelim', 1 if report_data.get('p_mix', 'לא') == 'כן' else 0),
            ('kashrut_cert', 1 if report_data.get('k_cert', 'לא') == 'לא' else 0),
            ('eruv_status', 1 if report_data.get('e_status', 'תקין') == 'פסול' else 0),
            ('shabbat_supervisor', 1 if report_data.get('k_shabbat_supervisor', 'כן') == 'לא' else 0),
        ]
        
        for deficit_type, current_count in deficit_checks:
            # ✅ בדיקה אם יש חוסר פתוח מסוג זה עבור אותו מוצב
            existing = supabase.table("deficit_tracking")\
                .select("*")\
                .eq("unit", unit)\
                .eq("base", base)\
                .eq("deficit_type", deficit_type)\
                .eq("status", "open")\
                .execute()
            
            if current_count > 0:
                # ✅ יש חוסר בדוח הנוכחי
                if existing.data:
                    # ✅ עדכון חוסר קיים אם הכמות השתנתה
                    existing_deficit = existing.data[0]
                    if existing_deficit['deficit_count'] != current_count:
                        supabase.table("deficit_tracking").update({
                            'deficit_count': current_count,
                            'updated_at': current_date,
                            'last_report_id': report_id
                        }).eq("id", existing_deficit['id']).execute()
                else:
                    # ✅ יצירת רשומת חוסר חדשה
                    supabase.table("deficit_tracking").insert({
                        'unit': unit,
                        'base': base,
                        'deficit_type': deficit_type,
                        'deficit_count': current_count,
                        'status': 'open',
                        'detected_date': current_date,
                        'report_id': report_id,
                        'last_report_id': report_id
                    }).execute()
            else:
                # ✅ אין חוסר בדוח הנוכחי - סגירה אוטומטית!
                if existing.data:
                    for deficit in existing.data:
                        supabase.table("deficit_tracking").update({
                            'status': 'closed',
                            'resolved_date': current_date,
                            'updated_at': current_date,
                            'resolution_report_id': report_id,
                            'notes': f'✅ החוסר הושלם אוטומטית - דווח 0 בדוח מתאריך {current_date[:10]}'
                        }).eq("id", deficit['id']).execute()
        
    except Exception as e:
        print(f"⚠️ שגיאה במעקב חוסרים: {e}")


def calculate_total_deficits_from_reports(df):
    """
    ✅ חישוב מדויק של סך החוסרים מהדוחות
    לוקח את הדוח האחרון לכל מוצב ומסכם
    """
    import pandas as pd
    
    if df.empty or 'date' not in df.columns:
        return {'mezuzot': 0, 'eruv_kelim': 0, 'kashrut_cert': 0, 'eruv_broken': 0, 'no_supervisor': 0}
    
    # המרת תאריכים אם צריך
    if not pd.api.types.is_datetime64_any_dtype(df['date']):
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
    
    # ✅ קבלת הדוח האחרון לכל מוצב
    latest_reports = df.sort_values('date').groupby('base').tail(1)
    
    # ✅ חישוב סך החוסרים מהדוחות האחרונים
    total_mezuzot = latest_reports['r_mezuzot_missing'].sum() if 'r_mezuzot_missing' in latest_reports.columns else 0
    total_eruv_kelim = len(latest_reports[latest_reports['p_mix'] == 'כן']) if 'p_mix' in latest_reports.columns else 0
    total_no_cert = len(latest_reports[latest_reports['k_cert'] == 'לא']) if 'k_cert' in latest_reports.columns else 0
    total_eruv_broken = len(latest_reports[latest_reports['e_status'] == 'פסול']) if 'e_status' in latest_reports.columns else 0
    total_no_supervisor = len(latest_reports[latest_reports['k_shabbat_supervisor'] == 'לא']) if 'k_shabbat_supervisor' in latest_reports.columns else 0
    
    return {
        'mezuzot': int(total_mezuzot),
        'eruv_kelim': total_eruv_kelim,
        'kashrut_cert': total_no_cert,
        'eruv_broken': total_eruv_broken,
        'no_supervisor': total_no_supervisor
    }


def get_open_deficits(units: list):
    """✅ קבלת חוסרים פתוחים - עם סינון נכון"""
    try:
        result = supabase.table("deficit_tracking")\
            .select("*")\
            .in_("unit", units)\
            .eq("status", "open")\
            .order("detected_date", desc=True)\
            .execute()
        
        import pandas as pd
        return pd.DataFrame(result.data) if result.data else pd.DataFrame()
    except Exception as e:
        print(f"❌ שגיאה בטעינת חוסרים: {e}")
        import streamlit as st
        st.error(f"❌ שגיאה בטעינת חוסרים: {e}")
        import pandas as pd
        return pd.DataFrame()


def get_deficit_statistics(units: list):
    """✅ סטטיסטיקות חוסרים - מדויקות ומסונכרנות"""
    try:
        import pandas as pd
        
        open_result = supabase.table("deficit_tracking")\
            .select("*", count="exact")\
            .in_("unit", units)\
            .eq("status", "open")\
            .execute()
        
        closed_result = supabase.table("deficit_tracking")\
            .select("*")\
            .in_("unit", units)\
            .eq("status", "closed")\
            .execute()
        
        avg_resolution_days = 0
        if closed_result.data:
            total_days, count = 0, 0
            for deficit in closed_result.data:
                if deficit.get('resolved_date') and deficit.get('detected_date'):
                    detected = pd.to_datetime(deficit['detected_date'])
                    resolved = pd.to_datetime(deficit['resolved_date'])
                    total_days += (resolved - detected).days
                    count += 1
            avg_resolution_days = total_days / count if count > 0 else 0
        
        return {
            'total_open': len(open_result.data) if open_result.data else 0,
            'total_closed': len(closed_result.data) if closed_result.data else 0,
            'avg_resolution_days': avg_resolution_days
        }
    except Exception as e:
        print(f"❌ שגיאה בחישוב סטטיסטיקות: {e}")
        import streamlit as st
        st.error(f"❌ שגיאה בחישוב סטטיסטיקות: {e}")
        return {'total_open': 0, 'total_closed': 0, 'avg_resolution_days': 0}


def update_deficit_status(deficit_id: str, status: str, notes: str = ""):
    """✅ עדכון סטטוס חוסר"""
    try:
        update_data = {'status': status, 'updated_at': datetime.datetime.now().isoformat()}
        if notes:
            update_data['notes'] = notes
        if status == 'closed':
            update_data['resolved_date'] = datetime.datetime.now().isoformat()
        
        supabase.table("deficit_tracking").update(update_data).eq("id", deficit_id).execute()
        return True
    except Exception as e:
        print(f"❌ שגיאה בעדכון סטטוס: {e}")
        import streamlit as st
        st.error(f"❌ שגיאה בעדכון סטטוס: {e}")
        return False


# ===== Audit Logging =====

def log_audit_event(action: str, target: str = "", details: dict = None, severity: str = "info"):  # type: ignore[assignment]
    """
    🔍 תיעוד אירועי מערכת ב-Supabase.
    אם טבלת audit_logs לא קיימת — פשוט מדפיס ללוג ולא קורס.
    """
    try:
        unit = getattr(st.session_state, "selected_unit", "unknown") or "unknown"
        role = getattr(st.session_state, "role", "unknown") or "unknown"
        event = {
            "timestamp": datetime.datetime.now().isoformat(),
            "user": unit,
            "role": role,
            "action": action,
            "target": str(target),
            "details": json.dumps(details or {}),
            "severity": severity,
        }
        supabase.table("audit_logs").insert(event).execute()
    except Exception as e:
        # Silent fail — לא קורס אם הטבלה לא קיימת
        print(f"[audit_log] {action} | {target} | skip (table may not exist): {e}")


# ===== Risk Index =====

def calculate_operational_risk_index(unit: str, df: pd.DataFrame) -> dict:
    """
    🚨 מדד סיכון יחידה (0-100)
    משקולות: עירוב פסול (35), כשרות (25), מזוזות (2 כל אחת),
              ימי שתיקה (5/יום), חוסרים ישנים > 7 ימים (10 כל אחד)
    """
    unit_df = df[df['unit'] == unit] if not df.empty else pd.DataFrame()
    if unit_df.empty:
        return {"risk_score": 0, "level": "🟢 נמוך", "color": "#10b981", "details": {}}

    risk_score = 0
    details = {}

    # 1. עירוב פסול
    eruv_invalid = len(unit_df[unit_df.get('e_status', pd.Series(dtype=str)) == 'פסול']) if 'e_status' in unit_df.columns else 0
    risk_score += eruv_invalid * 35
    details['eruv_invalid'] = eruv_invalid

    # 2. כשרות
    no_kashrut = len(unit_df[unit_df['k_cert'] == 'לא']) if 'k_cert' in unit_df.columns else 0
    risk_score += no_kashrut * 25
    details['no_kashrut'] = no_kashrut

    # 3. מזוזות חסרות
    if 'r_mezuzot_missing' in unit_df.columns:
        total_mezuzot = pd.to_numeric(unit_df['r_mezuzot_missing'], errors='coerce').fillna(0).sum()
        risk_score += int(total_mezuzot * 2)
        details['missing_mezuzot'] = int(total_mezuzot)

    # 4. ימים ללא דיווח
    if 'date' in unit_df.columns:
        try:
            last_report = pd.to_datetime(unit_df['date'], errors='coerce').max()
            if pd.notna(last_report):
                days_silent = (pd.Timestamp.now() - last_report).days
                risk_score += max(0, days_silent) * 5
                details['days_silent'] = days_silent
        except Exception:
            pass

    # 5. חוסרים פתוחים מעל 7 ימים
    try:
        open_defs = get_open_deficits([unit])
        if not open_defs.empty and 'detected_date' in open_defs.columns:
            old_defs = open_defs[
                pd.to_datetime(open_defs['detected_date'], errors='coerce') <
                pd.Timestamp.now() - pd.Timedelta(days=7)
            ]
            risk_score += len(old_defs) * 10
            details['overdue_deficits'] = len(old_defs)
    except Exception:
        pass

    risk_score = min(100, max(0, risk_score))

    if risk_score >= 80:
        level, color = "⚫ קריטי", "#dc2626"
    elif risk_score >= 50:
        level, color = "🔴 גבוה", "#ef4444"
    elif risk_score >= 25:
        level, color = "🟡 בינוני", "#f59e0b"
    else:
        level, color = "🟢 נמוך", "#10b981"

    return {"risk_score": risk_score, "level": level, "color": color, "details": details}


# ===== SLA Helpers =====

def count_overdue_deficits(units: list) -> int:
    """ספור חוסרים פתוחים שעברו 7 ימים (SLA רגיל)"""
    try:
        deficits = get_open_deficits(units)
        if deficits.empty or 'detected_date' not in deficits.columns:
            return 0
        deficits = deficits.copy()
        deficits['days_open'] = (
            pd.Timestamp.now() - pd.to_datetime(deficits['detected_date'], errors='coerce').dt.tz_localize(None)
        ).dt.days
        return int((deficits['days_open'] > 7).sum())
    except Exception:
        return 0


def count_silent_units(df: pd.DataFrame) -> int:
    """ספור יחידות שלא דיווחו ביותר מ-7 ימים"""
    if df.empty or 'date' not in df.columns or 'unit' not in df.columns:
        return 0
    silent = 0
    try:
        dates = pd.to_datetime(df['date'], errors='coerce')
        for unit in df['unit'].unique():
            last = dates[df['unit'] == unit].max()
            if pd.notna(last) and (pd.Timestamp.now() - last).days > 7:
                silent += 1
    except Exception:
        pass
    return silent


# ===== Inspector Credibility =====

def calculate_inspector_credibility(inspector: str, df: pd.DataFrame) -> dict:
    """
    🔍 מדד אמינות מתקדם: 
    1. הצלבה מול היסטוריית בסיס (2-3 דוחות אחרונים).
    2. השוואת התנהגות המבקר במוצבים שונים (Cross-location).
    3. ניתוח דפוסי 'סימון וי' (Variance check).
    """
    if df.empty: return {"score": 50, "credibility": "אין נתונים", "color": "#64748b", "defect_rate": 0}
    
    insp_df = df[df['inspector'] == inspector]
    if insp_df.empty: return {"score": 50, "credibility": "אין נתונים", "color": "#64748b", "defect_rate": 0}

    # 1. מדד זמן מילוי (משקל: 25%)
    avg_duration = insp_df['report_duration'].mean() if 'report_duration' in insp_df.columns else 180
    duration_score = 100 if avg_duration >= 180 else (0 if avg_duration < 45 else 50)

    # 2. ניתוח חריגות מול היסטוריית הבסיס (משקל: 35%)
    base_consistency_penalty: float = 0
    for base in insp_df['base'].unique():
        report_at_base = insp_df[insp_df['base'] == base].iloc[0]
        # שליפת 3 דוחות קודמים של אחרים באותו בסיס
        others_at_base = df[(df['base'] == base) & (df['inspector'] != inspector)].sort_values('date', ascending=False).head(3)
        
        if not others_at_base.empty:
            avg_others_defects = others_at_base.apply(lambda r: 1 if r.get('e_status') == 'פסול' or r.get('k_cert') == 'לא' else 0, axis=1).mean()
            current_defect = 1 if report_at_base.get('e_status') == 'פסול' or report_at_base.get('k_cert') == 'לא' else 0
            
            # אם כולם רואים 'תקין' והוא פתאום רואה '100% פסול' (או הפוך) ללא פירוט - זו חריגה
            if abs(current_defect - avg_others_defects) > 0.8:
                base_consistency_penalty += 15

    # 3. מדד 'טביעת אצבע' של המבקר (משקל: 30%)
    defect_rates_per_base = insp_df.apply(lambda r: 1 if r.get('e_status') == 'פסול' or r.get('k_cert') == 'לא' else 0, axis=1)
    if len(insp_df) >= 3 and defect_rates_per_base.std() == 0:
        pattern_score = 20 # חשד כבד לטייס אוטומטי
    else:
        pattern_score = 100

    # חישוב סופי
    final_score = (duration_score * 0.25) + (pattern_score * 0.4) + (max(0, 35 - base_consistency_penalty))
    
    # קביעת סטטוס
    if final_score >= 85: res = {"credibility": "מבקר אמין ויסודי", "color": "#10b981"}
    elif final_score >= 65: res = {"credibility": "אמינות טובה", "color": "#3b82f6"}
    elif final_score >= 45: res = {"credibility": "חשד למילוי שטחי / סימון וי", "color": "#f59e0b"}
    else: res = {"credibility": "חשד גבוה לדיווח פיקטיבי", "color": "#ef4444"}
    
    res.update({"score": float(final_score), "defect_rate": float(defect_rates_per_base.mean() * 100)})
    return res


# ===== QR Scanner =====

def parse_qr_data(qr_string: str) -> dict:
    """
    פענוח QR code של מוצב.
    פורמט: BASE|CAMP_NAME|COORDS|DATE
    """
    try:
        parts = qr_string.strip().split("|")
        if len(parts) >= 2 and parts[0] == "BASE":
            base_name = parts[1].replace("_", " ")
            coords = None
            if len(parts) >= 3:
                try:
                    lat, lon = map(float, parts[2].split(","))
                    coords = (lat, lon)
                except Exception:
                    pass
            return {"base_name": base_name, "coordinates": coords, "qr_raw": qr_string}
    except Exception:
        pass
    return {}


def render_qr_scanner():
    """📱 סריקת QR מוצב (אופציונלי) — מחזיר שם מוצב או None"""
    with st.expander("📱 סריקת QR מוצב (אופציונלי)", expanded=False):
        st.info("💡 הדבס את ה-QR של המוצב לזיהוי אוטומטי. פורמט: BASE|CAMP_NAME|...")
        qr_input = st.text_input("📷 ערך QR", key="qr_input_scanner", placeholder="BASE|CAMP_OFIR|31.76,35.21|2025-01")
        if qr_input:
            result = parse_qr_data(qr_input)
            if result:
                st.success(f"✅ זוהה: **{result['base_name']}**")
                return result['base_name']
            else:
                st.warning("⚠️ QR לא ידוע — הזן מוצב ידנית")
    return None


# --- 5. AI Logic ---
def calculate_operational_readiness(df_unit):
    if len(df_unit) == 0: return 0
    df_calc = df_unit.copy()
    WEIGHTS = {'kashrut': 0.35, 'eruv': 0.25, 'procedures': 0.20, 'logistics': 0.20}
    total_score = 0
    for _, row in df_calc.iterrows():
        k_score = 0 if row.get('k_cert') == 'לא' else 100
        e_score = 0 if row.get('e_status') == 'פסול' else (60 if row.get('e_status') == 'בטיפול' else 100)
        p_score = 100
        if row.get('r_sg') == 'לא': p_score -= 20
        l_score = 100
        if row.get('s_clean') == 'לא': l_score -= 40
        report_final = (k_score * WEIGHTS['kashrut'] + e_score * WEIGHTS['eruv'] + p_score * WEIGHTS['procedures'] + l_score * WEIGHTS['logistics'])
        total_score += max(0, report_final)
    return total_score / len(df_calc)

def analyze_readiness(df):
    alerts = []
    today = pd.Timestamp.now()
    if df.empty: return []
    if not pd.api.types.is_datetime64_any_dtype(df['date']): df['date'] = pd.to_datetime(df['date'], errors='coerce')
    active_units = df['unit'].unique()
    for unit in active_units:
        unit_df = df[df['unit'] == unit]
        if not unit_df.empty:
            last_report = unit_df['date'].max()
            days_silent = (today - last_report).days
            if days_silent > 7: alerts.append(f"⚠️ {unit} לא דיווח כבר {days_silent} ימים")
    last_30 = df[df['date'] > (today - pd.Timedelta(days=30))]
    if not last_30.empty:
        for unit in last_30['unit'].unique():
            u30 = last_30[last_30['unit'] == unit]
            if len(u30) >= 3:
                defects = u30.apply(lambda r: 1 if r.get('e_status') == 'פסול' or r.get('k_cert') == 'לא' else 0, axis=1).sum()
                if (defects / len(u30)) > 0.20: alerts.append(f"🔴 {unit} - ריבוי ליקויים בחודש האחרון")
    return alerts

def calculate_unit_score(df_unit):
    """חישוב ציון מקיף ליחידה (0-100)"""
    if len(df_unit) == 0: return 0
    
    total_score = 0
    for _, row in df_unit.iterrows():
        score = 100
        
        # כשרות (30%)
        if row.get('k_cert') == 'לא': score -= 30
        if row.get('k_bishul') == 'לא': score -= 5
        
        # עירוב (25%)
        if row.get('e_status') == 'פסול': score -= 25
        elif row.get('e_status') == 'בטיפול': score -= 10
        
        # נהלים (20%)
        if row.get('r_sg') == 'לא': score -= 10
        if row.get('r_hamal') == 'לא': score -= 5
        if row.get('r_netilot') == 'לא': score -= 5
        
        # בית כנסת (15%)
        if row.get('s_clean') == 'לא': score -= 10
        if row.get('s_board') == 'לא': score -= 5
        
        # מזוזות (10%)
        mezuzot = row.get('r_mezuzot_missing', 0)
        if mezuzot > 0: score -= min(10, mezuzot * 2)
        
        total_score += max(0, score)
    
    return total_score / len(df_unit)

def get_unit_badge(score):
    """החזרת תג וצבע לפי ציון"""
    if score >= 90: return "🏆 מצטיין", "#10b981"
    elif score >= 80: return "⭐ טוב מאוד", "#3b82f6"
    elif score >= 70: return "✓ טוב", "#f59e0b"
    elif score >= 60: return "⚠️ בינוני", "#f97316"
    else: return "❌ דורש שיפור", "#ef4444"

# ==========================================
# 🆕 תכונות אסטרטגיות (Strategic Features)
# ==========================================

# --- 1️⃣ Offline-First Drafts ---
def save_draft_locally(data, draft_key):
    """שמירת טיוטת דוח מקומית ב-Session State"""
    if 'drafts' not in st.session_state:
        st.session_state.drafts = {}
    st.session_state.drafts[draft_key] = {
        'data': data,
        'timestamp': datetime.datetime.now().isoformat(),
        'status': 'draft'
    }
    st.success(f"✅ הדוח שמור כטיוטה ב-{draft_key}")

def load_draft(draft_key):
    """טעינת טיוטה"""
    if 'drafts' in st.session_state and draft_key in st.session_state.drafts:
        return st.session_state.drafts[draft_key]['data']
    return None

# --- 2️⃣ Closed-Loop Ticketing ---
def create_maintenance_ticket(report_data, report_id):
    """יצירת כרטיס תיקון אוטומטי מדיווח שלילי"""
    try:
        # בדוק אם יש בעיות קריטיות
        critical_issues = []
        
        if report_data.get('e_status') == 'פסול':
            critical_issues.append('עירוב פסול')
        if report_data.get('k_cert') == 'לא':
            critical_issues.append('כשרות לא תקינה')
        try:
            mezuzot_missing = int(report_data.get('r_mezuzot_missing', 0))
            if mezuzot_missing > 5:
                critical_issues.append(f"חוסר קריטי: {mezuzot_missing} מזוזות")
        except: pass
        
        if critical_issues:
            # יצור כרטיס עבודה
            ticket = {
                'report_id': report_id,
                'unit': report_data.get('unit'),
                'base': report_data.get('base'),
                'status': 'open',
                'priority': 'high' if len(critical_issues) > 1 else 'medium',
                'issues': ', '.join(critical_issues),
                'created_at': datetime.datetime.now().isoformat(),
                'assigned_to': 'תחזוקה - להקצאה',
                'deadline': (datetime.datetime.now() + datetime.timedelta(days=3)).isoformat()
            }
            
            # בדיקה אם הטבלה קיימת, אם לא - ניסיון ליצור או דילוג
            try:
                supabase.table("maintenance_tickets").insert(ticket).execute()
                
                # שלח התראה למטה
                send_alert_to_command(
                    f"🚨 כרטיס עבודה חדש - {report_data.get('base')}",
                    f"בעיות: {', '.join(critical_issues)}",
                    'high'
                )
                return True
            except Exception as e:
                print(f"⚠️ טבלת maintenance_tickets לא קיימת או שגיאה אחרת: {e}")
                
    except Exception as e:
        print(f"❌ שגיאה ביצירת כרטיס: {e}")
    return None

def send_alert_to_command(title, message, priority):
    """שליחת התראה למפקדים"""
    try:
        # יצור ערך התראה
        alert = {
            'title': title,
            'message': message,
            'priority': priority,
            'created_at': datetime.datetime.now().isoformat(),
            'read': False
        }
        supabase.table("command_alerts").insert(alert).execute()
    except Exception as e:
        print(f"❌ שגיאה בשליחת התראה: {e}")

# --- 3️⃣ OCR for Kashrut Certificates ---
def extract_kashrut_cert_data(image_bytes):
    """חילוץ נתונים מתעודת כשרות באמצעות OCR"""
    try:
        import cv2
        import pytesseract
        import numpy as np
        
        # המר לתמונה
        image = Image.open(io.BytesIO(image_bytes))
        
        # עיבוד מקדים
        img_np = np.array(image)
        if len(img_np.shape) == 3:
            img_cv = cv2.cvtColor(img_np, cv2.COLOR_RGB2GRAY)
        else:
            img_cv = img_np
            
        # שיפור ניגודיות
        img_cv = cv2.adaptiveThreshold(img_cv, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                       cv2.THRESH_BINARY, 11, 2)
        
        # זיהוי טקסט
        # הערה: זה דורש התקנת Tesseract על המכונה
        try:
            text = pytesseract.image_to_string(img_cv, lang='heb+eng')
        except Exception as e:
            return {'error': 'Tesseract not installed or not found'}

        # חלץ נתונים (סימולציה בסיסית אם ה-OCR לא מושלם)
        extracted = {
            'raw_text': text,
            'supplier_name': extract_supplier_name(text),
            'expiry_date': extract_date(text),
            'certificate_number': extract_cert_number(text)
        }
        
        return extracted
    except Exception as e:
        st.error(f"❌ שגיאה ב-OCR: {e}")
        return None

def extract_date(text):
    """חילוץ תאריך תפוגה"""
    import re
    # חפש תבניות תאריך
    pattern = r'(\d{1,2}[./\\-]\d{1,2}[./\\-]\d{2,4})'
    matches = re.findall(pattern, text)
    return matches[0] if matches else None

def extract_supplier_name(text):
    """ניסיון לחלץ שם ספק"""
    lines = text.split('\n')
    for line in lines[:5]:  # בדרך כלל בהתחלה
        if len(line.strip()) > 3:
            return line.strip()
    return "לא זוהה"

def extract_cert_number(text):
    import re
    matches = re.findall(r'(\d{5,10})', text)
    return matches[0] if matches else "לא זוהה"

def validate_cert_status(expiry_date_str):
    """בדוק אם התעודה תקפה"""
    try:
        if not expiry_date_str: return '❓ לא נמצא תאריך', 'unknown'
        expiry_date = pd.to_datetime(expiry_date_str, dayfirst=True)
        today = pd.Timestamp.now()
        
        if expiry_date < today:
            return '❌ פגה', 'expired'
        elif expiry_date < today + pd.Timedelta(days=30):
            return '⚠️ עומדת לפוג', 'expiring_soon'
        else:
            return '✅ תקפה', 'valid'
    except:
        return '❓ לא בטוח', 'unknown'

# --- 4️⃣ Real-time Heatmap ---
def render_realtime_heatmap(df, accessible_units):
    """מפת חום בזמן אמת של מצב היחידות"""
    st.markdown("### 🌡️ מפת חום - סטטוס בזמן אמת")
    
    if df.empty:
        st.info("אין נתונים להצגה")
        return

    # עיבוד נתונים
    unit_status = []
    
    # סינון רק ליחידות נגישות
    relevant_df = df[df['unit'].isin(accessible_units)] if accessible_units else df
    
    for unit in accessible_units:
        unit_df = relevant_df[relevant_df['unit'] == unit]
        if not unit_df.empty:
            last_report = pd.to_datetime(unit_df['date']).max()
            hours_ago = (pd.Timestamp.now() - last_report).total_seconds() / 3600
            
            # קבע צבע לפי זמן ההמתנה
            if hours_ago < 24:
                color = '#10b981'  # ירוק - פעיל ביממה האחרונה
                status = '🟢 פעיל'
                readiness = 95
            elif hours_ago < 72:
                color = '#f59e0b'  # כתום - פעיל ב-3 ימים אחרונים
                status = '🟡 בטיפול'
                readiness = 70
            else:
                color = '#ef4444'  # אדום - לא פעיל
                status = '🔴 שקט'
                readiness = 40
            
            # חשב ציון כשרות
            kashrut_score = (len(unit_df[unit_df['k_cert'] == 'כן']) / len(unit_df) * 100) if len(unit_df) > 0 else 0
            
            unit_status.append({
                'unit': unit,
                'color': color,
                'status': status,
                'hours_ago': hours_ago,
                'readiness': readiness,
                'kashrut_score': kashrut_score,
                'reports': len(unit_df)
            })
    
    if not unit_status:
        st.info("אין מספיק נתונים ליצירת מפת חום")
        return

    status_df = pd.DataFrame(unit_status)
    
    # גרף חום
    try:
        fig = go.Figure(data=go.Heatmap(
            z=[status_df['readiness'].values],
            x=status_df['unit'].values,
            y=['כשירות'],
            colorscale='RdYlGn',
            text=[status_df['status'].values],
            texttemplate="%{text}",
            textfont={"size": 10},
            showscale=False
        ))
        
        fig.update_layout(
            height=150, 
            margin=dict(l=0, r=0, b=0, t=10),
            xaxis_title=None,
            yaxis_title=None
        )
        st.plotly_chart(fig, use_container_width=True)
    except Exception as e:
        st.error(f"שגיאה בהצגת גרף: {e}")
    
    # טבלה מפורטת (אופציונלי, כרגע נציג רק את הגרף)

# --- 5️⃣ Anomaly Detection ---
def detect_anomalies(df, unit_name):
    """זיהוי דפוסים חריגים בדוחות יחידה"""
    if df.empty: return []
    
    unit_df = df[df['unit'] == unit_name].sort_values('date').tail(10)  # 10 דוחות אחרונים
    
    if len(unit_df) < 3:
        return []
    
    anomalies = []
    
    # 1. כשרות 100% תמיד = חשוד (אם יש מספיק דוחות)
    # 1. כשרות 100% תמיד = חשוד (אם יש מספיק דוחות)
    is_suspicious_perfect = False
    
    if len(unit_df) >= 5:
        k_cert_series = unit_df.get('k_cert')
        e_status_series = unit_df.get('e_status')
        
        # בדיקה בטוחה שיש לנו Series ושכל הערכים תקינים
        if k_cert_series is not None and e_status_series is not None:
             try:
                 if (k_cert_series == 'כן').all() and (e_status_series == 'תקין').all():
                     is_suspicious_perfect = True
             except: pass

    if is_suspicious_perfect:
        anomalies.append({
            'type': 'suspicious_perfect',
            'severity': 'medium',
            'message': '⚠️ תקינות מלאה רצופה - שקול ביקורת עמוקה לבדיקת אמינות'
        })
    
    # 2. שינוי פתאומי בציון
    if len(unit_df) >= 6:
        recent_score = calculate_unit_score(unit_df.tail(3))
        old_score = calculate_unit_score(unit_df.head(3))
        
        if abs(recent_score - old_score) > 30:
            direction = '📈 שיפור' if recent_score > old_score else '📉 ירידה'
            anomalies.append({
                'type': 'score_jump',
                'severity': 'high',
                'message': f'🚨 שינוי דרמטי בציון: {direction} של {abs(recent_score - old_score):.0f} נקודות'
            })
    
    # 3. ערבוב בין דיווחים עם מבקרים שונים
    if 'inspector' in unit_df.columns:
        inspector_changes = unit_df['inspector'].nunique()
        if inspector_changes >= 3:
            anomalies.append({
                'type': 'multiple_inspectors',
                'severity': 'low',
                'message': f'📌 {inspector_changes} מבקרים שונים לאחרונה - שקול סדרת בדיקות אחידה'
            })
    
    return anomalies

def generate_ai_summary(df):
    """יצירת סיכום AI של המצב הכללי"""
    if df.empty:
        return {"overview": "אין נתונים זמינים לניתוח"}
    
    total_reports = len(df)
    active_units = df['unit'].nunique()
    
    # חישוב ממוצעים
    avg_score = sum([calculate_unit_score(df[df['unit']==u]) for u in df['unit'].unique()]) / active_units if active_units > 0 else 0
    
    # בעיות קריטיות
    critical_issues = 0
    if 'e_status' in df.columns:
        critical_issues += len(df[df['e_status'] == 'פסול'])
    if 'k_cert' in df.columns:
        critical_issues += len(df[df['k_cert'] == 'לא'])
    
    overview = f"""
    📊 **סיכום מצב פיקודי**
    
    - **{total_reports}** דוחות מ-**{active_units}** יחידות פעילות
    - ציון ממוצע: **{avg_score:.1f}/100**
    - בעיות קריטיות: **{critical_issues}**
    - מגמה: {"📈 שיפור" if avg_score > 75 else "📉 דורש תשומת לב"}
    """
    
    return {"overview": overview}

def generate_commander_alerts(df):
    """יצירת התראות חכמות למפקדים"""
    alerts = []
    
    if df.empty:
        return alerts
    
    # המרת תאריכים
    if 'date' in df.columns:
        if not pd.api.types.is_datetime64_any_dtype(df['date']):
            df['date'] = pd.to_datetime(df['date'], errors='coerce')
        
        # יחידות שלא דיווחו
        today = pd.Timestamp.now()
        for unit in df['unit'].unique():
            unit_df = df[df['unit'] == unit]
            last_report = unit_df['date'].max()
            days_silent = (today - last_report).days
            if days_silent > 7:
                alerts.append({
                    "icon": "⏰",
                    "title": "חוסר דיווח",
                    "message": f"{unit} לא דיווח כבר {days_silent} ימים"
                })
    
    # עירובין פסולים
    if 'e_status' in df.columns:
        invalid_eruv = df[df['e_status'] == 'פסול']
        if len(invalid_eruv) > 0:
            alerts.append({
                "icon": "🚧",
                "title": "עירובין פסולים",
                "message": f"{len(invalid_eruv)} מוצבים עם עירוב פסול: {', '.join(invalid_eruv['base'].unique()[:3])}"
            })
    
    # כשרות
    if 'k_cert' in df.columns:
        no_cert = df[df['k_cert'] == 'לא']
        if len(no_cert) > 0:
            alerts.append({
                "icon": "🍽️",
                "title": "בעיות כשרות",
                "message": f"{len(no_cert)} מוצבים ללא תעודת כשרות תקפה"
            })
    
    # מזוזות חסרות
    if 'r_mezuzot_missing' in df.columns:
        total_mezuzot = df['r_mezuzot_missing'].sum()
        if total_mezuzot > 0:
            alerts.append({
                "icon": "📜",
                "title": "מזוזות חסרות",
                "message": f"סה״כ {int(total_mezuzot)} מזוזות חסרות בכל היחידות"
            })
    
    return alerts

def analyze_unit_trends(df_unit):
    """ניתוח מגמות ליחידה ספציפית"""
    insights = []
    
    if df_unit.empty:
        return [{"icon": "📊", "title": "אין נתונים", "message": "לא נמצאו דוחות ליחידה זו"}]
    
    # ציון כללי
    score = calculate_unit_score(df_unit)
    badge, _ = get_unit_badge(score)
    insights.append({
        "icon": "🎯",
        "title": "ציון כללי",
        "message": f"היחידה קיבלה ציון {score:.1f}/100 - {badge}"
    })
    
    # תדירות דיווח
    insights.append({
        "icon": "📅",
        "title": "תדירות דיווח",
        "message": f"היחידה דיווחה {len(df_unit)} פעמים"
    })
    
    # נקודות חוזק
    strengths = []
    if 'k_cert' in df_unit.columns and (df_unit['k_cert'] == 'כן').all():
        strengths.append("כשרות מלאה")
    if 'e_status' in df_unit.columns and (df_unit['e_status'] == 'תקין').all():
        strengths.append("עירובין תקינים")
    if 's_clean' in df_unit.columns and (df_unit['s_clean'] == 'כן').all():
        strengths.append("ניקיון מצוין")
    
    if strengths:
        insights.append({
            "icon": "💪",
            "title": "נקודות חוזק",
            "message": ", ".join(strengths)
        })
    
    # נקודות לשיפור
    improvements = []
    if 'k_cert' in df_unit.columns and (df_unit['k_cert'] == 'לא').any():
        improvements.append("כשרות")
    if 'e_status' in df_unit.columns and (df_unit['e_status'] == 'פסול').any():
        improvements.append("עירובין")
    if 'r_mezuzot_missing' in df_unit.columns and df_unit['r_mezuzot_missing'].sum() > 0:
        improvements.append(f"מזוזות ({int(df_unit['r_mezuzot_missing'].sum())} חסרות)")
    
    if improvements:
        insights.append({
            "icon": "🔧",
            "title": "דורש שיפור",
            "message": ", ".join(improvements)
        })
    
    return insights

# --- פונקציות סטטיסטיקות מבקרים ---
def generate_inspector_stats(df):
    """יצירת סטטיסטיקות מבקרים"""
    if df.empty or 'inspector' not in df.columns:
        return None
    
    # סינון דוחות מהחודש הנוכחי
    today = pd.Timestamp.now()
    current_month = df[df['date'].dt.month == today.month]
    
    if current_month.empty:
        current_month = df  # אם אין דוחות החודש, קח הכל
    
    # ספירת דוחות לפי מבקר
    inspector_counts = current_month['inspector'].value_counts()
    
    # מיקומים פופולריים
    location_counts = current_month['base'].value_counts() if 'base' in current_month.columns else pd.Series()
    
    # שעות פעילות - בדיקה של עמודת time תחילה, אחר כך date
    if 'time' in current_month.columns:
        # אם יש עמודת time, השתמש בה
        def extract_hour_from_time(time_val):
            try:
                if pd.isna(time_val):
                    return None
                time_str = str(time_val)
                if ':' in time_str:
                    return int(time_str.split(':')[0])
                return None
            except:
                return None
        current_month['hour'] = current_month['time'].apply(extract_hour_from_time)
        peak_hours = current_month['hour'].dropna().value_counts().head(3)
    elif pd.api.types.is_datetime64_any_dtype(current_month['date']):
        current_month['hour'] = current_month['date'].dt.hour
        peak_hours = current_month['hour'].value_counts().head(3)
    else:
        peak_hours = pd.Series()
    
    return {
        'top_inspectors': inspector_counts.head(10),
        'top_locations': location_counts.head(5),
        'peak_hours': peak_hours,
        'total_reports': len(current_month),
        'unique_inspectors': current_month['inspector'].nunique()
    }


def create_full_report_excel(df):
    """
    יצירת קובץ Excel מעוצב - תואם לטבלת 'דוחות מפורטים' באתר
    כולל עיצוב, כיוון מימין לשמאל, ופילטרים
    """
    try:
        import io
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Side, Alignment, Border
        from openpyxl.utils import get_column_letter

        if df.empty:
            return None
            
        # 1. הגדרת העמודות
        column_mapping = {
            'date': 'תאריך',
            'base': 'מוצב',
            'inspector': 'מבקר',
            'e_status': 'סטטוס עירוב',
            'k_cert': 'תעודת כשרות',
            'k_issues_description': '📝 פירוט תקלות כשרות',
            'k_separation': 'הפרדת כלים',
            'p_mix': '🔴 ערבוב כלים',
            'k_products': 'רכש חוץ לא מאושר',
            'k_bishul': 'בישול ישראל',
            
            # טרקלין
            't_private': '☕ טרקלין - כלים פרטיים',
            't_kitchen_tools': '🥣 טרקלין - כלי מטבח',
            't_procedure': '🔒 טרקלין - נוהל סגירה',
            't_friday': '🛑 טרקלין - סגור בשבת',
            't_app': '📱 טרקלין - אפליקציה',
            
            # ויקוק
            'w_location': '📍 ויקוק - מיקום',
            'w_private': '🥤 ויקוק - כלים פרטיים',
            'w_kitchen_tools': '🍴 ויקוק - כלי מטבח',
            'w_procedure': '📜 ויקוק - עובד לפי פקודה',
            'w_guidelines': '📋 ויקוק - הנחיות',

            # שיעורי תורה
            'soldier_want_lesson': '💡 רצון לשיעור תורה',
            'soldier_has_lesson': '📚 יש שיעור במוצב?',
            'soldier_lesson_teacher': '👨‍🏫 שם מעביר השיעור',
            'soldier_lesson_phone': '📞 טלפון מעביר השיעור',
            'soldier_yeshiva': 'ימי ישיבה',
            'r_mezuzot_missing': '📜 מזוזות חסרות',
            'r_torah_missing': '📖 ספרי תורה חסרים',
            'missing_items': '⚠️ חוסרים כלליים',
            's_torah_id': "מס' צ' ספר תורה",
            's_torah_nusach': "נוסח ספר תורה",
            'free_text': '📝 הערות נוספות'
        }
        
        # 2. סינון ועיבוד נתונים
        available_cols = [col for col in column_mapping.keys() if col in df.columns]
        export_df = df[available_cols].copy()
        export_df.rename(columns=column_mapping, inplace=True)
        
        if 'תאריך' in export_df.columns:
            export_df['תאריך'] = pd.to_datetime(export_df['תאריך']).dt.strftime('%d/%m/%Y %H:%M')

        # 3. יצירת הקובץ עם עיצוב
        output = io.BytesIO()
        
        # שימוש ב-ExcelWriter עם openpyxl
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            export_df.to_excel(writer, index=False, sheet_name='דוחות רבנות')
            
            # קבלת הגיליון לעיצוב
            workbook = writer.book
            worksheet = writer.sheets['דוחות רבנות']
            
            # כיוון גיליון מימין לשמאל
            worksheet.sheet_view.rightToLeft = True
            
            # סגנונות
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid') # כחול כהה
            border_style = Side(border_style='thin', color='000000')
            thin_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
            alignment_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
            alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # עיצוב כותרות
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = alignment_center
                
            # עיצוב תאים והתאמת רוחב
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = alignment_right
                    
            # הוספת פילטרים
            worksheet.auto_filter.ref = worksheet.dimensions
            
            # התאמת רוחב עמודות
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                # חישוב אורך מקסימלי (עם גבול סביר)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                        
                adjusted_width = min(max_length + 2, 40) # מקסימום רוחב
                worksheet.column_dimensions[column_letter].width = adjusted_width

        return output.getvalue()
        
    except Exception as e:
        st.error(f"שגיאה ביצירת הקובץ: {e}")
        return None

def create_inspector_excel(df):
    """יצירת קובץ Excel עם סטטיסטיקות מבקרים (מוגבל ל-10 שורות)"""
    import io
    try:
        import openpyxl
    except ImportError:
        return None
        
    from datetime import datetime
    
    stats = generate_inspector_stats(df)
    if not stats:
        # יצירת מילון ריק כדי למנוע קריסה ולאפשר יצירת קובץ
        stats = {
            'top_inspectors': pd.Series(dtype='object'),
            'top_locations': pd.Series(dtype='object'),
            'peak_hours': pd.Series(dtype='object'),
            'total_reports': len(df),
            'unique_inspectors': 0
        }
    
    # יצירת DataFrame לייצוא
    export_data = []
    for idx, (inspector, count) in enumerate(stats['top_inspectors'].items(), 1):
        # מציאת המיקום הנפוץ ביותר של המבקר
        inspector_reports = df[df['inspector'] == inspector]
        top_location = inspector_reports['base'].mode()[0] if 'base' in inspector_reports.columns and not inspector_reports['base'].mode().empty else "לא ידוע"
        
        # שעה נפוצה
        if pd.api.types.is_datetime64_any_dtype(inspector_reports['date']):
            inspector_reports['hour'] = inspector_reports['date'].dt.hour
            peak_hour = inspector_reports['hour'].mode()[0] if not inspector_reports['hour'].mode().empty else 0
        else:
            peak_hour = 0
        
        export_data.append({
            'דירוג': idx,
            'שם המבקר': inspector,
            'מספר דוחות': count,
            'מיקום עיקרי': top_location,
            'שעת פעילות נפוצה': f"{peak_hour:02d}:00"
        })
    
    df_export = pd.DataFrame(export_data)
    
    # יצירת קובץ Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='סטטיסטיקות מבקרים')
    
    return output.getvalue()

def create_hierarchy_flowchart():
    """יצירת תרשים זרימה של מבנה היחידות"""
    try:
        hierarchy_data = supabase.table("hierarchy").select("*").execute().data
        
        if not hierarchy_data:
            return "```mermaid\ngraph TD\n    PIKUD[\"🎖️ פיקוד מרכז\"]\n    U1[\"⭐ אוגדת 877\"]\n    U2[\"⭐ אוגדת 96\"]\n    U3[\"⭐ אוגדת 98\"]\n    PIKUD --> U1\n    PIKUD --> U2\n    PIKUD --> U3\n    \n    style PIKUD fill:#1e3a8a,stroke:#1e40af,stroke-width:3px,color:#fff\n    style U1 fill:#3b82f6,stroke:#2563eb,stroke-width:2px,color:#fff\n    style U2 fill:#3b82f6,stroke:#2563eb,stroke-width:2px,color:#fff\n    style U3 fill:#059669,stroke:#047857,stroke-width:2px,color:#fff\n```"
        
        # בניית הגרף
        mermaid_code = "```mermaid\ngraph TD\n"
        mermaid_code += "    PIKUD[\"🎖️ פיקוד מרכז\"]\n"
        
        # קבוצות לפי אוגדה
        ugdot = {}
        for h in hierarchy_data:
            parent = h['parent_unit']
            child = h['child_unit']
            if parent not in ugdot:
                ugdot[parent] = []
            ugdot[parent].append(child)
        
        # הוספת אוגדות
        ugda_ids = {}
        for idx, ugda in enumerate(ugdot.keys(), 1):
            ugda_id = f"U{idx}"
            ugda_ids[ugda] = ugda_id
            mermaid_code += f"    {ugda_id}[\"⭐ {ugda}\"]\n"
            mermaid_code += f"    PIKUD --> {ugda_id}\n"
        
        # הוספת חטמ"רים
        for ugda, hatmarim in ugdot.items():
            ugda_id = ugda_ids[ugda]
            for idx, hatmar in enumerate(hatmarim, 1):
                hatmar_id = f"{ugda_id}_H{idx}"
                mermaid_code += f"    {hatmar_id}[\"🏛️ {hatmar}\"]\n"
                mermaid_code += f"    {ugda_id} --> {hatmar_id}\n"
        
        # עיצוב
        mermaid_code += "\n    style PIKUD fill:#1e3a8a,stroke:#1e40af,stroke-width:4px,color:#fff,font-size:16px\n"
        for ugda_id in ugda_ids.values():
            mermaid_code += f"    style {ugda_id} fill:#3b82f6,stroke:#2563eb,stroke-width:3px,color:#fff,font-size:14px\n"
        
        mermaid_code += "```"
        return mermaid_code
    except:
        return """```mermaid
graph TD
    C["⚠️ טרם הוגדרה היררכיה"]
    style C fill:#3b82f6,color:#fff
```"""

# --- 6. CSS (עיצוב רספונסיבי מושלם למובייל ומחשב) ---
st.markdown(f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Rubik:wght@300;400;500;600;700;800&display=swap');
    
    html, body, .stApp {{ 
        direction: rtl; 
        text-align: right; 
        font-family: 'Rubik', sans-serif !important; 
        background: linear-gradient(135deg, {COLORS['bg']} 0%, #e0e7ff 100%);
        color: {COLORS['dark']}; 
    }}
    
    /* תיקון לאייקונים של Streamlit */
    .st-emotion-cache-1p1m4ay, .st-emotion-cache-12fmjuu {{
        font-family: "Source Sans Pro", sans-serif !important;
    }}
    
    /* כרטיס יחידה - רספונסיבי */
    .login-card {{
        background: white; 
        border-radius: 16px; 
        padding: 20px; 
        text-align: center; 
        border-top: 5px solid {COLORS['primary']};
        box-shadow: 0 8px 16px rgba(0,0,0,0.08); 
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1); 
        cursor: pointer; 
        min-height: 180px;
        display: flex; 
        flex-direction: column; 
        align-items: center; 
        justify-content: center;
        margin-bottom: 20px;
        position: relative;
        overflow: hidden;
    }}
    
    .login-card::before {{
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        height: 5px;
        background: linear-gradient(90deg, {COLORS['primary']}, {COLORS['secondary']});
        transform: scaleX(0);
        transition: transform 0.3s ease;
    }}
    
    .login-card:hover {{
        transform: translateY(-8px) scale(1.02); 
        box-shadow: 0 20px 40px rgba(30, 58, 138, 0.15); 
        border-color: {COLORS['secondary']}; 
    }}
    
    .login-card:hover::before {{
        transform: scaleX(1);
    }}
    
    .login-card img {{
        max-height: 90px !important;
        max-width: 100% !important;
        width: auto !important;
        height: auto !important;
        object-fit: contain;
        margin-bottom: 12px;
        filter: drop-shadow(0 2px 4px rgba(0,0,0,0.1));
        transition: transform 0.3s ease;
    }}
    
    .login-card:hover img {{
        transform: scale(1.1);
    }}
    
    .login-card h3 {{
        font-size: 1.1rem;
        margin: 8px 0 0 0;
        font-weight: 700;
        color: {COLORS['primary']};
        line-height: 1.4;
    }}
    
    /* כפתורים משופרים */
    div.stButton > button {{ 
        width: 100%; 
        border-radius: 12px; 
        font-weight: 700; 
        border: none; 
        padding: 0.75rem 1.5rem; 
        box-shadow: 0 4px 12px rgba(30, 58, 138, 0.2); 
        transition: all 0.3s ease;
        background: linear-gradient(135deg, {COLORS['primary']}, {COLORS['secondary']});
        color: white;
        font-size: 1rem;
    }}
    
    div.stButton > button:hover {{
        transform: translateY(-2px);
        box-shadow: 0 8px 20px rgba(30, 58, 138, 0.3);
    }}
    
    /* כותרות */
    h1, h2, h3 {{ 
        color: {COLORS['primary']}; 
        font-weight: 800; 
        text-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }}
    
    h1 {{
        font-size: clamp(1.8rem, 5vw, 3rem);
        margin-bottom: 0.5rem;
    }}
    
    /* כרטיסי סטטוס יחידות */
    .unit-status-card {{
        background: white;
        padding: 15px;
        border-radius: 12px;
        border-top: 4px solid {COLORS['primary']};
        text-align: center;
        margin-bottom: 15px;
        box-shadow: 0 4px 12px rgba(0,0,0,0.06);
        transition: all 0.3s ease;
    }}
    
    .unit-status-card:hover {{
        transform: translateY(-4px);
        box-shadow: 0 8px 20px rgba(0,0,0,0.1);
    }}
    
    .unit-status-card img {{
        max-height: 50px;
        margin-bottom: 8px;
        filter: drop-shadow(0 2px 4px rgba(0,0,0,0.1));
    }}
    
    /* רספונסיביות למובייל */
    @media (max-width: 768px) {{
        .login-card {{
            min-height: 160px;
            padding: 15px;
            margin-bottom: 15px;
        }}
        
        .login-card img {{
            max-height: 70px !important;
        }}
        
        .login-card h3 {{
            font-size: 0.95rem;
        }}
        
        h1 {{
            font-size: 1.8rem !important;
        }}
        
        div.stButton > button {{
            padding: 0.6rem 1rem;
            font-size: 0.95rem;
        }}
        
        .unit-status-card {{
            padding: 12px;
        }}
        
        .unit-status-card img {{
            max-height: 40px;
        }}
    }}
    
    /* אנימציות */
    @keyframes fadeIn {{
        from {{ opacity: 0; transform: translateY(20px); }}
        to {{ opacity: 1; transform: translateY(0); }}
    }}
    
    .login-card {{
        animation: fadeIn 0.5s ease-out;
    }}
    
    /* שיפור טפסים */
    .stTextInput > div > div > input,
    .stSelectbox > div > div > select {{
        border-radius: 8px;
        border: 2px solid #e2e8f0;
        transition: all 0.3s ease;
    }}
    
    .stTextInput > div > div > input:focus,
    .stSelectbox > div > div > select:focus {{
        border-color: {COLORS['primary']};
        box-shadow: 0 0 0 3px rgba(30, 58, 138, 0.1);
    }}
    
    /* Expander משופר */
    .streamlit-expanderHeader {{
        background: white;
        border-radius: 12px;
        border-left: 4px solid {COLORS['warning']};
        font-weight: 700;
    }}
</style>
""", unsafe_allow_html=True)

# --- 7. State ---
if "logged_in" not in st.session_state: st.session_state.logged_in = False
if "role" not in st.session_state: st.session_state.role = "hatmar"
if "selected_unit" not in st.session_state: st.session_state.selected_unit = None
if "login_stage" not in st.session_state: st.session_state.login_stage = "gallery"
# Rate limiting counters (per unit)
if "login_attempts" not in st.session_state: st.session_state.login_attempts = {}
if "login_locked_until" not in st.session_state: st.session_state.login_locked_until = {}

# --- 8. Login Screens (עיצוב מושלם) ---

def render_unit_card(unit_name):
    """פונקציית עזר לציור כרטיס יפה"""
    logo = get_logo_url(unit_name)
    st.markdown(f"""
    <div class="login-card">
        <img src="{logo}" alt="{unit_name}">
        <h3>{unit_name}</h3>
    </div>
    """, unsafe_allow_html=True)
    if st.button(f"כניסה", key=f"btn_{unit_name}", use_container_width=True):
        st.session_state.selected_unit = unit_name
        st.session_state.login_stage = "password"
        st.rerun()

def render_login_gallery():
    st.markdown("<h1 style='text-align: center; margin-bottom: 10px;'>🛡️ מערכת שליטה ובקרה פיקודית</h1>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center; color: gray; margin-bottom: 40px; font-size: 1.1rem;'>בחר יחידה לכניסה מאובטחת</p>", unsafe_allow_html=True)
    
    st.markdown("### 🏔️ חטיבות מרחביות")
    
    # רינדור שורות (4 עמודות בכל שורה) - מבטיח סדר גם במובייל
    for i in range(0, len(REGIONAL_UNITS), 4):
        row_units = REGIONAL_UNITS[i:i+4]
        cols = st.columns(4)
        for j, unit in enumerate(row_units):
            with cols[j]:
                render_unit_card(unit)
            
    st.markdown("---")
    st.markdown("### 🎖️ חטיבות סדירות")
    for i in range(0, len(REGULAR_UNITS), 4):
        row_units = REGULAR_UNITS[i:i+4]
        cols = st.columns(4)
        for j, unit in enumerate(row_units):
            with cols[j]:
                render_unit_card(unit)

    st.markdown("---")
    st.markdown("### 🏛️ מפקדות ושליטה")
    
    c_cols = st.columns(4)
    for i, cmd in enumerate(COMMAND_UNITS):
        with c_cols[i % 4]:
            render_unit_card(cmd)

MAX_LOGIN_ATTEMPTS = 5
LOCKOUT_MINUTES = 15

def render_login_password():
    unit = st.session_state.selected_unit
    c1, c2, c3 = st.columns([1, 2, 1])
    with c2:
        st.markdown(f"""
        <div style='text-align:center; margin-bottom:20px; padding: 30px; background: white; border-radius: 16px; box-shadow: 0 8px 24px rgba(0,0,0,0.1);'>
            <img src='{get_logo_url(unit)}' style='max-height: 120px; object-fit: contain; margin-bottom: 15px; filter: drop-shadow(0 4px 8px rgba(0,0,0,0.1));'>
            <h2 style='margin: 0; color: {COLORS['primary']};'>כניסה ל{unit}</h2>
        </div>
        """, unsafe_allow_html=True)

        # ─── בדיקת נעילה ───
        now = datetime.datetime.now()
        locked_until = st.session_state.login_locked_until.get(unit)
        if locked_until and now < locked_until:
            remaining = int((locked_until - now).total_seconds() / 60) + 1
            st.error(f"🔒 החשבון נעול לאחר {MAX_LOGIN_ATTEMPTS} ניסיונות כושלים. המתן {remaining} דקות.")
            with st.columns([2, 1])[1]:
                if st.button("↩️ חזור", use_container_width=True):
                    st.session_state.login_stage = "gallery"
                    st.rerun()
            return

        attempts_left = MAX_LOGIN_ATTEMPTS - st.session_state.login_attempts.get(unit, 0)
        if attempts_left < MAX_LOGIN_ATTEMPTS:
            st.warning(f"⚠️ נותרו {attempts_left} ניסיונות לפני נעילה")

        password = st.text_input("🔐 הזן סיסמה", type="password", key="pwd_input")

        col_login, col_back = st.columns([2, 1])
        with col_login:
            if st.button("🚀 התחבר", type="primary", use_container_width=True):
                stored = get_stored_password_hash_dummy(unit)
                if verify_password(stored, password):
                    # ✅ כניסה מוצלחת — אפס ניסיונות
                    st.session_state.login_attempts[unit] = 0
                    st.session_state.login_locked_until.pop(unit, None)
                    st.session_state.logged_in = True
                    st.session_state.role = get_user_role(unit)
                    log_audit_event("LOGIN_SUCCESS", unit, severity="info")
                    st.success("✅ התחברות בוצעה בהצלחה!")
                    time.sleep(0.5)
                    st.rerun()
                else:
                    # ❌ ניסיון כושל
                    current = st.session_state.login_attempts.get(unit, 0) + 1
                    st.session_state.login_attempts[unit] = current
                    log_audit_event("LOGIN_FAILURE", unit, details={"attempts": current}, severity="warning")
                    if current >= MAX_LOGIN_ATTEMPTS:
                        st.session_state.login_locked_until[unit] = (
                            datetime.datetime.now() + datetime.timedelta(minutes=LOCKOUT_MINUTES)
                        )
                        st.error(f"🔒 יותר מדי ניסיונות. החשבון נעול ל-{LOCKOUT_MINUTES} דקות.")
                    else:
                        st.error(f"❌ סיסמה שגויה ({current}/{MAX_LOGIN_ATTEMPTS} ניסיונות)")

        with col_back:
            if st.button("↩️ חזור", use_container_width=True):
                st.session_state.login_stage = "gallery"
                st.rerun()

def get_stored_password_hash_dummy(unit):
    """פונקציית עזר קטנה למניעת קריסה אם אין יוזר ב-DB"""
    try:
        res = supabase.table("unit_passwords").select("password").eq("unit_name", unit).execute()
        if res.data: return res.data[0]['password']
    except: pass
    return "INVALID"

# --- 9. Dashboards ---

# ════════════════════════════════════════════════════════════
# SLA Dashboard - מעקב זמן תגובה לחוסרים
# ════════════════════════════════════════════════════════════
def render_sla_dashboard(accessible_units_list: list):
    """מציג progress bars לחוסרים עם מעקב SLA 7 ימים"""
    try:
        result = supabase.table("deficit_tracking").select("*")            .in_("unit", accessible_units_list)            .eq("status", "open")            .execute()
        open_deficits = pd.DataFrame(result.data) if result.data else pd.DataFrame()
    except Exception as e:
        st.warning(f"⚠️ לא ניתן לטעון חוסרים: {e}")
        return

    if open_deficits.empty:
        st.success("✅ אין חוסרים פתוחים כרגע")
        return

    open_deficits = open_deficits.copy()
    open_deficits['days_open'] = (
        pd.Timestamp.now() -
        pd.to_datetime(open_deficits['detected_date'], errors='coerce').dt.tz_localize(None)
    ).dt.days.fillna(0).astype(int)
    open_deficits = open_deficits.sort_values('days_open', ascending=False)

    st.markdown("#### ⏱️ מעקב SLA - זמן עד סגירה")
    SLA_DAYS = 7
    deficit_names = {
        'mezuzot': '📜 מזוזות',
        'eruv_kelim': '🔴 ערבוב כלים',
        'kashrut_cert': '🍽️ כשרות',
        'eruv_status': '🚧 עירוב פסול',
        'shabbat_supervisor': '👤 נאמן שבת'
    }

    for _, deficit in open_deficits.iterrows():
        days = int(deficit['days_open'])
        pct = min(100, int((days / SLA_DAYS) * 100))
        if days > SLA_DAYS:
            bar_color = "#ef4444"
            status = f"⚠️ עבר SLA ב-{days - SLA_DAYS} ימים!"
        elif days >= SLA_DAYS - 1:
            bar_color = "#f59e0b"
            status = "⚡ נותרו פחות מיומיים"
        else:
            bar_color = "#10b981"
            status = f"✅ {SLA_DAYS - days} ימים עד SLA"

        type_label = deficit_names.get(deficit.get('deficit_type', ''), deficit.get('deficit_type', '?'))
        col_info, col_close = st.columns([5, 1])
        with col_info:
            st.markdown(f"""
            <div style='margin-bottom:12px;padding:10px;background:#f8fafc;
                        border-radius:8px;border-right:4px solid {bar_color};'>
                <div style='display:flex;justify-content:space-between;margin-bottom:6px;'>
                    <span><b>{type_label}</b> — {deficit.get('base','?')} ({deficit.get('unit','?')})</span>
                    <span style='color:{bar_color};font-weight:600;'>{status}</span>
                </div>
                <div style='background:#e2e8f0;height:10px;border-radius:5px;'>
                    <div style='width:{pct}%;height:100%;background:{bar_color};
                                border-radius:5px;'></div>
                </div>
                <div style='font-size:12px;color:#64748b;margin-top:4px;'>יום {days} מתוך {SLA_DAYS}</div>
            </div>""", unsafe_allow_html=True)
        with col_close:
            deficit_id = deficit.get('id', '')
            if deficit_id and st.button("✅", key=f"sla_{deficit_id}", use_container_width=True, help="סגור חוסר"):
                try:
                    supabase.table("deficit_tracking").update({"status": "closed"}).eq("id", deficit_id).execute()
                    st.rerun()
                except Exception as e:
                    st.error(f"שגיאה: {e}")
    st.markdown("---")


# ════════════════════════════════════════════════════════════
# AI Chatbot לשאילתות בעברית
# ════════════════════════════════════════════════════════════
def _build_data_context(df: pd.DataFrame, accessible_units: list) -> str:
    """
    🧠 המוח של המערכת (גרסת דיוק כמותי): 
    בונה תמונת מצב מלאה ל-AI תוך הפרדה מבנית בין חוסרים כמותיים לנתונים מזהים.
    """
    if df.empty: return "אין נתונים זמינים במערכת."
        
    # שליפת הדיווח האחרון מכל מוצב למניעת כפילויות מידע
    latest = df.sort_values('date').groupby('base').tail(1)
    
    lines = [
        f"--- 📊 תמונת מצב גזרתית כמותית (מבוסס על {len(latest)} מוצבים) ---",
        f"סה\"כ דוחות במאגר: {len(df)}"
    ]
    
    # 1. דוח חוסרי מזוזות (דיוק כמותי)
    lines.append("\n📜 [חוסרי מזוזות - פירוט ממוקד]:")
    total_mezuzot: int = 0
    mez_lines = []
    if 'r_mezuzot_missing' in latest.columns:
        for _, row in latest.iterrows():
            val = row.get('r_mezuzot_missing', 0)
            try:
                count = int(pd.to_numeric(val, errors='coerce')) if pd.notna(val) else 0
            except: count = 0
            
            if count > 0:
                base_name = row.get('base', 'לא ידוע')
                mez_lines.append(f"- מוצב {base_name}: חסרות {count} מזוזות")
                total_mezuzot += count
    
    lines.extend(mez_lines if mez_lines else ["- אין מזוזות חסרות מדווחות."])
    lines.append(f"💰 סך הכל מזוזות חסרות בכל הגזרה: {total_mezuzot}")

    # 🍴 2. עולם הכשרות והמטבח
    lines.append("\n🍴 [כשרות וטרקלין - חריגות]:")
    k_lines = []
    for _, row in latest.iterrows():
        issues = []
        # כשרות
        if row.get('k_cert') == 'לא': issues.append("חסרה תעודת כשרות")
        if row.get('p_mix') == 'כן': issues.append("ערבוב כלים")
        if row.get('k_bishul') == 'לא': issues.append("בעיה בבישול ישראל")
        # טרקלין
        if row.get('t_friday') == 'לא': issues.append("טרקלין פתוח בשבת")
        if row.get('t_kitchen_tools') == 'לא': issues.append("תקלה בכלי מטבח")
        
        if issues:
            k_lines.append(f"- מוצב {row.get('base', 'לא ידוע')}: {', '.join(issues)}")
    lines.extend(k_lines if k_lines else ["- הכל תקין בתחומי הכשרות והטרקלין."])

    # 📖 3. שיעורי תורה וקשר למרצים
    lines.append("\n📖 [מדור רוח ושיעורי תורה]:")
    for _, row in latest.iterrows():
        if row.get('soldier_has_lesson') == 'כן':
            instructor = row.get('soldier_lesson_teacher', 'לא צוין')
            phone = row.get('soldier_lesson_phone', 'אין טלפון')
            lines.append(f"- מוצב {row.get('base', 'לא ידוע')}: מעביר {instructor}, טל': {phone}")
        elif row.get('soldier_want_lesson') == 'כן':
            lines.append(f"- 💡 מוצב {row.get('base', 'לא ידוע')} מעוניין בשיעור.")

    # 🔴 4. חוסרים פתוחים (כרטיסי טיפול)
    lines.append("\n🔴 [משימות פתוחות בטיפול]:")
    try:
        open_defs = get_open_deficits(accessible_units)
        if not open_defs.empty:
            for _, row in open_defs.iterrows():
                lines.append(f"- {row.get('base', 'לא ידוע')}: {row.get('deficit_type', 'חוסר')} (סטטוס: פתוח)")
        else:
            lines.append("- אין משימות פתוחות.")
    except: pass

    # 👥 5. אמינות מבקרים
    lines.append("\n👥 [ניתוח אמינות]:")
    for insp in df['inspector'].dropna().unique():
        cred = calculate_inspector_credibility(insp, df)
        lines.append(f"- {insp}: ציון {cred['score']:.0f} ({cred['credibility']})")

    return "\u200f" + "\n".join(lines)


def render_ai_chatbot(df: pd.DataFrame, accessible_units: list):
    """צ'טבוט AI לשאלות על הנתונים - מבוסס מודל Gemini"""
    st.markdown("### 🤖 עוזר פיקודי AI (Powered by Gemini)")

    if "chat_history" not in st.session_state:
        st.session_state.chat_history = []

    # תצוגת היסטוריית השיחה
    for msg in st.session_state.chat_history[-6:]:
        role_icon = "👤" if msg["role"] == "user" else "🤖"
        bg = "#f1f5f9" if msg["role"] == "user" else "#eff6ff"
        st.markdown(f"""<div style='background:{bg};padding:10px;border-radius:8px;margin-bottom:6px;direction:rtl;'>
            <b>{role_icon}</b> {msg["content"]}</div>""", unsafe_allow_html=True)

    user_question = st.text_input(
        "שאל שאלה על הנתונים...",
        placeholder='לדוגמה: כמה מוצבים יש עם עירוב פסול? מה המגמה?',
        key="ai_chat_input"
    )
    col_send, col_clear = st.columns([3, 1])
    with col_send:
        send_pressed = st.button("📤 שלח", key="ai_chat_send", use_container_width=True)
    with col_clear:
        if st.button("🗑️ נקה", key="ai_chat_clear", use_container_width=True):
            st.session_state.chat_history = []
            st.rerun()

    if send_pressed and user_question:
        # בדיקת בטיחות מכסה לפני הכל
        if not check_api_quota_safety(daily_limit=500):
            st.error("🚨 הגענו למכסת ה-AI היומית של החטיבה (500 קריאות). המערכת נעולה עד מחר למניעת חיובי יתר.")
            return

        # עדכון מונה ב-Session ורישום ב-DB
        st.session_state.api_call_count += 1
        log_api_usage(st.session_state.get('selected_unit', 'system'), "chat")

        context = _build_data_context(df, accessible_units)
        try:
            import google.generativeai as genai

            # טעינת מפתח ה-API של גוגל
            api_key = st.secrets["gemini"]["api_key"]
            genai.configure(api_key=api_key)

            # הגדרת תפקיד המערכת והזרקת הנתונים
            system_instruction = (
                f"אתה עוזר AI קמב\"ץ רבנות וחוקר נתונים פורנזי. תפקידך לנתח את המידע ולהצביע על כשלים עמוקים.\n"
                f"יש לך גישה לפרטים מזהים כמו שמות מעבירי שיעורים וטלפונים - השתמש בהם כדי לעזור למפקד לסגור מעגלים.\n"
                f"חוקים לניתוח אמינות:\n"
                f"1. השוואה היסטורית: אם חייל מדווח על מוצב מסוים כ'מושלם' בזמן ש-3 דוחות קודמים של מבקרים אחרים הצביעו על ליקויים חמורים - ציין זאת כחשד לחוסר אמינות.\n"
                f"2. השוואה רוחבית: אם חייל מדווח את אותן תוצאות בדיוק (למשל 100% תקין או 100% פסול) ב-5 מוצבים שונים - הוא נחשב למי שעובד על 'טייס אוטומטי' ואינו אמין.\n"
                f"3. ניתוח יסודיות: חייל אמין הוא כזה שזמן המילוי שלו סביר והוא מפרט הערות חופשיות במקרה של תקלות.\n"
                f"4. כשרות וטרקלין: שים לב במיוחד לערבוב כלים, בישול ישראל, וסגירת טרקלין בשבת.\n"
                f"5. שיעורי תורה: ציין שמות של מרצים וטלפונים אם הם מופיעים בנתונים, כדי להקל על התיאום.\n"
                f"\nהמטרה שלך היא לענות בעברית תמציתית, צבאית ומדויקת. אל תסכם מה שהמשתמש רואה בעיניים - תן לו תובנות (Insights) שהוא לא יכול לראות לבד.\n"
                f"\n--- DATA CONTEXT ---\n{context}"
            )

            # המרת היסטוריית השיחה לפורמט שג'מיני דורש
            gemini_history = []
            for m in st.session_state.chat_history[-4:]:
                role = "user" if m["role"] == "user" else "model"
                gemini_history.append({"role": role, "parts": [m["content"]]})

            response = None

            # אתחול המודל - fallback אמיתי: השגיאה נתפסת ברגע השליחה עצמה
            models_to_try = ["gemini-2.5-flash", "gemini-2.0-flash"]

            last_error = ""  # שמירת השגיאה האחרונה מגוגל

            for _model_name in models_to_try:
                try:
                    model = genai.GenerativeModel(
                        model_name=_model_name,
                        system_instruction=system_instruction
                    )
                    chat = model.start_chat(history=gemini_history)
                    # כאן מתבצעת התקשורת בפועל - כאן השגיאה תקפוץ אם המודל חסום
                    response = chat.send_message(user_question)
                    break  # הצלחה - יוצאים מהלולאה
                except Exception as e:
                    last_error = str(e)  # שמירת השגיאה האמיתית מגוגל
                    print(f"⚠️ דילוג על מודל {_model_name}: {e}")
                    continue

            if response is None:
                st.error(f"❌ המודלים נכשלו. השגיאה שהתקבלה מגוגל: {last_error}")
            else:
                st.session_state.chat_history.append({"role": "user", "content": user_question})
                st.session_state.chat_history.append({"role": "assistant", "content": response.text})
                st.rerun()

        except KeyError:
            st.warning("⚠️ לא נמצא מפתח API של Gemini. הוסף [gemini] ו-api_key לקובץ הסודות (secrets.toml).")
        except ImportError:
            st.warning("⚠️ חסרה חבילת google-generativeai. הוסף אותה לקובץ requirements.txt")
        except Exception as e:
            st.error(f"❌ שגיאה: {e}")


def calculate_mri(df: pd.DataFrame) -> pd.DataFrame:
    """חישוב מדד מוכנות מבצעית-רוחנית (MRI) לכל יחידה"""
    if df.empty or 'unit' not in df.columns:
        return pd.DataFrame()

    mri_data = []
    for unit in df['unit'].unique():
        u = df[df['unit'] == unit].copy()
        total = len(u)

        # כשרות: תקלות אינן + תעודות
        k_issues_count = (u.get('k_issues', pd.Series()) == 'כן').sum() if 'k_issues' in u.columns else 0
        k_cert_fail   = (u.get('k_cert',   pd.Series()) == 'לא').sum()  if 'k_cert'   in u.columns else 0

        # עירוב: סטטוס פסול
        eruv_fail = (u.get('e_status', pd.Series()) == 'פסול').sum() if 'e_status' in u.columns else 0

        # סטייט סיכון
        eruv_open = int(u.get('e_check', pd.Series()).isin(['לא']).sum()) if 'e_check' in u.columns else 0

        # חישוב MRI
        penalty = (k_issues_count * 8) + (k_cert_fail * 6) + (eruv_fail * 10) + (eruv_open * 4)
        score = max(35, 100 - penalty)

        # מגמה לפי זמן
        trend = "יציב ➡️"
        if 'date' in u.columns:
            u['date'] = pd.to_datetime(u['date'], errors='coerce')
            recent   = u[u['date'] >= pd.Timestamp.now() - pd.Timedelta(days=7)]
            previous = u[(u['date'] >= pd.Timestamp.now() - pd.Timedelta(days=14)) &
                         (u['date'] <  pd.Timestamp.now() - pd.Timedelta(days=7))]
            if not recent.empty and not previous.empty:
                r_score = max(35, 100 - int((recent.get('k_issues', pd.Series()) == 'כן').sum() * 8))
                p_score = max(35, 100 - int((previous.get('k_issues', pd.Series()) == 'כן').sum() * 8))
                if r_score > p_score + 5:  trend = "עלייה ⬆️"
                elif r_score < p_score - 5: trend = "ירידה ⬇️"

        # אזור סיכון
        risk = "🟢 תקין" if score >= 80 else ("🟡 מעקב" if score >= 60 else "🔴 סיכון")

        mri_data.append({
            "יחידה": unit,
            "ציון MRI": score,
            "מדד %": f"{score}%",
            "סטטוס": risk,
            "מגמה": trend,
            "תקלות כשרות": int(k_issues_count),
            "עירוב פסול": int(eruv_fail),
            "דוחות": int(total),
        })

    mri_df = pd.DataFrame(mri_data)
    if not mri_df.empty:
        mri_df = mri_df.sort_values("ציון MRI").reset_index(drop=True)
    return mri_df


def render_executive_ai_brief(df: pd.DataFrame, accessible_units: list):
    """🧠 מוח פיקודי - MRI + AI Decision Brief"""

    # הזרתת סיכון פועמת לימי חמישי/שישי (CSS Injection)
    current_day = datetime.datetime.now().weekday()
    is_shabbat_risk = current_day in [3, 4]
    if is_shabbat_risk:
        st.markdown("""
        <style>
        @keyframes pulseRed { 0% {box-shadow:0 0 0 0 rgba(220,38,38,.5);} 70% {box-shadow:0 0 0 10px rgba(220,38,38,0);} 100% {box-shadow:0 0 0 0 rgba(220,38,38,0);}}
        .shabbat-alert { animation: pulseRed 2s infinite; border:2px solid #dc2626; border-radius:10px; padding:12px; background:#fef2f2; }
        </style>
        <div class='shabbat-alert'>
        ⚠️ <b>התראה שבתית פעילה</b> אותר במצב מוכנות מוגברת — בדוק העירוב ונאמני כשרות לפני שבת!
        </div>
        """, unsafe_allow_html=True)
        st.markdown("")

    st.markdown("### 🧠 מוח פיקודי — AI Decision Brief")

    # טבלת MRI
    mri_df = calculate_mri(df)
    if mri_df.empty:
        st.info("💭 אין נתונים מספיקים לחישוב MRI.")
        return

    st.markdown("#### 📊 מדד מוכנות מבצעית-רוחנית (MRI)")
    display_cols = ["יחידה", "מדד %", "סטטוס", "מגמה", "תקלות כשרות", "עירוב פסול", "דוחות"]
    st.dataframe(
        mri_df[display_cols],
        use_container_width=True,
        column_config={
            "מדד %": st.column_config.TextColumn("ציון MRI"),
            "סטטוס": st.column_config.TextColumn("סטטוס סיכון"),
        },
        hide_index=True,
    )

    # כפתור להפקת טקציר דיון
    st.markdown("---")
    if st.button("⚡ הפק פקודת יום (AI Executive Brief)", use_container_width=True, type="primary", key="exec_brief_btn"):
        with st.spinner("המוח הפיקודי מנתח את נתוני הגזרה..."):
            # בדיקת בטיחות מכסה
            if not check_api_quota_safety(daily_limit=500):
                st.error("🚨 הגענו למכסת ה-AI היומית. המערכת נעולה למניעת חיובי יתר.")
            else:
                # עדכון מונה ב-Session ורישום ב-DB
                st.session_state.api_call_count += 1
                log_api_usage(st.session_state.get('selected_unit', 'system'), "exec_brief")

                try:
                    import google.generativeai as genai
                    genai.configure(api_key=st.secrets["gemini"]["api_key"])

                    context = mri_df[['יחידה', 'מדד %', 'סטטוס', 'מגמה', 'תקלות כשרות', 'עירוב פסול']].to_string(index=False)

                    prompt = f"""
אתה מערכת השו"ב של רבנות פיקוד מרכז. לפניך מדדי מוכנות (MRI) של יחידות:

{context}

הפק “תקציר החלטות מנהלים” קצר, חד, צבאי וממוקד פעולה בדיוק במבנה הבא:

🎯 לרב חטמ"ר (3 פעולות מידיות לביצוע):
1. []
2. []
3. []

🎯 לרב אוגדה (2 מוקדי סיכון אזוריים):
1. []
2. []

🎯 לרב פיקוד (מגמה אסטרטגית אחת הדורשת הקצאת משאבים):
1. []
כתוב בשפה צבאית-עניינית, ללא הקדמות, והתבסס אך ורק על חומרת הנתונים.
"""
                    response = None
                    for _m in ["gemini-2.5-flash", "gemini-2.0-flash"]:
                        try:
                            model = genai.GenerativeModel(_m)
                            response = model.generate_content(prompt)
                            break
                        except Exception:
                            continue

                    if response:
                        st.success("פקודת יום מוכנה:")
                        st.markdown(response.text)
                    else:
                        st.error("כל מודלי Gemini נכשלו. בדוק מפתח API.")
                except Exception as e:
                    st.error(f"שגיאה אסטרטגית: {e}")

    # 🆕 חיזוי וסיכונים עתידיים (Predictive)
    st.markdown("---")
    st.markdown("#### 🔮 מנוע חיזוי סיכונים (שבוע קרוב)")
    mri_df = calculate_mri(df)
    c1, c2 = st.columns(2)
    with c1:
        unit_to_predict = st.selectbox("בחר יחידה לחיזוי:", ["כלל היחידות"] + list(mri_df["יחידה"].unique()))
        if unit_to_predict != "כלל היחידות":
            risk = predict_unit_risk_next_week(unit_to_predict, df)
            if risk.get("level") == "high":
                st.error(f"⚠️ {risk['prediction']}")
            else:
                st.success(f"✅ {risk['prediction']}")
    
    with c2:
        st.markdown("##### 🏹 סדר עדיפויות חכם (AI Priority Queue)")
        if st.toggle("הצג תור עדיפויות", value=False):
            with st.spinner("מדרג משימות..."):
                queue = generate_smart_priority_queue(df, accessible_units)
                for item in queue:
                    st.markdown(f"**{item.get('priority', '?')}. {item.get('issue', '???')}**")
                    st.caption(f"💡 סיבה: {item.get('reason', '')}")

    # 🆕 חיפוש סמנטי (Semantic Search)
    st.markdown("---")
    st.markdown("#### 🔍 חיפוש סמנטי חכם (AI Search)")
    search_q = st.text_input("חפש בדוחות (לדוגמה: 'מוצבים עם תקלות כשרות חוזרות', 'דיווחים על עירוב פסול השבוע'):")
    if search_q:
        with st.spinner("ה-AI סורק את הדוחות..."):
            results = semantic_search_reports(search_q, df)
            if not results.empty:
                st.dataframe(results[['unit', 'base', 'date', 'inspector', 'k_issues', 'e_status']], use_container_width=True)
            else:
                st.info("לא נמצאו תוצאות רלוונטיות.")



def analyze_report_with_ai(base_name: str, report_data: dict) -> dict:
    """המוח הפיקודי: מנתח דיווחי שטח בזמן אמת ומחזיר סיווג AI"""
    # בנה סיכום קריא מהנתונים
    parts = []
    field_labels = {
        "k_issues": "תקלות כשרות", "k_issues_description": "תיאור תקלה",
        "e_status": "סטטוס עירוב", "e_check": "בדיקת עירוב",
        "s_status": "מצב חשמל שבת", "missing_items": "חוסרים",
        "notes": "הערות", "general_notes": "הערות כלליות",
    }
    for key, label in field_labels.items():
        val = report_data.get(key)
        if val and str(val).strip() and str(val) not in ("None", "תקין", "כן"):
            parts.append(f"{label}: {val}")
    report_text = "; ".join(parts) if parts else "לא צוינו ליקויים"

    prompt = f"""אתה מערכת שו"ב חכמה של רבנות פיקוד מרכז.
התקבל דיווח מהשטח (מוצב/בסיס: {base_name}):
"{report_text}"

נתח את הדיווח והחזר JSON חוקי בלבד עם 3 שדות:
1. "risk_level": חומרת התקלה (נמוכה / בינונית / גבוהה)
2. "sla": זמן יעד לטיפול (72 שעות / 48 שעות / מיידי)
3. "recommended_action": משפט אחד קצר ותכליתי של פעולה מומלצת

החזר רק JSON נקי, ללא טקסט לפני או אחריו."""

    try:
        # בדיקת מכסה
        if not check_api_quota_safety(daily_limit=500):
            return {"risk_level": "לא סווג", "sla": "טרם נקבע", "recommended_action": "מכסת AI הסתיימה"}
        
        # עדכון מונה ורישום
        if "api_call_count" in st.session_state:
            st.session_state.api_call_count += 1
        log_api_usage(st.session_state.get('selected_unit', 'system'), "auto_analysis")

        import google.generativeai as genai
        import json as _json
        genai.configure(api_key=st.secrets["gemini"]["api_key"])
        response = None
        for _m in ["gemini-2.5-flash", "gemini-2.0-flash"]:
            try:
                model = genai.GenerativeModel(_m)
                response = model.generate_content(prompt)
                break
            except Exception:
                continue
        if response:
            clean = response.text.strip().replace("```json", "").replace("```", "").strip()
            return _json.loads(clean)
    except Exception:
        pass
    # ברירת מחדל בטוחה — אף פעם לא חוסם שמירה
    return {"risk_level": "לא סווג", "sla": "טרם נקבע", "recommended_action": "נדרשת בחינה ידנית"}


def analyze_photo_with_vision(image_bytes: bytes) -> dict:
    """שימוש ב-Gemini Vision לניתוח תמונות מהשטח"""
    # בדיקת בטיחות מכסה
    if not check_api_quota_safety(daily_limit=500):
        return {"error": "Quota exceeded"}
    
    # עדכון מונה ב-Session ורישום ב-DB
    if "api_call_count" in st.session_state:
        st.session_state.api_call_count += 1
    log_api_usage(st.session_state.get('selected_unit', 'system'), "vision")

    import google.generativeai as genai
    import base64
    import json as _json

    try:
        genai.configure(api_key=st.secrets["gemini"]["api_key"])
        model = genai.GenerativeModel("gemini-2.5-flash")
        image_data = base64.b64encode(image_bytes).decode()

        prompt = """
        אתה מומחה רבנות צבאית (מומחה כשרות ועירוב). נתח תמונה זו מהשטח ודווח:
        1. האם רואים בעיות כשרות גלויות? (ערבוב, אוכל לא כשר, חוסר הפרדה)
        2. מצב ניקיון המטבח (1-5)
        3. האם רואים מזוזות / היעדרן במקומות הנדרשים (דלתות)
        4. בעיות עירוב גלויות (חוט קרוע, עמוד עקום)
        5. ציון אמינות הדוח (האם התמונה תואמת את הדיווח המקובל בבסיס?)
        6. הסבר מפורט בעברית על מה שאתה רואה, במיוחד אם יש בעיה.
        
        ענה בפורמט JSON בלבד עם מפתחות באנגלית: kashrut_issues, kitchen_cleanliness, mezuzot, eruv_issues, reliability_score, detailed_finding.
        הסבר ה-detailed_finding חייב להיות בעברית שוטפת ומקצועית.
        """

        response = model.generate_content([
            {"mime_type": "image/jpeg", "data": image_data},
            prompt
        ])
        clean = response.text.strip().replace("```json", "").replace("```", "").strip()
        return _json.loads(clean)
    except Exception:
        return {}


def transcribe_voice_note(audio_bytes: bytes) -> str:
    """תמלול הקלטות קוליות מהשטח באמצעות Gemini"""
    import google.generativeai as genai
    import base64

    try:
        # בדיקת מכסה
        if not check_api_quota_safety(daily_limit=500):
            return "מכסת תמלול הסתיימה"
        
        # עדכון מונה ורישום
        if "api_call_count" in st.session_state:
            st.session_state.api_call_count += 1
        log_api_usage("system", "voice_transcription")

        genai.configure(api_key=st.secrets["gemini"]["api_key"])
        model = genai.GenerativeModel("gemini-2.5-flash")
        audio_b64 = base64.b64encode(audio_bytes).decode()

        response = model.generate_content([
            {"mime_type": "audio/mp3", "data": audio_b64},
            "תמלל את ההקלטה לעברית וחלץ ממנה ממצאי ביקורת כשרות בפורמט טקסט חופשי קצר"
        ])
        return response.text
    except Exception:
        return "שגיאה בתמלול"


def ml_anomaly_detection(df: pd.DataFrame) -> pd.DataFrame:
    """זיהוי דוחות חשודים (אנומליות) באמצעות Isolation Forest"""
    try:
        from sklearn.ensemble import IsolationForest
        import numpy as np

        if df.empty or len(df) < 10:
            return df

        # בחירת מאפיינים לניתוח
        features = []
        for _, row in df.iterrows():
            features.append([
                1 if row.get('k_cert') == 'כן' else 0,
                1 if row.get('e_status') == 'תקין' else 0,
                int(row.get('r_mezuzot_missing', 0) or 0),
                1 if row.get('p_mix') == 'לא' else 0,
            ])

        X = np.array(features)
        clf = IsolationForest(contamination=0.1, random_state=42)
        df['anomaly_score'] = clf.fit_predict(X)
        df['is_suspicious'] = df['anomaly_score'] == -1

        return df
    except Exception:
        return df


def semantic_search_reports(query: str, df: pd.DataFrame) -> pd.DataFrame:
    """חיפוש סמנטי חכם בדוחות באמצעות Gemini"""
    import google.generativeai as genai
    import json as _json

    try:
        if df.empty: return df
        # בדיקת מכסה
        if not check_api_quota_safety(daily_limit=500):
            return df.head(0)
            
        # עדכון מונה ורישום
        if "api_call_count" in st.session_state:
            st.session_state.api_call_count += 1
        log_api_usage("system", "semantic_search")

        genai.configure(api_key=st.secrets["gemini"]["api_key"])

        # תקציר דוחות (50 אחרונים)
        summary = df[['unit', 'base', 'date', 'k_issues', 'e_status']].tail(50).to_string()

        prompt = f"""
        שאילתה: {query}
        נתונים: {summary}
        החזר רשימת אינדקסים (indices) של השורות הרלוונטיות ביותר מתוך הנתונים לעיל.
        ענה בפורמט JSON בלבד: {{"relevant_indices": [0, 5, 23]}}
        """

        model = genai.GenerativeModel("gemini-2.5-flash")
        response = model.generate_content(prompt)
        clean = response.text.strip().replace("```json", "").replace("```", "").strip()
        result = _json.loads(clean)
        indices = result.get('relevant_indices', [])
        return df.iloc[indices] if indices else df.head(0)
    except Exception:
        return df.head(0)


def predict_unit_risk_next_week(unit: str, df: pd.DataFrame) -> dict:
    """ניבוי סיכונים לשבוע הקרוב מבוסס מגמות"""
    unit_df = df[df['unit'] == unit].copy()
    if len(unit_df) < 5:
        return {"prediction": "אין מספיק נתונים לחיזוי מהימן"}

    if 'date' not in unit_df.columns:
        return {"prediction": "חסר שדה תאריך"}

    unit_df['date'] = pd.to_datetime(unit_df['date'], errors='coerce')
    unit_df['day_of_week'] = unit_df['date'].dt.dayofweek

    # בדיקת דפוס חזרתי בשישי
    friday_issues = unit_df[unit_df['day_of_week'] == 4]
    if friday_issues.empty:
        return {"prediction": "לא זוהה דפוס סיכון שבועי מובהק"}

    eruv_fail_rate = (friday_issues['e_status'] == 'פסול').mean() if 'e_status' in friday_issues.columns else 0

    if eruv_fail_rate > 0.3:
        return {
            "prediction": f"סיכון גבוה ({eruv_fail_rate*100:.0f}%) לתקלת עירוב בשישי הקרוב",
            "level": "high",
            "type": "Eruv"
        }
    return {"prediction": "המערכת חוזה יציבות בשבוע הקרוב", "level": "low"}


def generate_smart_priority_queue(df: pd.DataFrame, accessible_units: list) -> list:
    """יצירת תור עדיפויות חכם לטיפול בליקויים"""
    import google.generativeai as genai
    import json as _json

    issues = []
    for unit in accessible_units:
        u_df = df[df['unit'] == unit] if not df.empty else pd.DataFrame()
        if not u_df.empty:
            latest = u_df.sort_values('date').iloc[-1]
            if latest.get('e_status') == 'פסול':
                issues.append(f"{unit}: עירוב פסול במוצב {latest.get('base')}")
            if latest.get('k_cert') == 'לא':
                issues.append(f"{unit}: חסרה תעודת כשרות ב{latest.get('base')}")
            if latest.get('k_issues') == 'כן':
                issues.append(f"{unit}: תקלת כשרות דווחה ב{latest.get('base')}")

    if not issues: return []

    try:
        # בדיקת בטיחות מכסה
        if not check_api_quota_safety(daily_limit=500):
            return [{"priority": i+1, "issue": iss} for i, iss in enumerate(issues[:5])]
        
        # עדכון מונה ב-Session ורישום ב-DB
        if "api_call_count" in st.session_state:
            st.session_state.api_call_count += 1
        log_api_usage(st.session_state.get('selected_unit', 'system'), "priority_queue")

        genai.configure(api_key=st.secrets["gemini"]["api_key"])
        model = genai.GenerativeModel("gemini-2.5-flash")
        prompt = f"""דרג את רשימת הליקויים הבאה לפי דחיפות טיפול (1 - הכי דחוף).
        שקול: השפעה על כשרות החיילים והלכה. החזר JSON: [{{"priority": 1, "issue": "...", "reason": "..."}}]
        ליקויים: {chr(10).join(issues[:10])}"""

        response = model.generate_content(prompt)
        clean = response.text.strip().replace("```json", "").replace("```", "").strip()
        return _json.loads(clean)
    except Exception:
        return [{"priority": i+1, "issue": iss} for i, iss in enumerate(issues[:5])]


# Consolidated with logic at line 1599


def count_unvisited_bases_this_week(df: pd.DataFrame, unit: str) -> int:
    """ספירת מוצבים שלא בוקרו ב-7 הימים האחרונים"""
    import datetime
    if df.empty or 'base' not in df.columns: return 0
    
    unit_df = df[df['unit'] == unit]
    bases = unit_df['base'].unique()
    
    unvisited = 0
    now = datetime.datetime.now()
    week_ago = now - datetime.timedelta(days=7)
    
    for base in bases:
        base_df = unit_df[unit_df['base'] == base]
        latest_visit = pd.to_datetime(base_df['date']).max()
        if pd.isna(latest_visit) or latest_visit < week_ago:
            unvisited += 1
    return unvisited


def render_bases_status_board(df: pd.DataFrame, unit: str):
    """לוח ירוק/צהוב/אדום לכל מוצב - פשוט ומהיר לקריאה"""
    st.markdown("### 🗺️ לוח מצב מוצבים (רמזור)")
    if df.empty or 'base' not in df.columns:
        st.info("אין מוצבים רשומים בחטיבה")
        return
    
    bases = df[df['unit'] == unit]['base'].unique() if unit else df['base'].unique()
    cols = st.columns(min(len(bases), 3) if len(bases) > 0 else 1)
    
    for i, base in enumerate(bases):
        base_df = df[df['base'] == base].sort_values('date', ascending=False)
        latest = base_df.iloc[0] if not base_df.empty else None
        if latest is None: continue
            
        try:
            last_date = pd.to_datetime(latest['date'])
            days_ago = (pd.Timestamp.now() - last_date).days
        except Exception: days_ago = 99
        
        has_eruv_fail = latest.get('e_status') == 'פסול'
        has_kashrut_fail = latest.get('k_cert') == 'לא'
        is_old = days_ago > 7
        
        if has_eruv_fail or has_kashrut_fail:
            border_color, status_icon, status_text, bg_color = "#dc2626", "🔴", "דורש טיפול", "#fee2e2"
        elif is_old:
            border_color, status_icon, status_text, bg_color = "#f59e0b", "🟡", f"לא בוקר {days_ago} ימים", "#fef3c7"
        else:
            border_color, status_icon, status_text, bg_color = "#10b981", "🟢", "תקין", "#d1fae5"
        
        with cols[i % 3]:
            st.markdown(f"""
            <div style='background:{bg_color};border-right:5px solid {border_color};
                        padding:14px;border-radius:10px;margin-bottom:12px;min-height:120px;'>
                <div style='font-size:18px;font-weight:800;'>{status_icon} {base}</div>
                <div style='color:#374151;font-size:13px;margin-top:6px;'>
                    📅 לפני {days_ago} ימים<br/>
                    👤 {latest.get('inspector','?')}<br/>
                    <b>{status_text}</b>
                </div>
            </div>
            """, unsafe_allow_html=True)


def render_shabbat_preparation_assistant(unit: str, df: pd.DataFrame):
    """עוזר הכנה לשבת (AI) - מייצר לו"ז משימות ליום שישי"""
    import google.generativeai as genai
    st.markdown("### 🕯️ עוזר הכנה לשבת (AI)")
    
    open_issues = []
    if not df.empty:
        latest_by_base = df[df['unit'] == unit].sort_values('date').groupby('base').tail(1)
        for _, row in latest_by_base.iterrows():
            if row.get('e_status') == 'פסול': open_issues.append(f"עירוב פסול ב{row['base']}")
            if row.get('k_cert') == 'לא': open_issues.append(f"כשרות חסרה ב{row['base']}")

    if st.button("⚡ ג'נרט רשימת משימות ליום שישי", type="primary", use_container_width=True):
        # בדיקת בטיחות מכסה
        if not check_api_quota_safety(daily_limit=500):
            st.error("🚨 הגענו למכסת ה-AI היומית.")
        else:
            with st.spinner("ה-AI בונה עבורך תכנית עבודה..."):
                # עדכון מונה ורישום
                st.session_state.api_call_count += 1
                log_api_usage(unit, "shabbat_advisor")

                try:
                    genai.configure(api_key=st.secrets["gemini"]["api_key"])
                    model = genai.GenerativeModel("gemini-2.5-flash")
                    context = f"רב חטמ״ר {unit}. בעיות: {', '.join(open_issues) if open_issues else 'אין'}"
                    response = model.generate_content(f"צור לו\"ז משימות צבאי ליום שישי לרב חטמ״ר: {context}. פורמט: שעה - משימה.")
                    st.info(response.text)
                except Exception as e:
                    st.error(f"שגיאה: {e}")


def render_halachic_advisor():
    """יועץ הלכתי מהיר (AI) לשאלות מהשטח"""
    import google.generativeai as genai
    st.markdown("### 📖 יועץ הלכתי מהיר (AI)")
    q = st.text_input("שאלת הלכה מהשטח (לדוגמה: עירוב שנקרע בשישי):")
    if q and st.button("💡 קבל מענה ראשוני"):
        # בדיקת מכסה
        if not check_api_quota_safety(daily_limit=500):
            st.error("🚨 הגענו למכסת ה-AI היומית.")
        else:
            with st.spinner("מעיין במקורות..."):
                # עדכון מונה ורישום
                st.session_state.api_call_count += 1
                log_api_usage("halacha", "halacha_advisor")

                try:
                    genai.configure(api_key=st.secrets["gemini"]["api_key"])
                    model = genai.GenerativeModel("gemini-2.5-flash")
                    res = model.generate_content(f"ענה בקצרה כרב צבאי על שאלת הלכה: {q}. ציין מקור והדגש שזו תשובה ראשונית בלבד.")
                    st.success(res.text)
                except Exception as e:
                    st.error(f"שגיאה: {e}")


def render_inspector_management(unit: str, df: pd.DataFrame):
    """ניהול מבקרים ואמינות דוחות - תצוגה משופרת"""
    st.markdown("### 👥 ניהול מבקרים ואמינות")
    if df.empty: return
    
    inspectors = df[df['unit'] == unit]['inspector'].unique() if unit else df['inspector'].unique()
    
    for insp in sorted(inspectors):
        cred = calculate_inspector_credibility(insp, df)
        col1, col2, col3, col4 = st.columns([2, 1, 1, 3])
        with col1:
            st.write(f"**{insp}**")
        with col2:
            st.metric("ציון", f"{cred['score']:.0f}")
        with col3:
            st.write(f"{cred['defect_rate']:.0f}% ליקויים")
        with col4:
            st.markdown(f"<span style='color:{cred['color']}; font-weight:bold;'>{cred['credibility']}</span>", unsafe_allow_html=True)
        st.divider()


def render_weekly_report_generator(unit: str, df: pd.DataFrame):
    """מחולל דוח שבועי (AI) מוכן לשליחה לאוגדה"""
    import google.generativeai as genai
    st.markdown("### 📤 מחולל דוח שבועי לאוגדה")
    if st.button("🤖 נסח דוח שבועי (שפה צבאית)", use_container_width=True):
        # בדיקת מכסה
        if not check_api_quota_safety(daily_limit=500):
            st.error("🚨 הגענו למכסת ה-AI היומית.")
        else:
            with st.spinner("מנסח דוח רשמי..."):
                # עדכון מונה ורישום
                st.session_state.api_call_count += 1
                log_api_usage(unit, "weekly_report")

                try:
                    genai.configure(api_key=st.secrets["gemini"]["api_key"])
                    model = genai.GenerativeModel("gemini-1.5-flash")
                    summary = df[df['unit'] == unit].tail(10).to_string() if not df.empty else "אין נתונים"
                    res = model.generate_content(f"כתוב דוח שבועי רשמי בשפה צבאית לרב האוגדה על חטיבת {unit} לפי הנתונים: {summary}")
                    st.code(res.text, language="markdown")
                except Exception as e:
                    st.error(f"שגיאה: {e}")


def render_detailed_unit_analysis(df, selected_unit):
    """מציג ניתוח מפורט ליחידה ספציפית (חלק מכמה דשבורדים)"""
    if df.empty or 'unit' not in df.columns:
        return
    
    unit_df = df[df['unit'] == selected_unit]
    
    if len(unit_df) > 0:
        # ציון ותג
        score = calculate_unit_score(unit_df)
        badge, color = get_unit_badge(score)
        
        # זיהוי חריגות ודפוסים חשודים (Anomaly Detection)
        anomalies = detect_anomalies(df, selected_unit)
        if anomalies:
            st.warning(f"⚠️ זוהו {len(anomalies)} דפוסים חריגים ביחידה זו")
            with st.expander("🚨 פירוט חריגות ודפוסים חשודים", expanded=True):
                for anomaly in anomalies:
                    severity_map = {'high': '🔴', 'medium': '🟠', 'low': '🔵'}
                    icon = severity_map.get(anomaly['severity'], '⚪')
                    st.markdown(f"**{icon} {anomaly['message']}**")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("ציון כללי", f"{score:.1f}/100")
        with col2:
            st.metric("סה״כ דוחות", len(unit_df))
        with col3:
            st.markdown(f"<div style='background:{color}; color:white; padding:15px; border-radius:10px; text-align:center; font-weight:700; font-size:1.1rem;'>{badge}</div>", unsafe_allow_html=True)
        
        st.markdown("---")
        
        # פרטי שאלון מפורטים
        st.markdown("### 📋 פירוט שאלון ביקורת")
        
        # קבלת הדוח האחרון והקודם לו למעקב שינויים
        latest_report = unit_df.sort_values('date', ascending=False).iloc[0] if len(unit_df) > 0 else None
        previous_report = unit_df.sort_values('date', ascending=False).iloc[1] if len(unit_df) > 1 else None
        
        # טאבים לקטגוריות שונות
        detail_tabs = st.tabs(["🔴 חוסרים ובעיות", "🍴 עירוב וכשרות", "🏗️ תשתיות ויומן ביקורת", "📊 סיכום כללי"])
        
        with detail_tabs[0]:  # חוסרים
            st.markdown("#### חוסרים שדווחו")
            col1, col2 = st.columns(2)
            
            with col1:
                # מזוזות
                mezuzot_missing = int(latest_report.get('r_mezuzot_missing', 0)) if latest_report is not None else 0
                prev_mezuzot = int(previous_report.get('r_mezuzot_missing', 0)) if previous_report is not None else mezuzot_missing
                
                if mezuzot_missing > 0:
                    if mezuzot_missing < prev_mezuzot:
                        diff = prev_mezuzot - mezuzot_missing
                        pct = (diff / prev_mezuzot * 100) if prev_mezuzot > 0 else 0
                        st.metric("📜 מזוזות חסרות", mezuzot_missing, f"-{diff} ({pct:.0f}%)", delta_color="inverse")
                    elif mezuzot_missing > prev_mezuzot:
                        diff = mezuzot_missing - prev_mezuzot
                        pct = (diff / prev_mezuzot * 100) if prev_mezuzot > 0 else 0
                        st.metric("📜 מזוזות חסרות", mezuzot_missing, f"+{diff} ({pct:.0f}%)")
                    else:
                        st.metric("📜 מזוזות חסרות", mezuzot_missing, "ללא שינוי")
                else:
                    st.metric("📜 מזוזות חסרות", "0 🟢", "תקין")
                
                # ספרי תורה
                torah_missing = int(latest_report.get('r_torah_missing', 0)) if latest_report is not None else 0
                if torah_missing > 0:
                    st.metric("📖 ספרי תורה חסרים", torah_missing, delta_color="inverse")
                else:
                    st.metric("📖 ספרי תורה", "תקין 🟢")
            
            with col2:
                # ציצית
                tzitzit_missing = int(latest_report.get('r_tzitzit_missing', 0)) if latest_report is not None else 0
                if tzitzit_missing > 0:
                    st.metric("🧵 ציציות חסרות", tzitzit_missing, delta_color="inverse")
                else:
                    st.metric("🧵 ציציות", "תקין 🟢")
                
                # תפילין
                tefillin_missing = int(latest_report.get('r_tefillin_missing', 0)) if latest_report is not None else 0
                if tefillin_missing > 0:
                    st.metric("📿 תפילין חסרים", tefillin_missing, delta_color="inverse")
                else:
                    st.metric("📿 תפילין", "תקין 🟢")
        
        with detail_tabs[1]:  # עירוב וכשרות
            st.markdown("#### סטטוס עירוב וכשרות")
            col1, col2 = st.columns(2)
            with col1:
                eruv_status = latest_report.get('e_status', 'לא ידוע') if latest_report is not None else 'לא ידוע'
                if eruv_status == 'תקין':
                    st.success(f"✅ **סטטוס עירוב:** {eruv_status}")
                elif eruv_status == 'פסול':
                    st.error(f"❌ **סטטוס עירוב:** {eruv_status}")
                else:
                    st.warning(f"⚠️ **סטטוס עירוב:** {eruv_status}")
                
                eruv_kelim = latest_report.get('k_eruv_kelim', 'לא') if latest_report is not None else 'לא'
                if eruv_kelim == 'כן':
                    st.error("🔴 **עירוב כלים:** קיים - דורש טיפול")
                else:
                    st.success("🟢 **עירוב כלים:** לא קיים")
            
            with col2:
                k_cert = latest_report.get('k_cert', 'לא') if latest_report is not None else 'לא'
                if k_cert == 'כן':
                    st.success("✅ **תעודת כשרות:** קיימת")
                else:
                    st.warning("⚠️ **תעודת כשרות:** חסרה")
                
                traklin_closed = latest_report.get('k_traklin_closed', 'לא') if latest_report is not None else 'לא'
                if traklin_closed == 'כן':
                    st.success("✅ **סגירת טרקלין:** מבוצעת")
                else:
                    st.warning("⚠️ **סגירת טרקלין:** לא מבוצעת")
        
        with detail_tabs[2]:  # תשתיות
            st.markdown("#### תשתיות ויומן ביקורת")
            col1, col2 = st.columns(2)
            with col1:
                pikubok = latest_report.get('k_pikubok', 'לא') if latest_report is not None else 'לא'
                if pikubok == 'כן':
                    st.success("✅ **יומן ביקורת:** קיים")
                else:
                    st.warning("⚠️ **יומן ביקורת:** לא קיים")
            with col2:
                notes = latest_report.get('notes', '') if latest_report is not None else ''
                if notes and notes.strip():
                    st.text_area("📝 הערות והמלצות", notes, height=100, disabled=True)
                else:
                    st.info("אין הערות נוספות")
        
        with detail_tabs[3]:  # סיכום
            st.markdown("#### סיכום מצב היחידה")
            st.info("כאן מופיע סיכום המדדים של היחידה.")
            
        st.markdown("---")
        # טבלה מפורטת
        st.markdown("#### 📋 דוחות מפורטים - תצוגה מלאה")
        display_cols = ['date', 'base', 'inspector']
        st.dataframe(unit_df[[c for c in display_cols if c in unit_df.columns]], use_container_width=True, hide_index=True)

def render_unit_map(df, unit):
    if df.empty or 'unit' not in df.columns:
        st.info("אין נתונים להצגה במפה")
        return

    unit_df = df[df['unit'] == unit].copy()
    if unit_df.empty:
        st.warning("⚠️ לא נמצאו נתונים ליחידה זו.")
        return

    # הוספת קואורדינטות מ-BASE_COORDINATES לפי שם מוצב
    def get_coords(base_name):
        coords = BASE_COORDINATES.get(str(base_name), None)
        return coords if coords else (None, None)

    unit_df[['latitude', 'longitude']] = unit_df['base'].apply(
        lambda b: pd.Series(get_coords(b))
    )
    unit_df = unit_df.dropna(subset=['latitude', 'longitude'])

    if unit_df.empty:
        st.warning("⚠️ לא נמצאו נתוני מיקום ליחידה זו.")
        return

    center_lat = unit_df['latitude'].mean()
    center_lon = unit_df['longitude'].mean()
    unit_color_map = {unit: "#1e3a8a"}
    m = create_street_level_map(center=(center_lat, center_lon), zoom_start=11)
    for _, row in unit_df.iterrows():
        add_unit_marker_to_folium(m, row, unit_color_map)
    st_folium(m, width=800, height=500, key=f"unit_map_{unit}")


def render_hatmar_management_tab(df: pd.DataFrame, unit: str):
    """
    טאב ניהול מרכזי לרב חטמ"ר — 4 חלקים:
    1. KPI שבועי
    2. תור עבודה (מה בוער)
    3. כרטיסי מבקרים
    4. מעקב סגירת ממצאים
    """

    # ══════════════════════════════════════════
    # חלק 1 — KPI שבועי (4 מספרים בולטים)
    # ══════════════════════════════════════════

    st.markdown("### 📊 מצב שבועי")

    if df.empty or "unit" not in df.columns:
        st.info("אין נתונים להצגת מדדי ניהול")
        return

    today = pd.Timestamp.now()
    week_ago = today - pd.Timedelta(days=7)
    two_weeks_ago = today - pd.Timedelta(days=14)

    df["date_dt"] = pd.to_datetime(df["date"], errors="coerce")
    df_unit = df[df["unit"] == unit].copy()
    
    # חישוב אמינות ממוצעת לפי האלגוריתם המתקדם
    inspectors = df_unit["inspector"].dropna().unique()
    if len(inspectors) > 0:
        rel_scores = [calculate_inspector_credibility(i, df_unit)['score'] for i in inspectors]
        avg_reliability = sum(rel_scores) / len(rel_scores)
    else:
        avg_reliability = None

    total_bases = df_unit["base"].nunique()
    visited_this_week = df_unit[df_unit["date_dt"] >= week_ago]["base"].nunique()
    not_visited_14d = df_unit.groupby("base")["date_dt"].max()
    overdue_bases = (not_visited_14d < two_weeks_ago).sum()

    c1, c2, c3, c4 = st.columns(4)

    with c1:
        pct = int(visited_this_week / total_bases * 100) if total_bases else 0
        color = "#10b981" if pct >= 80 else "#f59e0b" if pct >= 50 else "#ef4444"
        st.markdown(f"""
        <div style='background:white; border-radius:12px; padding:16px; text-align:center;
                    border-top:4px solid {color}; box-shadow:0 2px 8px rgba(0,0,0,0.08);'>
            <div style='font-size:36px; font-weight:800; color:{color};'>{pct}%</div>
            <div style='font-size:13px; color:#64748b; margin-top:4px;'>כיסוי שבועי</div>
            <div style='font-size:11px; color:#94a3b8;'>{visited_this_week}/{total_bases} מוצבים</div>
        </div>
        """, unsafe_allow_html=True)

    with c2:
        color = "#ef4444" if overdue_bases > 0 else "#10b981"
        st.markdown(f"""
        <div style='background:white; border-radius:12px; padding:16px; text-align:center;
                    border-top:4px solid {color}; box-shadow:0 2px 8px rgba(0,0,0,0.08);'>
            <div style='font-size:36px; font-weight:800; color:{color};'>{overdue_bases}</div>
            <div style='font-size:13px; color:#64748b; margin-top:4px;'>מוצבים פגי תוקף</div>
            <div style='font-size:11px; color:#94a3b8;'>לא בוקרו 14+ יום</div>
        </div>
        """, unsafe_allow_html=True)

    with c3:
        open_findings = len(df_unit[
            (df_unit.get("e_status", pd.Series()) == "פסול") |
            (df_unit.get("k_cert", pd.Series()) == "לא")
        ]) if not df_unit.empty else 0
        color = "#ef4444" if open_findings > 3 else "#f59e0b" if open_findings > 0 else "#10b981"
        st.markdown(f"""
        <div style='background:white; border-radius:12px; padding:16px; text-align:center;
                    border-top:4px solid {color}; box-shadow:0 2px 8px rgba(0,0,0,0.08);'>
            <div style='font-size:36px; font-weight:800; color:{color};'>{open_findings}</div>
            <div style='font-size:13px; color:#64748b; margin-top:4px;'>ממצאים פתוחים</div>
            <div style='font-size:11px; color:#94a3b8;'>עירוב/כשרות חריג</div>
        </div>
        """, unsafe_allow_html=True)

    with c4:
        rel = f"{avg_reliability:.0f}" if avg_reliability else "—"
        color = "#10b981" if avg_reliability and avg_reliability >= 75 else "#f59e0b" if avg_reliability and avg_reliability >= 50 else "#ef4444"
        st.markdown(f"""
        <div style='background:white; border-radius:12px; padding:16px; text-align:center;
                    border-top:4px solid {color}; box-shadow:0 2px 8px rgba(0,0,0,0.08);'>
            <div style='font-size:36px; font-weight:800; color:{color};'>{rel}</div>
            <div style='font-size:13px; color:#64748b; margin-top:4px;'>ממוצע אמינות</div>
            <div style='font-size:11px; color:#94a3b8;'>כל המבקרים</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ══════════════════════════════════════════
    # חלק 2 — תור עבודה (מה בוער)
    # ══════════════════════════════════════════

    st.markdown("### 🔥 תור עבודה — ממוין לפי דחיפות")

    # בניית רשימת פריטים שדורשים טיפול
    work_items = []

    # מוצבים שלא בוקרו
    for base, last_date in not_visited_14d.items():
        days_overdue = (today - last_date).days
        if days_overdue > 14:
            work_items.append({
                "סוג": "⏰ לא בוקר",
                "מוצב": base,
                "פירוט": f"לא בוקר {days_overdue} יום",
                "דחיפות": days_overdue,
                "צבע": "#ef4444" if days_overdue > 21 else "#f59e0b"
            })

    # ממצאי עירוב פסול
    latest_per_base = df_unit.sort_values("date_dt").groupby("base").last()
    for base, row in latest_per_base.iterrows():
        if row.get("e_status") == "פסול":
            days = (today - row["date_dt"]).days
            work_items.append({
                "סוג": "🔴 עירוב פסול",
                "מוצב": base,
                "פירוט": f"פסול כבר {days} ימים",
                "דחיפות": days + 50,  # עירוב תמיד דחוף יותר
                "צבע": "#dc2626"
            })
        if row.get("k_cert") == "לא":
            days = (today - row["date_dt"]).days
            work_items.append({
                "סוג": "🟡 כשרות",
                "מוצב": base,
                "פירוט": f"אין תעודה — {days} ימים",
                "דחיפות": days + 30,
                "צבע": "#d97706"
            })

    if work_items:
        work_items.sort(key=lambda x: x["דחיפות"], reverse=True)
        for item in work_items[:10]:  # מציג עד 10 הדחופים
            col_info, col_btn = st.columns([5, 1])
            with col_info:
                st.markdown(f"""
                <div style='background:white; border-right:4px solid {item["צבע"]};
                            border-radius:8px; padding:10px 14px; margin:4px 0;
                            box-shadow:0 1px 4px rgba(0,0,0,0.06);'>
                    <span style='font-weight:700; color:{item["צבע"]};'>{item["סוג"]}</span>
                    <span style='margin-right:10px; color:#1e3a8a; font-weight:600;'>{item["מוצב"]}</span>
                    <span style='color:#64748b; font-size:13px;'>{item["פירוט"]}</span>
                </div>
                """, unsafe_allow_html=True)
            with col_btn:
                if st.button("✅ טופל", key=f"handled_{item['מוצב']}_{item['סוג']}", use_container_width=True):
                    # שמירה ב-Supabase כממצא שטופל
                    try:
                        supabase.table("handled_findings").insert({
                            "base": item["מוצב"],
                            "finding_type": item["סוג"],
                            "unit": unit,
                            "handled_by": st.session_state.get("inspector", "רב חטמ״ר"),
                            "handled_at": datetime.datetime.now().isoformat()
                        }).execute()
                        st.toast(f"✅ {item['מוצב']} — סומן כטופל", icon="✅")
                        st.rerun()
                    except Exception as e:
                        st.error(f"שגיאה: {e}")
    else:
        st.success("✅ אין פריטים פתוחים — הכל מטופל!", icon="🎉")

    st.markdown("<br>", unsafe_allow_html=True)

    # ══════════════════════════════════════════
    # חלק 3 — כרטיסי מבקרים
    # ══════════════════════════════════════════

    st.markdown("### 👤 מבקרים — מי עובד, מי מחפף")

    # בדיקה בטוחה של המבקרים
    inspectors = df_unit["inspector"].dropna().unique() if "inspector" in df_unit.columns else []

    # רק אם יש לפחות מבקר אחד - נייצר עמודות
    if len(inspectors) > 0:
        cols = st.columns(min(len(inspectors), 3))
        for idx, inspector in enumerate(inspectors):
            df_insp = df_unit[df_unit["inspector"] == inspector]
            visits_total = len(df_insp)
            visits_month = len(df_insp[df_insp["date_dt"] >= today - pd.Timedelta(days=30)])
            
            # שליפת ציון אמינות אם קיים
            if "reliability_score" in df_insp.columns:
                avg_rel = df_insp["reliability_score"].mean()
            else:
                # שימוש בפונקציה שלנו כגיבוי
                cred = calculate_inspector_credibility(inspector, df_unit)
                avg_rel = cred['score']
                
            is_p1 = (df_insp["e_status"] == "תקין") if "e_status" in df_insp.columns else pd.Series([True]*len(df_insp))
            is_p2 = (df_insp["k_cert"] == "כן") if "k_cert" in df_insp.columns else pd.Series([True]*len(df_insp))
            is_p3 = (df_insp["k_issues"] == "לא") if "k_issues" in df_insp.columns else pd.Series([True]*len(df_insp))
            
            all_perfect = False
            if len(df_insp) >= 3:
                try:
                    all_perfect = bool((is_p1 & is_p2 & is_p3).all())
                except:
                    all_perfect = False
                    
            last_visit = df_insp["date_dt"].max()
            days_since = (today - last_visit).days if pd.notna(last_visit) else 999

            # צבע כרטיס לפי מצב
            if avg_rel and avg_rel < 50:
                card_color = "#fef2f2"
                border_color = "#ef4444"
                status_label = "⚠️ חשוד בחיפוף"
                status_color = "#ef4444"
            elif all_perfect:
                card_color = "#fff7ed"
                border_color = "#f97316"
                status_label = "🟠 הכל תקין תמיד — לבדוק"
                status_color = "#f97316"
            elif days_since > 14:
                card_color = "#fafafa"
                border_color = "#94a3b8"
                status_label = f"💤 לא יצא {days_since} ימים"
                status_color = "#94a3b8"
            else:
                card_color = "#f0fdf4"
                border_color = "#10b981"
                status_label = "✅ פעיל ותקין"
                status_color = "#10b981"

            with cols[idx % 3]:
                st.markdown(f"""
                <div style='background:{card_color}; border:1px solid {border_color};
                            border-radius:12px; padding:14px; margin:6px 0;'>
                    <div style='font-weight:700; font-size:15px; color:#1e3a8a;'>👤 {inspector}</div>
                    <div style='font-size:12px; color:{status_color}; font-weight:600; margin:4px 0;'>{status_label}</div>
                    <div style='font-size:12px; color:#64748b;'>
                        📋 {visits_month} ביקורות החודש ({visits_total} סה״כ)<br/>
                        🛡️ אמינות ממוצעת: {f"{avg_rel:.0f}" if avg_rel else "—"}<br/>
                        📅 ביקור אחרון: לפני {days_since} ימים
                    </div>
                </div>
                """, unsafe_allow_html=True)
    else:
        # הודעה חלופית אם אין מבקרים (ומונעת את הקריסה!)
        st.info("📭 טרם הוזנו דוחות ומבקרים ביחידה זו.")

    st.markdown("<br>", unsafe_allow_html=True)

    # ══════════════════════════════════════════
    # חלק 4 — מעקב סגירת ממצאים (Ticket System)
    # ══════════════════════════════════════════

    st.markdown("### 🎫 מעקב ממצאים פתוחים")

    try:
        handled = supabase.table("handled_findings") \
            .select("*") \
            .eq("unit", unit) \
            .is_("handled_at", "null") \
            .execute()
        open_tickets = handled.data or []
    except Exception:
        open_tickets = []

    if open_tickets:
        for ticket in open_tickets:
            age_days = (today - pd.Timestamp(ticket["created_at"])).days if ticket.get("created_at") else "?"
            st.markdown(f"""
            <div style='background:#fefce8; border-right:3px solid #eab308;
                        border-radius:8px; padding:10px 14px; margin:4px 0; font-size:13px;'>
                🎫 <b>{ticket.get("base","?")}</b> —
                {ticket.get("finding_type","?")} —
                <span style='color:#92400e;'>פתוח {age_days} ימים</span>
                <span style='color:#64748b; font-size:11px;'> | נפתח ע״י {ticket.get("opened_by","?")}</span>
            </div>
            """, unsafe_allow_html=True)
    else:
        st.info("אין ממצאים פתוחים לטיפול כרגע")

# ══════════════════════════════════════════════════════
# 1. בריפינג בוקר — מסך פתיחה לרב חטמ"ר
# ══════════════════════════════════════════════════════

def render_morning_briefing(df: pd.DataFrame, unit: str):
    """
    נטען לפני כל הטאבים — 3 שורות, 30 שניות, תמונה מלאה.
    """
    if df.empty or "unit" not in df.columns:
        return
    today = pd.Timestamp.now()
    df_unit = df[df["unit"] == unit].copy()
    df_unit["date_dt"] = pd.to_datetime(df_unit["date"], errors="coerce")

    # --- חישובים ---
    # מה בוער - איחוד ממצאים
    latest = df_unit.sort_values("date_dt").groupby("base").last()
    burning_items = []
    for b, r in latest.iterrows():
        if r.get("e_status") == "פסול":
            burning_items.append({"label": "🔴 עירוב פסול", "base": b, "priority": 1})
        if r.get("k_cert") == "לא":
            burning_items.append({"label": "🟡 אין תעודה", "base": b, "priority": 2})
        days = (today - r["date_dt"]).days if pd.notna(r["date_dt"]) else 999
        if days > 14:
            burning_items.append({"label": f"⏰ לא בוקר {days} יום", "base": b, "priority": 3})
    
    burning_items.sort(key=lambda x: x["priority"])

    # מי לא דיווח השבוע - ניקוי שמות
    week_ago = today - pd.Timedelta(days=7)
    df_unit["inspector_clean"] = df_unit["inspector"].astype(str).str.strip()
    all_inspectors = [i for i in df_unit["inspector_clean"].unique() if i and i.lower() != "nan" and len(i) > 2]
    
    inactive = []
    for i in all_inspectors:
        has_recent = not df_unit[
            (df_unit["inspector_clean"] == i) &
            (df_unit["date_dt"] >= week_ago)
        ].empty
        if not has_recent:
            inactive.append(i)
    
    inactive = sorted(list(set(inactive)))

    # כמה שאלות הלכתיות פתוחות (אם יש טבלה)
    try:
        halachic_q = supabase.table("halachic_questions") \
            .select("id") \
            .eq("unit", unit) \
            .is_("answered_at", "null") \
            .execute()
        open_questions = len(halachic_q.data or [])
    except Exception:
        open_questions = 0

    # --- תצוגה ---
    greeting_hour = today.hour
    greeting = "בוקר טוב" if greeting_hour < 12 else "צהריים טובים" if greeting_hour < 17 else "ערב טוב"

    html_content = f"""
    <style>
        .briefing-scroll::-webkit-scrollbar {{
            width: 4px;
        }}
        .briefing-scroll::-webkit-scrollbar-track {{
            background: rgba(255,255,255,0.05);
            border-radius: 10px;
        }}
        .briefing-scroll::-webkit-scrollbar-thumb {{
            background: rgba(255,255,255,0.3);
            border-radius: 10px;
        }}
    </style>
    <div style='background: linear-gradient(135deg, #1e3a8a 0%, #1d4ed8 100%);
                border-radius: 18px; padding: 24px; margin-bottom: 24px;
                color: white; box-shadow: 0 10px 25px rgba(30,58,138,0.2); font-family: Arial, sans-serif; direction: rtl;'>
        <div style='display: flex; justify-content: space-between; align-items: flex-start;'>
            <div>
                <div style='font-size: 13px; opacity: 0.85; margin-bottom: 4px; letter-spacing: 0.5px;'>
                    {today.strftime("%A, %d/%m/%Y")} · {greeting}
                </div>
                <div style='font-size: 24px; font-weight: 800; margin-bottom: 20px; text-shadow: 0 2px 4px rgba(0,0,0,0.1);'>
                    📋 בריפינג יומי — {unit}
                </div>
            </div>
            <div style='background: rgba(255,255,255,0.15); padding: 8px 12px; border-radius: 10px; font-size: 12px;'>
                ⚡ LIVE
            </div>
        </div>
        
        <div style='display: grid; grid-template-columns: 1.2fr 1fr 0.8fr; gap: 16px;'>

            <div style='background: rgba(255,255,255,0.1); border: 1px solid rgba(255,255,255,0.1); 
                        border-radius: 14px; padding: 16px; backdrop-filter: blur(4px);'>
                <div style='font-size: 11px; opacity: 0.8; margin-bottom: 10px; font-weight: 600; text-transform: uppercase;'>🔥 דורש טיפול מיידי</div>
                <div class='briefing-scroll' style='max-height: 160px; overflow-y: auto;'>
                {"".join(f"<div style='font-size:13px; margin-bottom:6px; display:flex; align-items:center;'> <span style='margin-left:8px;'>{item['label']}</span> <span style='background:rgba(255,255,255,0.2); padding:2px 6px; border-radius:4px; font-size:11px;'>{item['base']}</span></div>" for item in burning_items[:10]) or
                 "<div style='font-size:13px; opacity:0.8;'>✅ הכל תקין בגזרה</div>"}
                </div>
            </div>

            <div style='background: rgba(255,255,255,0.1); border: 1px solid rgba(255,255,255,0.1); 
                        border-radius: 14px; padding: 16px; backdrop-filter: blur(4px);'>
                <div style='font-size: 11px; opacity: 0.8; margin-bottom: 10px; font-weight: 600; text-transform: uppercase;'>👤 לא דיווחו השבוע</div>
                <div class='briefing-scroll' style='max-height: 160px; overflow-y: auto;'>
                {"".join(f"<div style='font-size:13px; margin-bottom:4px; border-bottom:1px solid rgba(255,255,255,0.05); padding-bottom:2px;'>• {i}</div>" for i in inactive[:10]) or
                 "<div style='font-size:13px; opacity:0.8;'>✅ כל המבקרים פעילים</div>"}
                </div>
            </div>

            <div style='background: rgba(255,255,255,0.1); border: 1px solid rgba(255,255,255,0.1); 
                        border-radius: 14px; padding: 16px; text-align: center; backdrop-filter: blur(4px);'>
                <div style='font-size: 11px; opacity: 0.8; margin-bottom: 8px; font-weight: 600;'>📖 שאלות פתוחות</div>
                <div style='font-size: 34px; font-weight: 900; color: #facc15;'>{open_questions}</div>
                <div style='font-size: 11px; opacity: 0.7;'>שאלות הלכה</div>
            </div>

        </div>
    </div>
    """
    st.components.v1.html(html_content, height=280)


# ══════════════════════════════════════════════════════
# 2. תכנון מסלול + Waze
# ══════════════════════════════════════════════════════

def render_route_planner(df: pd.DataFrame, unit: str):
    """
    בוחר אוטומטית את 3 המוצבים הדחופים ביותר ומציג כפתורי Waze.
    """
    if df.empty or "unit" not in df.columns:
        st.info("אין נתונים זמינים לתכנון מסלול")
        return

    st.markdown("### 🛣️ מסלול הסיור הקרוב")

    today = pd.Timestamp.now()
    df_unit = df[df["unit"] == unit].copy()
    df_unit["date_dt"] = pd.to_datetime(df_unit["date"], errors="coerce")
    latest = df_unit.sort_values("date_dt").groupby("base").last().reset_index()

    if latest.empty:
        st.info("אין מוצבים שדורשים ביקור")
        return

    # חישוב ציון דחיפות לכל מוצב
    def urgency_score(row):
        score = 0
        days_since = (today - row["date_dt"]).days if pd.notna(row["date_dt"]) else 30
        score += min(days_since * 2, 60)          # עד 60 נק' על ימים
        if row.get("e_status") == "פסול":  score += 50
        if row.get("k_cert") == "לא":      score += 30
        mezuzot = int(row.get("r_mezuzot_missing") or 0)
        score += min(mezuzot * 3, 20)
        return score

    latest["urgency"] = latest.apply(urgency_score, axis=1)
    top3 = latest.nlargest(3, "urgency")

    if top3.empty:
        st.info("אין מוצבים שדורשים ביקור דחוף כרגע")
        return

    st.markdown("""
    <div style='background:#f0f9ff; border-right:4px solid #0ea5e9;
                border-radius:10px; padding:12px 16px; margin-bottom:16px; font-size:13px; color:#0c4a6e;'>
        🤖 המערכת בחרה את 3 המוצבים הדחופים ביותר לסיור הקרוב — לפי גיל ביקור, עירוב, כשרות ומזוזות.
    </div>
    """, unsafe_allow_html=True)

    priority_labels = ["🥇 עדיפות ראשונה", "🥈 עדיפות שנייה", "🥉 עדיפות שלישית"]

    for idx, (_, row) in enumerate(top3.iterrows()):
        base_name = row["base"]
        coords = BASE_COORDINATES.get(base_name)
        days_since = (today - row["date_dt"]).days if pd.notna(row["date_dt"]) else "?"

        # בניית רשימת ממצאים קיימים
        issues = []
        if row.get("e_status") == "פסול":           issues.append("🔴 עירוב פסול")
        if row.get("k_cert") == "לא":               issues.append("🟡 אין תעודת כשרות")
        mezuzot = int(row.get("r_mezuzot_missing") or 0)
        if mezuzot > 0:                              issues.append(f"🔵 {mezuzot} מזוזות חסרות")
        if not issues:                               issues.append(f"⏰ לא בוקר {days_since} ימים")

        col_info, col_waze = st.columns([4, 1])

        with col_info:
            issues_html = " · ".join(issues)
            st.markdown(f"""
            <div style='background:white; border-radius:10px; padding:14px 16px;
                        margin:6px 0; box-shadow:0 2px 6px rgba(0,0,0,0.07);
                        border-right:4px solid {"#ef4444" if idx==0 else "#f59e0b" if idx==1 else "#3b82f6"};'>
                <div style='font-size:11px; color:#64748b; margin-bottom:2px;'>{priority_labels[idx]}</div>
                <div style='font-size:16px; font-weight:700; color:#1e3a8a;'>📍 {base_name}</div>
                <div style='font-size:12px; color:#475569; margin-top:4px;'>{issues_html}</div>
                <div style='font-size:11px; color:#94a3b8; margin-top:4px;'>
                    ביקור אחרון: {row["date_dt"].strftime("%d/%m/%Y") if pd.notna(row["date_dt"]) else "לא ידוע"}
                    · מבקר: {row.get("inspector","?")}
                </div>
            </div>
            """, unsafe_allow_html=True)

        with col_waze:
            if coords:
                lat, lon = coords
                waze_url = f"https://waze.com/ul?ll={lat},{lon}&navigate=yes"
                st.link_button("🗺️ Waze", waze_url, use_container_width=True)
            else:
                st.caption("אין\nקואורדינטות")

    # כפתור שיתוף המסלול בוואטסאפ
    st.markdown("<br>", unsafe_allow_html=True)
    bases_list = " ← ".join(top3["base"].tolist())
    whatsapp_text = f"מסלול סיור רבנות: {bases_list}"
    wa_url = f"https://wa.me/?text={whatsapp_text.replace(' ', '%20')}"
    st.link_button("📤 שתף מסלול בוואטסאפ", wa_url, use_container_width=True)


# ══════════════════════════════════════════════════════
# 3. הודעת שבת אוטומטית לוואטסאפ
# ══════════════════════════════════════════════════════

def render_shabbat_whatsapp_generator(df: pd.DataFrame, unit: str):
    """
    מייצר הודעת שבת מותאמת אישית לפי נתוני השבוע — לשליחה למפקדים.
    """
    st.markdown("### 📱 הודעת שבת אוטומטית למפקדים")

    if df.empty or "unit" not in df.columns:
        st.info("אין נתונים ליצירת הודעת שבת")
        return

    today = pd.Timestamp.now()
    df_unit = df[df["unit"] == unit].copy()
    df_unit["date_dt"] = pd.to_datetime(df_unit["date"], errors="coerce")

    # נתוני השבוע האחרון
    week_ago = today - pd.Timedelta(days=7)
    latest = df_unit.sort_values("date_dt").groupby("base").last()

    # בניית סיכום לכל מוצב
    base_summaries = []
    for base, row in latest.iterrows():
        issues = []
        if row.get("e_status") == "פסול":
            issues.append("העירוב פסול")
        if row.get("k_cert") == "לא":
            issues.append("אין תעודת כשרות")
        mezuzot = int(row.get("r_mezuzot_missing") or 0)
        if mezuzot > 0:
            issues.append(f"חסרות {mezuzot} מזוזות")
        if row.get("p_mix") == "כן":
            issues.append("ערבוב כלים")

        base_summaries.append({
            "base": base,
            "issues": issues,
            "status": "בעיות" if issues else "תקין",
            "inspector": row.get("inspector", ""),
        })

    # בניית הודעה
    shabbat_times = {
        "ירושלים": "16:45", "תל אביב": "17:10",
        "חיפה": "17:00", "באר שבע": "17:15",
    }

    col_settings, col_preview = st.columns([1, 2])

    with col_settings:
        rabbi_name = st.text_input("שם הרב:", value="רב החטיבה", key="wa_rabbi_name")
        region = st.selectbox("עיר לזמן שבת:", list(shabbat_times.keys()), key="wa_region")
        include_issues_only = st.checkbox("כלול רק מוצבים עם בעיות", value=False, key="wa_issues_only")
        custom_note = st.text_area("הערה אישית (רשות):", key="wa_custom", height=80)

    with col_preview:
        shabbat_time = shabbat_times[region]
        next_friday = today + pd.Timedelta(days=(4 - today.dayofweek) % 7)
        date_str = next_friday.strftime("%d/%m/%Y")

        # בניית גוף ההודעה
        lines = [
            f"*מפקדים, שבת שלום!* 🕯️",
            f"להלן עדכון רבנות לשבת {date_str}:",
            f"כניסת שבת {region}: *{shabbat_time}*",
            "",
        ]

        shown_bases = [b for b in base_summaries if not include_issues_only or b["issues"]]

        for b in shown_bases:
            if b["issues"]:
                issues_str = " | ".join(b["issues"])
                lines.append(f"📍 *{b['base']}*: ⚠️ {issues_str}")
            else:
                lines.append(f"📍 *{b['base']}*: ✅ תקין")

        lines.append("")
        if custom_note.strip():
            lines.append(f"📝 {custom_note.strip()}")
            lines.append("")
        lines.append(f"זמין לכל שאלה. שבת שקטה,")
        lines.append(f"*{rabbi_name}*")

        message_text = "\n".join(lines)

        # תצוגה מקדימה
        st.markdown("""
        <div style='background:#e8f5e9; border-radius:12px; padding:4px 8px 2px 8px;
                    margin-bottom:8px; font-size:11px; color:#2e7d32;'>📱 תצוגה מקדימה</div>
        """, unsafe_allow_html=True)

        preview_html = message_text.replace("\n", "<br>").replace("*", "<b>").replace("</b><b>", "")
        # תיקון bold — החלף זוגות של ** ב-<b>...</b>
        import re
        preview_html = re.sub(r'\*([^*]+)\*', r'<b>\1</b>', message_text.replace("\n", "<br>"))

        st.markdown(f"""
        <div style='background:white; border-radius:12px; padding:14px 16px;
                    font-size:13px; line-height:1.7; color:#1a1a1a;
                    box-shadow:0 2px 8px rgba(0,0,0,0.08); direction:rtl;'>
            {preview_html}
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    col_copy, col_wa, col_send = st.columns(3)

    with col_copy:
        if st.button("📋 העתק הודעה", use_container_width=True):
            st.code(message_text, language=None)

    with col_wa:
        encoded = message_text.replace("\n", "%0A").replace(" ", "%20").replace("*", "")
        wa_url = f"https://wa.me/?text={encoded}"
        st.link_button("💬 פתח בוואטסאפ", wa_url, use_container_width=True)

    with col_send:
        if st.button("📤 שלח לכל מפקדי הגזרה", use_container_width=True, type="primary"):
            # שליחה אוטומטית לכל המספרים הרשומים
            try:
                commanders = supabase.table("whatsapp_numbers") \
                    .select("phone,callmebot_key,name") \
                    .eq("unit", unit) \
                    .eq("role", "commander") \
                    .eq("active", True) \
                    .execute()

                sent_count = 0
                for c in (commanders.data or []):
                    try:
                        import urllib.parse, urllib.request
                        encoded_msg = urllib.parse.quote(message_text.replace("*", ""))
                        url = (f"https://api.callmebot.com/whatsapp.php"
                               f"?phone={c['phone']}&text={encoded_msg}&apikey={c['callmebot_key']}")
                        urllib.request.urlopen(url, timeout=5)
                        sent_count += 1
                    except Exception:
                        pass

                if sent_count > 0:
                    st.success(f"✅ הודעה נשלחה ל-{sent_count} מפקדים!")
                    log_audit_event("SHABBAT_MESSAGE_SENT", unit,
                                    details={"recipients": sent_count}, severity="info")
                else:
                    st.warning("לא נמצאו מפקדים רשומים — העתק ושלח ידנית")
            except Exception as e:
                st.error(f"שגיאה: {e}")


def render_hatmar_rabbi_dashboard():
    """דשבורד רב חטמ״ר - הכל במקום אחד (Mobile Friendly)"""
    unit = st.session_state.selected_unit
    role = st.session_state.role
    accessible_units = get_accessible_units(unit, role)
    raw_data = load_reports_cached(accessible_units)
    df = pd.DataFrame(raw_data)

    # ← בריפינג בוקר תמיד מוצג בראש, לפני הטאבים (Wave 2.7)
    render_morning_briefing(df, unit)

    t1, t2, t3, t4, t5, t6, t7, t8, t9 = st.tabs([
        "🚦 סטטוס מוצבים",
        "🕯️ הכנה לשבת",
        "📖 הלכה",
        "📊 ניהול ודיווח",
        "🎯 ניתוח יחידה",
        "📜 היסטוריית דוחות",
        "🗺️ מפה",
        "🛣️ תכנון מסלול",
        "⚙️ ניהול מתקדם"
    ])

    with t1: render_bases_status_board(df, unit)
    with t2:
        render_shabbat_preparation_assistant(unit, df)
        st.divider()
        render_shabbat_whatsapp_generator(df, unit)  # (Wave 2.7)
    with t3: render_halachic_advisor()
    with t4:
        render_inspector_management(unit, df)
        st.divider()
        render_weekly_report_generator(unit, df)
    with t5:
        # ניתוח יחידה מעמיק
        st.markdown(f"### 📈 ניתוח מעמיק - {unit}")
        render_detailed_unit_analysis(df, unit)
    with t6:
        # היסטוריית דוחות (Wave 2.6)
        st.markdown(f"### 📜 היסטוריית דוחות מפורטת - {unit}")
        
        # ✅ התיקון: מוודאים שיש נתונים לפני שמנסים לשלוף עמודות
        if df.empty or 'unit' not in df.columns:
            st.info("📭 טרם הוזנו דוחות ביחידה זו. ברגע שיוגש הדוח הראשון, ההיסטוריה תופיע כאן.")
        else:
            # בניית רשימת עמודות בסדר לוגי
            base_columns = ['date', 'base', 'inspector', 'reliability_score']
            
            # עמודות מצב בסיסיות
            status_columns = ['e_status', 'k_cert'] if 'e_status' in df.columns else []
            
            # תקלות כשרות
            kashrut_issues_columns = [col for col in ['k_issues', 'k_issues_description', 'k_separation', 'p_mix', 'k_products', 'k_bishul'] if col in df.columns]
            
            # שיעורי תורה
            torah_columns = [col for col in ['soldier_want_lesson', 'soldier_has_lesson', 'soldier_lesson_teacher', 'soldier_lesson_phone', 'soldier_yeshiva', 's_torah_id', 's_torah_nusach'] if col in df.columns]
            
            # טרקלין וויקוק
            lounge_vikok_columns = [col for col in ['t_private', 't_kitchen_tools', 't_procedure', 't_friday', 't_app', 'w_location', 'w_private', 'w_kitchen_tools', 'w_procedure', 'w_guidelines'] if col in df.columns]
            
            # חוסרים ונוספות
            other_columns = [col for col in ['r_mezuzot_missing', 'r_torah_missing', 'missing_items', 'free_text'] if col in df.columns]
            
            # סינון לפי סוג יחידה (Wave 2.6)
            is_combat_brigade = unit in ["חטיבה 35", "חטיבה 89", "חטיבה 900"]
            if is_combat_brigade:
                all_cols = base_columns + status_columns + kashrut_issues_columns + torah_columns + other_columns
            else:
                all_cols = base_columns + status_columns + kashrut_issues_columns + torah_columns + lounge_vikok_columns + other_columns
                
            display_df = df[df['unit'] == unit][all_cols].copy()
            
            hebrew_columns = {
                'date': 'תאריך',
                'base': 'מוצב',
                'inspector': 'מבקר',
                'unit': 'יחידה',
                'reliability_score': '🛡️ אמינות',
                'e_status': 'עירוב',
                'k_cert': 'תעודת כשרות',
                'k_issues': 'תקלות כשרות',
                'k_issues_description': 'פירוט תקלה',
                'k_separation': 'הפרדה',
                'p_mix': 'ערבוב כלים',
                'k_products': 'מוצרים',
                'k_bishul': 'בישול',
                'soldier_want_lesson': 'רצון לשיעור',
                'soldier_has_lesson': 'יש שיעור',
                'soldier_lesson_teacher': 'מלמד',
                'soldier_lesson_phone': 'טלפון מלמד',
                'soldier_yeshiva': 'ישיבה',
                's_torah_id': 'ס״ת מזוהה',
                's_torah_nusach': 'נוסח',
                't_private': 'חדר פרטי',
                't_kitchen_tools': 'כלי טרקלין',
                't_procedure': 'נוהל טרקלין',
                't_friday': 'שישי טרקלין',
                't_app': 'אפליקציה',
                'w_location': 'מיקום ויקוק',
                'w_private': 'פרטיות ויקוק',
                'w_kitchen_tools': 'כלי ויקוק',
                'w_procedure': 'נוהל ויקוק',
                'w_guidelines': 'הנחיות ויקוק',
                'r_mezuzot_missing': 'מזוזות חסרות',
                'r_torah_missing': 'ס״ת חסרים',
                'missing_items': 'חוסרים',
                'free_text': 'הערות',
                'review_status': 'סטטוס בדיקה',
            }

            display_df = display_df.rename(columns={
                k: v for k, v in hebrew_columns.items() 
                if k in display_df.columns
            })
            st.dataframe(display_df, use_container_width=True, hide_index=True)

    with t7:
        # מפה
        st.markdown(f"### 🗺️ מפה מפורטת - {unit}")
        render_unit_map(df, unit)

    with t8:
        # תכנון מסלול (Wave 2.7)
        render_route_planner(df, unit)

    with t9:
        # ניהול מתקדם (Wave 2.7)
        render_hatmar_management_tab(df, unit)


def render_command_dashboard():
    role = st.session_state.role
    unit = st.session_state.selected_unit
    accessible_units = get_accessible_units(unit, role)
    raw_data = load_reports_cached(accessible_units)
    df = pd.DataFrame(raw_data)
    
    # כותרת הדף
    st.markdown(f"## 🎯 מרכז בקרה פיקודי - {unit}")
    
    # ✅ הכנת הקובץ מראש - לפני הטאבים (דוח ארצי מלא)
    all_data_for_excel = load_reports_cached(None) # None = כל הארץ
    df_full = pd.DataFrame(all_data_for_excel) if all_data_for_excel else pd.DataFrame()
    
    excel_file_ready = None
    if not df_full.empty:
        try:
            excel_file_ready = create_full_report_excel(df_full)
        except Exception as e:
            st.error(f"שגיאה ביצירת קובץ Excel: {e}")
    
    # ✅ כפתור הורדה בולט - מחוץ לכל לוגיקה מורכבת
    st.markdown("---")
    if excel_file_ready:
        st.download_button(
            label="📥 הורד דוח ארצי מלא (כל היחידות)",
            data=excel_file_ready,
            file_name=f"full_national_report_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
            key="national_excel_btn_stable"
        )
    else:
        if df.empty:
            st.info("📊 אין נתונים זמינים כרגע.")
        else:
            # st.warning("⚠️ לא ניתן ליצור קובץ Excel כרגע")
            pass
    
    st.markdown("---")
    
    # המשך הקוד הקיים עם הטאבים...
    if df.empty:
        return  # ✅ עצור כאן אם אין נתונים

    # טאבים לפי תפקיד
    if role == 'pikud':
        tab_names = ["📊 סקירה כללית", "🏆 ליגת יחידות", "🤖 תובנות AI", "📈 ניתוח יחידה", "📋 מעקב חוסרים", "🏆 Executive Summary", "🎯 Risk Center", "🔍 אמינות מבקרים", "🧠 מוח פיקודי", "💬 עוזר AI", "⚙️ ניהול"]
    elif role == 'ugda':
        tab_names = ["📊 סקירה כללית", "🏆 ליגת יחידות", "🤖 תובנות AI", "📈 ניתוח יחידה", "📋 מעקב חוסרים", "🏆 Executive Summary", "🗺️ Map", "🔍 אמינות מבקרים", "🧠 מוח פיקודי", "💬 עוזר AI"]
    else:
        tab_names = ["📊 סקירה כללית", "🏆 ליגת יחידות", "🤖 תובנות AI", "📋 מעקב חוסרים", "🏆 Executive Summary", "🔍 אמינות מבקרים", "🧠 מוח פיקודי", "💬 עוזר AI"]
    
    tabs_obj = st.tabs(tab_names)
    t_map = {name: tabs_obj[i] for i, name in enumerate(tab_names)}
    
    # ===== טאב 1: סקירה כללית =====
    if "📊 סקירה כללית" in t_map:
        with t_map["📊 סקירה כללית"]:
            st.markdown("### 📊 מדדים מרכזיים")
            
            # כרטיסי מדדים
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("📋 סה״כ דוחות", len(df))
            
            with col2:
                st.metric("🏢 יחידות פעילות", df['unit'].nunique())
            
            with col3:
                mezuzot_missing = int(df['r_mezuzot_missing'].sum()) if 'r_mezuzot_missing' in df.columns else 0
                st.metric("📜 מזוזות חסרות", mezuzot_missing)
            
            with col4:
                eruv_invalid = len(df[df['e_status'] == 'פסול']) if 'e_status' in df.columns else 0
                st.metric("🚧 עירובין פסולים", eruv_invalid, delta=None if eruv_invalid == 0 else f"-{eruv_invalid}", delta_color="inverse")
            
            st.markdown("---")
            
            # מדדי בקרה חשובים
            st.markdown("### 📋 מדדי בקרה מרכזיים")
            
            metric_cols = st.columns(4)
            
            with metric_cols[0]:
                # אחוז כשרות תקין
                if 'k_cert' in df.columns:
                    kosher_ok = len(df[df['k_cert'] == 'כן']) / len(df) * 100 if len(df) > 0 else 0
                    st.metric("✅ כשרות תקינה", f"{kosher_ok:.0f}%", 
                             delta=f"+{kosher_ok-85:.0f}%" if kosher_ok > 85 else f"{kosher_ok-85:.0f}%",
                             delta_color="normal" if kosher_ok > 85 else "inverse")
            
            with metric_cols[1]:
                # אחוז עירובין תקינים
                if 'e_status' in df.columns:
                    eruv_ok = len(df[df['e_status'] == 'תקין']) / len(df) * 100 if len(df) > 0 else 0
                    st.metric("🔵 עירובין תקינים", f"{eruv_ok:.0f}%",
                             delta=f"+{eruv_ok-90:.0f}%" if eruv_ok > 90 else f"{eruv_ok-90:.0f}%",
                             delta_color="normal" if eruv_ok > 90 else "inverse")
            
            with metric_cols[2]:
                # ממוצע ניקיון
                if 's_clean' in df.columns:
                    clean_avg = df['s_clean'].apply(lambda x: {'מצוין': 5, 'טוב': 4, 'בינוני': 3, 'גרוע': 2}.get(x, 0)).mean()
                    st.metric("🧹 ממוצע ניקיון", f"{clean_avg:.1f}/5",
                             delta=f"+{clean_avg-4:.1f}" if clean_avg > 4 else f"{clean_avg-4:.1f}",
                             delta_color="normal" if clean_avg > 4 else "inverse")
            
            with metric_cols[3]:
                # מגמת דיווחים
                if 'date' in df.columns and len(df) > 1:
                    df_sorted = df.sort_values('date')
                    recent_reports = len(df_sorted.tail(7))
                    prev_reports = len(df_sorted.iloc[-14:-7]) if len(df_sorted) >= 14 else 0
                    trend = recent_reports - prev_reports
                    st.metric("📈 דיווחים (7 ימים)", recent_reports,
                             delta=f"+{trend}" if trend > 0 else f"{trend}" if trend < 0 else "ללא שינוי",
                             delta_color="normal" if trend >= 0 else "inverse")
            
            st.markdown("---")
            
            # 🆕 מפת חום זמן אמת (Real-Time Heatmap)
            render_realtime_heatmap(df, accessible_units)
            
            st.markdown("---")
    
            
            # גרפים
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("#### 📊 דוחות לפי יחידה")
                unit_counts = df['unit'].value_counts().reset_index()
                unit_counts.columns = ['unit', 'count']
                fig = px.bar(
                    unit_counts, 
                    x='unit', 
                    y='count', 
                    color='count',
                    color_continuous_scale='Blues',
                    labels={'unit': 'יחידה', 'count': 'מספר דוחות'}
                )
                fig.update_layout(showlegend=False, height=350, xaxis_tickangle=-45)
                st.plotly_chart(fig, use_container_width=True)
            
            with col2:
                st.markdown("#### 🚧 סטטוס עירובין")
                if 'e_status' in df.columns:
                    eruv_counts = df['e_status'].value_counts()
                    colors_map = {'תקין': '#10b981', 'בטיפול': '#f59e0b', 'פסול': '#ef4444'}
                    fig = go.Figure(data=[go.Pie(
                        labels=eruv_counts.index, 
                        values=eruv_counts.values, 
                        hole=0.4,
                        marker=dict(colors=[colors_map.get(x, '#64748b') for x in eruv_counts.index]),
                        textfont=dict(color='#1e293b', size=14),
                        textposition='inside'
                    )])
                    fig.update_layout(
                        height=350,
                        paper_bgcolor='white',
                        plot_bgcolor='white',
                        font=dict(color='#1e293b')
                    )
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("אין נתוני עירוב זמינים")
            
            # גריד יחידות
            if role in ['pikud', 'ugda']:
                st.markdown("---")
                st.markdown("### 🎯 תמונת מצב יחידות")
                sub_units = [u for u in accessible_units if u != unit]
                
                if sub_units:
                    cols = st.columns(min(4, len(sub_units)))
                    for i, u_name in enumerate(sub_units):
                        with cols[i % len(cols)]:
                            u_df = df[df['unit'] == u_name]
                            score = int(calculate_unit_score(u_df)) if not u_df.empty else 0
                            reports_count = len(u_df)
                            badge, badge_color = get_unit_badge(score)
                            
                            st.markdown(f"""
                            <div class="unit-status-card" style="border-top-color: {badge_color};">
                                <img src="{get_logo_url(u_name)}">
                                <div style="font-weight:700; font-size:0.95rem; margin-bottom: 5px;">{u_name}</div>
                                <div style="font-size:0.85rem; color:gray;">ציון: <span style="color:{badge_color}; font-weight:700; font-size: 1.1rem;">{score}</span></div>
                                <div style="font-size:0.75rem; color:#888;">דוחות: {reports_count}</div>
                                <div style="font-size:0.7rem; margin-top:5px; padding:4px 8px; background:{badge_color}; color:white; border-radius:6px;">{badge}</div>
                            </div>
                            """, unsafe_allow_html=True)
    
    # ===== טאב 2: ליגת יחידות =====
    if "🏆 ליגת יחידות" in t_map:
        with t_map["🏆 ליגת יחידות"]:
            st.markdown("### 🏆 ליגת חטמ״רים - דירוג ביצועים")
            
            league = []
            for u in df['unit'].unique():
                unit_df = df[df['unit'] == u]
                if len(unit_df) > 0:
                    score = calculate_unit_score(unit_df)
                    badge, color = get_unit_badge(score)
                    league.append({
                        "יחידה": u,
                        "ציון": score,
                        "דוחות": len(unit_df),
                        "תג": badge,
                        "צבע": color
                    })
            
            league_df = pd.DataFrame(league).sort_values("ציון", ascending=False).reset_index(drop=True)
            
            for idx, row in league_df.iterrows():
                rank = idx + 1
                medal = "🥇" if rank == 1 else "🥈" if rank == 2 else "🥉" if rank == 3 else f"{rank}."
                
                st.markdown(f"""
                    <div style='background: white; border-radius: 14px; padding: 18px; margin-bottom: 12px;
                                box-shadow: 0 4px 15px rgba(0,0,0,0.08); display: flex; 
                                justify-content: space-between; align-items: center; border-right: 5px solid {row['צבע']};'>
                        <div style='display: flex; gap: 15px; align-items: center;'>
                            <span style='font-size: 1.8rem; font-weight: 800; min-width: 50px;'>{medal}</span>
                            <span style='font-size: 1.2rem; font-weight: 700;'>{row['יחידה']}</span>
                        </div>
                        <div style='display: flex; gap: 20px; align-items: center;'>
                            <div style='text-align: center;'>
                                <div style='font-size: 0.85rem; color: #64748b;'>ציון</div>
                                <div style='font-size: 1.8rem; font-weight: 800; color: {row['צבע']};'>
                                    {row['ציון']:.0f}
                                </div>
                            </div>
                            <div style='text-align: center;'>
                                <div style='font-size: 0.85rem; color: #64748b;'>דוחות</div>
                                <div style='font-size: 1.2rem; font-weight: 600;'>
                                    {row['דוחות']}
                                </div>
                            </div>
                            <div style='background: {row['צבע']}; color: white;
                                        padding: 8px 16px; border-radius: 8px; font-weight: 600; min-width: 120px; text-align: center;'>
                                {row['תג']}
                            </div>
                        </div>
                    </div>
                """, unsafe_allow_html=True)
            
            # גרף השוואתי
            st.markdown("---")
            st.markdown("### 📊 השוואת ציונים")
            fig = px.bar(
                league_df, 
                x='יחידה', 
                y='ציון',
                color='ציון',
                color_continuous_scale=['#ef4444', '#f59e0b', '#10b981'],
                range_color=[0, 100],
                labels={'ציון': 'ציון (0-100)'}
            )
            fig.update_layout(height=400, xaxis_tickangle=-45)
            st.plotly_chart(fig, use_container_width=True)
    
    # ===== טאב 3: תובנות AI =====
    if "🤖 תובנות AI" in t_map:
        with t_map["🤖 תובנות AI"]:
            st.markdown("### 🤖 ניתוח חכם")
            
            # סיכום AI
            summary = generate_ai_summary(df)
            st.info(summary["overview"])
            
            st.markdown("---")
            st.markdown("### 🚨 התראות והמלצות")
            
            # התראות מפקדים
            alerts = generate_commander_alerts(df)
            if alerts:
                for alert in alerts:
                    st.warning(f"{alert['icon']} **{alert['title']}**: {alert['message']}")
            else:
                st.success("✅ אין התראות קריטיות - המצב תקין!")
            
            # ניתוח מגמות
            st.markdown("---")
            st.markdown("### 📈 מגמות ותחזיות")
            
            if 'date' in df.columns:
                df_copy = df.copy()
                if not pd.api.types.is_datetime64_any_dtype(df_copy['date']):
                    df_copy['date'] = pd.to_datetime(df_copy['date'], errors='coerce')
                
                # דוחות לאורך זמן
                reports_over_time = df_copy.groupby(df_copy['date'].dt.to_period('W')).size().reset_index()
                reports_over_time.columns = ['week', 'count']
                reports_over_time['week'] = reports_over_time['week'].astype(str)
                
                fig = px.line(
                    reports_over_time, 
                    x='week', 
                    y='count',
                    markers=True,
                    labels={'week': 'שבוע', 'count': 'מספר דוחות'},
                    title='מגמת דיווחים שבועית'
                )
                fig.update_layout(height=300)
                st.plotly_chart(fig, use_container_width=True)
    
    # ===== טאב 4: ניתוח יחידה =====
    if "📈 ניתוח יחידה" in t_map:
        with t_map["📈 ניתוח יחידה"]:
            st.markdown("### 📈 ניתוח מעמיק ליחידה")
            
            selected_unit = st.selectbox("בחר יחידה לניתוח", sorted(df['unit'].unique()))
            unit_df = df[df['unit'] == selected_unit]
            
            if len(unit_df) > 0:
                # ציון ותג
                score = calculate_unit_score(unit_df)
                badge, color = get_unit_badge(score)
                
                # 🆕 זיהוי חריגות ודפוסים חשודים (Anomaly Detection)
                anomalies = detect_anomalies(df, selected_unit)
                if anomalies:
                    st.warning(f"⚠️ זוהו {len(anomalies)} דפוסים חריגים ביחידה זו")
                    with st.expander("🚨 פירוט חריגות ודפוסים חשודים", expanded=True):
                        for anomaly in anomalies:
                            severity_map = {'high': '🔴', 'medium': '🟠', 'low': '🔵'}
                            icon = severity_map.get(anomaly['severity'], '⚪')
                            st.markdown(f"**{icon} {anomaly['message']}**")
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("ציון כללי", f"{score:.1f}/100")
                with col2:
                    st.metric("סה״כ דוחות", len(unit_df))
                with col3:
                    st.markdown(f"<div style='background:{color}; color:white; padding:15px; border-radius:10px; text-align:center; font-weight:700; font-size:1.1rem;'>{badge}</div>", unsafe_allow_html=True)
                
                st.markdown("---")
            
                # פרטי שאלון מפורטים
                st.markdown("### 📋 פירוט שאלון ביקורת")
                
                # קבלת הדוח האחרון והקודם לו למעקב שינויים
                latest_report = unit_df.sort_values('date', ascending=False).iloc[0] if len(unit_df) > 0 else None
                previous_report = unit_df.sort_values('date', ascending=False).iloc[1] if len(unit_df) > 1 else None
                
                # טאבים לקטגוריות שונות
                detail_tabs = st.tabs(["🔴 חוסרים ובעיות", "🍴 עירוב וכשרות", "🏗️ תשתיות ויומן ביקורת", "📊 סיכום כללי"])
                
                with detail_tabs[0]:  # חוסרים
                    st.markdown("#### חוסרים שדווחו")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        # מזוזות
                        mezuzot_missing = int(latest_report.get('r_mezuzot_missing', 0)) if latest_report is not None else 0
                        prev_mezuzot = int(previous_report.get('r_mezuzot_missing', 0)) if previous_report is not None else mezuzot_missing
                        
                        if mezuzot_missing > 0:
                            if mezuzot_missing < prev_mezuzot:
                                diff = prev_mezuzot - mezuzot_missing
                                pct = (diff / prev_mezuzot * 100) if prev_mezuzot > 0 else 0
                                st.metric("📜 מזוזות חסרות", mezuzot_missing, f"-{diff} ({pct:.0f}%)", delta_color="inverse")
                                st.success(f"✅ שיפור! הושלמו {diff} מזוזות מהדוח הקודם")
                            elif mezuzot_missing > prev_mezuzot:
                                diff = mezuzot_missing - prev_mezuzot
                                pct = (diff / prev_mezuzot * 100) if prev_mezuzot > 0 else 0
                                st.metric("📜 מזוזות חסרות", mezuzot_missing, f"+{diff} ({pct:.0f}%)")
                                st.warning(f"⚠️ החוסר גדל ב-{diff} מזוזות")
                            else:
                                st.metric("📜 מזוזות חסרות", mezuzot_missing, "ללא שינוי")
                        else:
                            st.metric("📜 מזוזות חסרות", "0 🟢", "תקין")
                        
                        # ספרי תורה
                        torah_missing = int(latest_report.get('r_torah_missing', 0)) if latest_report is not None else 0
                        if torah_missing > 0:
                            st.metric("📖 ספרי תורה חסרים", torah_missing, delta_color="inverse")
                        else:
                            st.metric("📖 ספרי תורה", "תקין 🟢")
                    
                    with col2:
                        # ציצית
                        tzitzit_missing = int(latest_report.get('r_tzitzit_missing', 0)) if latest_report is not None else 0
                        if tzitzit_missing > 0:
                            st.metric("🧵 ציציות חסרות", tzitzit_missing, delta_color="inverse")
                        else:
                            st.metric("🧵 ציציות", "תקין 🟢")
                        
                        # תפילין
                        tefillin_missing = int(latest_report.get('r_tefillin_missing', 0)) if latest_report is not None else 0
                        if tefillin_missing > 0:
                            st.metric("📿 תפילין חסרים", tefillin_missing, delta_color="inverse")
                        else:
                            st.metric("📿 תפילין", "תקין 🟢")
                
                with detail_tabs[1]:  # עירוב וכשרות
                    st.markdown("#### סטטוס עירוב וכשרות")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        # סטטוס עירוב
                        eruv_status = latest_report.get('e_status', 'לא ידוע') if latest_report is not None else 'לא ידוע'
                        if eruv_status == 'תקין':
                            st.success(f"✅ **סטטוס עירוב:** {eruv_status}")
                        elif eruv_status == 'פסול':
                            st.error(f"❌ **סטטוס עירוב:** {eruv_status}")
                        else:
                            st.warning(f"⚠️ **סטטוס עירוב:** {eruv_status}")
                        
                        # עירוב כלים
                        eruv_kelim = latest_report.get('k_eruv_kelim', 'לא') if latest_report is not None else 'לא'
                        prev_eruv_kelim = previous_report.get('k_eruv_kelim', 'לא') if previous_report is not None else 'לא'
                        
                        if eruv_kelim == 'כן':
                            st.error("🔴 **עירוב כלים:** קיים - דורש טיפול")
                        else:
                            if prev_eruv_kelim == 'כן' and eruv_kelim == 'לא':
                                st.success("✅ **עירוב כלים:** תוקן מהדוח הקודם!")
                            else:
                                st.success("🟢 **עירוב כלים:** לא קיים")
                    
                    with col2:
                        # תעודת כשרות
                        k_cert = latest_report.get('k_cert', 'לא') if latest_report is not None else 'לא'
                        if k_cert == 'כן':
                            st.success("✅ **תעודת כשרות:** קיימת")
                        else:
                            st.warning("⚠️ **תעודת כשרות:** חסרה")
                        
                        # סגירת טרקלין
                        traklin_closed = latest_report.get('k_traklin_closed', 'לא') if latest_report is not None else 'לא'
                        if traklin_closed == 'כן':
                            st.success("✅ **סגירת טרקלין:** מבוצעת")
                        else:
                            st.warning("⚠️ **סגירת טרקלין:** לא מבוצעת")
                
                with detail_tabs[2]:  # תשתיות
                    st.markdown("#### תשתיות ויומן ביקורת")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        # יומן ביקורת
                        pikubok = latest_report.get('k_pikubok', 'לא') if latest_report is not None else 'לא'
                        if pikubok == 'כן':
                            st.success("✅ **יומן ביקורת:** קיים")
                        else:
                            st.warning("⚠️ **יומן ביקורת:** לא קיים")
                        
                        # נהלים
                        procedures = latest_report.get('k_streams', 'לא') if latest_report is not None else 'לא'
                        if procedures == 'כן':
                            st.info("📋 **נהלים מעודכנים:** קיימים")
                        else:
                            st.warning("⚠️ **נהלים מעודכנים:** לא קיימים")
                    
                    with col2:
                        # הערות כלליות
                        notes = latest_report.get('notes', '') if latest_report is not None else ''
                        if notes and notes.strip():
                            st.text_area("📝 הערות והמלצות", notes, height=100, disabled=True)
                        else:
                            st.info("אין הערות נוספות")
                
                with detail_tabs[3]:  # סיכום
                    st.markdown("#### סיכום מצב היחידה")
                    
                    # חישוב אחוזי תקינות
                    total_checks = 10  # סה"כ בדיקות
                    passed_checks = 0
                    
                    if mezuzot_missing == 0: passed_checks += 1
                    if torah_missing == 0: passed_checks += 1
                    if tzitzit_missing == 0: passed_checks += 1
                    if tefillin_missing == 0: passed_checks += 1
                    if eruv_status == 'תקין': passed_checks += 1
                    if eruv_kelim == 'לא': passed_checks += 1
                    if k_cert == 'כן': passed_checks += 1
                    if traklin_closed == 'כן': passed_checks += 1
                    if pikubok == 'כן': passed_checks += 1
                    if procedures == 'כן': passed_checks += 1
                    
                    compliance_pct = (passed_checks / total_checks) * 100
                    
                    st.metric("📊 אחוז תקינות כללי", f"{compliance_pct:.0f}%")
                    st.progress(compliance_pct / 100)
                    
                    if compliance_pct >= 90:
                        st.success("🌟 **מצוין!** היחידה במצב תקין מעולה")
                    elif compliance_pct >= 70:
                        st.info("👍 **טוב** - יש מקום לשיפור קל")
                    elif compliance_pct >= 50:
                        st.warning("⚠️ **בינוני** - דורש תשומת לב")
                    else:
                        st.error("🔴 **דורש טיפול דחוף** - נושאים רבים לטיפול")
                    
                    # רשימת נושאים לטיפול
                    issues = []
                    if mezuzot_missing > 0: issues.append(f"📜 {mezuzot_missing} מזוזות חסרות")
                    if torah_missing > 0: issues.append(f"📖 {torah_missing} ספרי תורה חסרים")
                    if tzitzit_missing > 0: issues.append(f"🧵 {tzitzit_missing} ציציות חסרות")
                    if tefillin_missing > 0: issues.append(f"📿 {tefillin_missing} תפילין חסרים")
                    if eruv_status != 'תקין': issues.append(f"⚠️ עירוב {eruv_status}")
                    if eruv_kelim == 'כן': issues.append("🔴 עירוב כלים קיים")
                    if k_cert != 'כן': issues.append("⚠️ תעודת כשרות חסרה")
                    if traklin_closed != 'כן': issues.append("⚠️ סגירת טרקלין לא מבוצעת")
                    if pikubok != 'כן': issues.append("⚠️ פיקבוק לא קיים")
                    
                    if issues:
                        st.markdown("**נושאים לטיפול:**")
                        for issue in issues:
                            st.markdown(f"- {issue}")
                    else:
                        st.success("✅ אין נושאים פתוחים לטיפול!")
            
                st.markdown("---")
                
                # תובנות
                st.markdown("### 💡 תובנות ומסקנות")
                
                # כפתור הורדה בסיכום הכללי
                enhanced_excel_tab = create_enhanced_excel_report(unit_df, unit_name=selected_unit)
                if enhanced_excel_tab:
                    st.download_button(
                        label="📥 הורד דוח מפורט משופר (Excel)",
                        data=enhanced_excel_tab,
                        file_name=f"detailed_report_{selected_unit}_{pd.Timestamp.now().strftime('%Y%m')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="dl_detailed_tab_main",
                        type="primary"
                    )
                    
                insights = analyze_unit_trends(unit_df)
                for ins in insights:
                    st.info(f"{ins['icon']} **{ins['title']}**: {ins['message']}")
                
                # פירוט נתונים
                st.markdown("---")
                st.markdown("### 📋 פירוט דוחות")
                
                # אפשרות מחיקה למנהלים בלבד
                if role in ['pikud', 'ugda']:
                    st.markdown("#### 🗑️ ניהול דוחות (מנהלים בלבד)")
                    
                    if not unit_df.empty and 'id' in unit_df.columns:
                        # בחירת דוח למחיקה
                        delete_options = []
                        for idx, row in unit_df.iterrows():
                            date_str = row['date'].strftime('%Y-%m-%d') if pd.notna(row['date']) else 'לא ידוע'
                            base = row.get('base', 'לא ידוע')
                            inspector = row.get('inspector', 'לא ידוע')
                            report_id = row.get('id', '')
                            delete_options.append(f"{date_str} | {base} | {inspector} (ID: {report_id})")
                        
                        selected_report = st.selectbox("בחר דוח למחיקה:", ["-- בחר דוח --"] + delete_options)
                        
                        if selected_report != "-- בחר דוח --":
                            # חילוץ ID מהבחירה
                            report_id = selected_report.split("ID: ")[1].rstrip(")")
                            
                            col1, col2 = st.columns([1, 4])
                            with col1:
                                if st.button("🗑️ מחק דוח", type="primary"):
                                    try:
                                        # מחיקה רכה (עוקף שגיאת DB Trigger)
                                        # מעדכנים את שם המפקח עם תחילית של מחיקה
                                        current_inspector = selected_report.split(" | ")[2].split(" (ID:")[0]
                                        supabase.table("reports").update({
                                            "inspector": f"🗑️ מחיקה | {current_inspector}"
                                        }).eq("id", report_id).execute()
                                        
                                        st.success("✅ הדוח הוסר מהמערכת!")
                                        clear_cache()
                                        time.sleep(1)
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"❌ שגיאה במחיקה: {e}")
                            with col2:
                                st.warning("⚠️ פעולה זו בלתי הפיכה!")
                
               # ===== קוד מעודכן לטבלה המפורטת =====
# החלף את החלק של display_df בטאב "ניתוח יחידה" עם הקוד הזה:

                st.markdown("---")
                
                # ===== טבלה מורחבת עם כל העמודות החדשות =====
                st.markdown("#### 📋 דוחות מפורטים - תצוגה מלאה")
                
                # בניית רשימת עמודות בסדר לוגי
                base_columns = ['date', 'base', 'inspector', 'reliability_score']
                
                # עמודות מצב בסיסיות
                status_columns = []
                if 'e_status' in unit_df.columns:
                    status_columns.append('e_status')
                if 'k_cert' in unit_df.columns:
                    status_columns.append('k_cert')
                
                # 🆕 עמודות תקלות כשרות (הכל!)
                kashrut_issues_columns = []
                if 'k_issues' in unit_df.columns:
                    kashrut_issues_columns.append('k_issues')
                if 'k_issues_description' in unit_df.columns:
                    kashrut_issues_columns.append('k_issues_description')
                if 'k_separation' in unit_df.columns:
                    kashrut_issues_columns.append('k_separation')
                if 'p_mix' in unit_df.columns:
                    kashrut_issues_columns.append('p_mix')
                if 'k_products' in unit_df.columns:
                    kashrut_issues_columns.append('k_products')
                if 'k_bishul' in unit_df.columns:
                    kashrut_issues_columns.append('k_bishul')
                
                # 🆕 עמודות שיעורי תורה (הכל!)
                torah_columns = []
                if 'soldier_want_lesson' in unit_df.columns:
                    torah_columns.append('soldier_want_lesson')
                if 'soldier_has_lesson' in unit_df.columns:
                    torah_columns.append('soldier_has_lesson')
                if 'soldier_lesson_teacher' in unit_df.columns:
                    torah_columns.append('soldier_lesson_teacher')
                if 'soldier_lesson_phone' in unit_df.columns:
                    torah_columns.append('soldier_lesson_phone')
                if 'soldier_yeshiva' in unit_df.columns:
                    torah_columns.append('soldier_yeshiva')
                    
                # 🆕 עמודות טופס בית כנסת
                if 's_torah_id' in unit_df.columns:
                    torah_columns.append('s_torah_id')
                if 's_torah_nusach' in unit_df.columns:
                    torah_columns.append('s_torah_nusach')
                
                # 🆕 עמודות טרקלין וויקוק
                lounge_vikok_columns = []
                # Lounge
                if 't_private' in unit_df.columns: lounge_vikok_columns.append('t_private')
                if 't_kitchen_tools' in unit_df.columns: lounge_vikok_columns.append('t_kitchen_tools')
                if 't_procedure' in unit_df.columns: lounge_vikok_columns.append('t_procedure')
                if 't_friday' in unit_df.columns: lounge_vikok_columns.append('t_friday')
                if 't_app' in unit_df.columns: lounge_vikok_columns.append('t_app')
                
                # Vikok
                if 'w_location' in unit_df.columns: lounge_vikok_columns.append('w_location')
                if 'w_private' in unit_df.columns: lounge_vikok_columns.append('w_private')
                if 'w_kitchen_tools' in unit_df.columns: lounge_vikok_columns.append('w_kitchen_tools')
                if 'w_procedure' in unit_df.columns: lounge_vikok_columns.append('w_procedure')
                if 'w_guidelines' in unit_df.columns: lounge_vikok_columns.append('w_guidelines')
        
                # 🆕 עמודות חוסרים ונוספות
                other_columns = []
                if 'r_mezuzot_missing' in unit_df.columns:
                    other_columns.append('r_mezuzot_missing')
                if 'missing_items' in unit_df.columns:
                    other_columns.append('missing_items')
                if 'free_text' in unit_df.columns:
                    other_columns.append('free_text')
                
                # איחוד כל העמודות
                all_columns = base_columns + status_columns + kashrut_issues_columns + torah_columns + lounge_vikok_columns + other_columns
                
                # סינון רק עמודות קיימות
                available_columns = [col for col in all_columns if col in unit_df.columns]
                
                # יצירת DataFrame לתצוגה
                if available_columns:
                    display_df = unit_df[available_columns].copy()
                    
                    # 🆕 מיפוי שמות עמודות לעברית - מלא ומפורט
                    column_mapping = {
                        # בסיסי
                        'date': 'תאריך',
                        'base': 'מוצב',
                        'inspector': 'מבקר',
                        
                        # מצב
                        'e_status': 'סטטוס עירוב',
                        'k_cert': 'תעודת כשרות',
                        
                        # תקלות כשרות
                        'k_issues': '❗ יש תקלות כשרות?',
                        'k_issues_description': '📝 פירוט תקלות כשרות',
                        'k_separation': 'הפרדת כלים',
                        'p_mix': '🔴 ערבוב כלים',
                        'k_products': 'רכש חוץ לא מאושר',
                        'k_bishul': 'בישול ישראל',
                        
                        # טרקלין
                        't_private': '☕ טרקלין - כלים פרטיים',
                        't_kitchen_tools': '🥣 טרקלין - כלי מטבח',
                        't_procedure': '🔒 טרקלין - נוהל סגירה',
                        't_friday': '🛑 טרקלין - סגור בשבת',
                        't_app': '📱 טרקלין - אפליקציה',
                        
                        # ויקוק
                        'w_location': '📍 ויקוק - מיקום',
                        'w_private': '🥤 ויקוק - כלים פרטיים',
                        'w_kitchen_tools': '🍴 ויקוק - כלי מטבח',
                        'w_procedure': '📜 ויקוק - עובד לפי פקודה',
                        'w_guidelines': '📋 ויקוק - הנחיות',
                        
                        # שיעורי תורה
                        'soldier_want_lesson': '💡 רצון לשיעור תורה',
                        'soldier_has_lesson': '📚 יש שיעור במוצב?',
                        'soldier_lesson_teacher': '👨‍🏫 שם מעביר השיעור',
                        'soldier_lesson_phone': '📞 טלפון מעביר השיעור',
                        'soldier_yeshiva': 'ימי ישיבה',
                        
                        # חוסרים ונוספים
                        'r_mezuzot_missing': '📜 מזוזות חסרות',
                        'missing_items': '⚠️ חוסרים כלליים',
                        'free_text': '📝 הערות נוספות',
                        
                        # Wave 2.5: Intelligence
                        'reliability_score': '🛡️ ציון אמינות',
                        'contradictions': '⚠️ סתירות שזוהו'
                    }
                    
                    # החלפת שמות העמודות
                    # display_df.columns = [column_mapping.get(col, col) for col in display_df.columns]
                    # Use rename instead to handle duplicates better if any, though map is safer
                    display_df.rename(columns=column_mapping, inplace=True)
                    
                    # הצגת הטבלה
                    st.dataframe(
                        display_df,
                        use_container_width=True,
                        hide_index=True,
                        height=400
                    )
                else:
                    st.warning("לא נמצאו עמודות להצגה")
        
                # 🆕 כפתור הורדה למפקדים
                st.markdown("---")
                
                try:
                    full_report_excel_cmd = create_full_report_excel(unit_df)
                    if full_report_excel_cmd:
                        st.download_button(
                            label="📥 לחץ כאן להורדת קובץ Excel מלא",
                            data=full_report_excel_cmd,
                            file_name=f"full_report_{selected_unit}_{datetime.date.today().strftime('%d%m%y')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            type="primary",
                            key=f"dl_excel_pikud_detailed_{selected_unit}_{int(time.time())}"
                        )
                    else:
                        st.info("ℹ️ לא ניתן ליצור קובץ Excel כרגע (אין נתונים מספיקים)")
                except Exception as e:
                    st.error(f"שגיאה ביצירת קובץ Excel: {e}")
                    
                st.caption("📊 הקובץ כולל את כל השאלות והתשובות מהשאלון")
                
                st.markdown("---")
                
                # 🆕 סיכומים מפורטים אחרי הטבלה
                st.markdown("### 📊 סיכומים מקיפים")
                
                # סיכום תקלות כשרות
                if kashrut_issues_columns:
                    st.markdown("#### 🔍 סיכום תקלות כשרות")
                    
                    cols = st.columns(min(4, len(kashrut_issues_columns)))
                    col_idx = 0
                    
                    if 'k_issues' in unit_df.columns:
                        has_issues = len(unit_df[unit_df['k_issues'] == 'כן'])
                        with cols[col_idx]:
                            st.metric("מוצבים עם תקלות", has_issues, 
                                     delta=f"-{len(unit_df) - has_issues}" if has_issues > 0 else "אין תקלות",
                                     delta_color="inverse" if has_issues > 0 else "off")
                        col_idx += 1
                    
                    if 'p_mix' in unit_df.columns:
                        mixing = len(unit_df[unit_df['p_mix'] == 'כן'])
                        with cols[col_idx % len(cols)]:
                            st.metric("🔴 ערבוב כלים", mixing, delta_color="inverse")
                        col_idx += 1
                    
                    if 'k_separation' in unit_df.columns:
                        no_sep = len(unit_df[unit_df['k_separation'] == 'לא'])
                        with cols[col_idx % len(cols)]:
                            st.metric("ללא הפרדה", no_sep, delta_color="inverse")
                        col_idx += 1
                    
                    if 'k_bishul' in unit_df.columns:
                        no_bishul = len(unit_df[unit_df['k_bishul'] == 'לא'])
                        with cols[col_idx % len(cols)]:
                            st.metric("ללא בי״ש", no_bishul, delta_color="inverse")
                    
                    # פירוט תקלות ספציפיות
                    if 'k_issues_description' in unit_df.columns:
                        issues_with_description = unit_df[unit_df['k_issues_description'].notna() & (unit_df['k_issues_description'] != '')]
                        if len(issues_with_description) > 0:
                            st.markdown("##### 📝 פירוט תקלות שדווחו:")
                            for idx, row in issues_with_description.iterrows():
                                base_name = row.get('base', 'לא ידוע')
                                description = row.get('k_issues_description', '')
                                date_str = row.get('date').strftime('%d/%m/%Y') if pd.notna(row.get('date')) else 'לא ידוע'
                                st.markdown(f"""
                                <div style='padding: 10px; background-color: #fee2e2; border-right: 4px solid #ef4444; 
                                            border-radius: 5px; margin-bottom: 10px;'>
                                    <div style='font-weight: 700;'>📍 {base_name} | 📅 {date_str}</div>
                                    <div style='margin-top: 5px; color: #475569;'>{description}</div>
                                </div>
                                """, unsafe_allow_html=True)
                
                # ==========================================
                # תצוגת שיעורי תורה (מפוצלת לפי סוג חטיבה)
                # ==========================================
                st.markdown("<br>", unsafe_allow_html=True)

                is_combat_unit = selected_unit in ["חטיבה 35", "חטיבה 89", "חטיבה 900"]

                if is_combat_unit:
                    st.markdown("#### 📚 יומן שיעורי תורה וימי ישיבה (סדיר)")

                    if 'lesson_qty' in unit_df.columns:
                        lessons_df = unit_df[pd.to_numeric(unit_df['lesson_qty'], errors='coerce').fillna(0) > 0].copy()

                        if not lessons_df.empty:
                            c1, c2, c3 = st.columns(3)
                            total_lessons = pd.to_numeric(lessons_df['lesson_qty'], errors='coerce').sum()
                            total_participants = pd.to_numeric(lessons_df['lesson_participants'], errors='coerce').sum() if 'lesson_participants' in lessons_df.columns else 0
                            c1.metric("סה\"כ שיעורים שהועברו", int(total_lessons))
                            c2.metric("סה\"כ משתתפים", int(total_participants))
                            c3.metric("מוקדי פעילות (מיקומים)", lessons_df['lesson_location'].nunique() if 'lesson_location' in lessons_df.columns else 0)

                            cols_to_show = {
                                'lesson_date': 'תאריך',
                                'lesson_location': 'מיקום',
                                'lesson_content': 'תוכן השיעור',
                                'lesson_instructors': 'מעבירי השיעור',
                                'lesson_qty': 'כמות שיעורים',
                                'lesson_participants': 'משתתפים',
                                'lesson_population': 'סוג אוכלוסייה'
                            }
                            available_cols = {k: v for k, v in cols_to_show.items() if k in lessons_df.columns}
                            if available_cols:
                                display_lessons = lessons_df[list(available_cols.keys())].rename(columns=available_cols)
                                st.dataframe(display_lessons, use_container_width=True, hide_index=True)
                        else:
                            st.info("טרם דווחו שיעורי תורה או ימי ישיבה ביחידה זו.")

                else:
                    # תצוגת שיעורי תורה רגילה לחטמ"ר
                    if torah_columns:
                        st.markdown("#### 📚 סיכום שיעורי תורה (חטמ\"ר)")

                        col1, col2, col3 = st.columns(3)

                        if 'soldier_want_lesson' in unit_df.columns:
                            want_lesson = len(unit_df[unit_df['soldier_want_lesson'] == 'כן'])
                            col1.metric("💡 מעוניינים בשיעור", want_lesson)

                        if 'soldier_has_lesson' in unit_df.columns:
                            has_lesson = len(unit_df[unit_df['soldier_has_lesson'] == 'כן'])
                            col2.metric("📚 יש שיעור פעיל", has_lesson)

                        if 'r_mezuzot_missing' in unit_df.columns:
                            total_mezuzot = int(unit_df['r_mezuzot_missing'].sum())
                            col3.metric("📜 סה\"כ מזוזות חסרות", total_mezuzot, delta_color="inverse" if total_mezuzot > 0 else "off")

                        if 'soldier_lesson_teacher' in unit_df.columns and 'soldier_has_lesson' in unit_df.columns:
                            active_lessons = unit_df[
                                (unit_df['soldier_has_lesson'] == 'כן') &
                                (unit_df['soldier_lesson_teacher'].notna()) &
                                (unit_df['soldier_lesson_teacher'] != '')
                            ]
                            if len(active_lessons) > 0:
                                st.markdown("##### 👨\u200d🏫 רשימת מעבירי שיעורים באוגדה:")
                                for idx, row in active_lessons.iterrows():
                                    teacher = row.get('soldier_lesson_teacher', 'לא ידוע')
                                    phone = row.get('soldier_lesson_phone', '')
                                    base_name = row.get('base', 'לא ידוע')
                                    phone_str = f" | 📞 {phone}" if phone else ""
                                    st.markdown(f"<div style='padding: 10px; background-color: #dbeafe; border-right: 4px solid #3b82f6; border-radius: 5px; margin-bottom: 8px;'><b>📍 {base_name}</b><br>👨\u200d🏫 {teacher}{phone_str}</div>", unsafe_allow_html=True)

                        if 'soldier_want_lesson' in unit_df.columns and 'soldier_has_lesson' in unit_df.columns:
                            want_but_no_lesson = unit_df[
                                (unit_df['soldier_want_lesson'] == 'כן') &
                                (unit_df['soldier_has_lesson'] == 'לא')
                            ]
                            if len(want_but_no_lesson) > 0:
                                st.markdown("##### ⚠️ מוצבים שמעוניינים בשיעור אך אין להם:")
                                bases_list = ", ".join(want_but_no_lesson['base'].unique())
                                st.warning(f"📍 {bases_list}")
                                st.info("💡 יש לתאם מעביר שיעור למוצבים אלו")

        
                # סיכום חוסרים כלליים
                if 'missing_items' in unit_df.columns:
                    items_with_missing = unit_df[unit_df['missing_items'].notna() & (unit_df['missing_items'] != '')]
                    if len(items_with_missing) > 0:
                        st.markdown("<br>", unsafe_allow_html=True)
                        st.markdown("#### ⚠️ חוסרים כלליים שדווחו")
                        
                        for idx, row in items_with_missing.iterrows():
                            base_name = row.get('base', 'לא ידוע')
                            missing = row.get('missing_items', '')
                            date_str = row.get('date').strftime('%d/%m/%Y') if pd.notna(row.get('date')) else 'לא ידוע'
                            
                            st.markdown(f"""
                            <div style='padding: 12px; background-color: #fef3c7; border-right: 4px solid #f59e0b; 
                                        border-radius: 5px; margin-bottom: 10px;'>
                                <div style='font-weight: 700;'>📍 {base_name} | 📅 {date_str}</div>
                                <div style='margin-top: 5px; color: #475569;'>{missing}</div>
                            </div>
                            """, unsafe_allow_html=True)
                
                st.markdown("---")
        

    
    # ===== טאב 5: מעקב חוסרים =====
    if "📋 מעקב חוסרים" in t_map:
        with t_map["📋 מעקב חוסרים"]:
            st.markdown("### 📋 מעקב חוסרים פתוחים")
            accessible_units_list = accessible_units if isinstance(accessible_units, list) else list(accessible_units)
            render_sla_dashboard(accessible_units_list)

            deficits_df = get_open_deficits(accessible_units_list)
            
            # ✅ קבלת סטטיסטיקות מדויקות
            stats = get_deficit_statistics(accessible_units_list)
            
            # ✅ חישוב נוסף מהדוחות עצמם (לאימות)
            total_from_reports = calculate_total_deficits_from_reports(df)
            
            # סטטיסטיקות - שורה עליונה
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("🔴 חוסרים פתוחים", stats['total_open'], 
                         help="מספר החוסרים פתוחים במערכת המעקב")
            with col2:
                # תצוגה של המזוזות החסרות מחישוב מהדוחות
                mezuzot_delta = total_from_reports['mezuzot'] - stats.get('total_mezuzot_tracked', 0) if 'total_mezuzot_tracked' in stats else None
                st.metric("📜 מזוזות (מדוחות)", total_from_reports['mezuzot'],
                         delta=f"+{mezuzot_delta}" if mezuzot_delta and mezuzot_delta > 0 else None,
                         help="חישוב מהדוח האחרון של כל מוצב")
            with col3:
                st.metric("✅ חוסרים שנסגרו", stats['total_closed'],
                         help="חוסרים שהושלמו ונסגרו")
            with col4:
                avg_days = stats['avg_resolution_days']
                st.metric("⏱️ זמן ממוצע לפתרון", 
                         f"{avg_days:.1f} ימים" if avg_days > 0 else "אין נתונים",
                         help="זמן ממוצע בימים עד סגירת חוסר")
            
            st.markdown("---")
            
            # סטטיסטיקות נוספות - שורה שנייה
            st.markdown("#### 📊 פירוט חוסרים לפי סוג")
            col1, col2, col3, col4, col5 = st.columns(5)
            
            with col1:
                st.metric("📜 מזוזות", total_from_reports['mezuzot'], 
                         help="סך כל המזוזות החסרות")
            with col2:
                st.metric("🔴 ערבוב כלים", total_from_reports['eruv_kelim'],
                         help="מוצבים עם ערבוב כלים")
            with col3:
                st.metric("📋 בלי תעודה", total_from_reports['kashrut_cert'],
                         help="מוצבים ללא תעודת כשרות")
            with col4:
                st.metric("🚧 עירוב פסול", total_from_reports['eruv_broken'],
                         help="מוצבים עם עירוב פסול")
            with col5:
                st.metric("👤 בלי נאמן", total_from_reports['no_supervisor'],
                         help="מוצבים ללא נאמן כשרות בשבת")
            
            st.markdown("---")
            
            # ✅ הצגת חוסרים לפי יחידה ומוצב
            if not deficits_df.empty:
                deficit_names = {
                    'mezuzot': 'מזוזות חסרות',
                    'eruv_kelim': 'ערבוב כלים',
                    'kashrut_cert': 'תעודת כשרות חסרה',
                    'eruv_status': 'עירוב פסול',
                    'shabbat_supervisor': 'נאמן כשרות חסר'
                }
                
                # קבוצה לפי יחידה
                for unit in sorted(deficits_df['unit'].unique()):
                    unit_deficits = deficits_df[deficits_df['unit'] == unit]
                    
                    # ספירת חוסרים לפי סוג
                    deficit_types_count = unit_deficits['deficit_type'].value_counts()
                    summary_text = ", ".join([f"{deficit_names.get(dt, dt)}: {count}" 
                                             for dt, count in deficit_types_count.items()])
                    
                    with st.expander(f"🔴 {unit} - {len(unit_deficits)} חוסרים פתוחים ({summary_text})"):
                        # ✅ קבוצה נוספת לפי מוצב
                        bases = unit_deficits['base'].unique() if 'base' in unit_deficits.columns else ['לא ידוע']
                        
                        for base in sorted(bases):
                            base_deficits = unit_deficits[unit_deficits['base'] == base] if 'base' in unit_deficits.columns else unit_deficits
                            
                            st.markdown(f"**📍 {base}:**")
                            
                            for _, deficit in base_deficits.iterrows():
                                deficit_type_he = deficit_names.get(deficit['deficit_type'], deficit['deficit_type'])
                                try:
                                    detected_dt = pd.to_datetime(deficit.get('detected_date'), errors='coerce')
                                    if pd.notna(detected_dt):
                                        detected_date = detected_dt.strftime('%d/%m/%Y')
                                        days_open = (pd.Timestamp.now() - detected_dt).days
                                    else:
                                        detected_date = 'לא ידוע'
                                        days_open = 0
                                except Exception:
                                    detected_date = 'לא ידוע'
                                    days_open = 0
                            
                            # צבע לפי חומרת החוסר
                            severity_color = "#ef4444" if days_open > 30 else "#f59e0b" if days_open > 14 else "#10b981"
                            
                            col1, col2 = st.columns([3, 1])
                            with col1:
                                st.markdown(f"""
                                <div style="padding: 10px; border-right: 4px solid {severity_color}; background-color: #f8fafc; border-radius: 5px; margin-bottom: 10px;">
                                    <div style="font-weight: 700; font-size: 1.1rem;">• {deficit_type_he}</div>
                                    <div style="color: #64748b; font-size: 0.9rem;">
                                        כמות: <b>{deficit['deficit_count']}</b> | 
                                        זוהה: {detected_date} | 
                                        פתוח: <span style="color: {severity_color}; font-weight: 600;">{days_open} ימים</span>
                                    </div>
                                    {f"<div style='color: #475569; font-size: 0.85rem; margin-top: 5px;'>💬 {deficit.get('notes', '')}</div>" if deficit.get('notes') else ""}
                                </div>
                                """, unsafe_allow_html=True)
                            
                            with col2:
                                if st.button("✅ סגור", key=f"close_{deficit['id']}", use_container_width=True):
                                    if update_deficit_status(deficit['id'], 'closed', notes="נסגר ידנית על ידי מפקד"):
                                        st.success("✅ החוסר סומן כסגור!")
                                        time.sleep(0.5)
                                        st.rerun()
                        
                        st.markdown("---")
            
                # כפתור רענון
                if st.button("🔄 רענן מעקב חוסרים", use_container_width=True):
                    clear_cache()
                    st.rerun()
            
            else:
                st.success("🎉 אין חוסרים פתוחים במערכת המעקב!")
                
                # בדיקה אם יש אי-התאמה
                if any(v > 0 for v in total_from_reports.values()):
                    st.warning("⚠️ **שים לב**: נמצאו חוסרים בדוחות האחרונים, אך הם עדיין לא במערכת המעקב.")
                    st.info("💡 חוסרים חדשים יווצרו אוטומטית בדוח הבא שיוגש.")
                    
                    # הצגת החוסרים שנמצאו בדוחות
                    st.markdown("**חוסרים שנמצאו בדוחות:**")
                    if total_from_reports['mezuzot'] > 0:
                        st.markdown(f"- 📜 **{total_from_reports['mezuzot']} מזוזות חסרות**")
                    if total_from_reports['eruv_kelim'] > 0:
                        st.markdown(f"- 🔴 **{total_from_reports['eruv_kelim']} מוצבים עם ערבוב כלים**")
                    if total_from_reports['kashrut_cert'] > 0:
                        st.markdown(f"- 📋 **{total_from_reports['kashrut_cert']} מוצבים ללא תעודת כשרות**")
                    if total_from_reports['eruv_broken'] > 0:
                        st.markdown(f"- 🚧 **{total_from_reports['eruv_broken']} מוצבים עם עירוב פסול**")
                    if total_from_reports['no_supervisor'] > 0:
                        st.markdown(f"- 👤 **{total_from_reports['no_supervisor']} מוצבים ללא נאמן כשרות**")
    
    # ===== טאב 6: Executive Summary =====
    if "🏆 Executive Summary" in t_map:
        with t_map["🏆 Executive Summary"]:
            if role == 'pikud':
                render_executive_summary_dashboard()
            elif role == 'ugda':
                render_ogda_summary_dashboard_v2()
            else:
                render_hatmar_summary_dashboard()

    # ===== טאב 7: מפה ארצית / Map =====
    map_tab_key = "🗺️ Map" if "🗺️ Map" in t_map else None
    if map_tab_key:
        with t_map[map_tab_key]:
            st.markdown("### 🛰️ תמונת מצב ארצית - כלל המגזרים")
            st.info("🔐 **ביטחון מידע:** המיקומים מוזזים 300 מטר מהמיקום המדויק לצורכי אבטחת מידע")
            
            # שליפת כל הנתונים ללא סינון (None)
            map_raw = load_reports_cached(None)
            full_map_df = pd.DataFrame(map_raw) if map_raw else pd.DataFrame()
            
            if not full_map_df.empty:
                # ניקוי וביטול סינונים גאוגרפיים
                v_map = full_map_df.dropna(subset=['latitude', 'longitude']).copy()
                # גבולות רחבים מאוד (כל ישראל)
                v_map = v_map[(v_map['latitude'] > 29) & (v_map['latitude'] < 34)]
                
                # יצירת מפת Folium
                center_lat = v_map['latitude'].mean()
                center_lon = v_map['longitude'].mean()
                
                # מיפוי צבעים לפי יחידה
                unit_color_map = {
                    "חטמ״ר בנימין": "#1e3a8a",
                    "חטמ״ר שומרון": "#60a5fa",
                    "חטמ״ר יהודה": "#22c55e",
                    "חטמ״ר עציון": "#fb923c",
                    "חטמ״ר אפרים": "#ef4444",
                    "חטמ״ר מנשה": "#a855f7",
                    "חטמ״ר הבקעה": "#db2777"
                }
                
                m = create_street_level_map(center=(center_lat, center_lon), zoom_start=8)
                
                for _, row in v_map.iterrows():
                    add_unit_marker_to_folium(m, row, unit_color_map)
                    
                st_folium(m, width=1200, height=700, key="global_dashboard_map", returned_objects=[])
                
                # מקרא
                st.markdown("#### 🔑 מקרא חטמ״רים")
                legend_html = "<div style='display: flex; flex-wrap: wrap; gap: 15px; margin-top: 10px;'>"
                for unit_name in sorted(v_map['unit'].unique()) if 'unit' in v_map.columns else []:
                    color = unit_color_map.get(unit_name, "#808080")
                    count = len(v_map[v_map['unit'] == unit_name])
                    legend_html += f"<div><span style='color: {color}; font-size: 1.5rem;'>●</span> {unit_name} ({count})</div>"
                legend_html += "</div>"
                st.markdown(legend_html, unsafe_allow_html=True)
    
            else:
                st.warning("⚠️ לא נמצאו נתוני מיקום")
    
    # ===== טאב 8: Risk Center (רק פיקוד) =====
    if "🎯 Risk Center" in t_map:
        with t_map["🎯 Risk Center"]:
            render_risk_command_center(df, accessible_units)

    # ===== טאב: אמינות מבקרים =====
    if "🔍 אמינות מבקרים" in t_map:
        with t_map["🔍 אמינות מבקרים"]:
            st.markdown("## 🔍 אמינות מבקרים")
            if not df.empty and 'inspector' in df.columns:
                inspectors = df['inspector'].dropna().unique()
                if len(inspectors) > 0:
                    for inspector in sorted(inspectors):
                        cred = calculate_inspector_credibility(inspector, df)
                        col1, col2, col3, col4 = st.columns([2, 1, 1, 3])
                        with col1:
                            st.markdown(f"**{inspector}**")
                        with col2:
                            st.metric("ציון", f"{cred['score']:.0f}")
                        with col3:
                            st.metric("% ליקויים", f"{cred['defect_rate']:.0f}%")
                        with col4:
                            st.markdown(
                                f"<span style='color:{cred['color']}; font-weight:bold;'>{cred['credibility']}</span>",
                                unsafe_allow_html=True
                            )
                        st.divider()
                else:
                    st.info("אין מבקרים רשומים")
            else:
                st.info("אין נתוני מבקרים")

    # ===== טאב 10: ניהול (רק פיקוד) =====
    if "⚙️ ניהול" in t_map:
        with t_map["⚙️ ניהול"]:
            management_tabs = st.tabs(["🔗 ניהול היררכיה", "🔑 ניהול סיסמאות", "📧 הגדרות מייל", "🖼️ ניהול לוגואים"])
            
            # ניהול היררכיה
            with management_tabs[0]:
                st.subheader("🔗 שיוך חטמ״רים לאוגדות")
                
                # הצגת שיוכים קיימים
                try:
                    current_hierarchy = supabase.table("hierarchy").select("*").execute().data
                    if current_hierarchy:
                        st.markdown("**שיוכים נוכחיים:**")
                        for h in current_hierarchy:
                            col1, col2 = st.columns([3, 1])
                            with col1:
                                st.info(f"📌 {h['child_unit']} ← {h['parent_unit']}")
                            with col2:
                                if st.button("🗑️ הסר", key=f"del_{h['child_unit']}"):
                                    try:
                                        supabase.table("hierarchy").delete().eq("child_unit", h['child_unit']).execute()
                                        st.success("✅ השיוך הוסר")
                                        time.sleep(0.5)
                                        st.rerun()
                                    except:
                                        st.error("❌ שגיאה בהסרת השיוך")
                except Exception as e:
                    st.warning(f"טבלת היררכיה טרם נוצרה. היא תיווצר אוטומטית בשיוך הראשון.")
                
                st.markdown("---")
                
                # טופס שיוך חדש
                with st.form("assign_hierarchy"):
                    col1, col2 = st.columns(2)
                    with col1:
                        parent = st.selectbox("אוגדה (Parent)", [u for u in COMMAND_UNITS if u not in ("פיקוד מרכז",)])
                    with col2:
                        child = st.selectbox("חטמ״ר (Child)", HATMAR_UNITS)
                    
                    if st.form_submit_button("✅ בצע שיוך", use_container_width=True):
                        try:
                            supabase.table("hierarchy").delete().eq("child_unit", child).execute()
                            supabase.table("hierarchy").insert({"parent_unit": parent, "child_unit": child}).execute()
                            st.success(f"✅ {child} שוייך בהצלחה ל-{parent}")
                            clear_cache()
                            time.sleep(1)
                            st.rerun()
                        except Exception as e:
                            error_msg = str(e)
                            st.error(f"❌ שגיאה: {error_msg}")
                            if "PGRST205" in error_msg or "hierarchy" in error_msg:
                                st.info("💡 **פתרון:** יש ליצור טבלה בשם `hierarchy` ב-Supabase עם העמודות:\n- `parent_unit` (text)\n- `child_unit` (text)")
            
            # ניהול סיסמאות
            with management_tabs[1]:
                st.subheader("🔑 עדכון סיסמאות יחידות")
                
                col1, col2 = st.columns(2)
                with col1:
                    selected_unit_pwd = st.selectbox("בחר יחידה", ALL_UNITS, key="pwd_unit")
                with col2:
                    new_pwd = st.text_input("סיסמה חדשה", type="password", key="new_pwd")
                
                if st.button("🔄 עדכן סיסמה", use_container_width=True):
                    if new_pwd and len(new_pwd) >= 4:
                        success, message = update_unit_password(selected_unit_pwd, new_pwd)
                        if success:
                            st.success(f"✅ {message} עבור {selected_unit_pwd}")
                            time.sleep(1)
                            st.rerun()
                        else:
                            st.error(f"❌ {message}")
                            st.info("💡 **אפשרויות פתרון:**\n- ודא שהטבלה `unit_passwords` קיימת ב-Supabase\n- בדוק שיש לך הרשאות כתיבה\n- נסה שוב או צור קשר עם התמיכה")
                    else:
                        st.warning("⚠️ הסיסמה חייבת להכיל לפחות 4 תווים")

                st.markdown("---")
                st.subheader("🔑 ניהול קודי גישה למפקדים (אור אישי)")
                st.caption("כאן ניתן לאפס או לשנות את קוד הגישה האישי של מפקדי היחידות")
                
                # הצגת המצב הקיים
                try:
                    all_cmd_codes = supabase.table("commander_settings").select("*").execute().data
                    if all_cmd_codes:
                        st.markdown("**קודי גישה פעילים:**")
                        for item in all_cmd_codes:
                            c_u, c_c = item['unit'], item['access_code']
                            col_info, col_reset = st.columns([3, 1])
                            with col_info:
                                st.code(f"{c_u}: {c_c}")
                            with col_reset:
                                if st.button(f"🔄 אפס ל-1111", key=f"reset_code_{c_u}"):
                                    if set_commander_code(c_u, "1111"):
                                        st.success(f"✅ הקוד של {c_u} אופס")
                                        time.sleep(0.5)
                                        st.rerun()
                except:
                    pass

                with st.form("update_commander_code"):
                    st.write("עדכון ידני של קוד מפקד")
                    c1, c2 = st.columns(2)
                    with c1:
                        target_u = st.selectbox("בחר יחידה", ALL_UNITS, key="target_cmd_unit")
                    with c2:
                        target_code = st.text_input("קוד חדש (לפחות 4 ספרות)", type="password", key="target_cmd_code")
                    
                    if st.form_submit_button("💾 שמור קוד מפקד", use_container_width=True):
                        if len(target_code) >= 4:
                            if set_commander_code(target_u, target_code):
                                st.success(f"✅ הקוד של {target_u} עודכן בהצלחה")
                                time.sleep(1)
                                st.rerun()
                            else:
                                st.error("❌ שגיאה בעדכון הקוד")
                        else:
                            st.warning("⚠️ הקוד חייב להכיל לפחות 4 תווים")

            # הגדרות אימייל להתראות
            with management_tabs[2]: # I will move logos to [3] and add email to [2]
                st.subheader("📧 הגדרות מייל להתראות אוטומטיות")
                st.info("כאן ניתן להגדיר את כתובת המייל שאליה יישלחו התראות קריטיות עבור כל יחידה.")
                
                # בחירת יחידה להגדרת מייל
                selected_email_unit = st.selectbox("בחר יחידה להגדרה", ALL_UNITS, key="alert_email_unit")
                
                # שליפת מייל קיים
                current_email = ""
                try:
                    res = supabase.table("unit_emails").select("email").eq("unit", selected_email_unit).execute()
                    if res.data:
                        current_email = res.data[0]['email']
                except:
                    pass
                
                with st.form("set_unit_email"):
                    new_email = st.text_input("כתובת אימייל להתראות", value=current_email)
                    if st.form_submit_button("💾 שמור הגדרות מייל", use_container_width=True):
                        try:
                            supabase.table("unit_emails").upsert({"unit": selected_email_unit, "email": new_email}).execute()
                            st.success(f"✅ כתובת המייל עבור {selected_email_unit} עודכנה בהצלחה")
                            clear_cache()
                            time.sleep(1)
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ שגיאה בשמירה: {e}")
                            st.info("💡 ודא שטבלת `unit_emails` קיימת ב-Supabase עם העמודות `unit` (PK, text) ו-`email` (text).")

            # ניהול לוגואים (מוזז לטאב 4)
            with management_tabs[3]:
                st.subheader("🖼️ העלאת לוגואים")
                
                selected_logo_unit = st.selectbox("בחר יחידה", ALL_UNITS, key="logo_unit")
                
                col_preview, col_upload = st.columns(2)
                with col_preview:
                    st.markdown("**לוגו נוכחי:**")
                    st.image(get_logo_url(selected_logo_unit), width=150)
                
                with col_upload:
                    st.markdown("**העלאת לוגו חדש:**")
                    uploaded_logo = st.file_uploader("בחר קובץ תמונה", type=['png', 'jpg', 'jpeg'], key="logo_file")
                    
                    if uploaded_logo and st.button("📤 העלה לוגו", use_container_width=True):
                        if upload_logo_to_supabase(selected_logo_unit, uploaded_logo.getvalue()):
                            st.success(f"✅ הלוגו עודכן בהצלחה עבור {selected_logo_unit}")
                            time.sleep(1)
                            st.rerun()
                        else:
                            st.error("❌ שגיאה בהעלאת הלוגו")

    # ===== טאב 🧠 מוח פיקודי (AI Brain) =====
    if "🧠 מוח פיקודי" in t_map:
        with t_map["🧠 מוח פיקודי"]:
            render_executive_ai_brief(df, accessible_units if isinstance(accessible_units, list) else list(accessible_units))

    # ===== טאב 📱 דשבורד חטמ״ר =====
    if "📱 דשבורד חטמ״ר" in t_map:
        with t_map["📱 דשבורד חטמ״ר"]:
            render_hatmar_rabbi_dashboard()

    # ===== טאב 💬 עוזר AI - האחרון =====
    if "💬 עוזר AI" in t_map:
        with t_map["💬 עוזר AI"]:
            render_ai_chatbot(df, accessible_units if isinstance(accessible_units, list) else list(accessible_units))

def create_enhanced_excel_report(df, unit_name=""):
    """
    🔧 תיקון: יצירת Excel מוגן משגיאות 'No visible sheets'
    """
    try:
        import io
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        if df.empty:
            return None
            
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # --- גיליון 1: סיכום מנהלים ---
            # אנחנו יוצרים את הגיליון הזה קודם כדי לוודא שתמיד יש לפחות גיליון אחד
            summary_data = {
                'מדד': ['שם היחידה', 'סה"כ דוחות', 'נוצר בתאריך'],
                'ערך': [unit_name, len(df), datetime.datetime.now().strftime('%d/%m/%Y %H:%M')]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='סיכום', index=False)
            
            # וידוא שהגיליון הראשון תמיד גלוי
            writer.book.active = 0
            writer.sheets['סיכום'].sheet_state = 'visible'

            # --- גיליון 2: נתונים מפורטים ---
            column_mapping = {
                'date': 'תאריך', 'base': 'מוצב', 'inspector': 'מבקר',
                'e_status': 'סטטוס עירוב', 'k_cert': 'תעודת כשרות', 
                'k_issues': 'תקלות כשרות', 'k_issues_description': 'פירוט תקלות',
                't_private': 'טרקלין-פרטי', 't_kitchen_tools': 'טרקלין-מטבח', 
                't_procedure': 'טרקלין-נוהל', 't_friday': 'טרקלין-שבת', 't_app': 'טרקלין-אפליקציה',
                'w_location': 'ויקוק-מיקום', 'w_private': 'ויקוק-פרטי', 
                'w_kitchen_tools': 'ויקוק-מטבח', 'w_procedure': 'ויקוק-פקודה', 'w_guidelines': 'ויקוק-הנחיות',
                'soldier_want_lesson': 'שיעור תורה-רצון', 'soldier_has_lesson': 'שיעור תורה-קיים',
                's_torah_id': "מס' צ' ספר תורה", 's_torah_nusach': "נוסח ספר תורה", 'r_torah_missing': 'ספרי תורה חסרים',
                'free_text': 'הערות'
            }
            
            existing_cols = [col for col in column_mapping.keys() if col in df.columns]
            if existing_cols:
                details_df = df[existing_cols].copy()
                details_df.rename(columns=column_mapping, inplace=True)
                details_df.to_excel(writer, sheet_name='נתונים מפורטים', index=False)
                writer.sheets['נתונים מפורטים'].sheet_state = 'visible'
        
        return output.getvalue()
        
    except Exception as e:
        st.error(f"שגיאה ביצירת אקסל: {e}")
        return None

def radio_with_explanation(label, key, horizontal=True, col=None):
    """
    Helper function to create a radio button with an optional explanation field
    for "Don't Know" answers.
    """
    # Use the provided column or default to streamlit main container
    container = col if col else st
    
    options = ["כן", "לא", "לא יודע / לא בדקתי"]
    # Use a unique key for the radio based on the provided key
    selected = container.radio(label, options, horizontal=horizontal, key=f"radio_{key}")
    
    final_answer = selected
    if selected == "לא יודע / לא בדקתי":
        # Show text input if "Don't Know" is selected
        reason = container.text_input(f"פרט מדוע ({label})", key=f"reason_{key}")
        if reason:
            final_answer = f"לא יודע ({reason})"
        else:
            # Return sentinel value for validation
            final_answer = f"__MISSING_EXPLANATION__:{label}"
            
    return final_answer


def get_session_seed() -> int:
    """🎲 מחזיר seed אקראי קבוע לכל session (למניעת אוטומציה)"""
    if "form_shuffle_seed" not in st.session_state:
        import random as _r
        st.session_state.form_shuffle_seed = _r.randint(1000, 9999)
    return st.session_state.form_shuffle_seed


def get_flip_week() -> int:
    """🔄 מחזיר 0 או 1 לפי שבוע (סיבוב שאלות קונטרול כל שבוע)"""
    return datetime.date.today().isocalendar()[1] % 2


def flipped_radio(label, key, col=None, flip_week_target: int = 0) -> str:
    """🔄 כמו radio_with_explanation אך מסמן אם הוא בשבוע הפוך"""
    flip_week = get_flip_week()
    if flip_week == flip_week_target:
        flipped_label = f"{label}  ✅[תשובה צפויה: לא]"
    else:
        flipped_label = label
    return radio_with_explanation(flipped_label, key, col=col)


COMPACT_FORM_CSS = """
<style>
  /* עיצוב קומפקטי לטפסים */
  div[data-testid="stRadio"] { margin-bottom: 2px !important; }
  div[data-testid="stRadio"] > div { gap: 4px !important; }
  div[data-testid="stRadio"] label p { font-size: 13px !important; line-height: 1.3 !important; }
  div[data-testid="column"] { padding: 2px 6px !important; }
  .element-container { margin-bottom: 3px !important; }
  div[data-testid="stMarkdownContainer"] h3 { 
    margin-top: 8px !important; margin-bottom: 2px !important; 
    font-size: 16px !important; padding: 4px 0 !important;
  }
  div[data-testid="stMarkdownContainer"] h4 { 
    margin-top: 6px !important; margin-bottom: 2px !important; 
    font-size: 14px !important;
  }
  div[data-testid="stTabs"] button { font-size: 13px !important; padding: 4px 8px !important; }
  div[data-testid="stNumberInput"] { margin-bottom: 2px !important; }
  div[data-testid="stSelectbox"] { margin-bottom: 2px !important; }
</style>
"""

def render_unit_report():
    """הטופס המלא"""
    unit = st.session_state.selected_unit
    
    # אתחול משתנה שליחה
    submitted = False
    
    # ⏱️ הכנסת CSS קומפקטי
    st.markdown(COMPACT_FORM_CSS, unsafe_allow_html=True)
    
    # ⏱️ אתחול seed לסדר אקראי
    _seed = get_session_seed()
    
    # ⏱️ אתחול טיימר דיווח (למדידת אמינות)
    if "report_start_time" not in st.session_state:
        st.session_state.report_start_time = time.time()
    
    # ✅ ניקוי cache בכל טעינה כדי למנוע שגיאות schema
    clear_cache()
    unit = st.session_state.selected_unit
    
    # כפתור קוד גישה לרב חטמ"ר
    st.markdown("---")
    st.markdown("### 🔑 כניסה לניתוח יחידה מפורט (רב חטמ\"ר)")
    
    # בדיקה אם כבר מחובר כמפקד
    if 'commander_authenticated' not in st.session_state:
        st.session_state.commander_authenticated = False
    
    if not st.session_state.commander_authenticated:
        st.info("הזן את קוד הגישה האישי שלך כדי לצפות בניתוח מפורט של היחידה")
        
        col1, col2 = st.columns([3, 1])
        with col1:
            access_code = st.text_input("קוד גישה", type="password", key="commander_code_input")
        with col2:
            st.write("")  # spacing
            st.write("")  # spacing
            if st.button("🔓 כניסה", use_container_width=True):
                # בדיקת קוד גישה דינמי
                correct_code = get_commander_code(unit)
                if access_code == correct_code:
                    st.session_state.commander_authenticated = True
                    st.session_state.commander_unit = unit
                    st.success("✅ קוד גישה נכון! מעביר לניתוח יחידה...")
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error(f"❌ קוד גישה שגוי")
                    st.info(f"💡 רמז: הקוד הנכון מתחיל ב-'{correct_code[:1]}...'")
    else:
        # מפקד מחובר - הצג ניתוח יחידה + אפשרות לשינוי קוד
        st.success(f"✅ מחובר כרב חטמ\"ר - {unit}")
        
        col_back, col_pwd, col_spacer = st.columns([1, 1, 2])
        with col_back:
            if st.button("🔙 חזרה לדשבורד", use_container_width=True):
                st.session_state.commander_authenticated = False
                st.rerun()
        
        with col_pwd:
            with st.popover("🔐 שינוי קוד גישה"):
                st.write("שנה את קוד הגישה האישי שלך")
                new_code = st.text_input("קוד חדש (לפחות 4 תווים)", type="password", key="new_commander_code")
                confirm_code = st.text_input("אימות קוד", type="password", key="confirm_commander_code")
                if st.button("שמור קוד חדש", use_container_width=True):
                    if len(new_code) < 4:
                        st.error("הקוד חייב להכיל לפחות 4 תווים")
                    elif new_code != confirm_code:
                        st.error("הקודים אינם תואמים")
                    else:
                        if set_commander_code(unit, new_code):
                            st.success("✅ הקוד עודכן בהצלחה!")
                            time.sleep(1)
                            st.rerun()
                        else:
                            st.error("שגיאה בעדכון הקוד בבסיס הנתונים")
        
        # דשבורד מפקד מאוחד (Wave 2.6) - כולל היסטוריה וניתוחים בתוך טאבים
        st.markdown("---")
        render_hatmar_rabbi_dashboard()
    
    # טופס דיווח (רק אם לא במצב מפקד)
    if not st.session_state.commander_authenticated:
        st.markdown("### 📋 דיווח ביקורת חדש")
        
        # כותרת הטופס
        col_logo, col_title = st.columns([1, 7])
        with col_logo:
            st.image(get_logo_url(unit), width=80)
        with col_title:
            st.title(f"📋 דיווח ביקורת - {unit}")
    
    # Removed st.form to allow dynamic UI updates
    # with st.form("report"):
    
    # 🆕 כפתור טעינת טיוטה (Drafts)
    if st.button("📂 טען טיוטה אחרונה", key="load_last_draft", help="טען את הנתונים מהטיוטה האחרונה שנשמרה"):
        draft_key = f"{unit}_last_draft"
        draft_data = load_draft(draft_key)
        if draft_data:
            # עדכון Session State כדי שהטופס יתמלא
            # הערה: זה דורש מיפוי חכם של מפתחות, כרגע נציג הודעה
            st.success(f"✅ טיוטה נטענה מ-{draft_data.get('timestamp', 'לא ידוע')}")
            st.json(draft_data) # זמני - להצגת הנתונים
            st.info("מנגנון מילוי אוטומטי מלא בבנייה...")
        else:
            st.warning("⚠️ לא נמצאה טיוטה שמורה")

    st.markdown("### 📍 מיקום ותאריך")
    # קריאה לכפתור המיקום שלנו!
    loc = custom_location_button(key="main_form_gps_btn")

    if loc and loc.get('lat'):
        gps_lat = loc['lat']
        gps_lon = loc['lon']
    else:
        gps_lat = None
        gps_lon = None
    
    if gps_lat:
        # ✅ הצגת המיקום המדויק שנקלט
        st.success(f"✅ מיקום GPS נקלט: {gps_lat:.6f}, {gps_lon:.6f}")
        
        # ✅ הדפסה ללוג (תוכל לראות בקונסול של Streamlit)
        print(f"🔍 DEBUG - GPS נקלט: lat={gps_lat}, lon={gps_lon}, base={base if 'base' in locals() else 'לא הוגדר'}")
        
        # ✅ בדיקה אם המיקום בגבולות ישראל
        if not (29.5 <= gps_lat <= 33.5 and 34.2 <= gps_lon <= 35.9):
            st.error(f"🚨 **שגיאה:** המיקום ({gps_lat:.4f}, {gps_lon:.4f}) מחוץ לגבולות ישראל!")
            st.warning("💡 ייתכן שהמכשיר שלך נותן מיקום שגוי. נסה להפעיל מחדש את ה-GPS")
            st.info("📍 **למידע:** ירושלים היא בערך lat=31.7683, lon=35.2137")
        else:
            st.info(f"✅ המיקום תקין - בגבולות ישראל")
        
        # בדיקת מרחק מבסיסים ידועים
        nearest_base, distance = find_nearest_base(gps_lat, gps_lon)
        
        if distance < 2.0:
            st.info(f"📍 **מיקום מזוהה:** {nearest_base} ({distance:.1f} ק\"מ)")
        elif distance < 5.0:
            st.warning(f"⚠️ **מרחק בינוני:** {nearest_base} ({distance:.1f} ק\"מ) - וודא שהמיקום נכון")
        else:
            st.error(f"🚨 **התראה:** {distance:.1f} ק\"מ מ-{nearest_base} - מיקום חריג!")
    else:
        st.warning("📡 מחפש מיקום GPS... אנא המתן עד להופעת אישור ירוק לפני השליחה")
        st.caption("ירושלים: lat ~31.7, lon ~35.2")
    
    # --- 🆕 ניהול ברקוד למציאת מיקום ---
    if 'barcode_manual_input' in st.session_state and st.session_state.barcode_manual_input:
        scanned_val = st.session_state.barcode_manual_input
        # חיפוש הבסיס לפי הברקוד
        found_base = None
        for b_name, b_code in BASE_BARCODES.items():
            if b_code == scanned_val:
                found_base = b_name
                break
        if found_base and st.session_state.get('base_input') != found_base:
            st.session_state.base_input = found_base
            st.toast(f"📍 נמצא מיקום: {found_base}", icon="✅")

    if 'barcode_from_image_input' in st.session_state and st.session_state.barcode_from_image_input:
        scanned_val = st.session_state.barcode_from_image_input
        found_base = None
        for b_name, b_code in BASE_BARCODES.items():
            if b_code == scanned_val:
                found_base = b_name
                break
        if found_base and st.session_state.get('base_input') != found_base:
            st.session_state.base_input = found_base
            st.toast(f"📍 נמצא מיקום: {found_base}", icon="✅")

    # --- 🆕 ניהול רשימת מוצבים (Wave 2.19) ---
    known_bases = []
    try:
        # Get unique bases for this unit from previous reports
        base_res = supabase.table("reports").select("base").eq("unit", unit).execute()
        if base_res.data:
            known_bases = sorted(list(set([r['base'] for r in base_res.data if r.get('base')])))
    except Exception:
        pass
    
    # הוספת הבסיסים הקבועים מהקואורדינטות (אם הם שייכים לחטיבה הזו - כרגע נוסיף את כולם כהצעות)
    suggested_bases = sorted(list(set(known_bases + list(BASE_COORDINATES.keys()))))
    
    c1, c2, c3 = st.columns(3)
    date = c1.date_input("תאריך", datetime.date.today())
    time_v = c2.time_input("שעה", datetime.datetime.now().time())
    inspector = c3.text_input("מבקר *")
    
    # בחירת מוצב (Selectbox + Manual input)
    st.markdown("---")
    base_selection = st.selectbox(
        "בחר מוצב / מיקום *", 
        ["-- בחר --", "📝 מוצב אחר (הזנה ידנית)"] + suggested_bases,
        key="base_selection_dropdown"
    )
    
    base = ""
    if base_selection == "📝 מוצב אחר (הזנה ידנית)":
        base = st.text_input("הזן שם מוצב / מיקום ידנית *", placeholder="לדוגמה: מוצב מנורה", key="base_input_manual")
    elif base_selection != "-- בחר --":
        base = base_selection
        # Update session state for consistency with other parts of the app
        st.session_state.base_input = base
    else:
        # If no selection, check if barcode found something
        base = st.session_state.get('base_input', '')
        if not base:
            st.info("💡 בחר מוצב מהרשימה או בחר 'הזנה ידנית'")
    
    render_base_history_card(base, unit)

    # ===== סריקת ברקוד מוצב =====
    with st.expander("📷 סריקת ברקוד מוצב (רשות)"):
        barcode_tab_cam, barcode_tab_img = st.tabs(["📷 סריקה חיה", "🖼️ העלאת תמונה"])
        with barcode_tab_cam:
            expected_barcode = BASE_BARCODES.get(base, "NONE")
            scanner_js = """
            <div id='barcode-scanner-container' style='font-family:sans-serif;direction:rtl;'>
                <button onclick='startCamera()' id='start-btn' style='padding:10px 20px;background:#1e3a8a;color:white;border:none;border-radius:8px;font-size:16px;cursor:pointer;margin-bottom:10px;'>
                    📷 הפעל מצלמה לסריקה
                </button>
                <video id='barcode-video' width='100%' style='max-height:260px;border-radius:8px;background:#000;display:none;'></video>
                <div id='barcode-feedback' style='padding:10px; border-radius:8px; margin-top:8px; background:#f1f5f9;display:none;'>
                    <p id='barcode-result' style='font-size:18px;font-weight:bold;color:#1e3a8a;margin:0;'>🔍 מחפש ברקוד...</p>
                    <p id='verification-status' style='font-size:14px;margin:4px 0 0 0;color:#64748b;'>סטטוס אימות: טרם נסרק</p>
                </div>
            </div>
            <script>
            var cameraStarted = false;
            function startCamera() {
                if (cameraStarted) return;
                cameraStarted = true;
                var btn = document.getElementById('start-btn');
                var video = document.getElementById('barcode-video');
                var resultEl = document.getElementById('barcode-result');
                var statusEl = document.getElementById('verification-status');
                var feedbackEl = document.getElementById('barcode-feedback');
                var expected = "{{EXPECTED}}";
                btn.style.display = 'none';
                video.style.display = 'block';
                feedbackEl.style.display = 'block';
                if ('BarcodeDetector' in window) {
                    var barcodeDetector = new BarcodeDetector({ formats: ['qr_code', 'data_matrix', 'pdf417', 'code_128', 'code_39', 'ean_13'] });
                    navigator.mediaDevices.getUserMedia({ video: { facingMode: 'environment' } }).then(function(stream) {
                        video.srcObject = stream;
                        video.play();
                        function scan() {
                            barcodeDetector.detect(video).then(function(barcodes) {
                                if (barcodes.length > 0) {
                                    var val = barcodes[0].rawValue;
                                    resultEl.textContent = '✅ נסרק: ' + val;
                                    if (expected !== "NONE") {
                                        if (val === expected) {
                                            statusEl.textContent = '🌟 אימות הצליח: ברקוד מיקום תקין!';
                                            feedbackEl.style.background = '#dcfce7';
                                            statusEl.style.color = '#166534';
                                        } else {
                                            statusEl.textContent = '❌ אימות נכשל: הברקוד אינו תואם למיקום זה';
                                            feedbackEl.style.background = '#fee2e2';
                                            statusEl.style.color = '#991b1b';
                                        }
                                    } else {
                                        statusEl.textContent = '⚠️ לא הוגדר ברקוד אימות למיקום זה';
                                    }
                                    stream.getTracks().forEach(function(t) { t.stop(); });
                                    btn.style.display = 'block';
                                    btn.textContent = '🔄 סרוק שוב';
                                    cameraStarted = false;
                                } else {
                                    requestAnimationFrame(scan);
                                }
                            }).catch(function() { requestAnimationFrame(scan); });
                        }
                        scan();
                    }).catch(function(err) {
                        resultEl.textContent = 'אין גישה למצלמה: ' + err.message;
                        resultEl.style.color = '#ef4444';
                        btn.style.display = 'block';
                        cameraStarted = false;
                    });
                } else {
                    resultEl.textContent = '⚠️ הדפדפן אינו תומך בסריקה אוטומטית. נסה Chrome עדכני.';
                    resultEl.style.color = '#f59e0b';
                    btn.style.display = 'block';
                    cameraStarted = false;
                }
            }
            </script>
            """.replace("{{EXPECTED}}", expected_barcode)
            import streamlit.components.v1 as _components
            _components.html(scanner_js, height=350)
            barcode_manual = st.text_input("📟 או הזן ברקוד ידנית", placeholder="לדוגמא: ABC-12345", key="barcode_manual_input")
            if barcode_manual:
                st.success(f"📷 ברקוד: {barcode_manual}")
        with barcode_tab_img:
            st.caption("העלה תמונה של ברקוד – הזן את הערך ידנית למטה")
            barcode_image_file = st.file_uploader("🖼️ העלה תמונת ברקוד", type=['jpg', 'png', 'jpeg'], key="barcode_image_upload")
            if barcode_image_file:
                st.image(barcode_image_file, caption="תמונת ברקוד שהועלתה", use_column_width=True)
            barcode_from_image = st.text_input("הזן את ערך הברקוד מהתמונה", placeholder="לדוגמא: ABC-12345", key="barcode_from_image_input")
            if barcode_from_image:
                st.success(f"✅ ברקוד מתמונה: {barcode_from_image}")

    barcode_value = st.session_state.get('barcode_manual_input', '') or st.session_state.get('barcode_from_image_input', '')
        
    # ========================================
    # 📁 טאבי הטופס (5 טאבים)
    # ========================================
    import random as _rand
    _rand.seed(get_session_seed())
    _flip = get_flip_week()
    
    # Pre-initialize fields that live in Tab 5 (so submit handler always has them)
    missing = ""
    free_text = ""
    _mandatory_warnings = []
    hq_vars = {}  # Pre-initialize so Tab 2 can also populate it

    # בדיקה האם זו חטיבה סדירה/קומנדו
    is_combat_brigade = unit in ["חטיבה 35", "חטיבה 89", "חטיבה 900"]

    # Pre-initialize combat-specific lesson variables to avoid NameErrors in DB save
    lesson_date, lesson_location, lesson_qty, lesson_participants = "", "", 0, 0
    lesson_content, lesson_instructors, lesson_population = "", "", ""
    soldier_shabbat_training, soldier_knows_rabbi = "לא רלוונטי", "לא רלוונטי"
    soldier_prayers, soldier_talk_cmd = "לא רלוונטי", "לא רלוונטי"
    soldier_yeshiva, soldier_want_lesson, soldier_has_lesson = "לא רלוונטי", "לא רלוונטי", "לא רלוונטי"
    soldier_food, soldier_lesson_teacher, soldier_lesson_phone = "לא רלוונטי", "", ""
    
    # 🆕 Wave 2.5: Initializations
    continuity_answers = {}

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "🍽️ כשרות",
        "🕍 ביהכ\"נ ועירוב",
        "📜 נהלים ורוח",
        "📖 שיחת חתך",
        "⚠️ חוסרים ושליחה"
    ])

    # ===========================================
    # TAB 1: כשרות (Kitchen, Pillbox, WeCook)
    # ===========================================
    with tab1:
        # 🆕 Wave 2.5: Continuity & GPS Checkpoint 1
        if base and len(base) > 2:
            continuity_answers = render_continuity_questions(base, unit)
        render_gps_checkpoint(1, base)
        
        # ⏱️ Tab timer
        if "tab1_start" not in st.session_state:
            st.session_state["tab1_start"] = time.time()
        # 🎤 דווח בקול (AI)
        with st.expander("🎤 דווח בקול (AI Transcription)", expanded=False):
            audio_file = st.file_uploader("העלה הקלטה קולית (mp3) לדיווח מהיר", type=['mp3'], key="voice_report")
            if audio_file:
                with st.spinner("מחלץ נתונים מהקול..."):
                    voice_text = transcribe_voice_note(audio_file.getvalue())
                    st.success("✅ תמלול בוצע!")
                    st.info(voice_text)
                    if st.button("📝 העתק להערות הכלליות"):
                        st.session_state.voice_note_transcription = voice_text
        _show_pillbox = unit not in NO_LOUNGE_WECOOK_UNITS
        if _show_pillbox:
            st.markdown("#### 🏠 פילבוקס / הגנ״ש")
            c1, c2 = st.columns(2)
            p_pakal = radio_with_explanation("האם יש פק״ל רבנות?", "p1", col=c1)
            p_marked = radio_with_explanation("האם הכלים מסומנים?", "p2", col=c2)
            c1, c2 = st.columns(2)
            # שאלת קונטרול – מתחלפת כל 3 שבועות
            if _flip == 0:
                p_mix = radio_with_explanation("האם זוהה ערבוב כלים?", "p3", col=c1)
            else:
                p_mix = radio_with_explanation("האם זוהה ערבוב כלים?", "p3", col=c1)
            p_kasher = radio_with_explanation("האם נדרשת הכשרה כלים?", "p4", col=c2)
        else:
            p_pakal = p_marked = p_mix = p_kasher = "לא רלוונטי"

        st.markdown("#### 🍽️ מטבח")
        k_cook_type = st.selectbox("סוג מטבח", ["מבשל", "מחמם"])
        c1, c2 = st.columns(2)
        k_cert = radio_with_explanation("תעודת כשרות מתוקפת?", "k7", col=c1)
        k_bishul = radio_with_explanation("האם יש בישול ישראל?", "k8", col=c2)

        # 🍯 Honeypot question (Wave 2.15)
        if base and len(base) > 2:
            render_honeypot_question(base)

        # OCR
        with st.expander("📄 סריקת תעודת כשרות (OCR)", expanded=False):
            cert_photo_ocr = st.file_uploader("העלה תמונה לחילוץ נתונים אוטומטי", type=['jpg', 'png', 'jpeg'], key="cert_ocr")
            if cert_photo_ocr:
                with st.spinner("מפענח תעודה..."):
                    extracted = extract_kashrut_cert_data(cert_photo_ocr.getvalue())
                    if extracted and 'error' not in extracted:
                        st.success("✅ נתונים חולצו בהצלחה!")
                        col_ocr1, col_ocr2 = st.columns(2)
                        with col_ocr1:
                            st.info(f"📌 ספק: {extracted.get('supplier_name')}")
                            st.info(f"🔢 מספר: {extracted.get('certificate_number')}")
                        with col_ocr2:
                            st.info(f"📅 תוקף: {extracted.get('expiry_date')}")
                            status, status_type = validate_cert_status(extracted.get('expiry_date'))
                            st.write(f"**סטטוס:** {status}")
                    elif extracted and 'error' in extracted:
                        st.warning(f"⚠️ {extracted['error']}")

        st.markdown("#### 📸 תקלות ונאמן כשרות")
        c1, c2 = st.columns(2)
        k_issues = radio_with_explanation("יש תקלות כשרות?", "k_issues", col=c1)
        k_shabbat_supervisor = radio_with_explanation("יש נאמן כשרות בשבת?", "k_shabbat_sup", col=c2)
        k_issues_description = ""
        k_issues_photo = None
        if k_issues == "כן":
            k_issues_description = st.text_area("פרט את תקלות הכשרות שנמצאו", key="k_issues_desc")
            k_issues_photo = st.file_uploader(
                "📷 צלם תמונה של תקלת הכשרות (חובה)", 
                type=['jpg', 'png', 'jpeg'], 
                key="k_issues_photo_upload",
                help="חובה לצלם את התקלה לפני שליחת הדוח"
            )
            if not k_issues_photo:
                st.warning("⚠️ נא לצלם תמונה של תקלת הכשרות לפני השליחה")
                _mandatory_warnings.append("📷 חובה לצלם תמונה של תקלת כשרות לפני שליחה")
        k_shabbat_supervisor_name = ""
        k_shabbat_supervisor_phone = ""
        k_shabbat_photo = None
        if k_shabbat_supervisor == "כן":
            with c2:
                col_sup_name, col_sup_phone = st.columns(2)
                k_shabbat_supervisor_name = col_sup_name.text_input("שם נאמן כשרות", key="k_sup_name")
                k_shabbat_supervisor_phone = col_sup_phone.text_input("טלפון נאמן", key="k_sup_phone")
                current_day = datetime.datetime.now().weekday()
                if current_day in [3, 4]:
                    k_shabbat_photo = st.file_uploader("📷 תמונת נאמן ⚠️ (חובה בחמישי-שישי)", type=['jpg', 'png', 'jpeg'], key="k_shabbat_photo_tab1")
                else:
                    k_shabbat_photo = st.file_uploader("📷 תמונת נאמן (אופציונלי)", type=['jpg', 'png', 'jpeg'], key="k_shabbat_photo_tab1")
        # (Photos in Tab 1)

        # רשימת שאלות כשרות לשאפל — מותאמת לסוג יחידה
        if is_combat_brigade:
            kashrut_questions = [
                ("האם יש דף תאריכים לתבלינים?", "k4", "k_dates"),
                ("האם יש הפרדה?", "k1", "k_separation"),
                ("האם יש חימום נפרד בין בשר ודגים?", "k11", "k_heater"),
                ("האם יש שטיפת ירק?", "k5", "k_leafs"),
                ("בוצע חירור גסטרונומים?", "k6", "k_holes"),
                ("האם רכש חוץ מתנהל לפי פקודה?", "k3", "k_products"),
                ("האם מבוצעת בדיקת ביצים?", "k9", "k_eggs"),
                ("האם מולאה אפליקציה?", "k12", "k_app"),
                ("האם בוצע תדריך טבחים?", "k2", "k_briefing"),
            ]
        else:
            kashrut_questions = [
                ("האם יש הפרדה?", "k1", "k_separation"),
                ("האם בוצע תדריך טבחים?", "k2", "k_briefing"),
                ("האם רכש חוץ מתנהל לפי פקודה?", "k3", "k_products"),
                ("האם יש דף תאריכים לתבלינים?", "k4", "k_dates"),
                ("האם יש שטיפת ירק?", "k5", "k_leafs"),
                ("בוצע חירור גסטרונומים?", "k6", "k_holes"),
                ("האם מבוצעת בדיקת ביצים?", "k9", "k_eggs"),
                ("האם יש חדר מכ״ש במפג״ד?", "k10", "k_machshir"),
                ("האם יש חימום נפרד בין בשר ודגים?", "k11", "k_heater"),
                ("האם מולאה אפליקציה?", "k12", "k_app"),
            ]
        _rand.shuffle(kashrut_questions)
        kashrut_answers = {}
        for i in range(0, len(kashrut_questions), 2):
            c1, c2 = st.columns(2)
            label, key, var = kashrut_questions[i]
            kashrut_answers[var] = radio_with_explanation(label, key, col=c1)
            if i + 1 < len(kashrut_questions):
                label2, key2, var2 = kashrut_questions[i + 1]
                kashrut_answers[var2] = radio_with_explanation(label2, key2, col=c2)
        k_separation = kashrut_answers.get('k_separation', 'לא יודע')
        k_briefing = kashrut_answers.get('k_briefing', 'לא יודע')
        k_products = kashrut_answers.get('k_products', 'לא יודע')
        k_dates = kashrut_answers.get('k_dates', 'לא יודע')
        k_leafs = kashrut_answers.get('k_leafs', 'לא יודע')
        k_holes = kashrut_answers.get('k_holes', 'לא יודע')
        k_eggs = kashrut_answers.get('k_eggs', 'לא יודע')
        k_machshir = kashrut_answers.get('k_machshir', 'לא יודע')
        k_heater = kashrut_answers.get('k_heater', 'לא יודע')
        k_app = kashrut_answers.get('k_app', 'לא יודע')

        # טרקלין ויקוק
        _show_lounge_wecook = unit not in NO_LOUNGE_WECOOK_UNITS
        if _show_lounge_wecook:
            st.markdown("#### ☕ טרקלין")
            c1, c2 = st.columns(2)
            t_private = radio_with_explanation("האם יש כלים פרטיים?", "t1", col=c1)
            t_kitchen_tools = radio_with_explanation("האם יש כלי מטבח?", "t2", col=c2)
            c1, c2 = st.columns(2)
            t_procedure = radio_with_explanation("האם נשמר נוהל סגירה?", "t3", col=c1)
            t_friday = radio_with_explanation("האם הכלים החשמליים סגורים בשבת?", "t4", col=c2)
            t_app = radio_with_explanation("האם מולאה אפליקציה לטרקלין?", "t5")
            st.markdown("#### 🍳 WeCook / ויקוק")
            w_location = st.text_input("מיקום הוויקוק")
            c1, c2 = st.columns(2)
            w_private = radio_with_explanation("האם יש כלים פרטיים בוויקוק?", "w1", col=c1)
            w_kitchen_tools = radio_with_explanation("האם יש כלי מטבח בוויקוק?", "w2", col=c2)
            c1, c2 = st.columns(2)
            w_procedure = radio_with_explanation("האם עובד לפי פקודה?", "w3", col=c1)
            w_guidelines = radio_with_explanation("האם יש הנחיות?", "w4", col=c2)
        else:
            t_private = t_kitchen_tools = t_procedure = t_friday = t_app = "לא רלוונטי"
            w_location = ""
            w_private = w_kitchen_tools = w_procedure = w_guidelines = "לא רלוונטי"

        st.components.v1.html("""<div style='text-align:center;margin-top:8px;'>
            <button onclick="window.parent.document.querySelectorAll('[data-baseweb=tab]')[3].click()" 
                style='background:#1e3a8a;color:white;border:none;border-radius:10px;padding:12px 28px;font-size:17px;font-weight:700;cursor:pointer;box-shadow:0 2px 8px rgba(0,0,0,0.2);direction:rtl;'>
                ⬅️ עבור לטאב הבא: 🕍 בית כנסת ועירוב
            </button></div>""", height=70)

    # ===========================================
    # TAB 2: בית כנסת ועירוב
    # ===========================================
    with tab2:
        # ⏱️ Tab timer
        if "tab2_start" not in st.session_state:
            st.session_state["tab2_start"] = time.time()
        st.markdown("#### 🕍 בית כנסת")
        c_torah1, c_torah2 = st.columns(2)
        s_torah_id = c_torah1.text_input("מס' צ' של ספר התורה", placeholder="לדוגמה: 12345", help="הזן את המספר הצה''לי של הספר")
        s_torah_nusach = c_torah2.selectbox("נוסח ספר התורה", ["ספרדי", "אשכנז", "תימן", "אחר", "לא ידוע"])
        c1, c2 = st.columns(2)
        s_board = radio_with_explanation("האם לוח רבנות מעודכן?", "s1", col=c1)
        s_clean = radio_with_explanation("האם בית הכנסת נקי?", "s7", col=c2)
        s_books = st.multiselect("ספרי יסוד קיימים:", ["תורת המחנה", "לוח דינים", "הלכה כסדרה", "שו״ת משיב מלחמה"])
        c1, c2 = st.columns(2)
        s_havdala = radio_with_explanation("האם יש ערכת הבדלה והדלקת נרות שבת?", "s3", col=c1)
        s_gemach = radio_with_explanation("האם יש גמ״ח טלית ותפילין?", "s4", col=c2)
        c1, c2 = st.columns(2)
        s_smartbis = radio_with_explanation("האם יש תקלת בינוי (אם כן עדכנת בסמארט-ביס)?", "s5", col=c1)
        s_geniza = radio_with_explanation("האם יש פח גניזה?", "s6", col=c2)

        # שדות נוספים לחטיבות 35/89/900
        _show_brigade_shul = unit in NO_LOUNGE_WECOOK_UNITS
        hq_board_info = "לא רלוונטי"
        hq_tefillin_stand = "לא רלוונטי"
        if _show_brigade_shul:
            st.markdown("#### 📋 לוח רבנות מורחב (חטיבות 35/89/900)")
            c1, c2 = st.columns(2)
            hq_board_info = radio_with_explanation(
                "לוח הרבנות מכיל: לוז שבת, דרכי תקשורת, מפת עירוב, לוז ישיבה?", "hq_board_info", col=c1)
            hq_tefillin_stand = radio_with_explanation(
                "האם קיימת עמדת טלית ותפילין?", "hq_tefillin_stand", col=c2)

        st.markdown("#### 🚧 עירוב")
        c1, c2 = st.columns(2)
        e_status = c1.selectbox("סטטוס עירוב", ["תקין", "פסול", "בטיפול"])
        e_check = radio_with_explanation("האם בוצעה בדיקה?", "e1", col=c2)
        c1, c2 = st.columns(2)
        e_doc = radio_with_explanation("האם בוצע תיעוד?", "e2", col=c1)
        e_photo = radio_with_explanation("האם קיימת תצ״א?", "e3", col=c2)

        # שדות עירוב מורחבים לחטיבות 35/89/900
        hq_eruv_door_shape = "לא רלוונטי"
        hq_eruv_fence_work = "לא רלוונטי"
        hq_shabbat_device_board = "לא רלוונטי"
        if _show_brigade_shul:
            st.markdown("#### 🔗 עירוב מורחב (35/89/900)")
            c1, c2 = st.columns(2)
            hq_eruv_door_shape = radio_with_explanation(
                "האם קיימת צורת הפתח לעירוב?", "hq_eruv_door", col=c1)
            hq_eruv_fence_work = radio_with_explanation(
                "האם יש עבודת גדר פתוחה שפוגעת בעירוב?", "hq_eruv_fence", col=c2)
            c1, c2 = st.columns(2)
            hq_shabbat_device_board = radio_with_explanation(
                "האם יש שילוט על התקני שבת הזמינים?", "hq_sdb", col=c1)

        # 🍯 Honeypot Question (33% chance)
        import random as _hp_rand
        _show_honeypot = _hp_rand.random() < 0.33
        honeypot_answer = "לא מוצג"
        if _show_honeypot:
            # שאלת מלכודת שונה לפי סוג היחידה
            if unit in NO_LOUNGE_WECOOK_UNITS:
                hp_text = "האם מכונת הברד בטרקלין הוכשרה לפסח?"
            else:
                hp_text = "האם פח הגניזה ב'חדר השמרים' ריק?"
            
            honeypot_answer = st.radio(
                hp_text, 
                options=["כן", "לא", "לא רלוונטי / לא קיים במוצב"],
                key=f"honeypot_lounge_q_{base}",
                index=None,
                horizontal=True
            )

        st.components.v1.html("""<div style='text-align:center;margin-top:8px;'>
            <button onclick="window.parent.document.querySelectorAll('[data-baseweb=tab]')[4].click()" 
                style='background:#1e3a8a;color:white;border:none;border-radius:10px;padding:12px 28px;font-size:17px;font-weight:700;cursor:pointer;box-shadow:0 2px 8px rgba(0,0,0,0.2);direction:rtl;'>
                ⬅️ עבור לטאב הבא: 📜 נהלים ורוח
            </button></div>""", height=70)

    # ===========================================
    # TAB 3: נהלים ורוח (Procedures, Torah, Shichat Chetek)
    # ===========================================
    with tab3:
        # 🆕 Wave 2.5: GPS Checkpoint 2
        render_gps_checkpoint(2, base)
        
        # ⏱️ Tab timer
        if "tab3_start" not in st.session_state:
            st.session_state["tab3_start"] = time.time()
        st.markdown("#### 📜 נהלים")
        c1, c2 = st.columns(2)
        r_sg = radio_with_explanation("האם יש הוראות רבנות בש.ג?", "r1", col=c1)
        r_hamal = radio_with_explanation("האם יש הוראות רבנות בחמ״ל?", "r2", col=c2)
        c1, c2 = st.columns(2)
        r_sign = radio_with_explanation("האם יש שילוט על מתקנים שיש בהם חילול שבת (כגון תמי 4)?", "r3", col=c1)
        r_netilot = radio_with_explanation("האם קיימות נטלות?", "r4", col=c2)
        c1, c2 = st.columns(2)
        r_mezuzot_missing = c1.number_input("כמה מזוזות חסרות?", 0)
        r_shabbat_device = c2.radio("האם קיימים התקני שבת?", ["כן", "לא", "חלקי"], horizontal=True, key="r5")

        # שדות נהלים מורחבים לחטיבות
        hq_shabbat_conduct = "לא רלוונטי"
        if unit in NO_LOUNGE_WECOOK_UNITS:
            st.markdown("#### 🕊️ התנהלות שבת (35/89/900)")
            hq_shabbat_conduct = st.radio(
                "התנהלות בשבת תקינה? (נוהל שבת, מד\"סים, שירות משותף)",
                ["כן", "חלקי", "לא", "לא בדקתי"], horizontal=True, key="hq_shabbat_conduct"
            )

        st.markdown("#### 📖 רוח ושיעורי תורה")
        if is_combat_brigade:
            st.markdown("""
            <div style='background:#f1f5f9; padding:15px; border-radius:8px; border:1px solid #cbd5e1; margin-bottom: 15px;'>
            <h5 style='margin-top: 0; color:#1e3a8a;'>📚 מעקב שיעורים וימי ישיבה (טופס מורחב סדיר)</h5>
            """, unsafe_allow_html=True)
            c_l1, c_l2 = st.columns(2)
            lesson_date = c_l1.date_input("תאריך השיעור", key="l_date")
            lesson_location = c_l2.text_input("מיקום", key="l_loc")
            c_l3, c_l4 = st.columns(2)
            lesson_qty = c_l3.number_input("כמות שיעורים", min_value=0, key="l_qty")
            lesson_participants = c_l4.number_input("כמות משתתפים משוערת", min_value=0, key="l_part")
            lesson_content = st.text_input("תוכן השיעורים", key="l_cont")
            lesson_instructors = st.text_area("מעבירי השיעורים (הפרד בפסיקים)", key="l_inst")
            lesson_population = st.selectbox("סוג האוכלוסיה", ["חיילי חובה", "קצינים", "מילואים", "מעורב", "אחר"], key="l_pop")
            st.markdown("</div>", unsafe_allow_html=True)
            # שיחת חתך לא רלוונטית לחטיבות סדירות
            soldier_shabbat_training = soldier_knows_rabbi = soldier_prayers = soldier_talk_cmd = "לא רלוונטי"
            soldier_yeshiva = soldier_want_lesson = soldier_has_lesson = soldier_food = "לא רלוונטי"
            soldier_lesson_teacher = soldier_lesson_phone = ""
        else:
            c1, c2 = st.columns(2)
            soldier_yeshiva = radio_with_explanation("האם יש ימי ישיבה?", "so1", col=c1)
            soldier_want_lesson = radio_with_explanation("האם יש רצון לשיעור תורה?", "so_want_lesson", col=c2)
            c1, c2 = st.columns(2)
            soldier_has_lesson = radio_with_explanation("יש שיעור תורה במוצב?", "so_has_lesson", col=c1)
            soldier_food = radio_with_explanation("האם המענה הכשרותי מספק?", "so2", col=c2)
            soldier_lesson_teacher = ""
            soldier_lesson_phone = ""
            if soldier_has_lesson == "כן":
                col_teacher, col_phone = st.columns(2)
                with col_teacher:
                    soldier_lesson_teacher = st.text_input("שם מעביר השיעור", key="so_lesson_teacher", placeholder="לדוגמה: הרב כהן")
                with col_phone:
                    soldier_lesson_phone = st.text_input("טלפון מעביר השיעור", key="so_lesson_phone", placeholder="לדוגמה: 050-1234567")

            st.markdown("#### 💬 שיחת חתך חיילים")
            c1, c2 = st.columns(2)
            if _flip == 1:
                soldier_shabbat_training = radio_with_explanation("האם יש אימונים בשבת? ✅[תשובה שלילית = תקין]", "so3", col=c1)
            else:
                soldier_shabbat_training = radio_with_explanation("האם יש אימונים בשבת?", "so3", col=c1)
            soldier_knows_rabbi = radio_with_explanation("האם מכיר את הרב?", "so4", col=c2)
            c1, c2 = st.columns(2)
            soldier_prayers = radio_with_explanation("האם יש זמני תפילות?", "so5", col=c1)
            if _flip == 2:
                soldier_talk_cmd = radio_with_explanation("האם יש שיח מפקדים? ✅[תשובה שלילית = בעיה]", "so6", col=c2)
            else:
                soldier_talk_cmd = radio_with_explanation("האם יש שיח מפקדים?", "so6", col=c2)

        # שיעורי תורה ואיש קשר רוחני — לחטיבות 35/89/900 (הועבר מ-שיחת חתך)
        if unit in NO_LOUNGE_WECOOK_UNITS:
            st.markdown("#### 📖 שיעורי תורה ואיש קשר רוחני")
            hq_vars['hq_want_torah_lesson'] = st.selectbox(
                "האם יש רצון לשיעור תורה?",
                ["כן", "לא", "לא בדקתי"],
                key="hq_want_torah_lesson_sel"
            )
            c1, c2 = st.columns(2)
            hq_vars['hq_contact_name'] = c1.text_input("איש קשר – שם", key="hq_contact_name_inp", placeholder="לדוגמה: הרב כהן")
            hq_vars['hq_contact_phone'] = c2.text_input("איש קשר – מספר פלאפון", key="hq_contact_phone_inp", placeholder="לדוגמה: 050-1234567")

        st.components.v1.html("""<div style='text-align:center;margin-top:8px;'>
            <button onclick="window.parent.document.querySelectorAll('[data-baseweb=tab]')[5].click()" 
                style='background:#1e3a8a;color:white;border:none;border-radius:10px;padding:12px 28px;font-size:17px;font-weight:700;cursor:pointer;box-shadow:0 2px 8px rgba(0,0,0,0.2);direction:rtl;'>
                ⬅️ עבור לטאב הבא: 📖 שיחת חתך
            </button></div>""", height=70)

    # ===========================================
    # TAB 4: שיחת חתך (35/89/900 only)
    # ===========================================
    with tab4:
        _show_halacha = unit in NO_LOUNGE_WECOOK_UNITS
        # hq_vars was pre-initialized before tabs - do NOT reset here to preserve Tab 2 entries
        if not _show_halacha:
            st.info("📌 שיחת חתך   , ,  .")
        else:
            st.info("📌 פנה לטאב 'נהלים ורוח' לעדכון שיעורי תורה ואיש קשר.")

        st.markdown("#### 👮 שאלון חיילים – רבנות היחידה ונושאים נוספים")

        # צום ותפילות – מועבר מה-heading הנפרד
        st.markdown("##### 🙏 תפילות")
        c1, c2 = st.columns(2)
        hq_vars['hq_prayer_times'] = radio_with_explanation("החיילים מקבלים זמני תפילות לפי פקודות?", "hq42", col=c1)
        hq_vars['hq_pre_prayer_act'] = radio_with_explanation("עושים פעילות לפני זמן תפילת בוקר?", "hq43", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_post_prayer_meal'] = radio_with_explanation("החיילים מקבלים ארוחת בוקר לאחר תפילת הבוקר?", "hq44", col=c1)
        hq_vars['hq_minyan'] = radio_with_explanation("מאפשרים לחיילים להתפלל במניין (ביחידה בה אפשרי)?", "hq45", col=c2)

        st.markdown("##### 📅 צומות")
        c1, c2 = st.columns(2)
        hq_vars['hq_rosh_shofar'] = radio_with_explanation("מאפשרים לכל חייל לשמוע קול שופר בראש השנה?", "hq28", col=c1)
        hq_vars['hq_fast_shoes'] = radio_with_explanation("אפשרו לצמים לנעול נעליים ללא עור ביו\"כ ות\"ב (מלבד פעילות מבצעית)?", "hq29", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_fast_meals'] = radio_with_explanation("לכל צם הוגשה ארוחה חמה לפני ואחרי הצום?", "hq30", col=c1)
        hq_vars['hq_yom_kippur_closed'] = radio_with_explanation("קנטינות, מזנונים וחד\"א היו סגורים במהלך יום כיפור?", "hq31", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_tisha_bav_events'] = radio_with_explanation("התקיימו בתשעה באב פעילות בידור / הווי / תרבות?", "hq32", col=c1)
        hq_vars['hq_fast_exempt'] = radio_with_explanation("חיילים צמים שוחררו מפעילות (כולל הוראת קרפ\"ר) לפני ואחרי הצום?", "hq33", col=c2)

        st.markdown("##### 🧑‍💼 כשרות ויהדות")

        c1, c2 = st.columns(2)
        hq_vars['hq_know_rabbi'] = radio_with_explanation("מכירים את סגל הדת ביחידה (רב / נגד רבנות)?", "hq46", col=c1)
        hq_vars['hq_kashrut_gaps'] = radio_with_explanation("ישנם פערי כשרות ביחידה בשגרה?", "hq47", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_mehadrin_req'] = radio_with_explanation("ביקשתם מוצרי מהדרין / חלק וקיבלתם?", "hq48", col=c1)
        hq_vars['hq_six_hours'] = radio_with_explanation("יש הפרדה של 6 שעות בין ארוחה בשרית לחלבית?", "hq49", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_tools_marked'] = radio_with_explanation("הכלים מסומנים ויש הפרדה בין בשר לחלב?", "hq50", col=c1)
        hq_vars['hq_field_cooking'] = radio_with_explanation("מתקיים בישול בשטח / על האש עם פיקוח כשרותי?", "hq51", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_shabbat_comms'] = radio_with_explanation("ישנן פניות ברשת הקשר / טלפוניות לצרכים לא מבצעיים בשבת?", "hq52", col=c1)
        hq_vars['hq_shabbat_logistics'] = radio_with_explanation("מתקיים ניוד מזון / לוגיסטי בשבת?", "hq53", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_shabbat_movement'] = radio_with_explanation("מתקיים ניוד אנשים לעמדות / שמירות בשבת שלא לצורך מבצעי?", "hq54", col=c1)
        hq_vars['hq_shabbat_vehicles'] = radio_with_explanation("נסיעות ביחידה בשבת שלא לצרכים מבצעיים?", "hq55", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_shabbat_entry'] = radio_with_explanation("קיימים מנגנוני בקרת כניסה מותאמים לשבת?", "hq56", col=c1)
        hq_vars['hq_shabbat_pen'] = radio_with_explanation("התאפשר לקבל עט שבת / מקלדת / עכבר שבת?", "hq57", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_shabbat_procedure'] = radio_with_explanation("קיים נוהל שבת – שחרור שעה לפני כניסה, חזרה חצי שעה אחרי?", "hq58", col=c1)
        hq_vars['hq_shabbat_return'] = radio_with_explanation("בחזרה ממוצ\"ש – לא נדרשו לצאת פחות משעה אחרי השבת?", "hq59", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_shabbat_kiddush'] = radio_with_explanation("התקיים קידוש וסעודת ליל שבת לכל חיילי היחידה?", "hq60", col=c1)
        hq_vars['hq_shabbat_meal_timing'] = radio_with_explanation("סעודת שבת מתקיימת לאחר סיום התפילה (כשעה ורבע אחרי כניסת שבת)?", "hq61", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_challot'] = radio_with_explanation("קיבלתם חלות / לחמניות שלמות ויין בליל שבת ובשחרית?", "hq62", col=c1)
        hq_vars['hq_candles'] = radio_with_explanation("יש מקום ונרות להדלקת נרות שבת / ערכת הבדלה?", "hq63", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_shabbat_drills'] = radio_with_explanation("מבוצעים תרגילים ותרגולות בשבת?", "hq64", col=c1)
        hq_vars['hq_food_warming'] = radio_with_explanation("נהלי חימום מזון בשבת מתקיימים במטבח ללא פערים?", "hq65", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_forced_reg'] = radio_with_explanation("מתבצע רישום מחייב שאינו חיוני בשבת (חתימות שומרים, מרפאה וכד')?", "hq66", col=c1)
        hq_vars['hq_eruv_problem'] = radio_with_explanation("קיימת בעיה עם עירוב שבת ביחידה / בשטח וקיבלתם מענה?", "hq67", col=c2)
        hq_vars['hq_shabbat_violation'] = radio_with_explanation("מתקיים חילול שבת יחידתי לצורך שאינו מבצעי?", "hq68")
        c1, c2 = st.columns(2)
        hq_vars['hq_soldier_prayer_allowed'] = radio_with_explanation("מתאפשר לכם להתפלל ומקבלים זמן מוקצה (כולל הליכה וחזרה)?", "hq69", col=c1)
        hq_vars['hq_soldier_minyan'] = radio_with_explanation("מתאפשר להתפלל במניין?", "hq70", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_pre_shacharit'] = radio_with_explanation("נדרשים להשתתף במד\"סים / פעילות לפני תפילת שחרית?", "hq71", col=c1)
        hq_vars['hq_breakfast_after_prayer'] = radio_with_explanation("יש מענה לארוחת בוקר בסיום שחרית?", "hq72", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_sleep_prayer'] = radio_with_explanation("זמן תפילת שחרית הוא חלק מזמן השינה?", "hq73", col=c1)
        hq_vars['hq_mincha_arvit_time'] = radio_with_explanation("מוקצה זמן נפרד לתפילות מנחה וערבית מהארוחה?", "hq74", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_arvit_tash'] = radio_with_explanation("זמן ערבית הוא חלק משעת ת\"ש? אם כן – קיבלתם תוספת זמן?", "hq75", col=c1)
        hq_vars['hq_fast_exempt_soldier'] = radio_with_explanation("במהלך הצומות – הצמים פטורים מכל פעילות?", "hq76", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_fast_office'] = radio_with_explanation("נדרשתם לעבודה משרדית / אחרת בצום?", "hq77", col=c1)
        hq_vars['hq_fast_break_meal'] = radio_with_explanation("יש ארוחה חמה בסיום הצום לצמים?", "hq78", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_intense_pre_fast'] = radio_with_explanation("התקיימה פעילות גופנית עצימה לפני / בסיום הצום (בניגוד לנהלי קרפ\"ר)?", "hq79", col=c1)
        hq_vars['hq_drills_in_fast'] = radio_with_explanation("התקיימה פעילות חריגה לא מבצעית בצום (תרגילים, מטווחים, מד\"סים)?", "hq80", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_holiday_equipment'] = radio_with_explanation("ניתן מענה בחגים (מגילה, חנוכיות וכד')?", "hq81", col=c1)
        hq_vars['hq_mezuzot_gap'] = radio_with_explanation("ישנו פער במזוזות ביחידה?", "hq82", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_holiday_equip_recv'] = radio_with_explanation("מקבלים ציוד מותאם לחגים?", "hq83", col=c1)
        hq_vars['hq_religion_equip_req'] = radio_with_explanation("פניתם וביקשתם ציוד דת ולא קיבלתם?", "hq84", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_shul_clean'] = radio_with_explanation("בית הכנסת מטופל / עובר ניקיון שוטף?", "hq85", col=c1)
        hq_vars['hq_shul_equip_missing'] = radio_with_explanation("ישנו ציוד חסר בבית הכנסת?", "hq86", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_shul_sefer_torah'] = radio_with_explanation("בשגרה: ספר תורה, תפילין, טליתות, כיפות, נרות – תקינים?", "hq87", col=c1)
        hq_vars['hq_yeshiva_days'] = radio_with_explanation("מתקיימים ימי ישיבה ביחידה ומאפשרים לדתיים להשתתף?", "hq88", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_torah_lessons'] = radio_with_explanation("מתקיימים שיעורי תורה קבועים / רב מגיע לפחות פעם בחודש?", "hq89", col=c1)
        hq_vars['hq_spiritual_shabbat'] = radio_with_explanation("ישנו ליווי רוחני בשבתות?", "hq90", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_culture_exemption'] = radio_with_explanation("מאפשרים לדתיים להשתחרר מפעילויות תרבות שאינן מתאימות?", "hq91", col=c1)
        hq_vars['hq_gym_separate'] = radio_with_explanation("ישנן שעות נפרדות בחדר כושר / בריכה?", "hq92", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_sport_gender'] = radio_with_explanation("מאפשרים פעילות ספורטיבית מגדרית?", "hq93", col=c1)
        hq_vars['hq_yichud'] = radio_with_explanation("שמירות / סיורים / תורנויות הגורמות למצבי ייחוד? פנייה קיבלה מענה?", "hq94", col=c2)
        c1, c2 = st.columns(2)
        hq_vars['hq_alt_activity'] = radio_with_explanation("ישנה פעילות אלטרנטיבית לאוכלוסייה הדתית כשלא ניתן להשתתף בפעילות היחידה?", "hq95", col=c1)
        hq_vars['hq_cmd_sensitivity'] = radio_with_explanation("המפקדים רגישים לצרכים הדתיים (תפילות ועוד)?", "hq96", col=c2)
        st.components.v1.html("""<div style='text-align:center;margin-top:8px;'>
            <button onclick="window.parent.document.querySelectorAll('[data-baseweb=tab]')[6].click()" 
                style='background:#1e3a8a;color:white;border:none;border-radius:10px;padding:12px 28px;font-size:17px;font-weight:700;cursor:pointer;box-shadow:0 2px 8px rgba(0,0,0,0.2);direction:rtl;'>
                ⬅️ עבור לטאב הבא: ⚠️ חוסרים ושליחה
            </button></div>""", height=70)

    # ===========================================
    # TAB 5: חוסרים ושליחה (Deficits + Submit)
    # ===========================================
    with tab5:
        # 🆕 Wave 2.5: GPS Checkpoint 3
        render_gps_checkpoint(3, base)
        st.markdown("### ⚠️ חוסרים")
        missing = st.text_area("פירוט חוסרים")
        st.markdown("### 💬 הערות נוספות")
        free_text = st.text_area("הערות נוספות")

        # 💡 Inspector Tip — Crowdsourced Intel ("העברת מקל")
        st.markdown("### 💡 טיפ למבקר הבא (רשות)")
        inspector_tip = st.text_input(
            "השאר טיפ קצר למי שיבקר כאן בשבוע הבא:",
            placeholder="לדוגמה: שים לב לברז המים החמים ליד המקפיא — נשכח לסגור לפני שבת.",
            key="inspector_tip_input",
            max_chars=200
        )

        st.markdown("### 📸 צילומים וחתימה")
        photo = st.file_uploader("📸 תמונה כללית (חובה)", type=['jpg', 'png', 'jpeg'], key="main_report_photo")

        # --- בדיקת תנאי חובה לפני שליחה ---
        _mandatory_warnings = []
        
        # 📍 טאב מיקום (חובה)
        if not inspector:
            _mandatory_warnings.append("⚠️ חובה להזין שם מבקר (לשונית 📍 מיקום)")
        if not base:
            _mandatory_warnings.append("⚠️ חובה להזין מוצב / מיקום (לשונית 📍 מיקום)")
            
        # 🍽️ טאב כשרות (חובה)
        if k_cert == "לא יודע / לא בדקתי":
            _mandatory_warnings.append("⚠️ חובה לבדוק תעודת כשרות (לשונית 🍽️ כשרות)")
            
        # 🕍 טאב ביהכ"נ ועירוב (חובה)
        if e_status not in ["תקין", "פסול", "בטיפול"]:
            _mandatory_warnings.append("⚠️ חובה לציין סטטוס עירוב (לשונית 🕍 ביהכ\"נ ועירוב)")
            
        # 📜 טאב נהלים ורוח (חובה)
        if r_sg == "לא יודע / לא בדקתי" and r_hamal == "לא יודע / לא בדקתי":
            _mandatory_warnings.append("⚠️ חובה לבדוק לפחות נוהל אחד (לשונית 📜 נהלים ורוח)")
            
        # ⚠️ טאב חוסרים ושליחה (חובה)
        if not photo:
            _mandatory_warnings.append("⚠️ חובה להעלות תמונה כללית של הביקורת")
        if not missing:
            _mandatory_warnings.append("⚠️ חובה למלא את שדה הפירוט (במידה ואין חוסרים כתוב 'אין')")

        # 🧠 ניתוח ראייה ממוחשבת (Vision AI)
        if photo:
            with st.spinner("🧠 המוח הפיקודי מנתח את התמונה..."):
                vision_analysis = analyze_photo_with_vision(photo.getvalue())
                st.session_state["last_vision_analysis"] = vision_analysis # Store for submission
                if vision_analysis:
                    st.markdown("##### 👓 ממצאי AI מהתמונה:")
                    c_v1, c_v2 = st.columns(2)
                    with c_v1:
                        st.write(f"🧼 ניקיון מטבח: {vision_analysis.get('kitchen_cleanliness', '?')}/5")
                        st.write(f"📜 מזוזות: {vision_analysis.get('mezuzot', 'לא אותר')}")
                    with c_v2:
                        st.write(f"🥩 כשרות: {vision_analysis.get('kashrut_issues', 'תקין')}")
                        st.write(f"🔍 אמינות דיווח: {vision_analysis.get('reliability_score', '?')}")
                    
                    if vision_analysis.get('detailed_finding'):
                        st.info(f"🧠 **ניתוח המוח הפיקודי:** {vision_analysis.get('detailed_finding')}")

                    # 🔍 Vision-Text Crosscheck — "פוליגרף AI"
                    _contradictions = []
                    try:
                        cleanliness = int(vision_analysis.get('kitchen_cleanliness', 5))
                        if cleanliness < 3 and k_issues == "לא":
                            _contradictions.append("📸 AI מזהה לכלוך חריג — אך דיווחת 'אין תקלות כשרות'")
                        mezuzot_vision = str(vision_analysis.get('mezuzot', '')).lower()
                        if ('חסר' in mezuzot_vision or 'missing' in mezuzot_vision) and int(r_mezuzot_missing or 0) == 0:
                            _contradictions.append("📸 AI מזהה מזוזות חסרות — אך דיווחת '0 מזוזות חסרות'")
                        kashrut_vision = str(vision_analysis.get('kashrut_issues', 'תקין'))
                        if kashrut_vision not in ['תקין', 'תקין.', 'ok', 'none', ''] and k_cert == "כן":
                            _contradictions.append(f"📸 AI מזהה בעיות כשרות ({kashrut_vision}) — אך דיווחת תעודת כשרות תקפה")
                    except Exception:
                        pass
                    if _contradictions:
                        st.warning("⚠️ **המוח הפיקודי זיהה סתירות אפשריות (למידע בלבד, לא חוסם שליחה):**")
                        for c in _contradictions:
                            st.write(f"🔍 {c}")
                        st.session_state["vision_contradictions"] = _contradictions
                    else:
                        st.session_state["vision_contradictions"] = []
                    
                    st.caption("ℹ️ ממצאי ה-AI נועדו לסייע ולשפר את איכות הדיווח. ניתן לשלוח את הדוח גם אם ישנן אזהרות.")
            # 🎉 Micro-interaction: הודיה על העלאת תמונה
            st.toast("📸 בום! תמונה פצצה, איזה תיעוד!", icon="📸")


        
        # תמונות כשרות מועברות לטאב 1
        # k_issues_photo (Tab 1), k_shabbat_photo (Tab 1)
        # רק לוודא שהמשתנים מאותחלים
        if 'k_issues_photo' not in dir():
            k_issues_photo = None


        sig_url = True  # מאפשר שליחה ללא חתימה

        
        # 🎉 Micro-interactions: תגובות מידיות לפעילות חייל
        if free_text and len(free_text.split()) > 10:
            st.toast("📝 פירוט של מקצוענים! זה יעזור מאוד לחטמ\"ר!", icon="📝")
        if e_status == "פסול":
            st.toast("🚧 זיהוי קריטי! הצלת את השבת של המוצב!", icon="🚧")
        if r_mezuzot_missing and int(r_mezuzot_missing) > 0:
            st.toast(f"📜 דיווחת מזוזות חסרות — כשר שיפתח מעקב אוטומטי בדוחה!", icon="📜")

        # 🚨 Emergency WhatsApp Button - רק לחטמרים (לא לחטיבות סדירות)
        if not is_combat_brigade:
            st.markdown("")
            import urllib.parse as _urlparse
            emergency_msg = _urlparse.quote(
                f"🚨 הקפץ רב חטמ\"ר | {unit} | מוצב: {base} | ביצוע: {inspector}\nפרטים: ",
                safe=""
            )
            whatsapp_url = f"https://wa.me/?text={emergency_msg}"
            st.markdown(f"""
            <a href='{whatsapp_url}' target='_blank' style='
                display:block; text-align:center; background:linear-gradient(135deg,#dc2626,#991b1b);
                color:white; padding:14px; border-radius:12px; font-size:17px; font-weight:700;
                text-decoration:none; margin:10px 0; box-shadow:0 4px 12px rgba(220,38,38,0.4);
                letter-spacing:0.5px;'>
                🚨 הקפץ את הרב עכשיו! &nbsp;&nbsp; מקרה קטסטרופאלי או דורש תגובה מידית
            </a>
            """, unsafe_allow_html=True)


        # כפתורי שליחה וטיוטה - תמיד מוצגים (השליחה תהיה חסומה אם יש בעיה)
        st.markdown("---")
        col_submit, col_draft = st.columns([3, 1])
        
        with col_draft:
            if st.button("💾 שמור טיוטה", key="save_draft_btn"):
                draft_data = {
                    "unit": unit, "base": base, "inspector": inspector,
                    "date": str(date), "time": str(time_v),
                    "timestamp": datetime.datetime.now().isoformat()
                }
                save_draft_locally(draft_data, f"{unit}_last_draft")
    
        with col_submit:
            # הלחצן תמיד פעיל כדי לאפשר למשתמש לקבל משוב על מה חסר
            submitted = st.button("🚀 שגר דיווח", type="primary", use_container_width=True, key="submit_new_report")


        if submitted:
            # חישוב משך הדיווח
            report_duration = 0
            if "report_start_time" in st.session_state:
                report_duration = int(time.time() - st.session_state.report_start_time)
                # איפוס הטיימר לדיווח הבא
                del st.session_state.report_start_time

            # ===== 🍯 Honeypot Check =====
            honeypot_failed = False
            if 'honeypot_answer' in dir() and honeypot_answer == "כן":
                honeypot_failed = True
                st.error("🚨 **נכשלת בשאלת מלכודת!** ציון האמינות שלך אופס. הדוח יסומן לבדיקת מפקד.")

            # ===== ⏱️ Tab Speed Check =====
            _quick_fill_flags = []
            _now = time.time()
            _tab_thresholds = {
                "tab1": ("כשרות (12+ שאלות)", 90),
                "tab2": ("בית כנסת ועירוב (8 שאלות)", 30),
                "tab3": ("נהלים ורוח (10 שאלות)", 30),
            }
            for _tkey, (_tlabel, _tmin) in _tab_thresholds.items():
                _tstart = st.session_state.get(f"{_tkey}_start")
                if _tstart:
                    _telapsed = _now - _tstart
                    if _telapsed < _tmin:
                        _quick_fill_flags.append(f"{_tlabel} — {_telapsed:.1f} שניות")
            if _quick_fill_flags:
                # Note: We will handle this in the final check below
                pass

            # בדיקת יום בשבוע - חמישי (3) ושישי (4) ב-Python weekday
            current_weekday = datetime.datetime.now().weekday()
            is_thursday_or_friday = current_weekday in [3, 4]

            # 🆕 בדיקת הסברים חסרים עבור "לא יודע / לא בדקתי"
            missing_explanations = []

            # This list must match the variables used above.
            answers_to_check = {
                "פק״ל רבנות": p_pakal, "כלים מסומנים": p_marked, "ערבוב כלים": p_mix, "הכשרת כלים": p_kasher,
                "הוראות בש.ג": r_sg, "הוראות בחמ״ל": r_hamal, "שילוט שבת": r_sign, "נטלות": r_netilot,
                "לוח רבנות": s_board, "ניקיון בית כנסת": s_clean, "ערכת הבדלה": s_havdala, "גמ״ח טלית ותפילין": s_gemach,
                "תקלת בינוי": s_smartbis, "פח גניזה": s_geniza,
                "בדיקת עירוב": e_check, "תיעוד עירוב": e_doc, "תצ״א עירוב": e_photo,
                "תעודת כשרות": k_cert, "בישול ישראל": k_bishul, "תקלות כשרות": k_issues, "נאמן שבת": k_shabbat_supervisor,
                "הפרדה במטבח": k_separation, "תדריך טבחים": k_briefing, "רכש חוץ": k_products, "דף תאריכים": k_dates,
                "שטיפת ירק": k_leafs, "חירור גסטרונומים": k_holes, "בדיקת ביצים": k_eggs, "חדר מכ״ש": k_machshir,
                "חימום נפרד": k_heater, "אפליקציה במטבח": k_app,
                # טרקלין ויקוק – רק אם רלוונטי ליחידה
                **({
                    "כלים פרטיים טרקלין": t_private, "כלי מטבח טרקלין": t_kitchen_tools,
                    "נוהל סגירה טרקלין": t_procedure, "סגור בשבת טרקלין": t_friday, "אפליקציה טרקלין": t_app,
                    "כלים פרטיים ויקוק": w_private, "כלי מטבח ויקוק": w_kitchen_tools,
                    "נהלים ויקוק": w_procedure, "הנחיות ויקוק": w_guidelines,
                } if _show_lounge_wecook else {}),
                "ימי ישיבה": soldier_yeshiva, "רצון לשיעור": soldier_want_lesson, "שיעור קיים": soldier_has_lesson,
                "מענה כשרותי": soldier_food, "אימונים בשבת": soldier_shabbat_training, "מכיר את הרב": soldier_knows_rabbi,
                "זמני תפילות": soldier_prayers, "שיח מפקדים": soldier_talk_cmd
            }
            
            for label, value in answers_to_check.items():
                if isinstance(value, str) and value.startswith("__MISSING_EXPLANATION__"):
                    missing_explanations.append(label)
            
            if missing_explanations:
                st.error("❌ לא ניתן לשלוח את הדוח! חסר פירוט עבור התשובות 'לא יודע / לא בדקתי':")
                for item in missing_explanations:
                    st.warning(f"⚠️ {item} - חובה לפרט סיבה בתיבת הטקסט")
            
            # 🆕 בדיקת תנאי חובה (אזהרות שהוגדרו למעלה)
            elif _mandatory_warnings:
                st.error("❌ **לא ניתן לשלוח את הדוח! חסרים פרטי חובה:**")
                for w in _mandatory_warnings:
                    st.error(w)
                st.info("💡 אנא השלם את כל השדות המסומנים ב-⚠️ ועבור שוב על הלשוניות.")

            # 🆕 בדיקת מילוי מהיר (speed-check)
            elif _quick_fill_flags:
                st.error("⚡ **לא ניתן לשלוח - זיהינו מילוי מהיר מדי!**")
                st.info("אנא וודא שעברת על כל הסעיפים ביסודיות ולא רק סימנת 'וי'.")
                for f in _quick_fill_flags:
                    st.warning(f"⚠️ {f}")

            # בדיקת חובת תמונת נאמן כשרות בחמישי-שישי
            elif is_thursday_or_friday and k_shabbat_supervisor == "כן" and not k_shabbat_photo:
                st.error("⚠️ **חובה להעלות תמונת נאמן כשרות בימי חמישי ושישי!**")
                st.warning("💡 נא להעלות תמונה של נאמן הכשרות בשדה המתאים למעלה")
            
            # בדיקת מיקום חובה (פטור לחטיבות סדירות)
            elif not is_combat_brigade and not (gps_lat and gps_lon):
                 st.error("❌ חובה להפעיל מיקום (GPS) כדי לשלוח את הדוח בחטמ\"ר!")
                 st.warning("💡 אנא וודא שה-GPS דולק ואישרת לדפדפן לגשת למיקום")
                 
            elif base and inspector and photo:
                photo_url = upload_report_photo(photo.getvalue(), unit, base)
                
                # העלאת תמונות נוספות (תקלות כשרות ונאמן כשרות)
                k_issues_photo_url = None
                k_shabbat_photo_url = None
                
                if k_issues_photo:
                    k_issues_photo_url = upload_report_photo(k_issues_photo.getvalue(), unit, f"{base}_kashrut_issue")
                
                if k_shabbat_photo:
                    k_shabbat_photo_url = upload_report_photo(k_shabbat_photo.getvalue(), unit, f"{base}_shabbat_supervisor")
                
                data = {
                    "unit": st.session_state.selected_unit, "date": datetime.datetime.now().isoformat(),
                    "base": base, "inspector": inspector, "photo_url": photo_url,
                    "k_cert": k_cert, "k_dates": k_dates,
                    "e_status": e_status,
                    "s_clean": s_clean,
                    "t_private": t_private, "t_kitchen_tools": t_kitchen_tools, "t_procedure": t_procedure,
                    "t_friday": t_friday, "t_app": t_app, "w_location": w_location, "w_private": w_private,
                    "w_kitchen_tools": w_kitchen_tools, "w_procedure": w_procedure, "w_guidelines": w_guidelines,
                    "soldier_yeshiva": soldier_yeshiva,
                    "soldier_want_lesson": soldier_want_lesson,
                    "soldier_has_lesson": soldier_has_lesson,
                    "soldier_lesson_teacher": soldier_lesson_teacher,
                    "soldier_lesson_phone": soldier_lesson_phone,
                    "soldier_food": soldier_food,
                    "soldier_shabbat_training": soldier_shabbat_training, "soldier_knows_rabbi": soldier_knows_rabbi,
                    "soldier_prayers": soldier_prayers, "soldier_talk_cmd": soldier_talk_cmd, 
                    "free_text": free_text + (f"\n[תמלול קולי]: {st.session_state.get('voice_note_transcription', '')}" if st.session_state.get('voice_note_transcription') else ""),
                    "time": str(time_v), "p_pakal": p_pakal, "missing_items": missing,
                    "r_mezuzot_missing": r_mezuzot_missing, "k_cook_type": k_cook_type,
                    "p_marked": p_marked, "p_mix": p_mix, "p_kasher": p_kasher,
                    "r_sg": r_sg, "r_hamal": r_hamal, "r_sign": r_sign, "r_netilot": r_netilot,
                    "r_shabbat_device": r_shabbat_device, "s_board": s_board, "s_books": str(s_books),
                    "s_havdala": s_havdala, "s_gemach": s_gemach, "s_smartbis": s_smartbis, "s_geniza": s_geniza,
                    "s_torah_id": s_torah_id, "s_torah_nusach": s_torah_nusach,
                    "e_check": e_check, "e_doc": e_doc, "e_photo": e_photo,
                    "k_separation": k_separation, "k_briefing": k_briefing, "k_products": k_products,
                    "k_leafs": k_leafs, "k_holes": k_holes, "k_bishul": k_bishul,
                    "k_eggs": k_eggs, "k_machshir": k_machshir, "k_heater": k_heater, "k_app": k_app,
                    "k_issues": k_issues,
                    "k_issues_description": k_issues_description,
                    "k_shabbat_supervisor": k_shabbat_supervisor,
                    "k_shabbat_supervisor_name": k_shabbat_supervisor_name,
                    "k_shabbat_supervisor_phone": k_shabbat_supervisor_phone,
                    "k_issues_photo_url": k_issues_photo_url,
                    "k_shabbat_photo_url": k_shabbat_photo_url,
                    "report_duration": report_duration,
                    "barcode_verified": (barcode_value == BASE_BARCODES.get(base)) if base in BASE_BARCODES else False,
                    "signature_url": sig_url or "",
                    "honeypot_triggered": honeypot_failed,
                    "quick_fill_flags": str(_quick_fill_flags) if _quick_fill_flags else "",
                    "vision_contradictions": str(st.session_state.get("vision_contradictions", [])),
                    "vision_detailed_finding": st.session_state.get("last_vision_analysis", {}).get("detailed_finding", ""),
                    "inspector_tip": inspector_tip,
                    **continuity_answers
                }
                
                # 🆕 Wave 3 integration: חטיבות סדירות
                if is_combat_brigade:
                    data.update({
                        "lesson_date": str(lesson_date),
                        "lesson_location": lesson_location,
                        "lesson_qty": lesson_qty,
                        "lesson_participants": lesson_participants,
                        "lesson_content": lesson_content,
                        "lesson_instructors": lesson_instructors,
                        "lesson_population": lesson_population,
                        "hq_shabbat_conduct": hq_shabbat_conduct
                    })
                
                # הוספת שאלות הלכה לחטיבות 35/89/900
                if hq_vars:
                    data.update(hq_vars)
                
                # 🆕 Wave 2.5: שדרוג שמירה עם ניתוח אמינות
                if save_report_with_analysis(data, unit):
                    time.sleep(2)
                    st.rerun()
                    data.update(hq_vars)
                
                # הוספת מיקום רק אם קיים ואם הטבלה תומכת בזה
                if gps_lat and gps_lon:
                    # ✅ בדיקה נוספת שהמיקום תקין
                    if 29.5 <= gps_lat <= 33.5 and 34.2 <= gps_lon <= 35.9:
                        # הוספת רעש למיקום GPS לצורכי אבטחה (~500 מטר)
                        # ✅ שימוש ב-secure_location_offset עם ID יציב
                        unique_id_for_offset = f"{unit}_{base}"
                        lat_with_offset, lon_with_offset = secure_location_offset(gps_lat, gps_lon, unique_id_for_offset, offset_meters=500)
                        data["latitude"] = lat_with_offset
                        data["longitude"] = lon_with_offset
                        
                        # ✅ הדפסה ללוג
                        print(f"💾 שומר למסד נתונים: lat={lat_with_offset:.6f}, lon={lon_with_offset:.6f}")
                    else:
                        st.warning("⚠️ המיקום לא נשמר כי הוא מחוץ לגבולות ישראל")

                # ===== מוח פיקודי: ניתוח AI לפני השמירה =====
                ai_insight = {}
                try:
                    with st.spinner("🧠 המוח הפיקודי מסווג את הדיווח..."):
                        ai_insight = analyze_report_with_ai(base, data)
                    data["ai_risk_level"]     = ai_insight.get("risk_level", "לא סווג")
                    data["ai_sla"]            = ai_insight.get("sla", "טרם נקבע")
                    data["ai_action"]         = ai_insight.get("recommended_action", "נדרשת בחינה ידנית")
                except Exception:
                    pass  # אף פעם לא חוסם שמירה

                try:
                    # ניסיון לשמור את הדוח
                    try:
                        result = supabase.table("reports").insert(data).execute()
                    except Exception as e:
                        # טיפול בשגיאה אם העמודות החדשות עדיין לא קיימות במסד הנתונים
                        if "PGRST204" in str(e) or "Could not find" in str(e):
                            # ניסיון חוזר ללא השדות החדשים (שמירה שקטה של בסיס הדוח)
                            # רשימת כל השדות החדשים שאולי חסרים
                            new_fields = [
                                "k_issues", "k_issues_description", "k_shabbat_supervisor",
                                "k_shabbat_supervisor_name", "k_shabbat_supervisor_phone",
                                "k_issues_photo_url", "k_shabbat_photo_url",
                                "soldier_want_lesson", "soldier_has_lesson", "soldier_lesson_teacher", "soldier_lesson_phone",
                                "report_duration", "barcode_verified", "signature_url",
                                "ai_risk_level", "ai_sla", "ai_action",
                                "vision_detailed_finding",
                                # שדות שיעורים של חטיבות סדירות
                                "lesson_date", "lesson_location", "lesson_qty", "lesson_participants",
                                "lesson_content", "lesson_instructors", "lesson_population",
                                # 🆕 Wave 2 Anti-Fraud fields
                                "honeypot_failed", "quick_fill_flags", "vision_contradictions", "inspector_tip",
                            ]
                            for field in new_fields:
                                data.pop(field, None)
                            result = supabase.table("reports").insert(data).execute()
                        else:
                            raise e


                    # מעקב אוטומטי אחר חוסרים
                    if result.data and len(result.data) > 0:
                        report_id = result.data[0].get('id')
                        if report_id:
                            detect_and_track_deficits(data, report_id, unit)
                            
                            # 🆕 יצירת כרטיס תקלה אוטומטי (Closed-Loop Ticketing)
                            create_maintenance_ticket(data, report_id)
                    
                    # ===== 🎊 Impact Screen — מסך ניצחון אחרי שגר =====
                    st.balloons()
                    # חשב רצף דיווחים
                    try:
                        streak_res = supabase.table("reports").select("date").eq("unit", unit).eq("inspector", inspector).order("date", desc=True).limit(10).execute()
                        streak_count = len(streak_res.data) if streak_res.data else 1
                    except Exception:
                        streak_count = 1
                    # חישוב אוכלוסייה בעלת השפע
                    soldiers_impacted = 150 if not is_combat_brigade else 800
                    reliability_score = min(100, 70 + streak_count * 3)
                    streak_emoji = "🔥" if streak_count >= 4 else "⭐" if streak_count >= 2 else ""
                    st.markdown(f"""
                    <div style='text-align:center; background:linear-gradient(135deg,#f0fdf4,#dcfce7); border:2px solid #10b981;
                                border-radius:16px; padding:28px 20px; margin:16px 0;'>
                        <div style='font-size:64px; margin-bottom:8px;'>🎯</div>
                        <h2 style='color:#10b981; margin:0 0 8px 0;'>הדוח שוגר לחפ"ק!</h2>
                        <p style='font-size:16px; color:#374151; margin:6px 0;'>
                            בזכותך, <b>{soldiers_impacted} לוחמים</b> במוצב יאכלו כשר ויוכלו לטלטל בשבת הקרובה.
                        </p>
                        <p style='font-size:15px; color:#374151; margin:4px 0;'>
                            {streak_emoji} <b>רצף דיווחים:</b> {streak_count} דוחות
                        </p>
                        <p style='font-size:15px; color:#374151; margin:4px 0;'>
                            📈 <b>ציון אמינות:</b> {reliability_score}/100 ({'!🏆 אלוף' if reliability_score >= 95 else '🔥 מקצוען' if reliability_score >= 85 else 'טוב מאוד'})
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
                    # 📨 התראות דוא"ל לבעיות קריטיות
                    send_email_alerts(data, unit)
                    
                    # 📊 הצגת מה השתנה מהפעם הקודמת
                    render_report_diff(data, unit, base)
                    
                    clear_cache()
                    st.session_state.voice_note_transcription = None # 🧼 ניקוי תמלול
                    time.sleep(4)  # תוספת זמן לקריאת ה-Diff
                    st.rerun()
                except Exception as e:
                    error_msg = str(e)
                    # אם השגיאה היא בגלל עמודות שלא קיימות, נסה בלעדיהן
                    if any(col in error_msg for col in ["latitude", "longitude", "photo_url"]):
                        try:
                            # הסרת עמודות שלא קיימות
                            data.pop("latitude", None)
                            data.pop("longitude", None)
                            data.pop("photo_url", None)
                            supabase.table("reports").insert(data).execute()
                            st.success("✅ הדוח נשלח בהצלחה!")
                            clear_cache()
                            time.sleep(2)
                            st.rerun()
                        except Exception as e2:
                            st.error(f"❌ שגיאה בשמירה: {e2}")
                    else:
                        st.error(f"❌ שגיאה בשמירה: {error_msg}")
            else:
                st.error("⚠️ חסרים פרטי חובה (מוצב, מבקר או תמונה)")
    
    # --- סטטיסטיקות מבקרים ---
    st.markdown("---")
    st.markdown("## 📊 סטטיסטיקות מבקרים")
    
    # טעינת דוחות של היחידה (ללא קאש)
    # ניקוי קאש לפני טעינה כדי להבטיח נתונים עדכניים
    clear_cache()
    res = supabase.table("reports").select("*").eq("unit", st.session_state.selected_unit).execute().data
    unit_reports_raw = [d for d in res if not str(d.get('inspector', '')).startswith('🗑️')] if res else []
    unit_df = pd.DataFrame(unit_reports_raw)
    
    if not unit_df.empty and 'date' in unit_df.columns:
        # המרת תאריכים
        unit_df['date'] = pd.to_datetime(unit_df['date'], errors='coerce')
        
        stats = generate_inspector_stats(unit_df)
        
        if stats:
            # מדדים עיקריים
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("📝 סה\"כ דוחות החודש", stats['total_reports'])
            with col2:
                st.metric("👥 מבקרים פעילים", stats['unique_inspectors'])
            with col3:
                if not stats['top_inspectors'].empty:
                    top_inspector = stats['top_inspectors'].index[0]
                    top_count = stats['top_inspectors'].iloc[0]
                    st.metric("🏆 מבקר מוביל", f"{top_inspector} ({top_count})")
            
            # הוספת בלוק ציון ומדד (חדש!)
            st.markdown("---")
            st.markdown("### 🎖️ מדד כשירות יחידה וסיכום פעילות")
            
            unit_score = calculate_unit_score(unit_df)
            unit_badge, badge_color = get_unit_badge(unit_score)
            
            col_s1, col_s2, col_s3 = st.columns([1, 1, 2])
            with col_s1:
                st.metric("ציון משוקלל", f"{unit_score:.1f}/100")
            with col_s2:
                st.markdown(f"<div style='background:{badge_color}; color:white; padding:10px; border-radius:8px; text-align:center; font-weight:bold; margin-top: 5px;'>{unit_badge}</div>", unsafe_allow_html=True)
            # טאבים לסטטיסטיקות
            stats_tabs = st.tabs(["🏆 טבלת מובילים", "📍 מיקומים", "⏰ שעות פעילות", "📈 התקדמות"])
            
            # טאב 1: טבלת מובילים
            with stats_tabs[0]:
                st.markdown("### 🏆 9 המבקרים המובילים")
                
                if not stats['top_inspectors'].empty:
                    # יצירת טבלה מעוצבת - 9 הראשונים
                    leaderboard_data = []
                    number_emojis = {
                        1: "🥇", 2: "🥈", 3: "🥉",
                        4: "4️⃣", 5: "5️⃣", 6: "6️⃣",
                        7: "7️⃣", 8: "8️⃣", 9: "9️⃣"
                    }
                    
                    for idx, (inspector, count) in enumerate(stats['top_inspectors'].head(9).items(), 1):
                        medal = number_emojis.get(idx, f"#{idx}")
                        leaderboard_data.append({
                            "מקום": medal,
                            "שם המבקר": inspector,
                            "דוחות": count
                        })
                    
                    leaderboard_df = pd.DataFrame(leaderboard_data)
                    
                    # תצוגה משופרת עם עיצוב ממורכז
                    # שימוש ב-HTML לעיצוב מדליות ממורכזות
                    html_table = "<table style='width:100%; text-align:center; border-collapse: collapse; color: #000000;'>"
                    html_table += "<thead><tr style='background-color: #f0f2f6;'>"
                    html_table += "<th style='padding: 12px; font-size: 16px; color: #000000;'>מקום</th>"
                    html_table += "<th style='padding: 12px; font-size: 16px; color: #000000;'>שם המבקר</th>"
                    html_table += "<th style='padding: 12px; font-size: 16px; color: #000000;'>דוחות</th>"
                    html_table += "</tr></thead><tbody>"
                    
                    for _, row in leaderboard_df.iterrows():
                        html_table += "<tr style='border-bottom: 1px solid #e0e0e0;'>"
                        html_table += f"<td style='padding: 10px; font-size: 24px; color: #000000;'>{row['מקום']}</td>"
                        html_table += f"<td style='padding: 10px; text-align: right; font-size: 16px; color: #000000;'>{row['שם המבקר']}</td>"
                        html_table += f"<td style='padding: 10px; font-size: 16px; color: #000000;'>{row['דוחות']}</td>"
                        html_table += "</tr>"
                    
                    html_table += "</tbody></table>"
                    st.markdown(html_table, unsafe_allow_html=True)
                    

                else:
                    st.info("אין נתונים זמינים")
            
            # טאב 2: מיקומים

            with stats_tabs[1]:
                st.markdown("### 📍 מפת מיקומים")
                st.info("🔐 **ביטחון מידע:** המיקומים מוזזים 300 מטר מהמיקום המדויק לצורכי אבטחת מידע")
                
                # בדיקה אם יש עמודות מיקום
                has_location_columns = not unit_df.empty and 'latitude' in unit_df.columns and 'longitude' in unit_df.columns
                
                if has_location_columns:
                    # ניקוי נתונים ריקים
                    valid_map = unit_df.dropna(subset=['latitude', 'longitude']).copy()
                    
                    if not valid_map.empty:
                        # מיפוי צבעים לפי יחידה (Folium format)
                        unit_color_map = {
                            "חטמ״ר בנימין": "#1e3a8a",
                            "חטמ״ר שומרון": "#60a5fa",
                            "חטמ״ר יהודה": "#22c55e",
                            "חטמ״ר עציון": "#fb923c",
                            "חטמ״ר אפרים": "#ef4444",
                            "חטמ״ר מנשה": "#a855f7",
                            "חטמ״ר הבקעה": "#db2777"
                        }
                        
                        # חישוב מרכז המפה
                        center_lat = valid_map['latitude'].mean()
                        center_lon = valid_map['longitude'].mean()
                        
                        # יצירת מפת Folium
                        m = create_street_level_map(center=(center_lat, center_lon), zoom_start=13)
                        
                        # הוספת כל הנקודות למפה
                        for _, row in valid_map.iterrows():
                            add_unit_marker_to_folium(m, row, unit_color_map)
                        
                        # הצגת המפה
                        st_folium(m, width=1200, height=500, returned_objects=[], key=f"map_hatmar_{unit}")
                        
                        # מקרא
                        st.markdown("#### 🔑 מקרא")
                        legend_html = "<div style='display: flex; flex-wrap: wrap; gap: 15px; margin-top: 10px;'>"
                        
                        # מקרא ייחודי ליחידה הנוכחית או כללי אם יש ערבוב
                        unique_units = sorted(valid_map['unit'].unique()) if 'unit' in valid_map.columns else [unit]
                        
                        for u in unique_units:
                            color = unit_color_map.get(u, "#808080")
                            legend_html += f"<div><span style='color: {color}; font-size: 1.5rem;'>●</span> {u}</div>"
                        legend_html += "</div>"
                        st.markdown(legend_html, unsafe_allow_html=True)
                        
                        st.success("✅ **מפה ברמת רחוב** - זום עד 20 | שמות רחובות בעברית | שכבות: רחובות + לווין")
                        st.info("💡 **נקודות גדולות** = בעיות (עירוב פסול או כשרות לא תקינה)")
                        
                    else:
                        st.info("אין נתונים עם מיקום GPS תקין להצגה.")
                else:
                    st.warning("⚠️ לא נמצאו נתוני מיקום (GPS) בדוחות היחידה.")
            
            # טאב 3: שעות פעילות
            with stats_tabs[2]:
                st.markdown("### ⏰ שעות פעילות")
                
                if not stats['peak_hours'].empty:
                    # יצירת תרשים עמודות אינטראקטיבי
                    hours_df = pd.DataFrame({
                        'שעה': [f"{int(h):02d}:00" for h in stats['peak_hours'].index],
                        'דוחות': stats['peak_hours'].values
                    })
                    
                    fig = px.bar(
                        hours_df,
                        x='שעה',
                        y='דוחות',
                        title="התפלגות דיווחים לפי שעות",
                        labels={'שעה': 'שעה ביום', 'דוחות': 'מספר דוחות'},
                        color='דוחות',
                        color_continuous_scale='Blues'
                    )
                    
                    fig.update_layout(
                        showlegend=False,
                        height=350,
                        xaxis_tickangle=-45,
                        paper_bgcolor='white',
                        plot_bgcolor='white',
                        font=dict(color='#1e293b')
                    )
                    
                    st.plotly_chart(fig, use_container_width=True)
                    
                    # פירוט מפורט של שעות פעילות
                    st.markdown("#### 📊 פירוט שעות פעילות")
                    
                    # יצירת DataFrame עם כל 24 השעות
                    all_hours = pd.DataFrame({'hour': range(24), 'count': 0})
                    activity_hours = stats['peak_hours'].reset_index()
                    activity_hours.columns = ['hour', 'count']
                    
                    # מיזוג עם כל השעות
                    hourly_data = all_hours.set_index('hour').combine_first(activity_hours.set_index('hour')).reset_index()
                    hourly_data = hourly_data.sort_values('hour')
                    
                    # הצגת גרף עמודות מפורט
                    fig_detailed = px.bar(
                        hourly_data,
                        x='hour',
                        y='count',
                        labels={'hour': 'שעה', 'count': 'מספר דוחות'},
                        title='התפלגות דוחות לפי שעה (24 שעות)',
                        color='count',
                        color_continuous_scale='Blues'
                    )
                    
                    fig_detailed.update_layout(
                        xaxis=dict(
                            tickmode='linear',
                            tick0=0,
                            dtick=1,
                            tickformat='%02d:00'
                        ),
                        showlegend=False,
                        height=400,
                        paper_bgcolor='white',
                        plot_bgcolor='white',
                        font=dict(color='#1e293b')
                    )
                    
                    st.plotly_chart(fig_detailed, use_container_width=True)
                    
                    # סטטיסטיקות מפורטות
                    active_hours = hourly_data[hourly_data['count'] > 0]
                    if len(active_hours) > 0:
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            peak_hour = active_hours.loc[active_hours['count'].idxmax(), 'hour']
                            peak_count = active_hours['count'].max()
                            st.metric("🔥 שעת שיא", f"{int(peak_hour):02d}:00", f"{int(peak_count)} דוחות")
                        with col2:
                            total_active_hours = len(active_hours)
                            st.metric("⏰ שעות פעילות", f"{total_active_hours} שעות")
                        with col3:
                            avg_per_active_hour = active_hours['count'].mean()
                            st.metric("📊 ממוצע לשעה פעילה", f"{avg_per_active_hour:.1f}")
                        
                        # רשימת שעות פעילות
                        st.markdown("**שעות עם דיווחים:**")
                        hours_list = ", ".join([f"{int(h):02d}:00 ({int(c)} דוחות)" for h, c in zip(active_hours['hour'], active_hours['count'])])
                        st.caption(hours_list)
                else:
                    st.info("אין מספיק נתונים להצגת שעות פעילות")
            
            # טאב 4: התקדמות
            with stats_tabs[3]:
                st.markdown("### 📈 גרף התקדמות")
                
                # התקדמות לפי תאריך
                daily_reports = unit_df.groupby(unit_df['date'].dt.date).size().reset_index()
                daily_reports.columns = ['תאריך', 'דוחות']
                
                fig = px.line(
                    daily_reports,
                    x='תאריך',
                    y='דוחות',
                    title="התקדמות דיווחים לאורך זמן",
                    markers=True
                )
                fig.update_layout(height=300)
                st.plotly_chart(fig, use_container_width=True)
                
                # סטטיסטיקה נוספת
                col1, col2 = st.columns(2)
                with col1:
                    avg_daily = daily_reports['דוחות'].mean()
                    st.metric("ממוצע דוחות ליום", f"{avg_daily:.1f}")
                with col2:
                    max_day = daily_reports.loc[daily_reports['דוחות'].idxmax()]
                    st.metric("יום שיא", f"{max_day['תאריך']} ({int(max_day['דוחות'])})")
        else:
            st.info("אין מספיק נתונים להצגת סטטיסטיקות")
    else:
        st.info("טרם הוגשו דוחות ליחידה זו")

# ===== Executive Summary Dashboards =====

def render_executive_summary_dashboard():
    """
    🎖️ דשבורד הפיקוד – תמונה שלמה בדקה אחת
    עיצוב: Dark Mode, גרפים יפים, מדדי ביצוע מרכזיים עם הסברים בעברית
    """
    import datetime as _dt
    
    # ===== Header Premium עם הסברים בעברית =====
    st.markdown(f"""
    <div style='background: linear-gradient(135deg, #1e3a8a 0%, #1d4ed8 100%);
                padding: 40px; border-radius: 16px; margin-bottom: 30px;
                box-shadow: 0 10px 30px rgba(30,58,138,0.2);'>
        <div style='display: flex; justify-content: space-between; align-items: center;'>
            <div>
                <h1 style='color: white; margin: 0; font-size: 36px; direction: ltr;'>
                    🎖️ Executive Summary – Pikud
                </h1>
                <p style='color: rgba(255,255,255,0.9); font-size: 16px; font-weight: 600; margin: 5px 0 10px 0; direction: rtl;'>
                    לוח מעקב פיקודי - תמונת מצב אוגדתית לאיתור מוקדי סיכון ניהוליים והלכתיים בחטיבות.
                </p>
                <p style='color: rgba(255,255,255,0.75); margin: 0; font-size: 14px; direction: ltr;'>
                    Command Operations Intelligence · {_dt.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
                </p>
            </div>
            <div style='text-align: right;'><div style='font-size: 48px;'>🎖️</div></div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ===== כרטיסי KPI פיקודיים =====
    st.markdown("""
    <h3 style='margin-bottom: 0;'>📊 Key Performance Indicators</h3>
    <p style='color: #64748b; font-size: 14px; margin-top: 0; margin-bottom: 15px;'>
        מדדי ביצוע מרכזיים - מספקים אינדיקציה מיידית לבריאות האוגדה: כמה דוחות מולאו, מה היקף החוסרים, ומה רמת הכשירות הממוצעת.
    </p>
    """, unsafe_allow_html=True)

    kpi_cols = st.columns(4)
    
    # חישוב אוגדות כפופות
    subordinate_divisions = ["אוגדת 877", "אוגדת 96", "אוגדת 98"]
    div_count = len(subordinate_divisions)
    
    all_reports = load_reports_cached(None)
    df = pd.DataFrame(all_reports) if all_reports else pd.DataFrame()
    accessible_units = get_accessible_units(st.session_state.selected_unit, st.session_state.role)

    # Calculate metrics
    issues_count = 0
    if not df.empty:
        if 'e_status' in df.columns: issues_count += len(df[df['e_status'] == 'פסול'])
        if 'k_cert' in df.columns: issues_count += len(df[df['k_cert'] == 'לא'])

    # Calculate average score (100 - risk)
    div_scores = []
    for d in subordinate_divisions:
         try:
             s = 100 - calculate_operational_risk_index(d, df)['risk_score']
             div_scores.append(s)
         except Exception:
             div_scores.append(0)
    
    avg_score = sum(div_scores) / len(div_scores) if div_scores else 0
    risk_status = "🔴 CRITICAL" if (100 - avg_score) >= 50 else "🟡 HIGH" if (100 - avg_score) >= 25 else "🟢 NORMAL"

    # Recalculate metrics for display
    critical_bases = len(df[df['e_status'] == 'פסול']) if not df.empty and 'e_status' in df.columns else 0
    no_kashrut = len(df[df['k_cert'] == 'לא']) if not df.empty and 'k_cert' in df.columns else 0
    open_deficits = get_deficit_statistics(accessible_units)['total_open'] if not df.empty else 0
    silent_units = count_silent_units(df) if not df.empty else 0
    units_list = df['unit'].unique().tolist() if not df.empty else []

    with kpi_cols[0]:
        st.metric(label="🏛️ Subordinate Divisions", value=div_count, 
                  delta="אוגדות תחת הפיקוד", delta_color="off")

    with kpi_cols[1]:
        total_reports = len(df)
        avg_reports = total_reports / div_count if div_count > 0 else 0
        st.metric(label="📋 Total Reports", value=total_reports, 
                  delta=f"ממוצע {avg_reports:.1f} לאוגדה", delta_color="normal")

    with kpi_cols[2]:
        st.metric(label="🔴 Open Issues", value=issues_count,
                  delta="בעיות ליבה הדורשות התערבות" if issues_count > 0 else "הכל תקין",
                  delta_color="inverse" if issues_count > 0 else "normal")

    with kpi_cols[3]:
        score_status = "🟢 מצוין" if avg_score >= 80 else "🟡 סביר" if avg_score >= 60 else "🔴 טעון שיפור"
        st.metric(label="📊 Avg Performance", value=f"{avg_score:.0f}/100", delta=score_status,
                  delta_color="normal" if avg_score >= 70 else "inverse")
        
    st.markdown("---")

    # ===== Subordinate Divisions Grid =====
    st.markdown("""
    <h3 style='margin-bottom: 0;'>🏛️ Subordinate Divisions Status</h3>
    <p style='color: #64748b; font-size: 14px; margin-top: 0; margin-bottom: 15px;'>
        כרטיסיית ניתוח אוגדתית - פירוט מדויק של כמות הדוחות, הציון, והחוסרים הפתוחים לכל אוגדה כפופה.
    </p>
    """, unsafe_allow_html=True)

    grid_cols = st.columns(3)
    for idx, div_name in enumerate(subordinate_divisions):
        div_df = df[df['unit'].str.contains(div_name, na=False)] if not df.empty else pd.DataFrame()
        score = 100 - calculate_operational_risk_index(div_name, df)['risk_score']
        reports = len(div_df)
        try:
             # This is a bit complex as it needs unit-level access
             # For now, we'll estimate or filter
             open_def_count = len(get_open_deficits([u for u in ALL_UNITS if div_name in u]))
        except:
             open_def_count = 0
             
        border_color = "#10b981" if score >= 80 else "#f59e0b" if score >= 60 else "#ef4444"
        badge = "✅ תקין ומבצעי" if score >= 80 else "👍 דורש השלמות" if score >= 60 else "⚠️ טעון התערבות"
        
        with grid_cols[idx % 3]:
            st.markdown(f"""
            <div style='background:linear-gradient(135deg,#f8fafc 0%,#e2e8f0 100%);
                padding:20px;border-radius:12px;border-left:4px solid {border_color};
                box-shadow:0 4px 12px rgba(0,0,0,0.1);margin-bottom:12px;'>
                <h4 style='margin:0 0 12px 0;color:#0f172a;'>{div_name}</h4>
                <div style='display:grid;grid-template-columns:1fr 1fr;gap:12px;'>
                    <div><div style='color:#64748b;font-size:12px;'>ציון כשירות</div>
                        <div style='font-size:24px;font-weight:bold;color:#0f172a;'>{score:.0f}</div></div>
                    <div><div style='color:#64748b;font-size:12px;'>דוחות במערכת</div>
                        <div style='font-size:24px;font-weight:bold;color:#0f172a;'>{reports}</div></div>
                    <div><div style='color:#64748b;font-size:12px;'>חוסרים פתוחים</div>
                        <div style='font-size:24px;font-weight:bold;color:{border_color};'>{open_def_count}</div></div>
                </div>
                <div style='margin-top:12px;padding-top:12px;border-top:1px solid #cbd5e1;
                    font-size:12px;color:{border_color};font-weight:bold;'>{badge}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("---")

    # ===== Row 1: Gauge + Critical Issues =====
    col_gauge, col_issues = st.columns([1, 2])

    with col_gauge:
        st.markdown("### 🎯 Overall Risk Status")
        st.markdown("""
        <p style='color: #64748b; font-size: 14px; margin-top: -10px;'>
            מדד סיכון כולל - משקלל את כלל הדיווחים מהשבועיים האחרונים לכדי ציון אחד. 
            מדד זה הפוך לציון הכשירות (ציון גבוה = סיכון נמוך).
        </p>
        """, unsafe_allow_html=True)
        fig_gauge = go.Figure(data=[go.Indicator(
            mode="gauge+number",
            value=avg_score,
            title={'text': "ציון פיקודי משוקלל"},
            domain={'x': [0, 1], 'y': [0, 1]},
            gauge={
                'axis': {'range': [None, 100]},
                'bar': {'color': "#1e3a8a"},
                'steps': [
                    {'range': [0, 25], 'color': "#fee2e2"},
                    {'range': [25, 50], 'color': "#fef3c7"},
                    {'range': [50, 100], 'color': "#d1fae5"}
                ],
                'threshold': {'line': {'color': "green", 'width': 4}, 'thickness': 0.75, 'value': 70}
            }
        )])
        fig_gauge.update_layout(paper_bgcolor='rgba(0,0,0,0)', height=350)
        st.plotly_chart(fig_gauge, use_container_width=True)

    with col_issues:
        st.markdown("### 🚨 Critical Issues (Top Priority)")
        st.markdown("""
        <p style='color: #64748b; font-size: 14px; margin-top: -10px;'>
            סוגיות קריטיות - ריכוז אוטומטי של תקלות חמורות הדורשות התערבות פיקודית מיידית.
        </p>
        """, unsafe_allow_html=True)
        critical_issues = []

        if not df.empty and 'e_status' in df.columns and 'base' in df.columns:
            for base in df[df['e_status'] == 'פסול']['base'].unique()[:3]:
                critical_issues.append({"icon": "🚧", "title": f"עירוב פסול – {base}", "color": "#dc2626"})

        if not df.empty and 'k_cert' in df.columns and 'base' in df.columns:
            for base in df[df['k_cert'] == 'לא']['base'].unique()[:2]:
                critical_issues.append({"icon": "🍽️", "title": f"ללא תעודת כשרות – {base}", "color": "#f59e0b"})

        overdue_count = count_overdue_deficits(accessible_units)
        if overdue_count > 0:
            critical_issues.append({"icon": "⏰", "title": f"{overdue_count} חוסרים בחריגת SLA", "color": "#f59e0b"})
        if silent_units > 0:
            critical_issues.append({"icon": "📡", "title": f"{silent_units} יחידות שלא דיווחו", "color": "#3b82f6"})

        if critical_issues:
            for issue in critical_issues[:5]:
                st.markdown(f"""
                <div style='background:{issue["color"]}20; border-left:4px solid {issue["color"]};
                            padding:12px; border-radius:6px; margin-bottom:10px;'>
                    <strong>{issue["icon"]} {issue["title"]}</strong>
                </div>""", unsafe_allow_html=True)
        else:
            st.success("✅ לא זוהו בעיות קריטיות")
    
    st.markdown("---")

    # ===== Compliance Heatmap =====
    st.markdown("""
    <h3 style='margin-bottom: 0;'>🌡️ Compliance Matrix</h3>
    <p style='color: #64748b; font-size: 14px; margin-top: 0; margin-bottom: 15px;'>
        מפת חום ציות - מראה אחוז עמידה בדרישות לפי קטגוריה (כשרות, עירוב, ניקיון, לוח) עבור כל אוגדה. ירוק = תקין, אדום = דורש טיפול.
    </p>
    """, unsafe_allow_html=True)

    heatmap_data = []
    compliance_metrics = {'k_cert': 'כשרות', 'e_status': 'עירוב', 's_clean': 'ניקיון', 's_board': 'לוח'}
    for div_name in subordinate_divisions:
        div_df = df[df['unit'].str.contains(div_name, na=False)] if not df.empty else pd.DataFrame()
        # Initialize with Any typed dict to avoid lint errors
        row_data: dict[str, Any] = {"Division": str(div_name)}
        for col_name, metric_name in compliance_metrics.items():
            val = 0
            if not div_df.empty and col_name in div_df.columns and len(div_df) > 0:
                if col_name == 'k_cert':
                    ok = len(div_df[div_df[col_name] == 'כן'])
                elif col_name == 'e_status':
                    ok = len(div_df[div_df[col_name] == 'תקין'])
                elif col_name == 's_clean':
                    ok = len(div_df[div_df[col_name].isin(['טוב', 'מצוין'])])
                else:
                    ok = len(div_df[div_df[col_name] == 'כן'])
                val = int(ok / len(div_df) * 100)
            row_data[str(metric_name)] = val
        heatmap_data.append(row_data)

    if heatmap_data:
        heatmap_df = pd.DataFrame(heatmap_data).set_index('Division')
        fig_hm = px.imshow(heatmap_df, color_continuous_scale='RdYlGn', text_auto='.0f',
                           aspect='auto', color_continuous_midpoint=50, labels={'color': 'ציות %', 'x': 'קטגוריה', 'y': 'אוגדה'})
        fig_hm.update_layout(height=350, font=dict(color='#1e293b'),
                              paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
        st.plotly_chart(fig_hm, use_container_width=True)

    st.markdown("---")

    # ===== Row 2: Risk bar + 30-day trend =====
    col_risk, col_trend = st.columns([1.2, 1])

    with col_risk:
        st.markdown("### 🌡️ Unit Risk Matrix")
        st.markdown("""
        <p style='color: #64748b; font-size: 14px; margin-top: -10px;'>
            מטריצת סיכון יחידתית - השוואה בין היחידות השונות תחת הפיקוד לפי רמת הסיכון המחושבת שלהן.
        </p>
        """, unsafe_allow_html=True)
        if units_list:
            risk_rows = [{"unit": u,
                          "risk": calculate_operational_risk_index(u, df)['risk_score']}
                         for u in sorted(units_list)]
            risk_df = pd.DataFrame(risk_rows).sort_values('risk', ascending=False)
            fig_risk = px.bar(risk_df, x='unit', y='risk',
                              color='risk',
                              color_continuous_scale=['#10b981', '#f59e0b', '#ef4444'],
                              range_color=[0, 100])
            fig_risk.update_layout(height=350, xaxis_tickangle=-45, showlegend=False,
                                   paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
            st.plotly_chart(fig_risk, use_container_width=True)

    with col_trend:
        st.markdown("### 📈 30-Day Trend")
        st.markdown("""
        <p style='color: #64748b; font-size: 14px; margin-top: -10px;'>
            מגמת 30 יום - מעקב אחר כמות הליקויים לאורך זמן כדי לזהות שיפור או הידרדרות.
        </p>
        """, unsafe_allow_html=True)
        if not df.empty and 'date' in df.columns:
            try:
                df_sorted = df.copy()
                df_sorted['date'] = pd.to_datetime(df_sorted['date'], errors='coerce')
                df_30 = df_sorted[df_sorted['date'] > pd.Timestamp.now() - pd.Timedelta(days=30)]
                trend_data = []
                for d in pd.date_range(df_30['date'].min(), df_30['date'].max(), freq='D'):
                    day_df = df_30[df_30['date'].dt.date == d.date()]
                    issues = 0
                    if 'e_status' in day_df.columns:
                        issues += int((day_df['e_status'] == '\u05e4\u05e1\u05d5\u05dc').sum())
                    if 'k_cert' in day_df.columns:
                        issues += int((day_df['k_cert'] == '\u05dc\u05d0').sum())
                    trend_data.append({"date": d, "issues": issues})
                trend_df = pd.DataFrame(trend_data)
                fig_trend = px.line(trend_df, x='date', y='issues', markers=True,
                                    labels={'issues': 'כמות ליקויים', 'date': 'תאריך'})
                fig_trend.update_layout(height=350, showlegend=False, hovermode='x unified',
                                        paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
                fig_trend.update_traces(line_color='#3b82f6', marker_color='#60a5fa')
                st.plotly_chart(fig_trend, use_container_width=True)
            except Exception as e:
                st.warning(f"\u05dc\u05d0 \u05e0\u05d9\u05ea\u05df \u05dc\u05d8\u05e2\u05d5\u05df \u05de\u05d2\u05de\u05d5\u05ea: {e}")

    st.markdown("---")

    # ===== Row 3: Detailed deficit table =====
    st.markdown("### 📋 Detailed Deficit Breakdown")
    st.markdown("""
    <p style='color: #64748b; font-size: 14px; margin-top: -10px;'>
        פירוט חוסרים מעמיק - רשימת כלל הפריטים החסרים בשטח, ממויינים לפי דחיפות (ימים מאז הגילוי).
    </p>
    """, unsafe_allow_html=True)
    all_open_deficits = get_open_deficits(accessible_units)
    if not all_open_deficits.empty and 'detected_date' in all_open_deficits.columns:
        all_open_deficits = all_open_deficits.copy()
        all_open_deficits['days_open'] = (
            pd.Timestamp.now() - pd.to_datetime(all_open_deficits['detected_date'], errors='coerce').dt.tz_localize(None)
        ).dt.days.fillna(0).astype(int)
        all_open_deficits = all_open_deficits.sort_values('days_open', ascending=False)

        table_html = """
        <table style='width:100%;border-collapse:collapse;color:#0f172a;'>
        <thead><tr style='background:#1e3a8a;color:white;'>
            <th style='padding:10px;border:1px solid #334155;'>\u05e2\u05d3\u05d9\u05e4\u05d5\u05ea</th>
            <th style='padding:10px;border:1px solid #334155;'>\u05e1\u05d5\u05d2</th>
            <th style='padding:10px;border:1px solid #334155;'>\u05d9\u05d7\u05d9\u05d3\u05d4</th>
            <th style='padding:10px;border:1px solid #334155;'>\u05de\u05d5\u05e6\u05d1</th>
            <th style='padding:10px;border:1px solid #334155;text-align:center;'>\u05db\u05de\u05d5\u05ea</th>
            <th style='padding:10px;border:1px solid #334155;text-align:center;'>\u05d9\u05de\u05d9\u05dd</th>
            <th style='padding:10px;border:1px solid #334155;text-align:center;'>\u05e1\u05d8\u05d8\u05d5\u05e1</th>
        </tr></thead><tbody>"""
        for _, deficit in all_open_deficits.head(10).iterrows():
            days = int(deficit.get('days_open', 0))
            priority = 1 if days > 14 else 2 if days > 7 else 3
            priority_icon = "🔴" if priority == 1 else "🟡" if priority == 2 else "🟢"
            row_bg = "#fee2e2" if priority == 1 else "#fef3c7" if priority == 2 else "#f0fdf4"
            status = "⚠️ חריגת SLA" if days > 14 else "⏰ דחוף" if days > 7 else "✅ חדש"
            table_html += f"""
            <tr style='background:{row_bg};'>
                <td style='padding:10px;border:1px solid #e2e8f0;'>{priority_icon}</td>
                <td style='padding:10px;border:1px solid #e2e8f0;'>{deficit.get('deficit_type','')}</td>
                <td style='padding:10px;border:1px solid #e2e8f0;'>{deficit.get('unit','')}</td>
                <td style='padding:10px;border:1px solid #e2e8f0;'>{deficit.get('base','')}</td>
                <td style='padding:10px;border:1px solid #e2e8f0;text-align:center;'>{int(deficit.get('deficit_count',1))}</td>
                <td style='padding:10px;border:1px solid #e2e8f0;text-align:center;'><strong>{days}</strong></td>
                <td style='padding:10px;border:1px solid #e2e8f0;text-align:center;'>{status}</td>
            </tr>"""
        table_html += "</tbody></table>"
        st.markdown(table_html, unsafe_allow_html=True)
    else:
        st.success("✅ No open deficits")

    st.markdown("---")
    # 🤖 Weekly Insights Manager - Command Panel
    render_weekly_insights_control_panel(key_suffix="_pikud")
    st.markdown("---")
    c1, c2, c3 = st.columns(3)
    with c1: st.metric("📊 סה\"כ דוחות", len(df) if not df.empty else 0)
    with c2: st.metric("🎯 יחידות פעילות", df['unit'].nunique() if not df.empty else 0)
    with c3: st.metric("🔄 רענון אחרון", _dt.datetime.now().strftime('%H:%M:%S'))


def render_ogda_summary_dashboard():
    """
    🎯 \u05d3\u05e9\u05d1\u05d5\u05e8\u05d3 \u05d0\u05d5\u05d2\u05d3\u05d4 \u2013 \u05e1\u05e7\u05d9\u05e8\u05d4 \u05e9\u05dc \u05db\u05dc \u05d4\u05d7\u05d8\u05de\u05e8\u05d9\u05dd \u05d1\u05ea\u05d7\u05ea\u05d5\u05e0\u05d9\u05d5\u05ea
    """
    st.markdown("""
    <div style='background:linear-gradient(90deg,#059669 0%,#10b981 100%);
                padding:30px;border-radius:12px;margin-bottom:30px;'>
        <h1 style='color:white;margin:0;'>🎯 Ogda Dashboard \u2013 Summary</h1>
        <p style='color:rgba(255,255,255,0.9);margin:8px 0 0 0;'>Subordinate Units Status \u00b7 """ +
    datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S') + """</p>
    </div>""", unsafe_allow_html=True)

    unit = st.session_state.selected_unit
    role = st.session_state.role
    accessible_units = get_accessible_units(unit, role)
    raw_data = load_reports_cached(accessible_units)
    df = pd.DataFrame(raw_data) if raw_data else pd.DataFrame()

    kpi_cols = st.columns(4)
    with kpi_cols[0]:
        subordinate_count = max(0, len(accessible_units) - 1)
        st.metric("🏗\ufe0f Subordinate Units", subordinate_count)
    with kpi_cols[1]:
        st.metric("📋 Total Reports", len(df))
    with kpi_cols[2]:
        issues_count = 0
        if not df.empty:
            if 'e_status' in df.columns: issues_count += int((df['e_status'] == '\u05e4\u05e1\u05d5\u05dc').sum())
            if 'k_cert' in df.columns: issues_count += int((df['k_cert'] == '\u05dc\u05d0').sum())
        st.metric("🔴 Open Issues", issues_count, delta_color="inverse")
    with kpi_cols[3]:
        active_units = [u for u in accessible_units if u != unit and not df[df['unit'] == u].empty] if not df.empty else []
        avg_score = (
            sum(calculate_unit_score(df[df['unit'] == u]) for u in active_units) / len(active_units)
            if active_units else 0
        )
        st.metric("📊 Avg Score", f"{avg_score:.0f}/100")

    st.markdown("---")
    st.markdown("### 📊 Units Comparison Matrix")

    comparison_data = []
    for unit_name in accessible_units:
        if unit_name == unit:
            continue
        unit_df = df[df['unit'] == unit_name] if not df.empty else pd.DataFrame()
        if not unit_df.empty:
            score = calculate_unit_score(unit_df)
            open_defs = len(get_open_deficits([unit_name]))
            comparison_data.append({"Unit": unit_name, "Score": score,
                                     "Reports": len(unit_df), "Deficits": open_defs})

    if comparison_data:
        comp_df = pd.DataFrame(comparison_data).sort_values('Score', ascending=False)
        fig_comp = px.bar(comp_df, x='Unit', y='Score', color='Score',
                          color_continuous_scale=['#ef4444', '#f59e0b', '#10b981'],
                          range_color=[0, 100], hover_data=['Reports', 'Deficits'],
                          title='Performance Comparison')
        fig_comp.update_layout(height=400, xaxis_tickangle=-45,
                                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
        st.plotly_chart(fig_comp, use_container_width=True)

        st.markdown("### 📋 Detailed Unit Metrics")
        table_html = """
        <table style='width:100%;border-collapse:collapse;'>
        <thead><tr style='background:#059669;color:white;'>
            <th style='padding:10px;border:1px solid #334155;'>Unit</th>
            <th style='padding:10px;border:1px solid #334155;text-align:center;'>Score</th>
            <th style='padding:10px;border:1px solid #334155;text-align:center;'>Reports</th>
            <th style='padding:10px;border:1px solid #334155;text-align:center;'>Deficits</th>
            <th style='padding:10px;border:1px solid #334155;text-align:center;'>Status</th>
        </tr></thead><tbody>"""
        for _, row in comp_df.iterrows():
            sc = row['Score']
            badge = "\u2705 Excellent" if sc >= 80 else "👍 Good" if sc >= 60 else "\u26a0\ufe0f Needs Work"
            bg = "#f0fdf4" if sc >= 80 else "#fefce8" if sc >= 60 else "#fef2f2"
            table_html += f"""
            <tr style='background:{bg};'>
                <td style='padding:10px;border:1px solid #e2e8f0;'>{row['Unit']}</td>
                <td style='padding:10px;border:1px solid #e2e8f0;text-align:center;'><strong>{sc:.0f}</strong></td>
                <td style='padding:10px;border:1px solid #e2e8f0;text-align:center;'>{row['Reports']}</td>
                <td style='padding:10px;border:1px solid #e2e8f0;text-align:center;'>{row['Deficits']}</td>
                <td style='padding:10px;border:1px solid #e2e8f0;text-align:center;'>{badge}</td>
            </tr>"""
        table_html += "</tbody></table>"
        st.markdown(table_html, unsafe_allow_html=True)
    else:
        st.info("\u05d0\u05d9\u05df \u05e0\u05ea\u05d5\u05e0\u05d9 \u05ea\u05ea\u05d5\u05e0\u05d9\u05d5\u05ea")




# ===== Weekly AI Insights Functions =====

@st.cache_data(ttl=1800, show_spinner=False)  # TTL 30 דקות — מונע קריאת AI בכל רינדור
def generate_weekly_questions(unit: str, accessible_units: list) -> dict:
    """
    🤖 יוצר שאלות מבוססות-נתונים אמיתיים — חריגות שנמצאו בשאלונים.
    מזהה בעיות חוזרות: מזוזות חסרות, ערבוב כלים, שיעורי תורה חסרים,
    עירוב פסול, כשרות ועוד — ומשתמש ב-AI לניסוח שאלות ממוקדות.
    """
    # סינון: רק יחידות כפופות (לא היחידה עצמה)
    subordinate_units = [u for u in accessible_units if u != unit]
    units_to_query = subordinate_units if subordinate_units else accessible_units

    all_reports = load_reports_cached(units_to_query)
    df = pd.DataFrame(all_reports) if all_reports else pd.DataFrame()

    # סינון נוסף: רק דוחות של יחידות כפופות
    if not df.empty and 'unit' in df.columns and subordinate_units:
        df = df[df['unit'].isin(subordinate_units)]

    insights = {}

    # ===== חילוץ חריגות אמיתיות מהנתונים =====
    anomaly_lines = []  # רשימת תיאורי חריגות לשלוח ל-AI

    if not df.empty:
        df_copy = df.copy()
        if 'date' in df_copy.columns:
            df_copy['date'] = pd.to_datetime(df_copy['date'], errors='coerce').dt.tz_localize(None)
        last_30 = df_copy[df_copy['date'] >= (pd.Timestamp.now() - pd.Timedelta(days=30))] if 'date' in df_copy.columns else df_copy

        # 1. מזוזות חסרות — חוזר באותה יחידה/מוצב
        if 'r_mezuzot_missing' in last_30.columns and 'unit' in last_30.columns:
            try:
                mezuzot_df = last_30[pd.to_numeric(last_30['r_mezuzot_missing'], errors='coerce') > 0]
                for grp_unit, grp in mezuzot_df.groupby('unit'):
                    total_missing = int(pd.to_numeric(grp['r_mezuzot_missing'], errors='coerce').sum())
                    weeks_count = grp['date'].dt.isocalendar().week.nunique() if 'date' in grp.columns else len(grp)
                    if weeks_count >= 2:
                        anomaly_lines.append(f"⚠️ {grp_unit}: חסרות מזוזות ({total_missing} סה\"כ) ב-{weeks_count} דוחות בחודש האחרון")
            except Exception:
                pass

        # 2. ערבוב כלים חוזר
        if 'p_mix' in last_30.columns:
            try:
                mix_df = last_30[last_30['p_mix'].isin(['כן', 'כן (חלקי)'])]
                for grp_unit, grp in mix_df.groupby('unit'):
                    bases = grp['base'].unique().tolist() if 'base' in grp.columns else []
                    base_str = ', '.join(str(b) for b in bases[:3]) if bases else ''
                    count = len(grp)
                    if count >= 2:
                        anomaly_lines.append(f"⚠️ {grp_unit}: ערבוב כלים זוהה {count} פעמים" + (f" במוצבים: {base_str}" if base_str else ''))
            except Exception:
                pass

        # 3. שיעורי תורה חסרים
        if 'soldier_has_lesson' in last_30.columns or 'hq_want_torah_lesson' in last_30.columns:
            try:
                lesson_col = 'soldier_has_lesson' if 'soldier_has_lesson' in last_30.columns else 'hq_want_torah_lesson'
                no_lesson_df = last_30[last_30[lesson_col].isin(['לא', 'לא בדקתי'])]
                for grp_unit, grp in no_lesson_df.groupby('unit'):
                    count = len(grp)
                    if count >= 2:
                        anomaly_lines.append(f"⚠️ {grp_unit}: אין שיעור תורה — דווח {count} פעמים בחודש האחרון")
            except Exception:
                pass

        # 4. כשרות (עירוב פסול וכשרות לא תקינה)
        if 'e_status' in last_30.columns:
            try:
                eruv_bad = last_30[last_30['e_status'] == 'פסול'].groupby('unit').size()
                for grp_unit, count in eruv_bad.items():
                    if count >= 2:
                        anomaly_lines.append(f"🚧 {grp_unit}: עירוב פסול {count} פעמים בחודש האחרון")
            except Exception:
                pass

        if 'k_cert' in last_30.columns:
            try:
                no_cert = last_30[last_30['k_cert'] == 'לא'].groupby('unit').size()
                for grp_unit, count in no_cert.items():
                    if count >= 2:
                        anomaly_lines.append(f"🍽️ {grp_unit}: תעודת כשרות לא תקפה {count} פעמים בחודש האחרון")
            except Exception:
                pass

        # 5. יחידות שלא דיווחו השבוע
        if 'date' in df_copy.columns:
            try:
                this_week = df_copy[df_copy['date'] >= (pd.Timestamp.now() - pd.Timedelta(days=7))]
                reported_this_week = set(this_week['unit'].unique()) if not this_week.empty else set()
                silent = [u for u in units_to_query if u not in reported_this_week]
                if silent:
                    anomaly_lines.append(f"📡 לא דיווחו השבוע: {', '.join(silent[:5])}")
            except Exception:
                pass

        # ניתוחים סטטיסטיים נוספים
        kashrut_insights = analyze_kashrut_trend(df, units_to_query)
        if kashrut_insights:
            insights['kashrut'] = kashrut_insights
        eruv_insights = analyze_eruv_trend(df, units_to_query)
        if eruv_insights:
            insights['eruv'] = eruv_insights
        performance_insights = analyze_unit_performance(df, units_to_query)
        if performance_insights:
            insights['performance'] = performance_insights
        anomaly_insights = detect_weekly_anomalies(df, units_to_query)
        if anomaly_insights:
            insights['anomalies'] = anomaly_insights

    # ===== מחולל AI — שאלות מבוססות חריגות אמיתיות =====
    import google.generativeai as genai
    import json as _json
    try:
        if check_api_quota_safety(daily_limit=500):
            genai.configure(api_key=st.secrets["gemini"]["api_key"])
            model = genai.GenerativeModel("gemini-2.5-flash")
            log_api_usage(unit, "weekly_insights")

            units_str = ', '.join(units_to_query) if units_to_query else 'אין יחידות'
            anomalies_str = '\n'.join(anomaly_lines) if anomaly_lines else 'לא זוהו חריגות חוזרות בנתונים.'

            prompt = f"""
אתה קצין מטה של {unit} (רב אוגדה/פיקוד). היחידות הכפופות: {units_str}.

חריגות שזוהו אוטומטית מהשאלונים בחודש האחרון:
{anomalies_str}

תפקידך: עזור לרב לנהל שיחות חתך ממוקדות עם כפופיו לקראת השבוע.
ספק 5-7 שאלות קצרות, חדות, מבוססות ישירות על החריגות שנמצאו.
כל שאלה תתייחס לחריגה ספציפית — יחידה, מוצב, או תופעה חוזרת.
אם אין חריגות — שאל על שמירת איכות ועל היחידה החלשה.

החזר JSON בלבד:
{{"question": "שאלות מיקוד שבועיות לקראת השבוע", "current": 0, "previous": 0, "trend": "ai_focus",
"suggestion": "1. [שאלה על חריגה ספציפית]\n2. [שאלה על חריגה ספציפית]\n3. ...\n4. ...\n5. ..."}}
"""
            response = model.generate_content(prompt)
            clean_json = response.text.replace("```json", "").replace("```", "").strip()
            ai_focus = _json.loads(clean_json)
            insights['ai_commander_focus'] = ai_focus
    except Exception:
        # Fallback: אם ה-AI לא זמין, בנה שאלות מהחריגות ישירות
        fallback_questions = anomaly_lines[:5] if anomaly_lines else [
            "האם כל היחידות הגישו דוח שבועי?",
            "מה מצב העירובין לקראת השבת?",
            "האם יש חוסרים פתוחים יותר מ-7 ימים?",
            "מה מצב שיעורי התורה בחטיבות?",
            "האם זוהו בעיות כשרות חוזרות?"
        ]
        numbered = '\n'.join(f"{i+1}. {q.lstrip('⚠️🚧🍽️📡 ')}" for i, q in enumerate(fallback_questions))
        insights['ai_commander_focus'] = {
            "question": "🤖 שאלות הכוונה שבועיות לקראת השבוע",
            "current": 0, "previous": 0, "trend": "ai_focus",
            "suggestion": numbered
        }

    # ניתוח חוסרים (על הכפופים בלבד)
    deficit_insights = analyze_deficit_progress(units_to_query)
    if deficit_insights and isinstance(deficit_insights, list):
        insights['deficits'] = {
            "question": deficit_insights[0].get("question", "בדוק את סטטוס החוסרים השבוע"),
            "current": len(get_open_deficits(units_to_query)),
            "previous": 0,
            "trend": "attention_needed",
            "suggestion": deficit_insights[0].get("suggestion", "עבור על טבלת החוסרים המלאה")
        }
    elif deficit_insights and isinstance(deficit_insights, dict):
        insights['deficits'] = deficit_insights

    return insights


def analyze_kashrut_trend(df: pd.DataFrame, units: list) -> dict:
    """🔍 ניתוח מגמת כשרות שבועית"""
    if 'date' not in df.columns:
        return {}
    df = df.copy()
    df['date'] = pd.to_datetime(df['date'], errors='coerce')
    current_week = df[df['date'] >= (pd.Timestamp.now() - pd.Timedelta(days=7))]
    previous_week = df[(df['date'] >= (pd.Timestamp.now() - pd.Timedelta(days=14))) &
                       (df['date'] < (pd.Timestamp.now() - pd.Timedelta(days=7)))]
    current_issues = 0
    prev_issues = 0
    if not current_week.empty and 'k_cert' in current_week.columns:
        current_issues = len(current_week[current_week['k_cert'] == 'לא'])
    if not previous_week.empty and 'k_cert' in previous_week.columns:
        prev_issues = len(previous_week[previous_week['k_cert'] == 'לא'])
    if current_issues == 0 and prev_issues == 0:
        question = "✅ כשרות במצב מצוין כל שבוע — האם זה בגלל אינספקשנים אקטיביים?"
        suggestion = "שמור על הרמה! המשך עם אותו קצב בדיקות"
        trend = "stable_good"
    elif current_issues < prev_issues:
        improvement = prev_issues - current_issues
        improvement_pct = (improvement / prev_issues * 100) if prev_issues > 0 else 0
        question = f"🎉 בשבוע האחרון הצטמצמו בעיות כשרות ב-{improvement} ({improvement_pct:.0f}%) — מה השתנה?"
        suggestion = "בדוק מה עשית אחרת השבוע — זה עובד!"
        trend = "improving"
    elif current_issues > prev_issues:
        worsening = current_issues - prev_issues
        question = f"⚠️ בעיות כשרות עלו ב-{worsening} בשבוע זה — למה?"
        suggestion = "אפשר: (1) מטבח חדש עם בעיה קבועה, (2) אינספקטור לא קפדן, (3) בחורים חדשים"
        trend = "worsening"
    else:
        question = f"🟡 כשרות: {current_issues} בעיות כל שבוע — זה נורמלי אבל בדוק הסיבה"
        suggestion = "יכול להיות: (1) יחידה גדולה, (2) תחלופה גבוהה, (3) אתגר ספציפי"
        trend = "stable"
    return {"question": question, "current": current_issues, "previous": prev_issues,
            "trend": trend, "suggestion": suggestion}


def analyze_eruv_trend(df: pd.DataFrame, units: list) -> dict:
    """🚧 ניתוח מגמת עירוב שבועית"""
    if 'date' not in df.columns:
        return {}
    df = df.copy()
    df['date'] = pd.to_datetime(df['date'], errors='coerce')
    current_week = df[df['date'] >= (pd.Timestamp.now() - pd.Timedelta(days=7))]
    previous_week = df[(df['date'] >= (pd.Timestamp.now() - pd.Timedelta(days=14))) &
                       (df['date'] < (pd.Timestamp.now() - pd.Timedelta(days=7)))]
    current_eruv_issues = 0
    prev_eruv_issues = 0
    if not current_week.empty and 'e_status' in current_week.columns:
        current_eruv_issues = len(current_week[current_week['e_status'] == 'פסול'])
    if not previous_week.empty and 'e_status' in previous_week.columns:
        prev_eruv_issues = len(previous_week[previous_week['e_status'] == 'פסול'])
    if current_eruv_issues == 0:
        if prev_eruv_issues > 0:
            question = f"✅ עירוב תוקן! משבוע שעבר היו {prev_eruv_issues} בעיות — מה עזר?"
            suggestion = "בדוק עם חטמ״ר / עיר בשביל ללמוד"
            trend = "resolved"
        else:
            question = "✅ עירובין במצב מצוין — המשך כך"
            suggestion = "זה תקין, לא צריך שינוי"
            trend = "stable_good"
    elif current_eruv_issues > prev_eruv_issues:
        question = f"🚨 עירוב: {current_eruv_issues} בעיות (מ-{prev_eruv_issues} בשבוע שעבר) — CRITICAL!"
        suggestion = "אפשרויות: (1) עירוב פסול חדש, (2) תקלה קבועה, (3) אי־דיווח של בעיה ישנה"
        trend = "worsening"
    else:
        question = f"🟡 עירוב: {current_eruv_issues} בעיות — בדוק אם זה בעיה קבועה או זמנית"
        suggestion = "כל שבוע חזור אל הצוות בשטח"
        trend = "stable"
    return {"question": question, "current": current_eruv_issues, "previous": prev_eruv_issues,
            "trend": trend, "suggestion": suggestion}


def analyze_recurring_base_issues(df: pd.DataFrame) -> list:
    """🔍 זיהוי בעיות חוזרות באותו בסיס"""
    if df.empty or 'base' not in df.columns:
        return []
    
    recurring = []
    recent_df = df.copy()
    recent_df['date'] = pd.to_datetime(recent_df['date'], errors='coerce')
    # רק 30 יום אחרונים
    recent_df = recent_df[recent_df['date'] >= (pd.Timestamp.now() - pd.Timedelta(days=30))]
    
    for base in recent_df['base'].unique():
        if not base or base == "לא ידוע": continue
        base_df = recent_df[recent_df['base'] == base]
        if len(base_df) < 2: continue
            
        # כשרות
        kashrut_fails = base_df[base_df['k_cert'] == 'לא']
        if len(kashrut_fails) >= 2:
            recurring.append({
                "type": "kashrut",
                "base": base,
                "unit": base_df.iloc[0]['unit'],
                "question": f"🔴 למה יש בעיות חוזרות בכשרות ב-{base}? (כבר {len(kashrut_fails)} דיווחים בפסול בחודש האחרון)",
                "suggestion": "בדוק אם הרב המקומי מכיר את הבעיה או שיש חוסר כוח אדם"
            })
            
        # עירוב
        eruv_fails = base_df[base_df['e_status'] == 'פסול']
        if len(eruv_fails) >= 2:
            recurring.append({
                "type": "eruv",
                "base": base,
                "unit": base_df.iloc[0]['unit'],
                "question": f"🚧 בעיות חוזרות בעירוב ב-{base} (פסול {len(eruv_fails)} פעמים בחודש האחרון) — למה?",
                "suggestion": "אולי צריך תקציב לתיקון תשתיתי ולא רק תיקון נקודתי"
            })
            
    return recurring


def analyze_critical_data_changes(df: pd.DataFrame) -> list:
    """🧐 זיהוי שינויים בנתונים קריטיים (כמו מס' ספר תורה)"""
    if df.empty or 'base' not in df.columns or 's_torah_id' not in df.columns:
        return []
        
    changes = []
    for base in df['base'].unique():
        if not base or base == "לא ידוע": continue
        base_df = df[df['base'] == base].sort_values('date', ascending=False)
        if len(base_df) < 2: continue
            
        latest_torah = str(base_df.iloc[0]['s_torah_id'])
        prev_torah = str(base_df.iloc[1]['s_torah_id'])
        
        if latest_torah and prev_torah and latest_torah != prev_torah and latest_torah != "nan" and prev_torah != "nan":
             changes.append({
                "type": "data_change",
                "base": base,
                "unit": base_df.iloc[0]['unit'],
                "question": f"🧐 בבסיס {base} ({base_df.iloc[0]['unit']}) השתנה מס' צ' של ספר התורה מ-{prev_torah} ל-{latest_torah} — מה הסיבה?",
                "suggestion": "בדוק אם הספר הוחלף, נשלח לתיקון או שמדובר בטעות הקלדה"
            })
    return changes


def analyze_deficit_progress(accessible_units: list) -> list:
    """📊 ניתוח התקדמות בסגירת חוסרים - מורחב"""
    insight_list = []
    try:
        open_now = get_open_deficits(accessible_units)
    except Exception:
        open_now = pd.DataFrame()
        
    try:
        closed_this_week = supabase.table("deficit_tracking") \
            .select("*").in_("unit", accessible_units).eq("status", "closed") \
            .gte("resolved_date", (pd.Timestamp.now() - pd.Timedelta(days=7)).isoformat()) \
            .execute()
        closed_count = len(closed_this_week.data) if closed_this_week.data else 0
    except Exception:
        closed_count = 0
        
    total_open = len(open_now) if not open_now.empty else 0
    
    # 1. ניתוח כללי
    if total_open > 0:
        overdue = count_overdue_deficits(accessible_units)
        if overdue > 0:
            insight_list.append({
                "type": "deficits",
                "question": f"🚨 {overdue} חוסרים עברו SLA (7 ימים) — למה הם לא מטופלים בדחיפות?",
                "suggestion": "בדוק איזה חוסרים פתוחים זמן רב וקדם אותם מול הלוגיסטיקה"
            })
            
        # 2. ניתוח בסיסים ספציפיים (החוסרים הכי ישנים)
        if 'detected_date' in open_now.columns and 'base' in open_now.columns:
            open_now['detected_date'] = pd.to_datetime(open_now['detected_date'], errors='coerce').dt.tz_localize(None)
            cutoff_14 = pd.Timestamp.now() - pd.Timedelta(days=14)
            old_items = open_now[open_now['detected_date'] < cutoff_14]
            if not old_items.empty:
                now_ts = pd.Timestamp.now()
                for _, row in old_items.sort_values('detected_date').head(2).iterrows():
                    days = (now_ts - row['detected_date']).days
                    insight_list.append({
                        "type": "deficits",
                        "base": row['base'],
                        "question": f"🔴 למה החוסרים ב-{row['base']} ({row['unit']}) לא הושלמו כבר {days} ימים?",
                        "suggestion": "יכול להיות שיש בעיית תקציב או חוסר במלאי ארצי?"
                    })
    
    elif closed_count > 0:
        insight_list.append({
            "type": "deficits",
            "question": f"🎉 סגרתם {closed_count} חוסרים בשבוע! — כל הכבוד על המעקב.",
            "suggestion": "שתף את שיטת העבודה עם יחידות אחרות"
        })
        
    return insight_list


def analyze_unit_performance(df: pd.DataFrame, units: list) -> dict:
    """📈 ניתוח ביצועים כללי יחידה"""
    scores = []
    for u in units:
        unit_df = df[df['unit'] == u] if not df.empty and 'unit' in df.columns else pd.DataFrame()
        if not unit_df.empty:
            score = calculate_unit_score(unit_df)
            scores.append((u, score))
    if not scores:
        return {}
    scores.sort(key=lambda x: x[1], reverse=True)
    best_unit, best_score = scores[0]
    worst_unit, worst_score = scores[-1] if len(scores) > 1 else (scores[0][0], scores[0][1])
    avg_score = sum(s for _, s in scores) / len(scores)
    if best_score >= 90:
        question = f"🏆 {best_unit} עלתה ל-{best_score:.0f}! — מה הם עושים נכון?"
        suggestion = "שתף את הניסיון שלהם עם יחידות אחרות"
        trend = "excellent"
    elif worst_score < 60:
        gap = avg_score - worst_score
        question = f"⚠️ {worst_unit} בציון {worst_score:.0f} (נמוך מממוצע ב-{gap:.0f} נקודות) — מה קורה?"
        suggestion = "בדוק: בעיה בטיים? בקצב? בהנהלה?"
        trend = "struggling"
    else:
        question = f"📊 ממוצע: {avg_score:.0f}/100 — טוב! מה {best_unit} עושה נכון?"
        suggestion = f"בדוק את {best_unit} ושתף תרגול"
        trend = "healthy"
    return {"question": question, "avg_score": avg_score, "best": (best_unit, best_score),
            "worst": (worst_unit, worst_score), "trend": trend, "suggestion": suggestion}


def detect_weekly_anomalies(df: pd.DataFrame, units: list) -> dict:
    """🔍 זיהוי חריגויות שבועיות"""
    if 'date' not in df.columns:
        return {}
    df = df.copy()
    df['date'] = pd.to_datetime(df['date'], errors='coerce')
    current_week = df[df['date'] >= (pd.Timestamp.now() - pd.Timedelta(days=7))]
    if current_week.empty:
        return {}
    anomalies = []
    unit_col = current_week['unit'].values if 'unit' in current_week.columns else []
    for u in units:
        if u not in unit_col:
            anomalies.append({"type": "no_report",
                               "question": f"📡 {u} לא דיווחה כל שבוע! — מה קרה?",
                               "severity": "high"})
    try:
        recurring_issues = supabase.table("deficit_tracking") \
            .select("*").in_("unit", units).eq("status", "open") \
            .gte("detected_date", (pd.Timestamp.now() - pd.Timedelta(days=30)).isoformat()) \
            .execute()
        if recurring_issues.data:
            old_issues = [d for d in recurring_issues.data
                         if (pd.Timestamp.now() - pd.to_datetime(d.get('detected_date', ''), errors='coerce')).days > 14]
            if len(old_issues) > 3:
                anomalies.append({"type": "chronic_deficits",
                                   "question": f"🔴 {len(old_issues)} חוסרים פתוחים יותר מ-14 ימים! — זה בעיה מבנית?",
                                   "severity": "critical"})
    except Exception:
        pass
    if anomalies:
        return {"anomalies": anomalies, "count": len(anomalies),
                "highest_severity": max(a['severity'] for a in anomalies)}
    return {}


def render_weekly_questions_widget():
    """🤖 תצוגת שאלות חכמות שבועיות — ממשק אינטראקטיבי"""
    st.markdown("""
    <div style='background: linear-gradient(135deg, #6366f1 0%, #8b5cf6 100%);
                padding: 30px; border-radius: 16px; margin: 30px 0;'>
        <h2 style='color: white; margin: 0;'>🤖 Weekly AI Insights</h2>
        <p style='color: rgba(255,255,255,0.9); margin: 8px 0 0 0;'>
            שאלות חכמות שמשתנות בהתאם לנתונים שלך
        </p>
    </div>
    """, unsafe_allow_html=True)

    unit = st.session_state.get('selected_unit', '')
    role = st.session_state.get('role', '')
    accessible_units = get_accessible_units(unit, role)

    insights = generate_weekly_questions(unit, accessible_units)

    if "error" in insights:
        st.warning("אין מספיק נתונים להצגת תובנות")
        return

    insight_tabs = st.tabs(["🍽️ כשרות", "🚧 עירוב", "📋 חוסרים", "📈 ביצועים", "🔍 חריגויות"])

    # === Tab 1: Kashrut ===
    with insight_tabs[0]:
        if 'kashrut' in insights:
            k = insights['kashrut']
            color_map = {"stable_good": "#10b981", "improving": "#3b82f6",
                         "worsening": "#ef4444", "stable": "#f59e0b"}
            color = color_map.get(k['trend'], "#64748b")
            st.markdown(f"""
            <div style='background:{color}20;border-left:4px solid {color};
                        padding:20px;border-radius:10px;margin-bottom:20px;'>
                <h3 style='color:{color};margin:0 0 10px 0;'>❓ {k['question']}</h3>
                <div style='background:white;padding:15px;border-radius:6px;margin-bottom:10px;'>
                    <strong>📊 נתונים:</strong><br/>
                    • שבוע זה: {k['current']} בעיות<br/>
                    • שבוע שעבר: {k['previous']} בעיות<br/>
                    • מגמה: <span style='color:{color};'>{k['trend'].replace('_',' ').upper()}</span>
                </div>
                <div style='background:#f0fdf4;padding:15px;border-radius:6px;border-left:3px solid #10b981;'>
                    <strong>💡 הצעה:</strong><br/>{k['suggestion']}
                </div>
            </div>""", unsafe_allow_html=True)
            col1, col2, col3 = st.columns(3)
            with col1:
                if st.button("✅ טופלנו", key="ai_kashrut_resolved"):
                    st.success("עודכן לשבוע הבא!")
            with col2:
                if st.button("📞 תזכורת", key="ai_kashrut_reminder"):
                    st.info("תזכורת תישלח לרב")
            with col3:
                if st.button("📝 הוסף הערה", key="ai_kashrut_note"):
                    st.text_input("הערה:", key="ai_kashrut_note_text")
        else:
            st.info("אין נתוני כשרות זמינים")

    # === Tab 2: Eruv ===
    with insight_tabs[1]:
        if 'eruv' in insights:
            e = insights['eruv']
            color_map = {"resolved": "#10b981", "stable_good": "#10b981",
                         "stable": "#f59e0b", "worsening": "#ef4444"}
            color = color_map.get(e['trend'], "#64748b")
            st.markdown(f"""
            <div style='background:{color}20;border-left:4px solid {color};
                        padding:20px;border-radius:10px;'>
                <h3 style='color:{color};margin:0 0 10px 0;'>❓ {e['question']}</h3>
                <div style='background:white;padding:15px;border-radius:6px;margin-bottom:10px;'>
                    <strong>📊 נתונים:</strong><br/>
                    • שבוע זה: {e['current']} בעיות עירוב<br/>
                    • שבוע שעבר: {e['previous']} בעיות
                </div>
                <div style='background:#f0fdf4;padding:15px;border-radius:6px;border-left:3px solid #10b981;'>
                    <strong>💡 הצעה:</strong><br/>{e['suggestion']}
                </div>
            </div>""", unsafe_allow_html=True)
            col1, col2 = st.columns(2)
            with col1:
                if st.button("✅ עירוב תוקן", key="ai_eruv_fixed"):
                    st.success("עדכון לשבוע הבא")
            with col2:
                if st.button("📞 התקשר לעיר", key="ai_eruv_call"):
                    st.info("רשום: התקשר לעיר בנוגע לעירוב")
        else:
            st.info("אין נתוני עירוב זמינים")

    # === Tab 3: Deficits ===
    with insight_tabs[2]:
        if 'deficits' in insights:
            d = insights['deficits']
            color_map = {"excellent": "#10b981", "healthy": "#3b82f6",
                         "high_volume": "#f59e0b", "critical": "#ef4444", "stable_good": "#10b981"}
            color = color_map.get(d['trend'], "#64748b")
            st.markdown(f"""
            <div style='background:{color}20;border-left:4px solid {color};
                        padding:20px;border-radius:10px;'>
                <h3 style='color:{color};margin:0 0 10px 0;'>❓ {d['question']}</h3>
                <div style='background:white;padding:15px;border-radius:6px;margin-bottom:10px;'>
                    <strong>📊 נתונים:</strong><br/>
                    • חוסרים פתוחים: {d['open']}<br/>
                    • סגורים בשבוע: {d['closed_this_week']}<br/>
                    • עברו SLA: {d['overdue']}
                </div>
                <div style='background:#f0fdf4;padding:15px;border-radius:6px;border-left:3px solid #10b981;'>
                    <strong>💡 הצעה:</strong><br/>{d['suggestion']}
                </div>
            </div>""", unsafe_allow_html=True)
            closure_rate = (d['closed_this_week'] / (d['open'] + d['closed_this_week']) * 100
                            if (d['open'] + d['closed_this_week']) > 0 else 0)
            st.progress(min(100, closure_rate) / 100, text=f"שיעור סגירה: {closure_rate:.0f}%")
        else:
            st.info("אין נתוני חוסרים זמינים")

    # === Tab 4: Performance ===
    with insight_tabs[3]:
        if 'performance' in insights:
            p = insights['performance']
            color_map = {"excellent": "#10b981", "healthy": "#3b82f6", "struggling": "#ef4444"}
            color = color_map.get(p['trend'], "#64748b")
            st.markdown(f"""
            <div style='background:{color}20;border-left:4px solid {color};
                        padding:20px;border-radius:10px;'>
                <h3 style='color:{color};margin:0 0 10px 0;'>❓ {p['question']}</h3>
                <div style='background:white;padding:15px;border-radius:6px;margin-bottom:10px;'>
                    <strong>📊 נתונים:</strong><br/>
                    • ממוצע אוגדה: {p['avg_score']:.0f}/100<br/>
                    • יחידה מובילה: {p['best'][0]} ({p['best'][1]:.0f})<br/>
                    • יחידה זקוקה לעזרה: {p['worst'][0]} ({p['worst'][1]:.0f})
                </div>
                <div style='background:#f0fdf4;padding:15px;border-radius:6px;border-left:3px solid #10b981;'>
                    <strong>💡 הצעה:</strong><br/>{p['suggestion']}
                </div>
            </div>""", unsafe_allow_html=True)
        else:
            st.info("אין נתוני ביצועים זמינים")

    # === Tab 5: Anomalies ===
    with insight_tabs[4]:
        if 'anomalies' in insights:
            anom = insights['anomalies']
            for a in anom.get('anomalies', []):
                severity_colors = {"high": "#ef4444", "critical": "#dc2626", "medium": "#f59e0b"}
                color = severity_colors.get(a['severity'], "#64748b")
                st.markdown(f"""
                <div style='background:{color}20;border-left:4px solid {color};
                            padding:15px;border-radius:10px;margin-bottom:12px;'>
                    <h4 style='color:{color};margin:0 0 8px 0;'>{a['question']}</h4>
                    <span style='background:{color};color:white;padding:3px 8px;
                                 border-radius:3px;font-size:12px;'>{a['severity'].upper()}</span>
                </div>""", unsafe_allow_html=True)
        else:
            st.success("✅ אין חריגויות שבועיות")

    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        st.info("השאלות הבאות תופעלנה בעוד **7 ימים**")
    with col2:
        if st.button("🔄 עדכן שאלות עכשיו", key="ai_refresh_questions"):
            st.success("שאלות עודכנו!")
            st.balloons()


def render_weekly_insights_control_panel(key_suffix: str = ""):
    """
    🎛️ פאנל שאלות שבועיות — מציג שאלות אוטומטית לפי הנתונים העדכניים
    """
    st.markdown("""
    <div style='background: linear-gradient(135deg, #6366f1 0%, #8b5cf6 100%);
                padding: 20px; border-radius: 12px; margin-bottom: 20px;'>
        <h2 style='color: white; margin: 0;'>🤖 שאלות הכוונה שבועיות לקראת השבוע</h2>
        <p style='color: rgba(255,255,255,0.85); margin: 6px 0 0 0; font-size: 14px;'>
            5-7 שאלות מיקוד שנוצרו על בסיס הנתונים שלך — לקראת שיחות החתך שלך עם הכפופים
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    unit = st.session_state.get('selected_unit', '')
    role = st.session_state.get('role', '')
    accessible_units = get_accessible_units(unit, role)
    
    # הצג ספינר קצר ואז את השאלות — אוטומטי, ללא לחיצת כפתור
    with st.spinner("⏳ מייצר שאלות הכוונה..."):
        insights = generate_weekly_questions(unit, accessible_units)
    
    # הצגת שאלות ה-AI המאוחדות
    if 'ai_commander_focus' in insights:
        ai = insights['ai_commander_focus']
        suggestion_text = ai.get('suggestion', '')
        # המר שאלות ממחרוזת לרשימה
        lines = [l.strip() for l in suggestion_text.split('\n') if l.strip()]
        st.markdown("""
        <div style='background: #f0fdf4; border-left: 4px solid #6366f1;
                    padding: 20px; border-radius: 10px; margin-bottom: 16px;'>
            <strong style='color: #6366f1; font-size: 16px;'>❓ שאלות מיקוד לשבוע:</strong>
        </div>
        """, unsafe_allow_html=True)
        for i, line in enumerate(lines, 1):
            # הסר מספור קיים מהטקסט אם יש
            clean_line = line.lstrip('0123456789. ').strip()
            if clean_line:
                st.markdown(f"""
                <div style='background: white; border: 1px solid #e2e8f0; border-right: 4px solid #6366f1;
                            padding: 14px 18px; border-radius: 8px; margin-bottom: 10px;
                            box-shadow: 0 1px 3px rgba(0,0,0,0.06);'>
                    <span style='color: #6366f1; font-weight: bold; font-size: 16px;'>{i}.</span>
                    <span style='margin-right: 8px; font-size: 15px; color: #1e293b;'>{clean_line}</span>
                </div>
                """, unsafe_allow_html=True)
    
    # כפתור ריענון בלבד
    st.markdown("")
    col1, col2 = st.columns([3, 1])
    with col2:
        if st.button("🔄 רענן שאלות", use_container_width=True, key=f"refresh_weekly{key_suffix}"):
            st.rerun()


def update_weekly_insights_now():
    """
    🚀 עדכן את כל השאלות עכשיו
    כל מה שצריך: לחיצה אחת על כפתור
    """
    with st.spinner("⏳ מעדכן שאלות... זה לוקח 10 שניות"):
        try:
            # 1. יצור טבלה אם לא קיימת
            create_weekly_insights_table()
            # 2. מחק שאלות של שבוע שעבר
            delete_old_insights()
            # 3. הכן את כל היחידות
            all_units = HATMAR_UNITS + COMMAND_UNITS
            new_insights_count = 0
            for unit in all_units:
                try:
                    role = get_user_role(unit)
                    accessible = get_accessible_units(unit, role)
                    insights_dict = generate_weekly_questions(unit, accessible)
                    # שמור כל insight מהמילון
                    for insight_data in insights_dict.values():
                        if isinstance(insight_data, dict) and 'question' in insight_data:
                            i_type = insight_data.get('type', 'general')
                            save_insight(unit, i_type, insight_data)
                            new_insights_count += 1
                except Exception as e:
                    print(f"⚠️ {unit}: {str(e)[:50]}")
            # 4. הצג הצלחה
            st.success(f"✅ עדכנו {new_insights_count} שאלות חדשות!")
            st.balloons()
            time.sleep(2)
            st.rerun()
        except Exception as e:
            st.error(f"❌ שגיאה: {str(e)}")


def create_weekly_insights_table():
    """
    🔧 יצור טבלה פעם אחת בלבד
    """
    sql = """
    CREATE TABLE IF NOT EXISTS weekly_insights (
        id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
        week_start DATE NOT NULL DEFAULT CURRENT_DATE,
        unit TEXT NOT NULL,
        insight_type TEXT NOT NULL,
        previous_value INT DEFAULT 0,
        current_value INT DEFAULT 0,
        change_direction TEXT,
        ai_question TEXT NOT NULL,
        ai_suggestion TEXT,
        created_at TIMESTAMP DEFAULT NOW(),
        UNIQUE(week_start, unit, insight_type)
    );
    CREATE INDEX IF NOT EXISTS idx_weekly_insights_unit 
    ON weekly_insights(unit, week_start DESC);
    """
    try:
        supabase.table("weekly_insights").select("id").limit(1).execute()
    except Exception:
        pass  # טבלה כבר קיימת או תיווצר דרך migration


def delete_old_insights():
    """
    🗑️ מחק שאלות מ-7 ימים אחרונים
    כדי לא להיות כפילויות
    """
    try:
        week_ago = (pd.Timestamp.now() - pd.Timedelta(days=7)).strftime('%Y-%m-%d')
        supabase.table("weekly_insights") \
            .delete() \
            .gte("week_start", week_ago) \
            .execute()
    except Exception as e:
        print(f"⚠️ לא הצלחנו למחוק: {str(e)[:50]}")


def save_insight(unit: str, insight_type: str, insight_data: dict):
    """
    💾 שמור insight אחד
    """
    try:
        supabase.table("weekly_insights").insert({
            "week_start": pd.Timestamp.now().strftime('%Y-%m-%d'),
            "unit": unit,
            "insight_type": insight_type,
            "previous_value": insight_data.get('previous', insight_data.get('open', 0)),
            "current_value": insight_data.get('current', insight_data.get('closed_this_week', 0)),
            "change_direction": insight_data.get('trend', 'unknown'),
            "ai_question": insight_data.get('question', ''),
            "ai_suggestion": insight_data.get('suggestion', '')
        }).execute()
    except Exception as e:
        if "violates unique constraint" not in str(e):
            print(f"⚠️ שגיאה בשמירה: {str(e)[:50]}")


def show_last_weekly_questions():
    """
    📋 הצג את השאלות האחרונות שנוצרו
    """
    try:
        result = supabase.table("weekly_insights") \
            .select("*") \
            .order("created_at", desc=True) \
            .limit(20) \
            .execute()
        if result.data:
            df = pd.DataFrame(result.data)
            # עצב את התצוגה
            st.markdown("### 📊 שאלות זה עתה:")
            for _, row in df.iterrows():
                unit = row['unit']
                insight_type = row['insight_type']
                question = row['ai_question']
                suggestion = row.get('ai_suggestion', '')
                # צבע לפי סוג
                color_map = {
                    'kashrut': '#f59e0b',
                    'eruv': '#ef4444',
                    'deficits': '#3b82f6',
                    'performance': '#10b981',
                    'anomalies': '#ec4899'
                }
                color = color_map.get(insight_type, '#64748b')
                st.markdown(f"""
                <div style='background: {color}15; border-left: 4px solid {color};
                            padding: 15px; border-radius: 8px; margin-bottom: 12px;'>
                    <strong style='color: {color};'>{unit} • {insight_type.upper()}</strong><br/>
                    <div style='margin-top: 8px; font-size: 15px;'>{question}</div>
                    {f'<div style="margin-top: 8px; color: #64748b; font-size: 13px;">💡 {suggestion}</div>' if suggestion else ''}
                </div>
                """, unsafe_allow_html=True)
        else:
            st.info("📭 אין שאלות עדיין. לחץ על 'עדכן שאלות עכשיו'")
    except Exception as e:
        st.warning(f"⚠️ לא ניתן לטעון: {str(e)}")




def render_ogda_summary_dashboard_v2():
    """Ogda Dashboard v2 - עם הסברים בעברית ותמיכה באוגדות ללא נתונים"""
    import datetime as _dt

    # ===== Header Premium עם הסברים בעברית =====
    st.markdown("""
    <div style='background: linear-gradient(135deg, #059669 0%, #10b981 100%);
                padding: 40px; border-radius: 16px; margin-bottom: 30px;
                box-shadow: 0 10px 30px rgba(0,0,0,0.2);'>
        <div style='display: flex; justify-content: space-between; align-items: center;'>
            <div>
                <h1 style='color: white; margin: 0; font-size: 36px;'>
                    🎯 Ogda Dashboard – Summary
                </h1>
                <p style='color: #d1fae5; font-size: 16px; font-weight: 600; margin: 5px 0 10px 0;'>
                    לוח מעקב פיקודי - תמונת מצב אוגדתית לאיתור מוקדי סיכון ניהוליים והלכתיים בחטיבות.
                </p>
                <p style='color: rgba(255,255,255,0.9); margin: 0; font-size: 14px;'>
                    Subordinate Units Status · """ + _dt.datetime.now().strftime('%d/%m/%Y %H:%M:%S') + """
                </p>
            </div>
            <div style='text-align: right;'><div style='font-size: 48px;'>🎖️</div></div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    unit = st.session_state.selected_unit
    role = st.session_state.role
    accessible_units = get_accessible_units(unit, role)

    raw_data = load_reports_cached(accessible_units)
    df = pd.DataFrame(raw_data) if raw_data else pd.DataFrame()

    subordinate_units = [u for u in accessible_units if u != unit]
    subordinate_count = len(subordinate_units)

    # ===== KPI Cards =====
    st.markdown("""
    <h3 style='margin-bottom: 0;'>📊 Key Performance Indicators</h3>
    <p style='color: #64748b; font-size: 14px; margin-top: 0; margin-bottom: 15px;'>
        מדדי ביצוע מרכזיים - מספקים אינדיקציה מיידית לבריאות האוגדה: כמה דוחות מולאו, מה היקף החוסרים, ומה רמת הכשירות הממוצעת.
    </p>
    """, unsafe_allow_html=True)

    kpi_cols = st.columns(4)
    with kpi_cols[0]:
        st.metric(label="🏛️ Subordinate Units", value=subordinate_count, delta="חטיבות מרחביות תחת הפיקוד", delta_color="off")

    with kpi_cols[1]:
        total_reports = len(df)
        avg_reports = total_reports / subordinate_count if subordinate_count > 0 else 0
        st.metric(label="📋 Total Reports", value=total_reports, delta=f"ממוצע {avg_reports:.1f} לחטיבה", delta_color="normal")

    with kpi_cols[2]:
        issues_count = 0
        if not df.empty:
            if 'e_status' in df.columns: issues_count += len(df[df['e_status'] == 'פסול'])
            if 'k_cert' in df.columns: issues_count += len(df[df['k_cert'] == 'לא'])
        st.metric(label="🔴 Open Issues", value=issues_count,
                  delta="בעיות ליבה הדורשות התערבות" if issues_count > 0 else "הכל תקין",
                  delta_color="inverse" if issues_count > 0 else "normal")

    with kpi_cols[3]:
        scores = [calculate_unit_score(df[df['unit'] == u]) for u in subordinate_units
                  if not df.empty and len(df[df['unit'] == u]) > 0]
        avg_score = sum(scores) / len(scores) if scores else 0
        score_status = "🟢 מצוין" if avg_score >= 80 else "🟡 סביר" if avg_score >= 60 else "🔴 טעון שיפור"
        st.metric(label="📊 Avg Score", value=f"{avg_score:.0f}/100", delta=score_status,
                  delta_color="normal" if avg_score >= 70 else "inverse")

    st.markdown("---")

    # ===== Performance Comparison =====
    st.markdown("""
    <h3 style='margin-bottom: 0;'>📊 Performance Comparison</h3>
    <p style='color: #64748b; font-size: 14px; margin-top: 0; margin-bottom: 15px;'>
        השוואת ביצועים בין חטיבות - מאפשר למפקד האוגדה לזהות אילו חטיבות מתפקדות כראוי ואילו דורשות שיחת חתך ומיקוד משאבים. ציון 0 פירושו שהחטיבה טרם הגישה דוחות.
    </p>
    """, unsafe_allow_html=True)

    col_chart, col_gauge = st.columns([2, 1])

    # תמיד כולל את כל החטיבות — גם ללא נתונים מציג 0
    comparison_data = []
    for u in subordinate_units:
        unit_df = df[df['unit'] == u] if not df.empty else pd.DataFrame()
        score = calculate_unit_score(unit_df) if not unit_df.empty else 0
        try:
            open_deficits = len(get_open_deficits([u]))
        except Exception:
            open_deficits = 0
        comparison_data.append({"Unit": u, "Score": score, "Reports": len(unit_df), "Deficits": open_deficits})

    with col_chart:
        if comparison_data:
            comp_df = pd.DataFrame(comparison_data).sort_values('Score', ascending=False)
            fig_comp = px.bar(comp_df, x='Unit', y='Score', color='Score',
                              color_continuous_scale=['#ef4444', '#f59e0b', '#10b981'],
                              range_color=[0, 100], hover_data=['Reports', 'Deficits'], text='Score')
            fig_comp.update_traces(textposition='outside')
            fig_comp.update_layout(height=400, font=dict(color='#1e293b', size=12),
                                   paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(240,244,248,0.5)',
                                   xaxis_tickangle=-45, showlegend=False, hovermode='x unified')
            st.plotly_chart(fig_comp, use_container_width=True)
        else:
            st.info("📊 לא נמצאו חטיבות כפופות לאוגדה זו.")

    with col_gauge:
        fig_gauge = go.Figure(data=[go.Indicator(
            mode="gauge+number", value=avg_score,
            title={'text': "ציון אוגדתי משוקלל"},
            domain={'x': [0, 1], 'y': [0, 1]},
            gauge={
                'axis': {'range': [None, 100]},
                'bar': {'color': "#059669"},
                'steps': [
                    {'range': [0, 25], 'color': "#fee2e2"},
                    {'range': [25, 50], 'color': "#fef3c7"},
                    {'range': [50, 100], 'color': "#d1fae5"}
                ],
                'threshold': {'line': {'color': "green", 'width': 4}, 'thickness': 0.75, 'value': 70}
            }
        )])
        fig_gauge.update_layout(font=dict(color='#1e293b', size=14), paper_bgcolor='rgba(0,0,0,0)', height=400)
        st.plotly_chart(fig_gauge, use_container_width=True)

    st.markdown("---")

    # ===== Detailed Metrics Grid =====
    st.markdown("""
    <h3 style='margin-bottom: 0;'>📋 Detailed Unit Metrics</h3>
    <p style='color: #64748b; font-size: 14px; margin-top: 0; margin-bottom: 15px;'>
        כרטיסיית ניתוח יחידתית - פירוט מדויק של כמות הדוחות, הציון, והחוסרים הפתוחים (Deficits) לכל חטמ"ר. חטמ"ר עם 0 דוחות מוצג בצבע אדום לתשומת לב.
    </p>
    """, unsafe_allow_html=True)

    if comparison_data:
        comp_df = pd.DataFrame(comparison_data).sort_values('Score', ascending=False)
        grid_cols = st.columns(min(3, max(1, len(comp_df))))
        for idx, (_, row) in enumerate(comp_df.iterrows()):
            score = row['Score']
            has_data = row['Reports'] > 0
            border_color = "#10b981" if score >= 80 else "#f59e0b" if score >= 60 else "#ef4444"
            badge = "✅ תקין ומבצעי" if score >= 80 else "👍 דורש השלמות" if score >= 60 else ("⚠️ טעון התערבות" if has_data else "📭 טרם הוגשו דוחות")
            deficit_color = "#10b981" if row['Deficits'] == 0 else "#ef4444"
            with grid_cols[idx % len(grid_cols)]:
                st.markdown(f"""
                <div style='background:linear-gradient(135deg,#f8fafc 0%,#e2e8f0 100%);
                    padding:20px;border-radius:12px;border-left:4px solid {border_color};
                    box-shadow:0 4px 12px rgba(0,0,0,0.1);margin-bottom:12px;'>
                    <h4 style='margin:0 0 12px 0;color:#0f172a;'>{row["Unit"]}</h4>
                    <div style='display:grid;grid-template-columns:1fr 1fr;gap:12px;'>
                        <div><div style='color:#64748b;font-size:12px;'>ציון כשירות</div>
                            <div style='font-size:24px;font-weight:bold;color:#0f172a;'>{score:.0f}</div></div>
                        <div><div style='color:#64748b;font-size:12px;'>דוחות במערכת</div>
                            <div style='font-size:24px;font-weight:bold;color:#0f172a;'>{row["Reports"]}</div></div>
                        <div><div style='color:#64748b;font-size:12px;'>חוסרים פתוחים</div>
                            <div style='font-size:24px;font-weight:bold;color:{deficit_color};'>{row["Deficits"]}</div></div>
                    </div>
                    <div style='margin-top:12px;padding-top:12px;border-top:1px solid #cbd5e1;
                        font-size:12px;color:{border_color};font-weight:bold;'>{badge}</div>
                </div>""", unsafe_allow_html=True)

    st.markdown("---")

    # ===== Compliance Heatmap =====
    st.markdown("""
    <h3 style='margin-bottom: 0;'>🌡️ Compliance Matrix</h3>
    <p style='color: #64748b; font-size: 14px; margin-top: 0; margin-bottom: 15px;'>
        מדד ציות לפקודות - מציג באחוזים את רמת העמידה של כל חטיבה בנהלי כשרות, עירוב וניקיון. אזורים אדומים מחייבים שיחת חתך.
    </p>
    """, unsafe_allow_html=True)

    heatmap_data = []
    compliance_metrics = {'k_cert': 'כשרות', 'e_status': 'עירוב', 's_clean': 'ניקיון', 's_board': 'לוח'}
    for u in subordinate_units:
        unit_df = df[df['unit'] == u] if not df.empty else pd.DataFrame()
        row_data = {"Unit": u}
        for col_name, metric_name in compliance_metrics.items():
            if not unit_df.empty and col_name in unit_df.columns and len(unit_df) > 0:
                if col_name == 'k_cert':
                    ok = len(unit_df[unit_df[col_name] == 'כן'])
                elif col_name == 'e_status':
                    ok = len(unit_df[unit_df[col_name] == 'תקין'])
                elif col_name == 's_clean':
                    ok = len(unit_df[unit_df[col_name].isin(['טוב', 'מצוין'])])
                else:
                    ok = len(unit_df[unit_df[col_name] == 'כן'])
                row_data[metric_name] = round(ok / len(unit_df) * 100, 0)
            else:
                row_data[metric_name] = 0
        heatmap_data.append(row_data)

    if heatmap_data:
        heatmap_df = pd.DataFrame(heatmap_data).set_index('Unit')
        fig_hm = px.imshow(heatmap_df, color_continuous_scale='RdYlGn', text_auto='.0f',
                           aspect='auto', color_continuous_midpoint=50, labels={'color': 'ציות %'})
        fig_hm.update_layout(height=350, font=dict(color='#1e293b'),
                              paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
        st.plotly_chart(fig_hm, use_container_width=True)

    st.markdown("---")

    # ===== Critical Issues =====
    st.markdown("""
    <h3 style='margin-bottom: 0;'>🚨 Critical Issues Summary</h3>
    <p style='color: #64748b; font-size: 14px; margin-top: 0; margin-bottom: 15px;'>
        סיכום בעיות קריטיות - מציג חטיבות עם ליקויי עירוב או כשרות שדורשים התערבות מיידית של הרב האוגדתי.
    </p>
    """, unsafe_allow_html=True)

    iss_col1, iss_col2 = st.columns(2)
    with iss_col1:
        st.markdown("#### ⚠️ יחידות עם ליקויים קריטיים")
        critical_units = []
        for u in subordinate_units:
            unit_df = df[df['unit'] == u] if not df.empty else pd.DataFrame()
            if not unit_df.empty:
                eruv_iss = len(unit_df[unit_df['e_status'] == 'פסול']) if 'e_status' in unit_df.columns else 0
                kash_iss = len(unit_df[unit_df['k_cert'] == 'לא']) if 'k_cert' in unit_df.columns else 0
                if eruv_iss > 0 or kash_iss > 0:
                    critical_units.append({"Unit": u, "Eruv": eruv_iss, "Kashrut": kash_iss, "Total": eruv_iss + kash_iss})
        if critical_units:
            for row in sorted(critical_units, key=lambda x: x['Total'], reverse=True):
                st.markdown(f"""
                <div style='background:#fee2e2;border-left:4px solid #ef4444;padding:12px;
                    border-radius:6px;margin-bottom:10px;'>
                    <strong>🔴 {row["Unit"]}</strong><br/>
                    🚧 עירוב: {row["Eruv"]} | 🍽️ כשרות: {row["Kashrut"]}
                </div>""", unsafe_allow_html=True)
        else:
            st.success("✅ אין ליקויים קריטיים מדווחים")

    with iss_col2:
        st.markdown("#### 📊 פירוט חוסרים פתוחים")
        try:
            all_open_deficits = get_open_deficits(accessible_units)
            if not all_open_deficits.empty and 'deficit_type' in all_open_deficits.columns:
                deficit_counts = all_open_deficits['deficit_type'].value_counts()
                fig_pie = px.pie(values=deficit_counts.values, names=deficit_counts.index, hole=0.4)
                fig_pie.update_layout(height=300, font=dict(color='#1e293b'), paper_bgcolor='rgba(0,0,0,0)')
                st.plotly_chart(fig_pie, use_container_width=True)
            else:
                st.success("✅ אין חוסרים פתוחים")
        except Exception:
            st.info("אין נתוני חוסרים")

    st.markdown("---")

    # ===== 30-Day Trend =====
    st.markdown("""
    <h3 style='margin-bottom: 0;'>📈 30-Day Trend</h3>
    <p style='color: #64748b; font-size: 14px; margin-top: 0; margin-bottom: 15px;'>
        מגמה של 30 יום - מראה את קצב הגשת הדוחות ואת מספר הבעיות לאורך זמן. עלייה חדה בבעיות (קו אדום) מסמנת אירוע שדורש בדיקה.
    </p>
    """, unsafe_allow_html=True)

    if not df.empty and 'date' in df.columns:
        try:
            df_copy = df.copy()
            df_copy['date'] = pd.to_datetime(df_copy['date'], errors='coerce')
            df_30 = df_copy[df_copy['date'] > (pd.Timestamp.now() - pd.Timedelta(days=30))].copy()
            if not df_30.empty:
                trend_data = []
                for tdate in pd.date_range(start=df_30['date'].min(), end=df_30['date'].max(), freq='D'):
                    day_reports = df_30[df_30['date'].dt.date == tdate.date()]
                    iss = 0
                    if 'e_status' in day_reports.columns:
                        iss += len(day_reports[day_reports['e_status'] == 'פסול'])
                    if 'k_cert' in day_reports.columns:
                        iss += len(day_reports[day_reports['k_cert'] == 'לא'])
                    trend_data.append({"date": tdate, "issues": iss, "reports": len(day_reports)})
                trend_df = pd.DataFrame(trend_data)
                fig_trend = make_subplots(specs=[[{"secondary_y": True}]])
                fig_trend.add_trace(go.Bar(x=trend_df['date'], y=trend_df['reports'],
                                           name='דוחות', marker_color='#3b82f6'), secondary_y=False)
                fig_trend.add_trace(go.Scatter(x=trend_df['date'], y=trend_df['issues'], name='בעיות',
                                               marker=dict(color='#ef4444'), mode='lines+markers'), secondary_y=True)
                fig_trend.update_layout(height=350, font=dict(color='#1e293b'),
                                        paper_bgcolor='rgba(0,0,0,0)',
                                        plot_bgcolor='rgba(240,244,248,0.5)', hovermode='x unified')
                fig_trend.update_yaxes(title_text="דוחות", secondary_y=False)
                fig_trend.update_yaxes(title_text="בעיות", secondary_y=True)
                st.plotly_chart(fig_trend, use_container_width=True)
        except Exception as e:
            st.info(f"לא ניתן לטעון טרנד: {e}")
    else:
        st.info("📈 טרם הוזנו דוחות - הגרף יוצג לאחר קבלת הדוחות הראשונים.")

    st.markdown("---")

    # 🤖 Weekly Insights Manager - Command Panel
    st.markdown("""
    <h3 style='margin-bottom: 0;'>🤖 Weekly Insights Manager</h3>
    <p style='color: #64748b; font-size: 14px; margin-top: 0; margin-bottom: 15px;'>
        מחולל התובנות השבועי - מערכת AI המייצרת עבורך באופן אוטומטי שאלות הכוונה ומוקדי שיח לשיחת החתך עם רבני החטיבות לקראת השבוע.
    </p>
    """, unsafe_allow_html=True)
    render_weekly_insights_control_panel(key_suffix="_exec")

    st.markdown("---")

    # ===== Footer Metrics =====
    footer_cols = st.columns(4)
    with footer_cols[0]:
        avg_rep = len(df) / subordinate_count if subordinate_count > 0 else 0
        st.metric("📊 דוחות ממוצע ליחידה", f"{avg_rep:.1f}")
    with footer_cols[1]:
        total_iss = 0
        if not df.empty:
            if 'e_status' in df.columns:
                total_iss += len(df[df['e_status'] == 'פסול'])
            if 'k_cert' in df.columns:
                total_iss += len(df[df['k_cert'] == 'לא'])
        st.metric("🔴 סה\"כ ליקויים", total_iss)
    with footer_cols[2]:
        compliant = sum(1 for u in subordinate_units if not df.empty and len(df[df['unit'] == u]) > 0
                        and calculate_unit_score(df[df['unit'] == u]) >= 80)
        st.metric("🟢 יחידות תקינות", compliant)
    with footer_cols[3]:
        st.metric("🔄 עדכון אחרון", _dt.datetime.now().strftime('%H:%M'))


def render_hatmar_summary_dashboard():
    """
    📊 \u05d3\u05e9\u05d1\u05d5\u05e8\u05d3 \u05e8\u05d1 \u05d7\u05d8\u05de\u05e8 \u2013 \u05e1\u05e7\u05d9\u05e8\u05d4 \u05de\u05d7\u05d8\u05de\u05e8 \u05e9\u05dc\u05d5
    """
    st.markdown("""
    <div style='background:linear-gradient(90deg,#7c3aed 0%,#a855f7 100%);
                padding:30px;border-radius:12px;margin-bottom:30px;'>
        <h1 style='color:white;margin:0;'>📊 My Unit Dashboard</h1>
        <p style='color:rgba(255,255,255,0.9);margin:8px 0 0 0;'>Comprehensive operational overview \u00b7 """ +
    datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S') + """</p>
    </div>""", unsafe_allow_html=True)

    unit = st.session_state.selected_unit
    all_data = load_reports_cached(None)
    df = pd.DataFrame(all_data) if all_data else pd.DataFrame()
    unit_df = df[df['unit'] == unit].copy() if not df.empty and 'unit' in df.columns else pd.DataFrame()

    kpi_cols = st.columns(5)
    with kpi_cols[0]:
        st.metric("📋 Total Reports", len(unit_df))
    with kpi_cols[1]:
        score = calculate_unit_score(unit_df) if not unit_df.empty else 0
        st.metric("🎯 Overall Score", f"{score:.0f}/100",
                 delta="Excellent" if score >= 80 else "Good" if score >= 60 else "Needs Work")
    with kpi_cols[2]:
        open_defs = len(get_open_deficits([unit]))
        st.metric("📋 Open Deficits", open_defs, delta_color="inverse" if open_defs > 0 else "off")
    with kpi_cols[3]:
        if not unit_df.empty and 'date' in unit_df.columns:
            last_report = pd.to_datetime(unit_df['date'], errors='coerce').max()
            days_ago = int((pd.Timestamp.now() - last_report).days) if pd.notna(last_report) else 999
            st.metric("📅 Last Report", f"{days_ago} days ago")
        else:
            st.metric("📅 Last Report", "Never")
    with kpi_cols[4]:
        unique_bases = unit_df['base'].nunique() if not unit_df.empty and 'base' in unit_df.columns else 0
        st.metric("📍 Bases Covered", unique_bases)

    st.markdown("---")

    col_comp, col_issues = st.columns([1.2, 1])

    with col_comp:
        st.markdown("### \u2705 Compliance Status")
        if not unit_df.empty:
            kashrut_ok = (
                int((unit_df['k_cert'] == '\u05db\u05df').sum()) / len(unit_df) * 100
                if 'k_cert' in unit_df.columns else 50
            )
            eruv_ok = (
                int((unit_df['e_status'] == '\u05ea\u05e7\u05d9\u05df').sum()) / len(unit_df) * 100
                if 'e_status' in unit_df.columns else 50
            )
            comp_data = pd.DataFrame([
                {"Category": "Kashrut", "Compliance": kashrut_ok},
                {"Category": "Eruv", "Compliance": eruv_ok},
            ])
            fig_comp = px.bar(comp_data, x='Category', y='Compliance',
                              color='Compliance', color_continuous_scale='RdYlGn', range_color=[0, 100])
            fig_comp.update_layout(height=350, showlegend=False,
                                   paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
            st.plotly_chart(fig_comp, use_container_width=True)

    with col_issues:
        st.markdown("### \u26a0\ufe0f Active Issues")
        issues_list = []
        if not unit_df.empty:
            if 'k_cert' in unit_df.columns and (unit_df['k_cert'] == '\u05dc\u05d0').any():
                issues_list.append(("🍽\ufe0f", "Missing Kashrut", "#f59e0b"))
            if 'e_status' in unit_df.columns and (unit_df['e_status'] == '\u05e4\u05e1\u05d5\u05dc').any():
                issues_list.append(("🚧", "Invalid Eruv", "#ef4444"))
            if 'r_mezuzot_missing' in unit_df.columns:
                total_mez = int(pd.to_numeric(unit_df['r_mezuzot_missing'], errors='coerce').fillna(0).sum())
                if total_mez > 0:
                    issues_list.append(("📜", f"{total_mez} Missing Mezuzot", "#3b82f6"))
        if issues_list:
            for icon, title, color in issues_list:
                st.markdown(f"""
                <div style='background:{color}20;border-left:4px solid {color};
                            padding:12px;border-radius:6px;margin-bottom:10px;'>
                    <strong>{icon} {title}</strong>
                </div>""", unsafe_allow_html=True)
        else:
            st.success("\u2705 No active issues")

    st.markdown("---")
    st.markdown("### 📍 Bases Overview")
    if not unit_df.empty and 'base' in unit_df.columns:
        bases_data = []
        for base in unit_df['base'].unique():
            base_reports = unit_df[unit_df['base'] == base]
            sc = calculate_unit_score(base_reports)
            bases_data.append({"Base": base, "Reports": len(base_reports), "Score": sc})
        bases_df = pd.DataFrame(bases_data).sort_values('Score', ascending=False)
        fig_bases = px.scatter(bases_df, x='Reports', y='Score', size='Score',
                               hover_name='Base', color='Score',
                               color_continuous_scale='RdYlGn', range_color=[0, 100])
        fig_bases.update_layout(height=350, showlegend=False,
                                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
        st.plotly_chart(fig_bases, use_container_width=True)
    else:
        st.info("\u05d0\u05d9\u05df \u05e0\u05ea\u05d5\u05e0\u05d9 \u05de\u05d5\u05e6\u05d1\u05d9\u05dd")


# ===== New Dashboard Pages =====

def render_risk_command_center(df: pd.DataFrame, accessible_units: list):
    """
    🎯 מרכז בקרה פיקודי — Risk Index מלא
    מציג: מדדים, התראות אדום, גרף Risk לפי יחידה, גרף מגמות
    """
    st.markdown("""
    <div style='background: linear-gradient(135deg, #0f172a 0%, #1e40af 100%);
                padding: 28px; border-radius: 12px; margin-bottom: 24px;'>
        <h1 style='color: white; margin: 0;'>🎖️ מרכז בקרה פיקודי — Risk Center</h1>
        <p style='color: rgba(255,255,255,0.8); margin: 8px 0 0 0;'>תמונת מצב סיכונים בזמן אמת</p>
    </div>
    """, unsafe_allow_html=True)

    # ── שורה 1: מדדים ──
    m1, m2, m3, m4, m5 = st.columns(5)
    units_list = [u for u in df['unit'].unique()] if not df.empty else []

    critical_bases = len(df[df['e_status'] == 'פסול']) if ('e_status' in df.columns and not df.empty) else 0
    open_defs = get_deficit_statistics(accessible_units)['total_open']
    overdue = count_overdue_deficits(accessible_units)
    silent = count_silent_units(df)
    avg_risk = 0
    if units_list:
        scores = [calculate_operational_risk_index(u, df)['risk_score'] for u in units_list]
        avg_risk = round(sum(scores) / len(scores), 1) if scores else 0

    with m1:
        st.metric("🔴 בסיסים קריטיים", critical_bases,
                  delta="עירוב פסול" if critical_bases > 0 else "תקין",
                  delta_color="inverse" if critical_bases > 0 else "off")
    with m2:
        st.metric("📋 חוסרים פתוחים", open_defs)
    with m3:
        st.metric("⏰ עברו SLA (7 ימים)", overdue,
                  delta="דחוף!" if overdue > 0 else "הכל בזמן",
                  delta_color="inverse" if overdue > 0 else "off")
    with m4:
        st.metric("📡 יחידות שקטות", silent,
                  delta=f"{silent} ≥7 ימים" if silent > 0 else "בקשר",
                  delta_color="inverse" if silent > 0 else "off")
    with m5:
        st.metric("📊 Risk ממוצע", f"{avg_risk}/100")

    st.markdown("---")

    # ── שורה 2: Red Alerts ──
    st.markdown("## 🚨 Red Alert – טיפול מיידי")
    alerts = []

    if not df.empty and 'e_status' in df.columns and 'base' in df.columns:
        for base in df[df['e_status'] == 'פסול']['base'].unique():
            alerts.append({"icon": "🚧", "color": "#dc2626",
                           "title": f"עירוב פסול – {base}",
                           "desc": "דורש טיפול מיידי"})

    if not df.empty and 'unit' in df.columns and 'date' in df.columns:
        dates = pd.to_datetime(df['date'], errors='coerce')
        for u in df['unit'].unique():
            last = dates[df['unit'] == u].max()
            if pd.notna(last):
                days = (pd.Timestamp.now() - last).days
                if days > 7:
                    alerts.append({"icon": "⏰", "color": "#f97316",
                                   "title": f"{u} לא דיווחה",
                                   "desc": f"שקטה {days} ימים"})

    if alerts:
        for alert in alerts:
            st.markdown(f"""
            <div style='background:{alert["color"]}18; border-left:4px solid {alert["color"]};
                        padding:14px; border-radius:8px; margin-bottom:10px;'>
                <strong style='color:{alert["color"]}'>{alert["icon"]} {alert["title"]}</strong>
                <p style='margin:6px 0 0 0; color:#64748b'>{alert["desc"]}</p>
            </div>
            """, unsafe_allow_html=True)
    else:
        st.success("✅ אין אזהרות קריטיות כרגע")

    st.markdown("---")

    # ── שורה 3: Risk Bar Chart ──
    st.markdown("## 📊 Risk Index לפי יחידה")
    if units_list:
        risk_rows = []
        for u in sorted(units_list):
            rd = calculate_operational_risk_index(u, df)
            risk_rows.append({"יחידה": u, "Risk": rd['risk_score'], "רמה": rd['level']})
        risk_df = pd.DataFrame(risk_rows).sort_values("Risk", ascending=False)
        fig = px.bar(
            risk_df, x="יחידה", y="Risk",
            color="Risk",
            color_continuous_scale=[[0, "#10b981"], [0.25, "#f59e0b"], [0.5, "#ef4444"], [1, "#dc2626"]],
            range_color=[0, 100],
            text="Risk",
            height=400
        )
        fig.update_traces(texttemplate='%{text:.0f}', textposition='outside')
        fig.update_layout(coloraxis_showscale=False, yaxis_range=[0, 110])
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("אין נתונים להצגה")

    st.markdown("---")

    # ── שורה 4: מגמות חוסרים 30 יום ──
    if not df.empty and 'date' in df.columns:
        st.markdown("## 📈 מגמות חוסרים – 30 ימים אחרונים")
        try:
            df_copy = df.copy()
            df_copy['date'] = pd.to_datetime(df_copy['date'], errors='coerce')
            cutoff = pd.Timestamp.now() - pd.Timedelta(days=30)
            df_month = df_copy[df_copy['date'] > cutoff]

            if not df_month.empty:
                rows = []
                for d in pd.date_range(df_month['date'].min(), df_month['date'].max(), freq='D'):
                    day_df = df_month[df_month['date'].dt.date == d.date()]
                    rows.append({
                        "תאריך": d,
                        "מזוזות": int(pd.to_numeric(day_df.get('r_mezuzot_missing', pd.Series()), errors='coerce').fillna(0).sum()),
                        "כשרות": int((day_df['k_cert'] == 'לא').sum()) if 'k_cert' in day_df.columns else 0,
                        "עירוב": int((day_df['e_status'] == 'פסול').sum()) if 'e_status' in day_df.columns else 0,
                    })
                time_df = pd.DataFrame(rows)
                fig2 = px.area(
                    time_df, x='תאריך', y=['מזוזות', 'כשרות', 'עירוב'],
                    labels={'value': 'מספר חוסרים'},
                    color_discrete_map={'מזוזות': '#3b82f6', 'כשרות': '#f59e0b', 'עירוב': '#ef4444'},
                    height=350
                )
                fig2.update_layout(hovermode='x unified')
                st.plotly_chart(fig2, use_container_width=True)
        except Exception as e:
            st.warning(f"לא ניתן להציג מגמות: {e}")


def render_deficit_tracker_pro(unit: str, accessible_units: list):
    """
    📊 תצוגת חוסרים מקצועית לחטמ"ר — טאבים, מדדים, progress bars
    """
    st.markdown("""
    <div style='background: linear-gradient(135deg, #10b981 0%, #059669 100%);
                padding: 28px; border-radius: 12px; margin-bottom: 24px;'>
        <h1 style='color: white; margin: 0;'>📋 מעקב חוסרים – תצוגה מקצועית</h1>
        <p style='color: rgba(255,255,255,0.9); margin: 8px 0 0 0;'>כל החוסרים שלך, מיוניים וברורים</p>
    </div>
    """, unsafe_allow_html=True)

    unit_deficits = get_open_deficits(accessible_units)

    # ── מדדי ראש ──
    mc1, mc2, mc3, mc4 = st.columns(4)
    total_open = len(unit_deficits)
    overdue_count = 0
    avg_age = 0
    if not unit_deficits.empty and 'detected_date' in unit_deficits.columns:
        ages = (pd.Timestamp.now() - pd.to_datetime(unit_deficits['detected_date'], errors='coerce').dt.tz_localize(None)).dt.days
        avg_age = int(ages.mean()) if not ages.isna().all() else 0
        overdue_count = int((ages > 7).sum())

    with mc1:
        st.metric("🔴 חוסרים פתוחים", total_open)
    with mc2:
        st.metric("⏰ עברו SLA", overdue_count,
                  delta="דחוף!" if overdue_count > 0 else "הכל בזמן",
                  delta_color="inverse" if overdue_count > 0 else "off")
    with mc3:
        st.metric("📅 ממוצע גיל חוסר", f"{avg_age} ימים")
    with mc4:
        try:
            closed = supabase.table("deficit_tracking").select("id").in_("unit", accessible_units).eq("status", "closed").execute()
            closed_count = len(closed.data) if closed.data else 0
        except Exception:
            closed_count = 0
        total_ever = closed_count + total_open
        closure_rate = int(closed_count / total_ever * 100) if total_ever > 0 else 0
        st.metric("🎖️ שיעור סגירה", f"{closure_rate}%")

    st.markdown("---")

    if unit_deficits.empty:
        st.success("✅ אין חוסרים פתוחים!")
        return

    # ── טאבים לפי סוג חוסר ──
    deficit_types = {
        'mezuzot': ("📜 מזוזות", "#3b82f6"),
        'eruv_status': ("🚧 עירוב פסול", "#ef4444"),
        'kashrut_cert': ("🍽️ כשרות", "#f59e0b"),
        'eruv_kelim': ("🔴 ערבוב כלים", "#dc2626"),
        'shabbat_supervisor': ("👤 נאמן שבת", "#8b5cf6"),
    }
    tab_labels = [v[0] for v in deficit_types.values()]
    dtabs = st.tabs(tab_labels)

    for tab_i, (def_type, (label, color)) in enumerate(deficit_types.items()):
        with dtabs[tab_i]:
            if 'deficit_type' in unit_deficits.columns:
                type_defs = unit_deficits[unit_deficits['deficit_type'] == def_type].copy()
            else:
                type_defs = pd.DataFrame()

            if not type_defs.empty:
                type_defs['days_open'] = (
                    pd.Timestamp.now() - pd.to_datetime(type_defs['detected_date'], errors='coerce').dt.tz_localize(None)
                ).dt.days if 'detected_date' in type_defs.columns else 0

                for _, deficit in type_defs.iterrows():
                    days = int(deficit.get('days_open', 0))
                    base_name = deficit.get('base', 'לא ידוע')
                    deficit_count = int(deficit.get('deficit_count', 1))
                    overdue_flag = days > 7

                    col_info, col_action = st.columns([4, 1])
                    with col_info:
                        sla_badge = "⚠️ עבר SLA!" if overdue_flag else ""
                        st.markdown(f"""
                        <div style='background:{color}15; border-left:4px solid {color};
                                    padding:12px; border-radius:6px; margin-bottom:8px;'>
                            <strong>📍 {base_name}</strong> — {deficit_count} יחידות<br/>
                            <small>פתוח {days} ימים {'<span style="color:red">' + sla_badge + '</span>' if overdue_flag else ''}</small>
                        </div>
                        """, unsafe_allow_html=True)
                    with col_action:
                        deficit_id = str(deficit.get('id', ''))
                        if deficit_id and st.button("✅ סגור", key=f"close_{deficit_id}", use_container_width=True):
                            if update_deficit_status(deficit_id, 'closed'):
                                log_audit_event("CLOSE_DEFICIT", deficit_id,
                                                details={"type": def_type, "base": base_name},
                                                severity="info")
                                st.success("✅ נסגר!")
                                st.rerun()
            else:
                st.success(f"✅ אין {label} פתוחים")

    st.markdown("---")

    # ── Progress bars לפי מוצב ──
    st.markdown("### 🎯 חוסרים לפי מוצב")
    if 'base' in unit_deficits.columns:
        for base in sorted(unit_deficits['base'].unique()):
            cnt = len(unit_deficits[unit_deficits['base'] == base])
            pct = min(100, cnt * 20)
            st.markdown(f"""
            <div style='margin-bottom:14px;'>
                <div style='display:flex; justify-content:space-between; margin-bottom:4px;'>
                    <strong>📍 {base}</strong><span>{cnt} חוסרים</span></div>
                <div style='width:100%; height:8px; background:#e2e8f0; border-radius:4px;'>
                    <div style='width:{pct}%; height:100%; background:linear-gradient(90deg,#ef4444,#f59e0b); border-radius:4px;'></div>
                </div>
            </div>
            """, unsafe_allow_html=True)


def render_deficit_heat_map(df: pd.DataFrame, accessible_units: list):
    """
    🌡️ מפת חוסרים צבעונית לפי יחידה ← גרף Heatmap + סטטיסטיקות
    """
    st.markdown("""
    <div style='background: linear-gradient(135deg, #ec4899 0%, #f43f5e 100%);
                padding: 28px; border-radius: 12px; margin-bottom: 24px;'>
        <h1 style='color: white; margin: 0;'>🌡️ Deficit Heat Map – כל הארץ</h1>
        <p style='color: rgba(255,255,255,0.9); margin: 8px 0 0 0;'>חוסרים לפי יחידה וסוג בעיה</p>
    </div>
    """, unsafe_allow_html=True)

    # ── Heatmap מטריצה: יחידה × סוג חוסר ──
    st.markdown("## 🔥 Heatmap – חוסרים לפי יחידה")
    open_defs = get_open_deficits(accessible_units)

    if not open_defs.empty and 'unit' in open_defs.columns and 'deficit_type' in open_defs.columns:
        hm_rows = []
        def_type_labels = {
            'mezuzot': '📜 מזוזות',
            'eruv_status': '🚧 עירוב',
            'kashrut_cert': '🍽️ כשרות',
            'eruv_kelim': '🔴 ערבוב',
            'shabbat_supervisor': '👤 נאמן'
        }
        for unit in sorted(open_defs['unit'].unique()):
            row = {'יחידה': unit}
            for dt, lbl in def_type_labels.items():
                row[lbl] = int((open_defs[(open_defs['unit'] == unit) & (open_defs['deficit_type'] == dt)]['deficit_count'].sum()
                               if 'deficit_count' in open_defs.columns else
                               len(open_defs[(open_defs['unit'] == unit) & (open_defs['deficit_type'] == dt)])))
            hm_rows.append(row)

        hm_df = pd.DataFrame(hm_rows).set_index('יחידה')
        fig = px.imshow(
            hm_df, color_continuous_scale='RdYlGn_r',
            text_auto=True, aspect='auto',
        )
        fig.update_layout(height=max(300, len(hm_rows) * 50))
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.success("✅ אין חוסרים פתוחים")

    st.markdown("---")

    # ── מקרא ──
    st.markdown("## 🎨 מקרא")
    l1, l2, l3, l4 = st.columns(4)
    legend_items = [
        (l1, "#dc2626", "🔴 קריטי", "עירוב פסול"),
        (l2, "#f59e0b", "🟠 חמור", "כשרות חסרה"),
        (l3, "#3b82f6", "🔵 בינוני", "מזוזות חסרות"),
        (l4, "#10b981", "🟢 תקין", "כל הנתונים טובים"),
    ]
    for col, color, title, sub in legend_items:
        with col:
            st.markdown(f"""
            <div style='background:{color}; padding:12px; border-radius:6px; color:white; text-align:center;'>
                <strong>{title}</strong><br/>{sub}
            </div>
            """, unsafe_allow_html=True)

    st.markdown("---")

    # ── סטטיסטיקות ──
    st.markdown("## 📊 סטטיסטיקות מרכזיות")
    s1, s2, s3, s4 = st.columns(4)

    eruv_inv = len(df[df['e_status'] == 'פסול']) if ('e_status' in df.columns and not df.empty) else 0
    no_kash = len(df[df['k_cert'] == 'לא']) if ('k_cert' in df.columns and not df.empty) else 0
    total_mez = int(pd.to_numeric(df.get('r_mezuzot_missing', pd.Series()), errors='coerce').fillna(0).sum()) if not df.empty else 0
    total_rep = len(df)

    with s1: st.metric("🚧 עירוב פסול", eruv_inv)
    with s2: st.metric("🍽️ כשרות חסרה", no_kash)
    with s3: st.metric("📜 מזוזות חסרות", total_mez)
    with s4: st.metric("📋 סה\"כ דוחות", total_rep)


def make_pwa_app():
    """
    הופך את האתר לאפליקציית רשת (PWA).
    מעלים את שורת הכתובת, מגדיר שם לאייקון וצבע לשורת הסטטוס.
    """
    import streamlit.components.v1 as components
    components.html("""
    <script>
        // גישה ל-head של הדף הראשי
        const head = window.parent.document.querySelector('head');
        
        // 1. הגדרת מסך מלא (ללא שורת כתובת של דפדפן) לאייפון ואנדרואיד
        if (!window.parent.document.querySelector('meta[name="apple-mobile-web-app-capable"]')) {
            const metaApple = window.parent.document.createElement('meta');
            metaApple.name = "apple-mobile-web-app-capable";
            metaApple.content = "yes";
            head.appendChild(metaApple);
            
            const metaAndroid = window.parent.document.createElement('meta');
            metaAndroid.name = "mobile-web-app-capable";
            metaAndroid.content = "yes";
            head.appendChild(metaAndroid);
        }
        
        // 2. הגדרת השם שיופיע מתחת לאייקון במסך הבית
        if (!window.parent.document.querySelector('meta[name="apple-mobile-web-app-title"]')) {
            const metaTitle = window.parent.document.createElement('meta');
            metaTitle.name = "apple-mobile-web-app-title";
            metaTitle.content = "רבנות פקמ״ז";
            head.appendChild(metaTitle);
        }
        
        // 3. הגדרת צבע שורת הסטטוס (השורה של הסוללה/שעון) לצבע של המערכת
        if (!window.parent.document.querySelector('meta[name="theme-color"]')) {
            const metaTheme = window.parent.document.createElement('meta');
            metaTheme.name = "theme-color";
            metaTheme.content = "#1e3a8a"; // הכחול המבצעי של המערכת
            head.appendChild(metaTheme);
        }
    </script>
    """, height=0)

# --- 10. Main ---
def main():
    # הגנה על תקציב API - מקסימום 50 קריאות ל-Session
    if "api_call_count" not in st.session_state:
        st.session_state.api_call_count = 0

    if st.session_state.api_call_count > 50:
        st.error("🚨 חרגת ממכסת השימוש המותרת לחיבור זה. המערכת ננעלה למניעת חיובי יתר.")
        st.stop()

    # הפעלת הגדרות האפליקציה למובייל
    make_pwa_app()

    # החלת עיצוב CSS גלובלי
    apply_custom_css()
    
    if not st.session_state.logged_in:
        if st.session_state.login_stage == "gallery": render_login_gallery()
        else: render_login_password()
    else:
        # 🆕 Header עליון במקום סיידבר (Wave 2.6)
        header_col1, header_col2, header_col3 = st.columns([1, 2, 1])
        
        with header_col1:
            st.image(get_logo_url(st.session_state.selected_unit), width=60)
            
        with header_col2:
            st.markdown(f"**{st.session_state.selected_unit}** | {st.session_state.role}")
            
        with header_col3:
            # 🌙 Dark Mode Toggle
            if 'dark_mode' not in st.session_state:
                st.session_state.dark_mode = False
            
            dark_mode = st.toggle("🌙", value=st.session_state.dark_mode, help="מצב כהה")
            if dark_mode != st.session_state.dark_mode:
                st.session_state.dark_mode = dark_mode
                st.rerun()
                
            if st.button("🚪 יציאה", use_container_width=True):
                st.session_state.logged_in = False
                st.session_state.login_stage = "gallery"
                st.rerun()

        if st.session_state.dark_mode:
            apply_dark_theme_css()

        st.markdown("---")
        
        if st.session_state.role in ['pikud', 'ugda']: render_command_dashboard()
        else: render_unit_report()

if __name__ == "__main__":
    main()
    
